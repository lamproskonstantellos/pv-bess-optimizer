"""House-style assertions for every workbook the tool writes.

The input polisher and ``pvbess_opt.io`` share one styler
(:mod:`pvbess_opt.io_style`), so input and output workbooks must look
identical: navy ``1F3864`` frozen header, white bold font, and AutoFit
widths within the clamp.  These tests open freshly-written workbooks and
assert that contract, iterating every produced sheet so new sheets are
covered automatically.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pvbess_opt.theme import (
    COL_WIDTH_MAX,
    COL_WIDTH_MIN,
    HEADER_FILL_HEX,
    HEADER_FONT_HEX,
)

ROOT = Path(__file__).resolve().parent.parent


def _highs_available() -> bool:
    try:
        import highspy
    except ImportError:
        return False
    return bool(highspy)


def _norm(rgb: object) -> str:
    if not isinstance(rgb, str):
        return ""
    return rgb.upper().lstrip("0").rjust(6, "0")[-6:]


def _assert_sheet_styled(ws) -> None:
    assert ws.freeze_panes == "A2", (
        f"{ws.title}: freeze_panes={ws.freeze_panes!r}, expected 'A2'"
    )
    for cell in ws[1]:
        if cell.value is None:
            continue
        assert cell.font.bold is True, f"{ws.title}!{cell.coordinate}: not bold"
        fill = _norm(
            getattr(cell.fill.fgColor, "rgb", None)
            or getattr(cell.fill.start_color, "rgb", None)
        )
        assert fill == HEADER_FILL_HEX, (
            f"{ws.title}!{cell.coordinate}: fill={fill!r}, "
            f"expected {HEADER_FILL_HEX!r}"
        )
        font = _norm(getattr(cell.font.color, "rgb", None))
        assert font == HEADER_FONT_HEX, (
            f"{ws.title}!{cell.coordinate}: font={font!r}, "
            f"expected {HEADER_FONT_HEX!r}"
        )
    for c in range(1, ws.max_column + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        assert dim is not None and dim.width is not None, (
            f"{ws.title}: column {get_column_letter(c)} has no explicit width"
        )
        assert COL_WIDTH_MIN <= float(dim.width) <= COL_WIDTH_MAX, (
            f"{ws.title}: column {get_column_letter(c)} width "
            f"{dim.width} outside [{COL_WIDTH_MIN}, {COL_WIDTH_MAX}]"
        )


def _minimal_typed() -> dict[str, object]:
    from pvbess_opt.io import (
        BALANCING_SHEET_DEFAULTS,
        BESS_SHEET_DEFAULTS,
        ECONOMICS_SHEET_DEFAULTS,
        PROJECT_SHEET_DEFAULTS,
        PV_SHEET_DEFAULTS,
        SIMULATION_SHEET_DEFAULTS,
    )

    ts = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=4, freq="15min"),
        "load_kwh": [1.0, 1.0, 1.0, 1.0],
        "pv_kwh": [0.0, 0.5, 0.5, 0.0],
        "dam_price_eur_per_mwh": [50.0, 60.0, 70.0, 80.0],
    })
    return {
        "ts": ts,
        "project": dict(PROJECT_SHEET_DEFAULTS),
        "pv": dict(PV_SHEET_DEFAULTS),
        "bess": dict(BESS_SHEET_DEFAULTS),
        "economics": dict(ECONOMICS_SHEET_DEFAULTS),
        "simulation": dict(SIMULATION_SHEET_DEFAULTS),
        "balancing": dict(BALANCING_SHEET_DEFAULTS),
    }


def test_write_workbook_output_is_styled(tmp_path):
    """write_workbook routes every sheet through the shared styler."""
    from pvbess_opt.io import write_workbook

    out = tmp_path / "wb.xlsx"
    write_workbook(_minimal_typed(), out)
    wb = load_workbook(out)
    assert wb.sheetnames
    for sn in wb.sheetnames:
        _assert_sheet_styled(wb[sn])


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_results_and_dispatch_workbooks_are_styled(tmp_path):
    """The results + dispatch workbooks from a real run are all styled."""
    from pvbess_opt import RunConfig, run
    from pvbess_opt.io import read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    short = tmp_path / "short.xlsx"
    write_workbook(typed, short)

    result = run(RunConfig(
        excel=short, solver="highs", outdir=tmp_path / "results",
        mip_gap=0.05, time_limit=180,
    ))

    results_xlsx = result.out_dir / "03_results.xlsx"
    assert results_xlsx.exists()
    rwb = load_workbook(results_xlsx)
    assert len(rwb.sheetnames) >= 3
    for sn in rwb.sheetnames:
        _assert_sheet_styled(rwb[sn])

    dispatch_xlsx = result.out_dir / "02_dispatch" / "dispatch_hourly.xlsx"
    assert dispatch_xlsx.exists()
    dwb = load_workbook(dispatch_xlsx)
    assert dwb.sheetnames
    for sn in dwb.sheetnames:
        _assert_sheet_styled(dwb[sn])
