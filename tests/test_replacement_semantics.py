"""Three-way BESS replacement semantics (v1.0.0).

``bess_replacement_year`` resolves one way across every layer:

* N (positive integer): scheduled replacement in year N;
  ``bess_eol_soh_pct`` is ignored completely.
* blank cell or ``auto``: automatic replacement in the first project
  year SOH falls to ``bess_eol_soh_pct`` — a REAL replacement (CAPEX
  charged, fade reset, lifetime projection reset, report swap), or no
  replacement if the threshold is never reached.
* 0: never replace, and the report shows SOH fading below the
  threshold without a swap.

The AUTO sentinel is resolved exactly once
(:func:`pvbess_opt.lifetime.resolve_bess_replacement_year`); consumers
read the stored effective year.
"""

from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from pvbess_opt.degradation import build_degradation_report
from pvbess_opt.economics import build_yearly_cashflow
from pvbess_opt.io import BESS_REPLACEMENT_AUTO, _parse_bess_replacement_year
from pvbess_opt.lifetime import (
    bess_capacity_factors,
    effective_bess_replacement_year,
    resolve_bess_replacement_year,
)

# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("raw", [None, float("nan"), "", "  ", "auto", "AUTO", " Auto "])
def test_parser_blank_and_auto_resolve_to_sentinel(raw):
    assert _parse_bess_replacement_year(raw, 0) == BESS_REPLACEMENT_AUTO


@pytest.mark.parametrize(
    ("raw", "expected"), [(0, 0), (10, 10), ("10", 10), (15.0, 15), ("0", 0)],
)
def test_parser_integers_pass_through(raw, expected):
    assert _parse_bess_replacement_year(raw, 0) == expected


@pytest.mark.parametrize("raw", [-1, "-3", 10.5, "later", "yes", object()])
def test_parser_rejects_negative_and_garbage(raw):
    with pytest.raises(ValueError, match="bess_replacement_year"):
        _parse_bess_replacement_year(raw, 0)


# ---------------------------------------------------------------------------
# Resolver
# ---------------------------------------------------------------------------


def _econ(replacement, **overrides):
    econ = {
        "project_lifecycle_years": 20,
        "project_start_year": 2026,
        "discount_rate_pct": 7.0,
        "opex_inflation_pct": 0.0,
        "retail_inflation_pct": 0.0,
        "dam_inflation_pct": 0.0,
        "aggregator_fee_pct_revenue": 0.0,
        "capex_pv_eur_per_kw": 0.0,
        "capex_bess_eur_per_kwh": 250.0,
        "devex_pv_eur_per_kw": 0.0,
        "devex_bess_eur_per_kw": 0.0,
        "opex_pv_eur_per_kwp": 0.0,
        "opex_bess_eur_per_kw": 0.0,
        "pv_degradation_year1_pct": 0.0,
        "pv_degradation_annual_pct": 0.0,
        "bess_degradation_annual_pct": 3.0,
        "bess_degradation_pct_per_cycle": 0.0,
        "bess_replacement_year": replacement,
        "bess_replacement_cost_pct": 50.0,
        "bess_eol_soh_pct": 80.0,
    }
    econ.update(overrides)
    return econ


def test_resolver_picks_first_crossing_of_hand_computed_curve():
    """3 %/yr calendar fade: SOH(y) = 0.97^(y-1); first <= 0.80 at y = 9
    (0.97^8 = 0.7837).  Hand-check: 0.97^7 = 0.8080 > 0.80."""
    econ = _econ("auto")
    effective, source, second = resolve_bess_replacement_year(
        econ, year1_discharge_mwh=0.0, capacity_mwh=60.0,
    )
    assert (effective, source) == (9, "soh_threshold")
    # Fresh pack from year 9: 0.97^(20-9) = 0.97^11 = 0.7153 <= 0.80 at
    # year 9 + 8 = 17 — the second crossing the model warns about.
    assert second == 17
    factors = bess_capacity_factors(
        20, d_bess_annual=0.03, replacement_year=0,
    )
    assert factors[8] <= 0.80 < factors[7]


def test_resolver_scheduled_ignores_threshold():
    econ = _econ(4)  # SOH crosses 80 % at year 9, schedule says 4
    effective, source, second = resolve_bess_replacement_year(
        econ, year1_discharge_mwh=0.0, capacity_mwh=60.0,
    )
    assert (effective, source, second) == (4, "scheduled", 0)


def test_resolver_never_mode():
    econ = _econ(0)
    effective, source, second = resolve_bess_replacement_year(
        econ, year1_discharge_mwh=0.0, capacity_mwh=60.0,
    )
    assert (effective, source, second) == (0, "never", 0)


def test_resolver_threshold_not_reached():
    econ = _econ("auto", bess_degradation_annual_pct=0.5)
    effective, source, second = resolve_bess_replacement_year(
        econ, year1_discharge_mwh=0.0, capacity_mwh=60.0,
    )
    assert (effective, source, second) == (0, "soh_threshold_not_reached", 0)


def test_accessor_raises_on_unresolved_auto_and_prefers_effective():
    econ = _econ("auto")
    with pytest.raises(ValueError, match="auto"):
        effective_bess_replacement_year(econ)
    econ["bess_replacement_year_effective"] = 9
    assert effective_bess_replacement_year(econ) == 9
    assert effective_bess_replacement_year(_econ(7)) == 7
    assert effective_bess_replacement_year(_econ(0)) == 0


# ---------------------------------------------------------------------------
# Cross-layer consistency (cashflow / lifetime factors / degradation report)
# ---------------------------------------------------------------------------


def _kpis():
    return {
        "profit_total_eur": 1_000_000.0,
        "bess_total_discharge_mwh": 0.0,
    }


def _caps():
    return {"pv_kwp": 0.0, "bess_kw": 15_000.0, "bess_kwh": 60_000.0}


def test_auto_mode_charges_capex_and_resets_fade_consistently():
    """Auto: the resolved year carries the replacement CAPEX in the
    cashflow, resets the capacity factor, and shows the swap in the
    degradation report — all in the same year."""
    econ = _econ("auto")
    effective, source, _second = resolve_bess_replacement_year(
        econ, year1_discharge_mwh=0.0, capacity_mwh=60.0,
    )
    assert source == "soh_threshold"
    econ["bess_replacement_year_effective"] = effective

    cf = build_yearly_cashflow(_kpis(), econ, _caps())
    capex_bess_y0 = -econ["capex_bess_eur_per_kwh"] * _caps()["bess_kwh"]
    charged = cf.loc[
        (cf["project_year"] >= 1) & (cf["capex_eur"] != 0.0), "project_year",
    ].tolist()
    assert charged == [effective]
    repl_row = float(
        cf.loc[cf["project_year"] == effective, "capex_eur"].iloc[0]
    )
    assert repl_row == pytest.approx(capex_bess_y0 * 0.5)
    # The capacity factor resets to 1.0 in the replacement year.
    factor = cf.set_index("project_year")["bess_capacity_factor"]
    assert float(factor.loc[effective]) == pytest.approx(1.0)
    assert float(factor.loc[effective - 1]) < 0.81

    soc = np.append(np.tile([0.0, 60_000.0], 20), 0.0)
    report = build_degradation_report(
        soc,
        capacity_kwh=60_000.0, soc_min_frac=0.2, soc_max_frac=0.95,
        degradation_pct_per_cycle=0.0, degradation_annual_pct=3.0,
        year1_discharge_mwh=0.0, project_years=20, start_year=2026,
        replacement_year=effective,
    )
    repl_years = report.loc[report["replacement"], "project_year"].tolist()
    assert repl_years == [effective]
    soh = report.set_index("project_year")["soh_pct"]
    assert float(soh.loc[effective]) == pytest.approx(100.0)
    # The cashflow factor and the report SOH agree year by year
    # (the report rounds SOH to 4 decimals).
    for y in range(1, 21):
        assert float(soh.loc[y]) == pytest.approx(
            100.0 * float(factor.loc[y]), abs=1e-4,
        )


def test_scheduled_mode_ignores_earlier_threshold_crossing():
    """Scheduled year 15 with a curve that crosses 80 % at year 9: no
    replacement before 15, CAPEX charged at 15 only, SOH dips below the
    threshold in between without a swap."""
    econ = _econ(15)
    econ["bess_replacement_year_effective"] = 15  # what the pipeline stores
    cf = build_yearly_cashflow(_kpis(), econ, _caps())
    charged = cf.loc[
        (cf["project_year"] >= 1) & (cf["capex_eur"] != 0.0), "project_year",
    ].tolist()
    assert charged == [15]
    factor = cf.set_index("project_year")["bess_capacity_factor"]
    assert float(factor.loc[9]) < 0.80  # below threshold, NOT replaced
    assert float(factor.loc[15]) == pytest.approx(1.0)


def test_never_mode_charges_nothing_and_soh_fades_through_threshold():
    econ = _econ(0)
    cf = build_yearly_cashflow(_kpis(), econ, _caps())
    assert (cf.loc[cf["project_year"] >= 1, "capex_eur"] == 0.0).all()

    soc = np.append(np.tile([0.0, 60_000.0], 20), 0.0)
    report = build_degradation_report(
        soc,
        capacity_kwh=60_000.0, soc_min_frac=0.2, soc_max_frac=0.95,
        degradation_pct_per_cycle=0.0, degradation_annual_pct=3.0,
        year1_discharge_mwh=0.0, project_years=20, start_year=2026,
        replacement_year=0,
    )
    assert not report["replacement"].any()
    # SOH keeps fading below the 80 % threshold with no swap.
    assert float(report["soh_pct"].iloc[-1]) < 80.0
    assert report["soh_pct"].is_monotonic_decreasing


def test_second_crossing_is_reported_but_not_charged():
    """A fast-fading pack crosses the threshold again after the auto
    replacement; the resolver reports it, and the cashflow still
    charges exactly one replacement."""
    econ = _econ("auto", bess_degradation_annual_pct=6.0)
    effective, source, second = resolve_bess_replacement_year(
        econ, year1_discharge_mwh=0.0, capacity_mwh=60.0,
    )
    assert source == "soh_threshold"
    assert second > effective
    econ["bess_replacement_year_effective"] = effective
    cf = build_yearly_cashflow(_kpis(), econ, _caps())
    charged = cf.loc[
        (cf["project_year"] >= 1) & (cf["capex_eur"] != 0.0), "project_year",
    ].tolist()
    assert charged == [effective]


# ---------------------------------------------------------------------------
# Workbook surface: blank cell -> auto, 'auto' token, and validation
# ---------------------------------------------------------------------------


def test_workbook_blank_cell_parses_to_auto(tmp_path):
    from openpyxl import load_workbook

    from pvbess_opt.io import read_workbook, write_workbook
    from tests.test_bess_capex_basis import _minimal_typed

    dst = tmp_path / "blank_repl.xlsx"
    typed = _minimal_typed()
    typed["bess"]["bess_replacement_year"] = 10
    write_workbook(typed, dst)
    wb = load_workbook(dst)
    ws = wb["bess"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "bess_replacement_year":
            row[1].value = None  # blank the cell
    wb.save(dst)
    loaded = read_workbook(dst)
    assert loaded["bess"]["bess_replacement_year"] == BESS_REPLACEMENT_AUTO


def test_workbook_garbage_replacement_year_raises(tmp_path):
    from openpyxl import load_workbook

    from pvbess_opt.io import read_workbook, write_workbook
    from tests.test_bess_capex_basis import _minimal_typed

    dst = tmp_path / "bad_repl.xlsx"
    write_workbook(_minimal_typed(), dst)
    wb = load_workbook(dst)
    ws = wb["bess"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "bess_replacement_year":
            row[1].value = "sometime"
    wb.save(dst)
    with pytest.raises(ValueError, match="bess_replacement_year"):
        read_workbook(dst)


def test_config_accepts_auto_token(tmp_path):
    from pvbess_opt.io_read import load_structured_config

    ts = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=24, freq="h"),
        "pv_kwh": [100.0] * 24,
        "load_kwh": [50.0] * 24,
        "dam_price_eur_per_mwh": [80.0] * 24,
    })
    ts.to_csv(tmp_path / "ts.csv", index=False)
    cfg = tmp_path / "run.yaml"
    cfg.write_text(
        "timeseries_path: ts.csv\n"
        "pv:\n"
        "  pv_nameplate_kwp: 1000\n"
        "bess:\n"
        "  bess_power_kw: 500\n"
        "  bess_capacity_kwh: 2000\n"
        "  bess_replacement_year: auto\n",
        encoding="utf-8",
    )
    typed = load_structured_config(cfg)
    assert typed["bess"]["bess_replacement_year"] == BESS_REPLACEMENT_AUTO
