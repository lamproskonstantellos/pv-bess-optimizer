"""Input surface of the per-year stream trajectories (Eq. E24).

Loader-level contract only (the engine application is covered by the
economics tests): the three input surfaces share ONE parse/validate
path, the default-off states are indistinguishable from a workbook that
predates the sheet, and every malformed block fails loudly before any
solver time is spent.

Locked properties:

1. Sheet absent, ``enabled`` = FALSE, and an omitted YAML block all
   resolve to ``typed["trajectories"] is None`` with every other
   section unchanged (bit-identity of the loader output).
2. A populated block round-trips exactly through
   ``write_workbook`` / ``read_workbook`` and through
   ``dump_structured_config`` / ``load_structured_config``.
3. Structural validation: unknown stream (with the known-stream hint),
   unknown mode, duplicate / non-contiguous years, non-numeric and
   boolean values, empty vectors.
4. Lifecycle-aware validation: coverage must equal
   ``project_lifecycle_years``, the year-1 anchor must be 1.0, the
   shared ``opex`` stream conflicts with the per-asset split, and a
   replace-mode trajectory over a non-zero ``*_inflation_pct`` warns.
5. ``scripts/polish_input_workbook.py`` creates the sheet once
   (disabled example) and never rewrites an existing one.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from pvbess_opt.io import (
    TRAJECTORIES_SHEET_COLUMNS,
    TRAJECTORY_STREAMS,
    _build_trajectories_sheet,
    _normalise_trajectories_block,
    _parse_trajectories_sheet,
    read_workbook,
    validate_workbook_params,
    write_workbook,
)

ROOT = Path(__file__).resolve().parent.parent
SHIPPED = ROOT / "inputs" / "input.xlsx"


def _sheet_df(rows: list[tuple]) -> pd.DataFrame:
    return pd.DataFrame(
        [dict(zip(TRAJECTORIES_SHEET_COLUMNS, r, strict=True)) for r in rows],
        columns=list(TRAJECTORIES_SHEET_COLUMNS),
    )


def _typed_with(trajectories, tmp_path: Path) -> dict:
    """Shipped typed dict with a trajectories block, written + re-read."""
    typed = read_workbook(SHIPPED)
    typed["trajectories"] = trajectories
    dst = tmp_path / "wb.xlsx"
    write_workbook(typed, dst)
    return read_workbook(dst)


def _n_years() -> int:
    typed = read_workbook(SHIPPED)
    return int(typed["project"]["project_lifecycle_years"])


# ---------------------------------------------------------------------------
# 1. Default-off bit-identity
# ---------------------------------------------------------------------------


def test_shipped_disabled_sheet_resolves_none():
    typed = read_workbook(SHIPPED)
    assert typed["trajectories"] is None


def test_sheet_absent_equals_disabled_sheet(tmp_path):
    """Deleting the sheet entirely changes NOTHING in the loader output."""
    stripped = tmp_path / "no_sheet.xlsx"
    shutil.copy(SHIPPED, stripped)
    wb = load_workbook(stripped)
    del wb["trajectories"]
    wb.save(stripped)

    with_sheet = read_workbook(SHIPPED)
    without = read_workbook(stripped)
    assert without["trajectories"] is None
    assert with_sheet["trajectories"] is None
    for section in ("project", "pv", "bess", "economics", "simulation",
                    "balancing", "ppa"):
        assert with_sheet[section] == without[section], section
    pd.testing.assert_frame_equal(with_sheet["ts"], without["ts"])


def test_enabled_sheet_with_no_data_rows_resolves_none():
    df = _sheet_df([("TRUE", None, None, None, None)])
    assert _parse_trajectories_sheet(df) is None


# ---------------------------------------------------------------------------
# 2. Round-trips
# ---------------------------------------------------------------------------


def test_workbook_round_trip_exact(tmp_path):
    n = _n_years()
    block = {
        "revenue_dam": {
            "mode": "overlay",
            "values": [1.0] + [round(1.0 - 0.01 * y, 4) for y in range(1, n)],
        },
        "opex": {
            "mode": "replace",
            "values": [1.0] * (n - 2) + [1.1, 1.1],
        },
    }
    reread = _typed_with(block, tmp_path)
    assert reread["trajectories"] == block


def test_blank_cell_inheritance_and_mode_default():
    """Blank stream cells inherit; a blank mode defaults to replace."""
    df = _sheet_df([
        ("TRUE", "revenue_retail", None, 1, 1.0),
        (None, None, None, 2, 1.05),
    ])
    parsed = _parse_trajectories_sheet(df)
    assert parsed == {
        "revenue_retail": {"mode": "replace", "values": [1.0, 1.05]},
    }


def test_yaml_round_trip_exact(tmp_path):
    from pvbess_opt.io_read import (
        dump_structured_config,
        load_structured_config,
    )

    typed = read_workbook(SHIPPED)
    n = _n_years()
    typed["trajectories"] = {
        "balancing_capacity": {
            "mode": "replace",
            "values": [1.0] + [0.9] * (n - 1),
        },
    }
    cfg = tmp_path / "cfg.yaml"
    dump_structured_config(typed, cfg)
    loaded = load_structured_config(cfg)
    assert loaded["trajectories"] == typed["trajectories"]


def test_yaml_list_shorthand_is_replace_mode(tmp_path):
    from pvbess_opt.io_read import load_structured_config

    ts_csv = tmp_path / "ts.csv"
    read_workbook(SHIPPED)["ts"].to_csv(ts_csv, index=False)
    cfg = tmp_path / "cfg.yaml"
    cfg.write_text(
        "timeseries_path: ts.csv\n"
        "trajectories:\n"
        "  revenue_dam: [1.0, 0.99, 0.98]\n"
    )
    loaded = load_structured_config(cfg)
    assert loaded["trajectories"] == {
        "revenue_dam": {"mode": "replace", "values": [1.0, 0.99, 0.98]},
    }


def test_dump_omits_block_when_unset(tmp_path):
    from pvbess_opt.io_read import dump_structured_config

    typed = read_workbook(SHIPPED)
    cfg = tmp_path / "cfg.yaml"
    dump_structured_config(typed, cfg)
    assert "trajectories" not in cfg.read_text()


# ---------------------------------------------------------------------------
# 3. Structural validation
# ---------------------------------------------------------------------------


def test_unknown_stream_rejected_with_hint():
    with pytest.raises(ValueError, match="revenue_dam"):
        _normalise_trajectories_block(
            {"dam_price": [1.0, 1.0]}, source="test",
        )


def test_unknown_mode_rejected():
    with pytest.raises(ValueError, match="overlay"):
        _normalise_trajectories_block(
            {"opex": {"mode": "multiply", "values": [1.0]}}, source="test",
        )


def test_boolean_value_rejected():
    with pytest.raises(ValueError, match="boolean"):
        _normalise_trajectories_block(
            {"opex": [1.0, True]}, source="test",
        )


@pytest.mark.parametrize("bad", [float("nan"), float("inf"), -0.5, "abc"])
def test_non_finite_or_negative_value_rejected(bad):
    with pytest.raises(ValueError):
        _normalise_trajectories_block(
            {"opex": [1.0, bad]}, source="test",
        )


def test_empty_values_rejected():
    with pytest.raises(ValueError, match="non-empty"):
        _normalise_trajectories_block({"opex": []}, source="test")


def test_duplicate_year_rejected():
    df = _sheet_df([
        ("TRUE", "opex", "replace", 1, 1.0),
        (None, None, None, 1, 1.1),
    ])
    with pytest.raises(ValueError, match="duplicate year 1"):
        _parse_trajectories_sheet(df)


def test_year_gap_rejected_naming_missing_years():
    df = _sheet_df([
        ("TRUE", "opex", "replace", 1, 1.0),
        (None, None, None, 4, 1.1),
    ])
    with pytest.raises(ValueError, match=r"missing \[2, 3\]"):
        _parse_trajectories_sheet(df)


def test_value_row_before_stream_rejected():
    df = _sheet_df([("TRUE", None, None, 1, 1.0)])
    with pytest.raises(ValueError, match="before any 'stream'"):
        _parse_trajectories_sheet(df)


# ---------------------------------------------------------------------------
# 4. Lifecycle-aware validation
# ---------------------------------------------------------------------------


def _validate_block(block) -> None:
    typed = read_workbook(SHIPPED)
    typed["trajectories"] = _normalise_trajectories_block(
        block, source="test",
    )
    validate_workbook_params(typed)


def test_short_coverage_rejected():
    with pytest.raises(ValueError, match="project_lifecycle_years"):
        _validate_block({"revenue_dam": [1.0, 0.99]})


def test_year1_anchor_enforced():
    n = _n_years()
    with pytest.raises(ValueError, match="anchor"):
        _validate_block({"revenue_dam": [0.9] + [1.0] * (n - 1)})


def test_opex_split_conflict_rejected():
    n = _n_years()
    ones = [1.0] * n
    with pytest.raises(ValueError, match="opex_pv"):
        _validate_block({"opex": ones, "opex_pv": ones})


def test_replace_over_nonzero_inflation_warns(caplog):
    n = _n_years()
    typed = read_workbook(SHIPPED)
    typed["economics"]["opex_inflation_pct"] = 1.0
    typed["trajectories"] = _normalise_trajectories_block(
        {"opex": {"mode": "replace", "values": [1.0] * n}}, source="test",
    )
    with caplog.at_level("WARNING"):
        validate_workbook_params(typed)
    assert any("opex_inflation_pct" in r.message for r in caplog.records)


def test_overlay_over_nonzero_inflation_does_not_warn(caplog):
    n = _n_years()
    typed = read_workbook(SHIPPED)
    typed["economics"]["opex_inflation_pct"] = 1.0
    typed["trajectories"] = _normalise_trajectories_block(
        {"opex": {"mode": "overlay", "values": [1.0] * n}}, source="test",
    )
    with caplog.at_level("WARNING"):
        validate_workbook_params(typed)
    assert not any(
        "trajectories stream" in r.message for r in caplog.records
    )


def test_workbook_read_rejects_invalid_sheet(tmp_path):
    """A workbook whose enabled sheet violates the anchor fails at read."""
    n = _n_years()
    typed = read_workbook(SHIPPED)
    typed["trajectories"] = {
        "revenue_dam": {"mode": "replace", "values": [1.0] * n},
    }
    dst = tmp_path / "wb.xlsx"
    write_workbook(typed, dst)
    wb = load_workbook(dst)
    ws = wb["trajectories"]
    ws.cell(row=2, column=5).value = 0.5  # year-1 anchor broken
    wb.save(dst)
    with pytest.raises(ValueError, match="anchor"):
        read_workbook(dst)


# ---------------------------------------------------------------------------
# 5. Builder + polish script
# ---------------------------------------------------------------------------


def test_builder_defaults_to_disabled_example():
    df = _build_trajectories_sheet(None)
    assert list(df.columns) == list(TRAJECTORIES_SHEET_COLUMNS)
    assert str(df.iloc[0]["enabled"]).upper() == "FALSE"


def test_all_streams_accepted_end_to_end(tmp_path):
    n = _n_years()
    # Aggregates conflict with their split legs (E24a/E60/E61), so the
    # sweep carries the SPLIT names and drops the aggregate aliases.
    conflicting_aggregates = {
        "opex", "revenue_dam", "balancing_capacity",
        "balancing_activation",
    }
    block = {
        stream: {"mode": "overlay", "values": [1.0] * n}
        for stream in TRAJECTORY_STREAMS
        if stream not in conflicting_aggregates
    }
    reread = _typed_with(block, tmp_path)
    assert set(reread["trajectories"]) == set(block)


def test_aggregate_and_split_leg_conflicts_rejected(tmp_path):
    n = _n_years()
    ones = [1.0] * n
    with pytest.raises(ValueError, match="revenue_dam_pv"):
        _validate_block({"revenue_dam": ones, "revenue_dam_pv": ones})
    with pytest.raises(ValueError, match="balancing_capacity_fcr"):
        _validate_block({
            "balancing_capacity": ones, "balancing_capacity_fcr": ones,
        })
    with pytest.raises(ValueError, match="balancing_activation_afrr_up"):
        _validate_block({
            "balancing_activation": ones,
            "balancing_activation_afrr_up": ones,
        })


def test_polish_creates_sheet_once(tmp_path):
    import sys

    sys.path.insert(0, str(ROOT / "scripts"))
    from polish_input_workbook import _ensure_trajectories_sheet

    target = tmp_path / "wb.xlsx"
    shutil.copy(SHIPPED, target)
    wb = load_workbook(target)
    del wb["trajectories"]
    assert _ensure_trajectories_sheet(wb) is True
    assert "trajectories" in wb.sheetnames
    # Second call is a no-op (user content is never rewritten).
    ws = wb["trajectories"]
    ws.cell(row=2, column=1).value = "TRUE"
    assert _ensure_trajectories_sheet(wb) is False
    assert ws.cell(row=2, column=1).value == "TRUE"
