"""Input-workbook guardrails: dropdowns, bounds, protection, dimming.

The shipped ``inputs/input.xlsx`` carries spreadsheet-level guardrails
generated from the loader's own schemas by
``scripts/polish_input_workbook.py``: TRUE/FALSE and enum dropdowns on
every constrained key, numeric-range validations mirrored from
``validate_workbook_params``, sheet protection with only the value
column editable, and dependency dimming for feature blocks whose toggle
is off.  These tests lock the workbook to the schemas — an enum value
added to the loader breaks them until the polisher is re-run — and lock
the polisher's numeric table to the validator by probing the loader
just outside every declared bound.  The guardrails are convenience
only: the loaders keep rejecting bad values on every surface.
"""

from __future__ import annotations

import copy
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pvbess_opt.io import (
    _ALLOWED_VALUES,
    _BOOL_KEYS,
    _KEY_TO_SHEET,
    _STR_KEYS,
    read_workbook,
    validate_workbook_params,
)

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "inputs" / "input.xlsx"

sys.path.insert(0, str(ROOT / "scripts"))
from polish_input_workbook import (  # noqa: E402
    _DIM_GROUPS,
    _NUMERIC_BOUNDS,
    _PARAMETER_SHEETS,
    _WARNING_BOUNDS,
    _WHOLE_MINIMUMS,
    polish_workbook,
)


def _kv_rows(ws) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value is not None and str(row[0].value).strip():
            rows[str(row[0].value).strip()] = row[0].row
    return rows


def _validation_for(ws, coord: str):
    for dv in ws.data_validations.dataValidation:
        # MultiCellRange membership is exact ("B5" never matches B55).
        if coord in dv.sqref:
            return dv
    return None


@pytest.fixture(scope="module")
def shipped():
    wb = load_workbook(WORKBOOK)
    return wb, {
        name: _kv_rows(wb[name])
        for name in _PARAMETER_SHEETS if name in wb.sheetnames
    }


def test_every_boolean_key_has_a_true_false_dropdown(shipped):
    wb, rows_by_sheet = shipped
    missing = []
    for key in sorted(_BOOL_KEYS):
        sheet = _KEY_TO_SHEET.get(key)
        row = rows_by_sheet.get(sheet, {}).get(key)
        if row is None:
            continue  # key not materialised on this workbook layout
        dv = _validation_for(wb[sheet], f"B{row}")
        if dv is None or dv.type != "list" or "TRUE" not in str(dv.formula1):
            missing.append(f"{sheet}.{key}")
    assert not missing, f"boolean keys without TRUE/FALSE dropdown: {missing}"


def test_every_enum_key_dropdown_matches_the_loader_schema(shipped):
    wb, rows_by_sheet = shipped
    bad = []
    for key in sorted(_STR_KEYS):
        allowed = _ALLOWED_VALUES.get(key)
        if not allowed:
            continue
        sheet = _KEY_TO_SHEET.get(key)
        row = rows_by_sheet.get(sheet, {}).get(key)
        if row is None:
            continue
        dv = _validation_for(wb[sheet], f"B{row}")
        if dv is None or dv.type != "list":
            bad.append(f"{sheet}.{key}: no list validation")
            continue
        listed = str(dv.formula1).strip('"').split(",")
        if sorted(listed) != sorted(allowed):
            bad.append(f"{sheet}.{key}: {sorted(listed)} != {sorted(allowed)}")
    assert not bad, f"enum dropdowns out of sync with the schema: {bad}"


def test_numeric_bound_cells_carry_range_validations(shipped):
    wb, rows_by_sheet = shipped
    missing = []
    for key in sorted(
        set(_NUMERIC_BOUNDS) | set(_WHOLE_MINIMUMS) | set(_WARNING_BOUNDS)
    ):
        sheet = _KEY_TO_SHEET.get(key)
        row = rows_by_sheet.get(sheet, {}).get(key)
        if row is None:
            continue
        dv = _validation_for(wb[sheet], f"B{row}")
        if dv is None or dv.type not in ("decimal", "whole"):
            missing.append(f"{sheet}.{key}")
    assert not missing, f"numeric keys without range validation: {missing}"


# Bounds whose validator check only runs once a feature toggle is on
# (e.g. the PPA block is skipped entirely while ppa_enabled = FALSE).
# The probe flips the toggle so the bound becomes reachable, and first
# proves the flipped baseline still validates cleanly — otherwise the
# expected raise could come from the context, not the probed value.
_PROBE_CONTEXT: dict[str, dict[tuple[str, str], object]] = {
    "ppa_volume_share_pct": {("ppa", "ppa_enabled"): True},
}


def test_numeric_bounds_table_locks_to_the_validator():
    """Every closed bound the polisher renders must be a REAL loader
    bound: mutating the shipped workbook's typed dict to just outside
    each declared range must make validate_workbook_params raise.  A
    polisher bound with no validator behind it fails here."""
    base = read_workbook(WORKBOOK)
    probes = {
        key: (lo - 1.0, hi + 1.0) for key, (lo, hi) in _NUMERIC_BOUNDS.items()
    } | {key: (lo - 1, None) for key, lo in _WHOLE_MINIMUMS.items()}
    for key, (below, above) in sorted(probes.items()):
        sheet = _KEY_TO_SHEET[key]
        baseline = copy.deepcopy(base)
        for (ctx_sheet, ctx_key), ctx_val in _PROBE_CONTEXT.get(key, {}).items():
            baseline[ctx_sheet][ctx_key] = ctx_val
        # The context alone must be valid, so the raise below is
        # attributable to the probed value and nothing else.
        validate_workbook_params(copy.deepcopy(baseline))
        for bad in (v for v in (below, above) if v is not None):
            typed = copy.deepcopy(baseline)
            typed[sheet][key] = bad
            with pytest.raises(ValueError):
                validate_workbook_params(typed)


def test_parameter_sheets_lock_everything_but_the_value_column(shipped):
    wb, rows_by_sheet = shipped
    for sheet, rows in rows_by_sheet.items():
        ws = wb[sheet]
        assert ws.protection.sheet, f"{sheet}: protection off"
        assert not ws.protection.password, f"{sheet}: unexpected password"
        some_row = next(iter(rows.values()))
        assert ws.cell(row=some_row, column=1).protection.locked
        assert not ws.cell(row=some_row, column=2).protection.locked
        assert ws.cell(row=1, column=2).protection.locked  # header row
    # Data-entry table sheets stay entirely unprotected.
    for free in ("timeseries", "scenarios", "sizing"):
        if free in wb.sheetnames:
            assert not wb[free].protection.sheet, f"{free}: must stay open"


def test_dependency_dimming_references_each_toggle(shipped):
    wb, rows_by_sheet = shipped
    for sheet, toggle, _template, _prefixes in _DIM_GROUPS:
        rows = rows_by_sheet.get(sheet, {})
        if toggle not in rows:
            continue
        anchor = f"$B${rows[toggle]}"
        formulas = [
            f
            for cf in wb[sheet].conditional_formatting
            for rule in cf.rules
            for f in (rule.formula or [])
        ]
        assert any(anchor in f for f in formulas), (
            f"{sheet}: no dimming rule anchored to {toggle} ({anchor})"
        )


def test_polisher_guardrails_are_idempotent(tmp_path):
    dst = tmp_path / "input.xlsx"
    shutil.copy(WORKBOOK, dst)

    def snapshot():
        wb = load_workbook(dst)
        return {
            name: (
                sorted(
                    (str(dv.sqref), dv.type, str(dv.formula1))
                    for dv in wb[name].data_validations.dataValidation
                ),
                sum(len(cf.rules) for cf in wb[name].conditional_formatting),
                bool(wb[name].protection.sheet),
            )
            for name in _PARAMETER_SHEETS if name in wb.sheetnames
        }

    polish_workbook(dst)
    first = snapshot()
    polish_workbook(dst)
    assert snapshot() == first


def test_loader_parses_the_guarded_workbook_unchanged():
    typed = read_workbook(WORKBOOK)
    validate_workbook_params(typed)
    assert typed["project"]["mode"] in ("self_consumption", "merchant")
    assert typed["project"]["project_lifecycle_years"] >= 1
