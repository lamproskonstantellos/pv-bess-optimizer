"""Regression tests: the ENTSO-E security token must never leak.

Two paths are covered:

* a transport-level failure during a fetch must not carry the token
  (which rides in the request URL) into the raised exception / traceback
  (would otherwise surface via ``cli.main``'s ``logger.exception``);
* the shareable ``01_inputs`` input snapshot must have the token cell
  blanked even when no fetch happened (all sources at their ``file``
  default), so a live secret never rides into a results directory.
"""

from __future__ import annotations

import traceback
from pathlib import Path

import openpyxl
import pytest

from pvbess_opt.marketdata import base as md_base
from pvbess_opt.marketdata import blank_entsoe_token, entsoe

_FAKE_TOKEN = "SECRETdeadbeef-0123-4567-89ab-cdef01234567"


# ---------------------------------------------------------------------------
# S1 — transport-error token scrub
# ---------------------------------------------------------------------------


def test_http_get_transport_error_scrubs_token(monkeypatch):
    """A requests transport error must not leak the token or its chain."""
    import requests

    def _boom(url, params=None, timeout=None):
        # requests embeds the full query string (incl. securityToken) in
        # the ConnectionError message — mirror that here.
        raise requests.ConnectionError(
            f"Max retries exceeded with url: /api?securityToken={_FAKE_TOKEN}"
        )

    monkeypatch.setattr(requests, "get", _boom)

    with pytest.raises(entsoe.MarketDataError) as excinfo:
        entsoe._http_get({"securityToken": _FAKE_TOKEN}, timeout=1.0)

    err = excinfo.value
    assert _FAKE_TOKEN not in str(err)
    # The token-bearing original exception must be dropped from the chain
    # so logger.exception / a traceback can never print it.
    assert err.__suppress_context__ is True
    assert err.__cause__ is None
    rendered = "".join(
        traceback.format_exception(type(err), err, err.__traceback__)
    )
    assert _FAKE_TOKEN not in rendered


# ---------------------------------------------------------------------------
# S2 — snapshot token scrub (no-fetch path)
# ---------------------------------------------------------------------------


def _make_market_workbook(path: Path, token: str | None) -> None:
    """A tiny workbook (timeseries + market_data) — no need for the 5 MB file."""
    wb = openpyxl.Workbook()
    ts = wb.active
    ts.title = "timeseries"
    ts.append(["timestamp", "dam_price_eur_per_mwh"])
    for i in range(4):
        ts.append([f"2026-01-01 0{i}:00", 50.0])
    md = wb.create_sheet("market_data")
    md.append(["key", "value"])
    md.append(["price_source", "file"])
    md.append(["entsoe_token", token])
    md.append(["bidding_zone", "gr"])
    wb.save(path)


def _read_token(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    md = wb["market_data"]
    for row in md.iter_rows(min_row=2, max_col=2):
        if isinstance(row[0].value, str) and row[0].value.strip() == "entsoe_token":
            return row[1].value
    return "MISSING"


def test_blank_entsoe_token_scrubs_no_fetch_snapshot(tmp_path):
    """A token-bearing, all-file-source snapshot must be scrubbed."""
    snap = tmp_path / "input_snapshot.xlsx"
    _make_market_workbook(snap, _FAKE_TOKEN)
    assert _read_token(snap) == _FAKE_TOKEN  # precondition

    blank_entsoe_token(snap)

    assert _read_token(snap) in (None, "")


def test_blank_entsoe_token_noop_when_empty_is_byte_identical(tmp_path):
    """An empty-token snapshot is left byte-for-byte untouched (no re-save)."""
    snap = tmp_path / "input_snapshot.xlsx"
    _make_market_workbook(snap, None)
    before = snap.read_bytes()

    blank_entsoe_token(snap)

    assert snap.read_bytes() == before


def test_blank_entsoe_token_ignores_non_workbook(tmp_path):
    """A non-xlsx snapshot (e.g. a YAML config) is a silent no-op."""
    cfg = tmp_path / "input_snapshot.xlsx"
    cfg.write_text("mode: merchant\n")  # not a real workbook
    blank_entsoe_token(cfg)  # must not raise
    assert cfg.read_text() == "mode: merchant\n"


def test_materialize_still_blanks_token(tmp_path):
    """Refactor guard: the fetch path still scrubs the token."""
    import pandas as pd

    snap = tmp_path / "input_snapshot.xlsx"
    _make_market_workbook(snap, _FAKE_TOKEN)

    # A trivial provenance record flipping the DAM source back to file.
    ts = pd.DataFrame({"dam_price_eur_per_mwh": [50.0] * 4})
    md_base.materialize_bypassed_workbook(
        snap, ts,
        [{"column": "dam_price_eur_per_mwh", "source_key": "price_source"}],
    )
    assert _read_token(snap) in (None, "")
    # And the flipped source came through.
    wb = openpyxl.load_workbook(snap, data_only=True)
    md = {r[0]: r[1] for r in wb["market_data"].iter_rows(min_row=2, values_only=True)}
    assert md["price_source"] == "file"
