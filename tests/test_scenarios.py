"""Batch scenario engine: inheritance, overrides, comparison, outputs."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pvbess_opt.scenarios import (
    _COMPARISON_COLUMNS,
    _apply_scenario_overrides,
    evaluate_scenario,
    read_scenarios_file,
    resolve_inheritance,
    run_scenarios,
    write_scenario_comparison_workbook,
)
from pvbess_opt.theme import COL_WIDTH_MAX, COL_WIDTH_MIN, HEADER_FILL_HEX

ROOT = Path(__file__).resolve().parent.parent


def _highs_available() -> bool:
    try:
        import highspy
    except ImportError:
        return False
    return bool(highspy)


def _base_typed() -> dict:
    from pvbess_opt.io import (
        BALANCING_SHEET_DEFAULTS,
        BESS_SHEET_DEFAULTS,
        ECONOMICS_SHEET_DEFAULTS,
        PROJECT_SHEET_DEFAULTS,
        PV_SHEET_DEFAULTS,
        SIMULATION_SHEET_DEFAULTS,
    )

    return {
        "pv": dict(PV_SHEET_DEFAULTS, pv_nameplate_kwp=10000.0, capex_pv_eur_per_kw=500.0),
        "bess": dict(
            BESS_SHEET_DEFAULTS, bess_power_kw=5000.0, bess_capacity_kwh=10000.0,
            capex_bess_eur_per_kwh=200.0,
        ),
        "project": dict(PROJECT_SHEET_DEFAULTS, site_capex_eur=1000.0),
        "economics": dict(ECONOMICS_SHEET_DEFAULTS),
        "simulation": dict(SIMULATION_SHEET_DEFAULTS),
        "balancing": dict(BALANCING_SHEET_DEFAULTS),
    }


def test_resolve_inheritance_clones_and_overrides():
    scns = [
        {"name": "A", "bess": {"power_kw": 1000}},
        {"name": "B", "inherits": "A", "balancing": True},
    ]
    resolved = {s["name"]: s for s in resolve_inheritance(scns)}
    assert resolved["B"]["bess"]["power_kw"] == 1000   # inherited
    # The bare balancing scalar is canonicalised to its dict form so it
    # deep-merges across inherits instead of replacing wholesale.
    assert resolved["B"]["balancing"] == {"balancing_enabled": True}
    assert "inherits" not in resolved["B"]


def test_resolve_inheritance_merges_bare_and_dotted_balancing():
    """A parent's bare ``balancing`` enable and a child's dotted
    ``balancing.<key>`` override must BOTH survive the inherits merge — a
    scalar/dict cross used to replace wholesale, silently dropping the
    parent's enable (the child then solved with balancing OFF under the
    requested label) or, in the reverse direction, the parent's dotted
    keys."""
    resolved = {s["name"]: s for s in resolve_inheritance([
        {"name": "bal_on", "balancing": True},
        {"name": "bal_blocks", "inherits": "bal_on",
         "balancing": {"bm_block_hours": 4}},
    ])}
    assert resolved["bal_blocks"]["balancing"] == {
        "balancing_enabled": True, "bm_block_hours": 4,
    }
    # Reverse direction: parent dotted, child bare — the child's enable
    # composes with (not clobbers) the parent's dotted keys.
    resolved2 = {s["name"]: s for s in resolve_inheritance([
        {"name": "blocks", "balancing": {"bm_block_hours": 4}},
        {"name": "on", "inherits": "blocks", "balancing": True},
    ])}
    assert resolved2["on"]["balancing"] == {
        "bm_block_hours": 4, "balancing_enabled": True,
    }


def test_apply_overrides_shorthand_and_capex_multiplier():
    base = _base_typed()
    typed = _apply_scenario_overrides(base, {
        "name": "x",
        "pv": {"nameplate_kwp": 7000, "source": "pvgis"},
        "bess": {"power_kw": 3000, "capacity_kwh": 6000},
        "balancing": True,
        "capex_multiplier": 0.5,
    })
    assert typed["pv"]["pv_nameplate_kwp"] == 7000
    assert typed["pv"]["pv_source"] == "file"  # forced after resolution
    assert typed["bess"]["bess_power_kw"] == 3000
    assert typed["bess"]["bess_capacity_kwh"] == 6000
    assert typed["balancing"]["balancing_enabled"] is True
    assert typed["pv"]["capex_pv_eur_per_kw"] == pytest.approx(250.0)
    assert typed["bess"]["capex_bess_eur_per_kwh"] == pytest.approx(100.0)
    assert typed["project"]["site_capex_eur"] == pytest.approx(500.0)
    assert base["pv"]["pv_nameplate_kwp"] == 10000.0  # base untouched


def test_balancing_off_shorthand():
    typed = _apply_scenario_overrides(
        _base_typed(), {"name": "x", "balancing": False},
    )
    assert typed["balancing"]["balancing_enabled"] is False


def test_read_scenarios_file(tmp_path):
    good = tmp_path / "s.yaml"
    good.write_text(
        "scenarios:\n  - name: A\n  - name: B\n    inherits: A\n",
        encoding="utf-8",
    )
    assert [s["name"] for s in read_scenarios_file(good)] == ["A", "B"]
    bad = tmp_path / "bad.yaml"
    bad.write_text("foo: 1\n", encoding="utf-8")
    with pytest.raises(ValueError):
        read_scenarios_file(bad)


# ---------------------------------------------------------------------------
# Excel-driven scenarios sheet (tidy/long, gated by an enabled toggle)
# ---------------------------------------------------------------------------


def _typed_for_write(scenarios=None) -> dict:
    import numpy as np

    from pvbess_opt.io import (
        BESS_SHEET_DEFAULTS,
        ECONOMICS_SHEET_DEFAULTS,
        PROJECT_SHEET_DEFAULTS,
        PV_SHEET_DEFAULTS,
        SIMULATION_SHEET_DEFAULTS,
    )
    ts = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=24, freq="h"),
        "pv_kwh": [100.0] * 24,
        "load_kwh": [50.0] * 24,
        "dam_price_eur_per_mwh": [80.0] * 24,
    })
    typed = {
        "ts": ts,
        "project": dict(PROJECT_SHEET_DEFAULTS),
        "pv": dict(PV_SHEET_DEFAULTS, pv_nameplate_kwp=1000.0),
        "bess": dict(
            BESS_SHEET_DEFAULTS, bess_power_kw=500.0, bess_capacity_kwh=2000.0,
        ),
        "economics": dict(ECONOMICS_SHEET_DEFAULTS),
        "simulation": dict(SIMULATION_SHEET_DEFAULTS),
        "max_injection_profile": np.full(24, 100.0),
    }
    if scenarios is not None:
        typed["scenarios"] = scenarios
    return typed


def test_parse_scenarios_sheet_groups_and_nests():
    from pvbess_opt.scenarios import _parse_scenarios_sheet
    df = pd.DataFrame({
        "enabled": ["TRUE", None, None, None],
        "name": ["A", "B", "B", "C"],
        "inherits": [None, "A", "A", "A"],
        "target": [
            "project.mode", "bess.power_kw", "bess.capacity_kwh",
            "capex_multiplier",
        ],
        "value": ["merchant", 0, 0, 0.8],
    })
    enabled, scns = _parse_scenarios_sheet(df)
    assert enabled is True
    assert [s["name"] for s in scns] == ["A", "B", "C"]  # order preserved
    by = {s["name"]: s for s in scns}
    assert by["A"]["project"] == {"mode": "merchant"}
    assert by["B"]["inherits"] == "A"
    assert by["B"]["bess"] == {"power_kw": 0, "capacity_kwh": 0}
    assert by["C"]["capex_multiplier"] == 0.8


def test_parse_scenarios_sheet_merges_bare_and_dotted_balancing():
    """The documented bare ``balancing`` shorthand is canonicalised to its
    dict form at parse time, so mixing it with dotted ``balancing.<key>``
    rows COMPOSES correctly in either order (previously one of the two rows
    was silently dropped; an intermediate fix raised instead)."""
    from pvbess_opt.scenarios import _parse_scenarios_sheet

    for targets, values in (
        (["balancing", "balancing.bm_block_hours"], ["TRUE", 4]),
        (["balancing.bm_block_hours", "balancing"], [4, "TRUE"]),
    ):
        df = pd.DataFrame({
            "enabled": ["TRUE", None],
            "name": ["s1", "s1"],
            "inherits": [None, None],
            "target": targets,
            "value": values,
        })
        _enabled, scns = _parse_scenarios_sheet(df)
        assert scns[0]["balancing"] == {
            "balancing_enabled": "TRUE", "bm_block_hours": 4,
        }, (targets, scns[0])


def test_parse_scenarios_sheet_rejects_nonbalancing_collision_when_enabled():
    """For sections WITHOUT a documented scalar shorthand, a bare+dotted mix
    is still a hard error on an ENABLED sheet (silently dropping a row would
    solve a different scenario than described) — but a DISABLED sheet is
    documented as inert, so its drafting mistakes only warn and the base run
    proceeds."""
    from pvbess_opt.scenarios import _parse_scenarios_sheet

    def _sheet(enabled: str) -> pd.DataFrame:
        return pd.DataFrame({
            "enabled": [enabled, None],
            "name": ["s1", "s1"],
            "inherits": [None, None],
            "target": ["project", "project.mode"],
            "value": ["x", "merchant"],
        })

    with pytest.raises(ValueError, match=r"s1.*project"):
        _parse_scenarios_sheet(_sheet("TRUE"))
    # Disabled: parses without raising (warning only), stays disabled.
    enabled, _scns = _parse_scenarios_sheet(_sheet("FALSE"))
    assert enabled is False


def test_scenario_nameplate_override_rescales_pv_profile():
    """A ``pv.nameplate_kwp`` override must rescale the resolved ``pv_kwh``
    profile (shape preserved) — the module contract and the sizing sweep's
    identical treatment; otherwise the scenario solves the BASE plant's
    generation against the OVERRIDDEN plant's CAPEX/OPEX."""
    from pvbess_opt.scenarios import _apply_scenario_overrides

    base = {
        "pv": {"pv_nameplate_kwp": 15000.0},
        "bess": {},
        "project": {},
        "economics": {},
        "simulation": {},
        "balancing": {},
        "ts": pd.DataFrame({
            "timestamp": pd.date_range("2026-01-01", periods=4, freq="h"),
            "pv_kwh": [0.0, 100.0, 200.0, 50.0],
        }),
    }
    out = _apply_scenario_overrides(
        base, {"name": "double", "pv": {"nameplate_kwp": 30000.0}},
    )
    assert out["pv"]["pv_nameplate_kwp"] == 30000.0
    assert list(out["ts"]["pv_kwh"]) == [0.0, 200.0, 400.0, 100.0]
    # The base frame is untouched (deepcopy semantics).
    assert list(base["ts"]["pv_kwh"]) == [0.0, 100.0, 200.0, 50.0]
    # No override -> no rescale.
    same = _apply_scenario_overrides(base, {"name": "plain"})
    assert list(same["ts"]["pv_kwh"]) == [0.0, 100.0, 200.0, 50.0]
    # Zero-nameplate scenario zeroes the profile (consistent with the MILP
    # pinning PV at zero nameplate).
    zero = _apply_scenario_overrides(
        base, {"name": "no pv", "pv": {"nameplate_kwp": 0.0}},
    )
    assert list(zero["ts"]["pv_kwh"]) == [0.0, 0.0, 0.0, 0.0]


def test_scenario_nameplate_override_nonnumeric_names_the_scenario():
    """A locale-formatted or garbage nameplate override must raise naming
    the scenario and value, not surface as a bare ``float()`` error deep in
    the batch loop."""
    from pvbess_opt.scenarios import _apply_scenario_overrides

    base = {
        "pv": {"pv_nameplate_kwp": 15000.0}, "bess": {}, "project": {},
        "economics": {}, "simulation": {}, "balancing": {},
        "ts": pd.DataFrame({
            "timestamp": pd.date_range("2026-01-01", periods=2, freq="h"),
            "pv_kwh": [0.0, 100.0],
        }),
    }
    with pytest.raises(ValueError, match=r"big pv.*30,000.*not a number"):
        _apply_scenario_overrides(
            base, {"name": "big pv", "pv": {"nameplate_kwp": "30,000"}},
        )


def test_scenario_nameplate_override_on_zero_base_warns(caplog):
    """Overriding the nameplate on a base WITHOUT one cannot rescale any
    shape — CAPEX/OPEX grow while generation stays at the base profile.
    That asymmetry must be loud, not silent."""
    import logging

    from pvbess_opt.scenarios import _apply_scenario_overrides

    base = {
        "pv": {}, "bess": {}, "project": {}, "economics": {},
        "simulation": {}, "balancing": {},
        "ts": pd.DataFrame({
            "timestamp": pd.date_range("2026-01-01", periods=2, freq="h"),
            "pv_kwh": [0.0, 0.0],
        }),
    }
    with caplog.at_level(logging.WARNING, logger="pvbess_opt.scenarios"):
        out = _apply_scenario_overrides(
            base, {"name": "adds pv", "pv": {"nameplate_kwp": 30000.0}},
        )
    assert out["pv"]["pv_nameplate_kwp"] == 30000.0
    assert any(
        "no PV shape to scale" in rec.getMessage() for rec in caplog.records
    ), caplog.records


def test_parse_scenarios_sheet_disabled_toggle():
    from pvbess_opt.scenarios import _parse_scenarios_sheet
    df = pd.DataFrame({
        "enabled": ["FALSE", None],
        "name": ["A", "A"],
        "inherits": [None, None],
        "target": ["project.mode", "bess.power_kw"],
        "value": ["merchant", 0],
    })
    enabled, scns = _parse_scenarios_sheet(df)
    assert enabled is False
    assert len(scns) == 1 and scns[0]["name"] == "A"


def test_read_scenarios_block_gated_by_enabled(tmp_path):
    from pvbess_opt.io import write_workbook
    from pvbess_opt.scenarios import read_scenarios_block
    # The shipped default example is disabled -> no batch.
    disabled = tmp_path / "disabled.xlsx"
    write_workbook(_typed_for_write(None), disabled)
    assert read_scenarios_block(disabled) is None
    # enabled=TRUE on the first row -> the scenarios are parsed.
    enabled = tmp_path / "enabled.xlsx"
    write_workbook(_typed_for_write([
        {"enabled": "TRUE", "name": "Merchant", "target": "project.mode",
         "value": "merchant"},
        {"name": "PV only", "inherits": "Merchant", "target": "bess.power_kw",
         "value": 0},
    ]), enabled)
    scns = read_scenarios_block(enabled)
    assert [s["name"] for s in scns] == ["Merchant", "PV only"]
    assert scns[0]["project"] == {"mode": "merchant"}
    assert scns[1]["inherits"] == "Merchant"


def test_repo_input_xlsx_ships_disabled_scenarios_sheet():
    from pvbess_opt.scenarios import read_scenarios_block
    sheets = pd.ExcelFile(ROOT / "inputs" / "input.xlsx").sheet_names
    assert "scenarios" in sheets
    # Shipped disabled, so a normal run is unaffected.
    assert read_scenarios_block(ROOT / "inputs" / "input.xlsx") is None


def test_cli_errors_when_both_sweeps_enabled(tmp_path, monkeypatch):
    """Enabling the sizing and scenarios sheets together is rejected before
    any run path executes."""
    from pvbess_opt import cli
    from pvbess_opt.io import write_workbook

    called: list[str] = []
    monkeypatch.setattr(cli, "run", lambda *_a, **_k: called.append("run"))
    monkeypatch.setattr(
        cli, "run_sizing", lambda *_a, **_k: called.append("sizing"),
    )
    monkeypatch.setattr(
        cli, "run_scenarios", lambda *_a, **_k: called.append("scenarios"),
    )
    typed = _typed_for_write([
        {"enabled": "TRUE", "name": "M", "target": "project.mode",
         "value": "merchant"},
    ])
    typed["sizing"] = [
        {"enabled": "TRUE", "pv_nameplate_kwp": 1000, "bess_power_kw": 500,
         "bess_duration_hours": 2},
    ]
    wb = tmp_path / "both.xlsx"
    write_workbook(typed, wb)
    assert cli.main([str(wb)]) == 1
    assert called == []  # guard fires before any run path


def test_cli_routes_enabled_scenarios_sheet_to_batch(tmp_path, monkeypatch):
    """An enabled scenarios sheet (and no sizing) runs the batch path with
    the parsed scenarios."""
    from pvbess_opt import cli
    from pvbess_opt.io import write_workbook

    captured: dict = {}
    monkeypatch.setattr(
        cli, "run", lambda *_a, **_k: captured.setdefault("run", True),
    )
    monkeypatch.setattr(
        cli, "run_sizing", lambda *_a, **_k: captured.setdefault("sizing", True),
    )
    monkeypatch.setattr(
        cli, "run_scenarios",
        lambda _config, scenarios: captured.update(scenarios=scenarios),
    )
    wb = tmp_path / "scn.xlsx"
    write_workbook(_typed_for_write([
        {"enabled": "TRUE", "name": "Merchant", "target": "project.mode",
         "value": "merchant"},
        {"name": "PV only", "inherits": "Merchant", "target": "bess.power_kw",
         "value": 0},
    ]), wb)
    assert cli.main([str(wb)]) == 0
    assert "run" not in captured and "sizing" not in captured
    assert [s["name"] for s in captured["scenarios"]] == ["Merchant", "PV only"]


def test_comparison_workbook_is_styled(tmp_path):
    comp = pd.DataFrame(
        [
            {c: ("A" if c == "name" else 0.0) for c in _COMPARISON_COLUMNS},
            {c: ("B" if c == "name" else 1.0) for c in _COMPARISON_COLUMNS},
        ],
        columns=list(_COMPARISON_COLUMNS),
    )
    out = write_scenario_comparison_workbook(tmp_path / "cmp.xlsx", comp)
    wb = load_workbook(out)
    assert wb.sheetnames
    for sn in wb.sheetnames:
        ws = wb[sn]
        assert ws.freeze_panes == "A2"
        for cell in ws[1]:
            if cell.value is None:
                continue
            rgb = (getattr(cell.fill.fgColor, "rgb", None) or "")
            assert rgb.upper().lstrip("0").rjust(6, "0")[-6:] == HEADER_FILL_HEX
        for c in range(1, ws.max_column + 1):
            dim = ws.column_dimensions.get(get_column_letter(c))
            assert dim is not None and dim.width is not None
            assert COL_WIDTH_MIN <= float(dim.width) <= COL_WIDTH_MAX


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_run_scenarios_three(tmp_path):
    from pvbess_opt import RunConfig
    from pvbess_opt.io import read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    short = tmp_path / "short.xlsx"
    write_workbook(typed, short)

    config = RunConfig(
        excel=short, solver="highs", outdir=tmp_path / "out",
        mip_gap=0.05, time_limit=180,
    )
    scns = [
        {"name": "SC hybrid", "project": {"mode": "self_consumption"}},
        {"name": "Merchant hybrid", "project": {"mode": "merchant"}},
        {"name": "Cheap CAPEX", "inherits": "Merchant hybrid",
         "capex_multiplier": 0.8},
    ]
    result = run_scenarios(config, scns)
    assert list(result.comparison["name"]) == [
        "SC hybrid", "Merchant hybrid", "Cheap CAPEX",
    ]
    assert result.comparison["npv_eur"].notna().all()
    npv = dict(zip(
        result.comparison["name"], result.comparison["npv_eur"], strict=False,
    ))
    assert npv["Cheap CAPEX"] > npv["Merchant hybrid"]  # lower CAPEX -> higher NPV
    runs = list((tmp_path / "out").glob("*_scenarios_*"))
    assert runs
    assert (runs[0] / "scenario_comparison.xlsx").exists()
    assert (runs[0] / "scenario_comparison.pdf").exists()
    assert (runs[0] / "scenario_revenue_bridge.pdf").exists()


def test_run_scenarios_applies_cli_mode_override(tmp_path, monkeypatch):
    """The CLI ``--mode`` override is applied to the scenario-batch base,
    mirroring ``pipeline.run`` and ``sizing.run_sizing`` (regression:
    ``run_scenarios`` previously ignored ``config.mode`` so
    ``--mode merchant --scenarios ...`` ran in the workbook's mode)."""
    from pvbess_opt import RunConfig
    from pvbess_opt import scenarios as scn_mod

    captured: dict[str, str] = {}

    class _Stop(Exception):
        pass

    def _fake_batch(base_typed, scenarios, **_kwargs):
        captured["mode"] = base_typed["project"]["mode"]
        raise _Stop

    monkeypatch.setattr(scn_mod, "run_scenario_batch", _fake_batch)

    config = RunConfig(
        excel=ROOT / "inputs" / "input.xlsx", solver="highs",
        outdir=tmp_path / "out", mode="merchant",
    )
    with pytest.raises(_Stop):
        scn_mod.run_scenarios(config, [{"name": "base"}])
    assert captured["mode"] == "merchant"


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_no_override_scenario_matches_standalone(tmp_path):
    from pvbess_opt.availability import apply_unavailability_derate
    from pvbess_opt.io import read_inputs, read_workbook, write_workbook
    from pvbess_opt.kpis import compute_kpis
    from pvbess_opt.optimization import run_scenario as solve
    from pvbess_opt.pipeline import _build_financials

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    short = tmp_path / "short.xlsx"
    write_workbook(typed, short)
    opts = {
        "solver_name": "highs", "mip_gap": 0.05,
        "time_limit_seconds": 180, "tee": False,
    }

    params, ts = read_inputs(short)
    res, _s, _f = solve(params, ts, return_unrounded=True, **opts)
    kpis = apply_unavailability_derate(
        compute_kpis(res, params, verify_balance=False),
        float(params.get("unavailability_pct", 0.0) or 0.0),
    )
    fin = _build_financials(short, params, ts, kpis, res)["fin_kpis"]

    row = evaluate_scenario(read_workbook(short), {"name": "base"}, solver_opts=opts)
    assert row["npv_eur"] == pytest.approx(fin["npv_eur"], rel=1e-9, abs=1e-6)
    assert row["profit_total_eur"] == pytest.approx(
        kpis["profit_total_eur"], rel=1e-9, abs=1e-6,
    )


def test_scenario_row_balancing_enabled_reflects_parsed_value(tmp_path):
    """Regression: an Excel scenarios-sheet override written as the dotted
    target ``balancing.balancing_enabled = FALSE`` reaches the scenario dict
    as the unparsed string 'FALSE'.  The comparison row's balancing_enabled
    column must report the PARSED value (False, what actually solved), not
    ``bool('FALSE')`` (True)."""
    from pvbess_opt.io import read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    short = tmp_path / "short.xlsx"
    write_workbook(typed, short)
    opts = {
        "solver_name": "highs", "mip_gap": 0.05,
        "time_limit_seconds": 180, "tee": False,
    }
    row = evaluate_scenario(
        read_workbook(short),
        {"name": "nobal", "balancing": {"balancing_enabled": "FALSE"}},
        solver_opts=opts,
    )
    assert row["balancing_enabled"] is False


# ---------------------------------------------------------------------------
# market_data / scenario_engine overrides + the materialise-time flip
# ---------------------------------------------------------------------------


def test_overrides_reach_market_and_scenario_engine_sheets():
    """The validator whitelists the two new sections; the apply loop
    must land them (a validated-but-dropped override is exactly the
    silent no-op the validator exists to prevent)."""
    from pvbess_opt.io import (
        MARKET_DATA_SHEET_DEFAULTS,
        SCENARIO_ENGINE_SHEET_DEFAULTS,
    )

    base = _base_typed()
    base["market_data"] = dict(MARKET_DATA_SHEET_DEFAULTS)
    base["scenario_engine"] = dict(SCENARIO_ENGINE_SHEET_DEFAULTS)
    typed = _apply_scenario_overrides(base, {
        "name": "armed",
        "scenario_engine": {"price_scenarios_enabled": True},
        "market_data": {"bidding_zone": "de_lu"},
    })
    assert typed["scenario_engine"]["price_scenarios_enabled"] is True
    assert typed["market_data"]["bidding_zone"] == "de_lu"


def test_materialised_scenario_disarms_market_fetch():
    """Without an explicit market_data override the materialised temp
    workbook must not re-trigger the bypass on re-read: the base read
    already resolved fetched columns into the frame, and a re-fetch
    would clobber a price_deck override. Sources flip to 'file', the
    token cell is blanked (the materialize_bypassed_workbook rule)."""
    from pvbess_opt.io import MARKET_DATA_SHEET_DEFAULTS

    base = _base_typed()
    base["market_data"] = dict(
        MARKET_DATA_SHEET_DEFAULTS,
        price_source="entsoe", entsoe_token="secret-token",
    )
    typed = _apply_scenario_overrides(base, {"name": "plain"})
    assert typed["market_data"]["price_source"] == "file"
    assert typed["market_data"]["balancing_source"] == "file"
    assert typed["market_data"]["imbalance_source"] == "file"
    assert typed["market_data"]["entsoe_token"] == ""
    # An explicit market_data override keeps its configuration —
    # the deliberate re-fetch is the scenario's own semantics.
    explicit = _apply_scenario_overrides(base, {
        "name": "refetch", "market_data": {"price_reference_year": 2024},
    })
    assert explicit["market_data"]["price_source"] == "entsoe"
    assert explicit["market_data"]["price_reference_year"] == 2024
