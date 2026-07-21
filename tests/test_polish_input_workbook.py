"""Unit tests for :mod:`scripts.polish_input_workbook`.

The polish script is the single source of truth for the canonical
workbook styling. These tests build a small temporary workbook with the
same sheet schema as the shipped one, run the polish helper, and assert
the four invariants the script promises:

* every sheet ends with ``freeze_panes = "A2"``;
* every column has a non-default (explicit) width within the AutoFit
  clamps;
* row 1 of every sheet carries the navy fill (``1F3864``) and a white
  bold font;
* a second polish run produces an identical per-sheet hash so the
  operation is idempotent.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pvbess_opt.io import (
    BALANCING_SHEET_DEFAULTS,
    BESS_SHEET_DEFAULTS,
    ECONOMICS_SHEET_DEFAULTS,
    PROJECT_SHEET_DEFAULTS,
    PV_SHEET_DEFAULTS,
    SIMULATION_SHEET_DEFAULTS,
    write_workbook,
)
from pvbess_opt.theme import (
    COL_WIDTH_MAX,
    COL_WIDTH_MIN,
    HEADER_FILL_HEX,
    HEADER_FONT_HEX,
)
from scripts.polish_input_workbook import polish_workbook


def _normalise_rgb(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return value.upper().lstrip("0").rjust(6, "0")[-6:]


def _build_minimal_typed() -> dict[str, object]:
    """Return a typed-dict fixture covering every parameter sheet."""
    import numpy as np

    ts = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=4, freq="15min"),
        "load_kwh": [1.0, 1.0, 1.0, 1.0],
        "pv_kwh": [0.0, 0.5, 0.5, 0.0],
        "dam_price_eur_per_mwh": [50.0, 60.0, 70.0, 80.0],
    })
    return {
        "ts": ts,
        "project": dict(PROJECT_SHEET_DEFAULTS),
        "pv": dict(PV_SHEET_DEFAULTS),
        "bess": dict(BESS_SHEET_DEFAULTS),
        "economics": dict(ECONOMICS_SHEET_DEFAULTS),
        "simulation": dict(SIMULATION_SHEET_DEFAULTS),
        "balancing": dict(BALANCING_SHEET_DEFAULTS),
        # Per-asset sub-cap sheets so the polish exercises the centered
        # header path on the same schema the shipped workbook carries.
        "max_injection_profile_pv": np.full(24, 100.0),
        "max_injection_profile_bess": np.full(24, 100.0),
    }


def _sheet_state_hash(path: Path) -> str:
    """Hash sheet-level state: values, fills, fonts, widths, freeze."""
    wb = load_workbook(path)
    h = hashlib.sha256()
    for sn in wb.sheetnames:
        ws: Worksheet = wb[sn]
        h.update(b"|sheet:" + sn.encode())
        h.update(b"|fp:" + str(ws.freeze_panes).encode())
        for col_letter, dim in sorted(ws.column_dimensions.items()):
            h.update(f"|w[{col_letter}]={dim.width:.6f}".encode())
        for row in ws.iter_rows():
            for c in row:
                h.update(b"|v:" + repr(c.value).encode())
                rgb = getattr(getattr(c.fill, "fgColor", None), "rgb", None)
                h.update(b"|fill:" + repr(rgb).encode())
                h.update(b"|bold:" + repr(c.font.bold).encode())
                h.update(
                    b"|font_color:"
                    + repr(getattr(c.font.color, "rgb", None)).encode(),
                )
                h.update(b"|wrap:" + repr(c.alignment.wrap_text).encode())
                h.update(b"|halign:" + repr(c.alignment.horizontal).encode())
    return h.hexdigest()


@pytest.fixture()
def polished_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "input.xlsx"
    write_workbook(_build_minimal_typed(), workbook_path)
    polish_workbook(workbook_path)
    return workbook_path


def test_every_sheet_frozen_at_a2(polished_workbook: Path) -> None:
    wb = load_workbook(polished_workbook)
    for sn in wb.sheetnames:
        assert wb[sn].freeze_panes == "A2", (
            f"{sn}: freeze_panes={wb[sn].freeze_panes!r}, expected 'A2'"
        )


def test_every_column_has_explicit_width(polished_workbook: Path) -> None:
    wb = load_workbook(polished_workbook)
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Every visible column must have an explicit width within the
        # AutoFit clamps; "default" widths (None) are rejected.
        from openpyxl.utils import get_column_letter
        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            dim = ws.column_dimensions.get(letter)
            assert dim is not None and dim.width is not None, (
                f"{sn}: column {letter} has no explicit width"
            )
            assert COL_WIDTH_MIN <= float(dim.width) <= COL_WIDTH_MAX, (
                f"{sn}: column {letter} width "
                f"{dim.width} outside [{COL_WIDTH_MIN}, {COL_WIDTH_MAX}]"
            )


def test_header_row_navy_white_bold(polished_workbook: Path) -> None:
    wb = load_workbook(polished_workbook)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for cell in ws[1]:
            if cell.value is None:
                continue
            fill_rgb = _normalise_rgb(
                getattr(cell.fill.fgColor, "rgb", None)
                or getattr(cell.fill.start_color, "rgb", None)
            )
            assert fill_rgb == HEADER_FILL_HEX, (
                f"{sn}!{cell.coordinate}: fill={fill_rgb!r}, "
                f"expected {HEADER_FILL_HEX!r}"
            )
            font_rgb = _normalise_rgb(getattr(cell.font.color, "rgb", None))
            assert font_rgb == HEADER_FONT_HEX, (
                f"{sn}!{cell.coordinate}: font color={font_rgb!r}, "
                f"expected {HEADER_FONT_HEX!r}"
            )
            assert cell.font.bold is True, (
                f"{sn}!{cell.coordinate}: header not bold"
            )


def test_per_asset_injection_headers_are_centered(
    polished_workbook: Path,
) -> None:
    """The polish center-aligns row 1 of the per-asset max-injection
    sheets (the combined sheet is already centered by the pandas header
    style at write time, so the trio reads consistently)."""
    wb = load_workbook(polished_workbook)
    for sn in ("max_injection_profile_pv", "max_injection_profile_bess"):
        ws = wb[sn]
        for cell in ws[1]:
            if cell.value is None:
                continue
            assert cell.alignment.horizontal == "center", (
                f"{sn}!{cell.coordinate}: header alignment "
                f"{cell.alignment.horizontal!r}, expected 'center'"
            )


def test_polish_is_idempotent(tmp_path: Path) -> None:
    workbook_path = tmp_path / "input.xlsx"
    write_workbook(_build_minimal_typed(), workbook_path)
    polish_workbook(workbook_path)
    first_hash = _sheet_state_hash(workbook_path)
    polish_workbook(workbook_path)
    second_hash = _sheet_state_hash(workbook_path)
    assert first_hash == second_hash, (
        "polish_workbook is not idempotent: sheet-state hash changed "
        "across two consecutive runs."
    )


def test_script_runs_standalone_without_install(tmp_path: Path) -> None:
    """The README-documented command ``python scripts/polish_input_workbook.py``
    must run from a checkout where the package is not pip-installed: the
    script puts the repo root on ``sys.path`` itself (the script's own
    directory, not the repo root, is ``sys.path[0]`` for a file
    invocation).  Regression for the missing bootstrap that raised
    ``ModuleNotFoundError: No module named 'pvbess_opt'``."""
    import os
    import subprocess
    import sys

    repo_root = Path(__file__).resolve().parent.parent
    workbook_path = tmp_path / "input.xlsx"
    write_workbook(_build_minimal_typed(), workbook_path)

    env = dict(os.environ)
    # Drop PYTHONPATH so the subprocess cannot inherit the repo root that
    # way — the script's own bootstrap must do the work.  cwd is tmp_path
    # (not the repo root) for the same reason.
    env.pop("PYTHONPATH", None)
    result = subprocess.run(
        [sys.executable, str(repo_root / "scripts" / "polish_input_workbook.py"),
         str(workbook_path)],
        capture_output=True, text=True, cwd=str(tmp_path), env=env,
    )
    assert result.returncode == 0, (
        "standalone `python scripts/polish_input_workbook.py` failed:\n"
        + result.stderr
    )


def test_parameter_sheets_match_row_templates() -> None:
    """Every key/value sheet with a canonical row template must be in the
    polish rebuild list — else a migrated (old) workbook silently never
    gains a newly added sheet.
    """
    from pvbess_opt.io import _SHEET_ROW_TEMPLATES
    from scripts.polish_input_workbook import _PARAMETER_SHEETS

    assert set(_PARAMETER_SHEETS) == set(_SHEET_ROW_TEMPLATES)
