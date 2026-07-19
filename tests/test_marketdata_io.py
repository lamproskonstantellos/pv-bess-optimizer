"""market_data sheet: loader wiring, bypass semantics, provenance, modes.

Uses the shipped workbook as the base (its 2026 15-minute grid is the
canonical full non-leap year) with the ``market_data`` cells edited via
openpyxl — kv sheets must never round-trip through pandas (the value
column's numeric zeros would come back as booleans).  No live network:
the ENTSO-E HTTP call is mocked or pre-seeded through the cache.
"""

from __future__ import annotations

import logging
import shutil
from datetime import UTC, datetime
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from pvbess_opt.io import (
    MARKET_DATA_SHEET_DEFAULTS,
    read_inputs,
    read_workbook,
)
from pvbess_opt.io_read import dump_structured_config, load_structured_config
from pvbess_opt.marketdata import (
    MarketDataCache,
    MarketDataUnavailableError,
    MarketSeries,
    PriceSegment,
    materialize_bypassed_workbook,
)
from pvbess_opt.marketdata import base as md_base
from pvbess_opt.marketdata import entsoe as entsoe_mod
from pvbess_opt.scenarios import validate_scenario_overrides

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "inputs" / "input.xlsx"

_CIM_NS = "urn:iec62325.351:tc57wg16:451-3:publicationdocument:7:3"
_TOKEN = "abcdef1234567890TOKEN"


@pytest.fixture(autouse=True)
def _clear_fetch_memo():
    """The process-level fetch memo must not leak between tests."""
    md_base._FETCH_MEMO.clear()
    yield
    md_base._FETCH_MEMO.clear()


def _year_window_xml(price: float = 150.0) -> bytes:
    """One PT60M document covering the padded GR 2025 fetch window."""
    return (
        f'<?xml version="1.0"?><Publication_MarketDocument xmlns="{_CIM_NS}">'
        "<TimeSeries>"
        "<currency_Unit.name>EUR</currency_Unit.name>"
        "<price_Measure_Unit.name>MWH</price_Measure_Unit.name>"
        "<Period><timeInterval><start>2024-12-31T21:00Z</start>"
        "<end>2026-01-01T00:00Z</end></timeInterval>"
        "<resolution>PT60M</resolution>"
        f"<Point><position>1</position><price.amount>{price}</price.amount>"
        "</Point></Period></TimeSeries></Publication_MarketDocument>"
    ).encode()


def _install_fake_get(monkeypatch, price: float = 150.0, counter=None):
    def fake_get(params, timeout):
        assert timeout is not None
        if counter is not None:
            counter["n"] += 1
        return 200, _year_window_xml(price)

    monkeypatch.setattr(entsoe_mod, "_http_get", fake_get)


def _forbid_network(monkeypatch):
    def no_network(params, timeout):
        raise AssertionError("this test must not hit the network")

    monkeypatch.setattr(entsoe_mod, "_http_get", no_network)


def _workbook_with_market_cells(tmp_path: Path, **cells: object) -> Path:
    """Copy the shipped workbook and set market_data values by key."""
    dst = tmp_path / "input_market.xlsx"
    shutil.copy(WORKBOOK, dst)
    wb = load_workbook(dst)
    ws = wb["market_data"]
    by_key = {
        str(row[0].value).strip(): row[1]
        for row in ws.iter_rows(min_row=2, max_col=2)
        if isinstance(row[0].value, str)
    }
    for key, value in cells.items():
        by_key[key].value = value
    wb.save(dst)
    return dst


# ---------------------------------------------------------------------------
# Bit-identity with defaults
# ---------------------------------------------------------------------------


def test_sheet_absent_matches_defaults(tmp_path):
    stripped = tmp_path / "input_no_sheet.xlsx"
    shutil.copy(WORKBOOK, stripped)
    wb = load_workbook(stripped)
    del wb["market_data"]
    wb.save(stripped)

    typed_without = read_workbook(stripped)
    typed_with = read_workbook(WORKBOOK)
    assert typed_without["market_data"] == MARKET_DATA_SHEET_DEFAULTS
    assert typed_without["market_data"] == typed_with["market_data"]
    pd.testing.assert_frame_equal(typed_without["ts"], typed_with["ts"])


def test_all_file_sources_are_inert(caplog):
    with caplog.at_level(logging.INFO):
        typed = read_workbook(WORKBOOK)
    assert "market_provenance" not in typed
    assert not [r for r in caplog.records if "[marketdata]" in r.message]


# ---------------------------------------------------------------------------
# Bypass semantics
# ---------------------------------------------------------------------------


def test_entsoe_bypass_replaces_workbook_dam_column(
    monkeypatch, tmp_path, caplog,
):
    _install_fake_get(monkeypatch, price=150.0)
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        entsoe_token=_TOKEN,
        market_cache_dir=str(tmp_path / "cache"),
    )
    original = pd.read_excel(WORKBOOK, sheet_name="timeseries",
                             usecols=["dam_price_eur_per_mwh"])
    with caplog.at_level(logging.INFO):
        params, ts = read_inputs(src)

    # Total override: every step carries the fetched price even though
    # the workbook column had (different) values.
    assert (ts["dam_price_eur_per_mwh"] == 150.0).all()
    assert not (original["dam_price_eur_per_mwh"] == 150.0).all()

    records = params["market_provenance"]
    assert len(records) == 1
    record = records[0]
    assert record["column"] == "dam_price_eur_per_mwh"
    assert record["bidding_zone"] == "GR"
    assert record["reference_year"] == 2025
    assert record["workbook_column_overridden"] is True
    assert record["cache_key"]

    consolidated = [
        r.message for r in caplog.records
        if "bypassing workbook prices" in r.message
    ]
    assert consolidated and "dam_price_eur_per_mwh" in consolidated[0]
    assert _TOKEN not in "".join(r.getMessage() for r in caplog.records)
    # The raw fetch landed in the on-disk cache.
    assert list((tmp_path / "cache").glob("dam-a44_gr_2025_*.json"))


def test_repeated_reads_fetch_once(monkeypatch, tmp_path):
    counter = {"n": 0}
    _install_fake_get(monkeypatch, counter=counter)
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        entsoe_token=_TOKEN,
        market_cache_dir=str(tmp_path / "cache"),
    )
    read_workbook(src)
    read_workbook(src)  # economics re-read pattern inside one run
    assert counter["n"] == 1


def test_offline_cache_miss_is_a_hard_error(monkeypatch, tmp_path):
    _forbid_network(monkeypatch)
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        market_cache_dir=str(tmp_path / "empty_cache"),
        market_fetch_mode="offline",
    )
    with pytest.raises(ValueError, match="offline"):
        read_workbook(src)


def test_offline_cache_hit_needs_no_token(monkeypatch, tmp_path):
    _forbid_network(monkeypatch)
    monkeypatch.delenv("ENTSOE_API_TOKEN", raising=False)
    cache = MarketDataCache(tmp_path / "cache")
    cache.save(
        MarketDataCache.key("dam-a44", "GR", 2025),
        MarketSeries(
            segments=[PriceSegment(
                datetime(2024, 12, 31, 21, tzinfo=UTC), 60, [88.0] * 8763,
            )],
            metadata={"fetched_at": "2026-07-17T00:00:00+00:00"},
        ),
    )
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        market_cache_dir=str(tmp_path / "cache"),
        market_fetch_mode="offline",
    )
    _params, ts = read_inputs(src)
    assert (ts["dam_price_eur_per_mwh"] == 88.0).all()


def test_partial_year_grid_rejected(monkeypatch, tmp_path):
    _install_fake_get(monkeypatch)
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        entsoe_token=_TOKEN,
        market_cache_dir=str(tmp_path / "cache"),
    )
    wb = load_workbook(src)
    ws = wb["timeseries"]
    ws.delete_rows(98, ws.max_row)  # keep the header + one day
    wb.save(src)
    with pytest.raises(ValueError, match="full non-leap model year"):
        read_workbook(src)


# ---------------------------------------------------------------------------
# Token policy
# ---------------------------------------------------------------------------


def test_workbook_token_wins_over_environment(monkeypatch, tmp_path):
    seen: dict[str, str] = {}

    def responder(params, timeout):
        seen.update(params)
        return 200, _year_window_xml()

    monkeypatch.setattr(entsoe_mod, "_http_get", responder)
    monkeypatch.setenv("ENTSOE_API_TOKEN", "env-token-should-lose")
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        entsoe_token=_TOKEN,
        market_cache_dir=str(tmp_path / "cache"),
    )
    read_workbook(src)
    assert seen["securityToken"] == _TOKEN


def test_env_token_used_when_workbook_blank(monkeypatch, tmp_path):
    seen: dict[str, str] = {}

    def responder(params, timeout):
        seen.update(params)
        return 200, _year_window_xml()

    monkeypatch.setattr(entsoe_mod, "_http_get", responder)
    monkeypatch.setenv("ENTSOE_API_TOKEN", "env-token-wins-here")
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        market_cache_dir=str(tmp_path / "cache"),
    )
    read_workbook(src)
    assert seen["securityToken"] == "env-token-wins-here"


def test_missing_token_error_names_the_env_var(monkeypatch, tmp_path):
    _forbid_network(monkeypatch)
    monkeypatch.delenv("ENTSOE_API_TOKEN", raising=False)
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        market_cache_dir=str(tmp_path / "empty"),
    )
    with pytest.raises(ValueError, match="ENTSOE_API_TOKEN"):
        read_workbook(src)


# ---------------------------------------------------------------------------
# Surfaces: scenarios + YAML
# ---------------------------------------------------------------------------


def test_scenario_overrides_accept_market_data_targets():
    validate_scenario_overrides({
        "name": "fetched prices",
        "market_data": {"price_source": "entsoe", "bidding_zone": "de_lu"},
    })
    with pytest.raises(ValueError, match="unknown"):
        validate_scenario_overrides({
            "name": "typo",
            "market_data": {"price_sourc": "entsoe"},
        })


def test_yaml_roundtrip_carries_market_data(tmp_path):
    typed = read_workbook(WORKBOOK)
    config_path = tmp_path / "config.yaml"
    dump_structured_config(typed, config_path)
    loaded = load_structured_config(config_path)
    assert loaded["market_data"] == MARKET_DATA_SHEET_DEFAULTS


# ---------------------------------------------------------------------------
# Materialised snapshot (offline reproducibility)
# ---------------------------------------------------------------------------


def test_materialized_snapshot_reruns_offline(monkeypatch, tmp_path):
    _install_fake_get(monkeypatch, price=150.0)
    src = _workbook_with_market_cells(
        tmp_path,
        price_source="entsoe",
        entsoe_token=_TOKEN,
        market_cache_dir=str(tmp_path / "cache"),
    )
    params, ts = read_inputs(src)

    snapshot = tmp_path / "input_snapshot.xlsx"
    shutil.copy(src, snapshot)
    materialize_bypassed_workbook(
        snapshot, ts, params["market_provenance"],
    )

    # Re-read the snapshot with the network forbidden, an empty cache
    # and no memo: the fetched values must come from the workbook itself
    # and the source key must be back at 'file'.
    md_base._FETCH_MEMO.clear()
    _forbid_network(monkeypatch)
    typed = read_workbook(snapshot)
    assert typed["market_data"]["price_source"] == "file"
    # The snapshot re-runs with 'file' sources, so the token cell is
    # blanked — results directories must never carry a live secret.
    assert typed["market_data"]["entsoe_token"] == ""
    assert "market_provenance" not in typed
    np.testing.assert_allclose(
        typed["ts"]["dam_price_eur_per_mwh"].to_numpy(dtype=float), 150.0,
    )


# ---------------------------------------------------------------------------
# Balancing / imbalance bypass (ADMIE + ENTSO-E + registry)
# ---------------------------------------------------------------------------

from pvbess_opt.marketdata import admie as admie_mod  # noqa: E402
from tests._marketdata_helpers import (  # noqa: E402
    ADMIE_IMBALANCE_HEADERS,
    balancing_year_responder,
    install_admie_year_fetch,
)


def test_gr_balancing_via_admie_replaces_eight_columns(
    monkeypatch, tmp_path, caplog,
):
    install_admie_year_fetch(monkeypatch, value=10.0)
    src = _workbook_with_market_cells(
        tmp_path,
        balancing_source="admie",
        market_cache_dir=str(tmp_path / "cache"),
    )
    original_fcr = pd.read_excel(
        WORKBOOK, sheet_name="timeseries",
        usecols=["fcr_capacity_price_eur_per_mwh"],
    )["fcr_capacity_price_eur_per_mwh"].to_numpy(dtype=float)
    with caplog.at_level(logging.INFO):
        params, ts = read_inputs(src)

    from pvbess_opt.marketdata.admie import BALANCING_HEADER_PATTERNS

    # The 30-min synthetic prices step-hold onto the 15-min grid.
    for column in BALANCING_HEADER_PATTERNS:
        assert (ts[column] == 10.0).all(), column
    # FCR is NOT served by ADMIE (no standalone Greek FCR): the column
    # keeps its workbook values and a WARNING says so.
    np.testing.assert_allclose(
        ts["fcr_capacity_price_eur_per_mwh"].to_numpy(dtype=float),
        original_fcr,
    )
    assert any(
        "no standalone FCR" in r.message for r in caplog.records
    )
    records = params["market_provenance"]
    assert len(records) == 8
    assert {r["source"] for r in records} == {"admie"}
    assert {r["source_key"] for r in records} == {"balancing_source"}
    assert all(r["workbook_column_overridden"] for r in records)


def test_gr_auto_routes_balancing_to_admie(monkeypatch, tmp_path):
    install_admie_year_fetch(monkeypatch, value=12.5)
    src = _workbook_with_market_cells(
        tmp_path,
        balancing_source="auto",
        market_cache_dir=str(tmp_path / "cache"),
    )
    _params, ts = read_inputs(src)
    assert (ts["afrr_up_capacity_price_eur_per_mwh"] == 12.5).all()


def test_gr_imbalance_via_admie(monkeypatch, tmp_path):
    install_admie_year_fetch(
        monkeypatch, value=33.0, headers=ADMIE_IMBALANCE_HEADERS,
        n_periods=96,
    )
    src = _workbook_with_market_cells(
        tmp_path,
        imbalance_source="admie",
        market_cache_dir=str(tmp_path / "cache"),
    )
    params, ts = read_inputs(src)
    assert (ts["imbalance_price_eur_per_mwh"] == 33.0).all()
    [record] = params["market_provenance"]
    assert record["source_key"] == "imbalance_source"
    assert record["workbook_column_overridden"] is False


def test_de_lu_balancing_via_entsoe(monkeypatch, tmp_path):
    def fake_get(params, timeout):
        assert timeout is not None
        return balancing_year_responder(params)

    monkeypatch.setattr(entsoe_mod, "_http_get", fake_get)
    src = _workbook_with_market_cells(
        tmp_path,
        bidding_zone="de_lu",
        balancing_source="entsoe",
        entsoe_token=_TOKEN,
        market_cache_dir=str(tmp_path / "cache"),
    )
    _params, ts = read_inputs(src)
    assert (ts["fcr_capacity_price_eur_per_mwh"] == 5.0).all()
    assert (ts["afrr_dn_capacity_price_eur_per_mwh"] == 7.0).all()
    assert (ts["mfrr_up_activation_price_eur_per_mwh"] == 90.0).all()


def test_de_lu_imbalance_via_entsoe_writes_dual_columns(
    monkeypatch, tmp_path,
):
    def fake_get(params, timeout):
        assert timeout is not None
        return balancing_year_responder(params)

    monkeypatch.setattr(entsoe_mod, "_http_get", fake_get)
    src = _workbook_with_market_cells(
        tmp_path,
        bidding_zone="de_lu",
        imbalance_source="entsoe",
        entsoe_token=_TOKEN,
        market_cache_dir=str(tmp_path / "cache"),
    )
    _params, ts = read_inputs(src)
    assert (ts["imbalance_price_long_eur_per_mwh"] == 30.0).all()
    assert (ts["imbalance_price_short_eur_per_mwh"] == 30.0).all()


def test_registry_rejects_entsoe_balancing_for_greece(tmp_path):
    src = _workbook_with_market_cells(
        tmp_path, balancing_source="entsoe",
    )
    with pytest.raises(
        MarketDataUnavailableError, match="Integrated Scheduling",
    ):
        read_workbook(src)


def test_registry_rejects_admie_outside_greece(tmp_path):
    src = _workbook_with_market_cells(
        tmp_path, bidding_zone="de_lu", balancing_source="admie",
    )
    with pytest.raises(MarketDataUnavailableError, match="Greek TSO"):
        read_workbook(src)


def test_snapshot_flips_balancing_source_too(monkeypatch, tmp_path):
    install_admie_year_fetch(monkeypatch, value=10.0)
    src = _workbook_with_market_cells(
        tmp_path,
        balancing_source="admie",
        market_cache_dir=str(tmp_path / "cache"),
    )
    params, ts = read_inputs(src)
    snapshot = tmp_path / "snap.xlsx"
    shutil.copy(src, snapshot)
    materialize_bypassed_workbook(
        snapshot, ts, params["market_provenance"],
    )
    md_base._FETCH_MEMO.clear()
    _forbid_network(monkeypatch)
    def no_admie(url, params, timeout):
        raise AssertionError("snapshot must not hit ADMIE")

    monkeypatch.setattr(admie_mod, "_http_get", no_admie)
    typed = read_workbook(snapshot)
    assert typed["market_data"]["balancing_source"] == "file"
    assert (
        typed["ts"]["afrr_up_capacity_price_eur_per_mwh"] == 10.0
    ).all()
