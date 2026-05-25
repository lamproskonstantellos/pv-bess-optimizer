"""Workbook-loader tests for the balancing market schema."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from pvbess_opt.io import (
    BALANCING_SHEET_DEFAULTS,
    read_workbook,
    write_workbook,
)


def _load_typed(repo_input_xlsx: Path) -> dict:
    return read_workbook(repo_input_xlsx)


def test_balancing_sheet_absent_falls_back_to_defaults(repo_input_xlsx):
    typed = _load_typed(repo_input_xlsx)
    assert "balancing" in typed
    # The shipped workbook does not carry the balancing sheet yet, so
    # the loader populates the section from BALANCING_SHEET_DEFAULTS.
    assert typed["balancing"]["balancing_enabled"] is False
    for key, default in BALANCING_SHEET_DEFAULTS.items():
        assert typed["balancing"][key] == default


def test_balancing_sheet_round_trips_through_writer(tmp_path, repo_input_xlsx):
    typed = _load_typed(repo_input_xlsx)
    typed["balancing"]["balancing_enabled"] = True
    dst = tmp_path / "with_balancing.xlsx"
    write_workbook(typed, dst)
    sheets = set(pd.ExcelFile(dst).sheet_names)
    assert "balancing" in sheets
    typed2 = read_workbook(dst)
    for key, value in typed["balancing"].items():
        assert typed2["balancing"][key] == value


def test_share_sum_exceeding_100_pct_raises(tmp_path, repo_input_xlsx):
    typed = _load_typed(repo_input_xlsx)
    typed["balancing"]["balancing_enabled"] = True
    typed["balancing"]["dam_capacity_share_pct"] = 80.0
    typed["balancing"]["fcr_capacity_share_pct"] = 30.0
    dst = tmp_path / "bad_share.xlsx"
    write_workbook(typed, dst)
    with pytest.raises(ValueError, match="capacity shares sum"):
        read_workbook(dst)


def test_probability_out_of_range_raises(tmp_path, repo_input_xlsx):
    typed = _load_typed(repo_input_xlsx)
    typed["balancing"]["balancing_enabled"] = True
    typed["balancing"]["fcr_bid_acceptance_pct"] = 150.0
    dst = tmp_path / "bad_prob.xlsx"
    write_workbook(typed, dst)
    with pytest.raises(ValueError, match="fcr_bid_acceptance_pct"):
        read_workbook(dst)


def test_settlement_minutes_mismatch_raises(tmp_path, repo_input_xlsx):
    typed = _load_typed(repo_input_xlsx)
    typed["balancing"]["balancing_enabled"] = True
    typed["balancing"]["bm_settlement_minutes"] = 30  # workbook is 15 min
    dst = tmp_path / "bad_settle.xlsx"
    write_workbook(typed, dst)
    with pytest.raises(ValueError, match="bm_settlement_minutes"):
        read_workbook(dst)


def test_soc_headroom_out_of_range_raises(tmp_path, repo_input_xlsx):
    typed = _load_typed(repo_input_xlsx)
    typed["balancing"]["balancing_enabled"] = True
    typed["balancing"]["bm_soc_headroom_pct"] = 80.0
    dst = tmp_path / "bad_headroom.xlsx"
    write_workbook(typed, dst)
    with pytest.raises(ValueError, match="bm_soc_headroom_pct"):
        read_workbook(dst)


def test_validation_skipped_when_balancing_disabled(tmp_path, repo_input_xlsx):
    typed = _load_typed(repo_input_xlsx)
    # Set an obviously bad value but keep balancing_enabled=False.
    typed["balancing"]["balancing_enabled"] = False
    typed["balancing"]["fcr_bid_acceptance_pct"] = 300.0
    dst = tmp_path / "disabled_bad.xlsx"
    write_workbook(typed, dst)
    # Should not raise — validation is gated on the master switch.
    typed2 = read_workbook(dst)
    assert typed2["balancing"]["balancing_enabled"] is False


def test_missing_timeseries_column_falls_back_to_default(
    tmp_path, repo_input_xlsx, caplog,
):
    typed = _load_typed(repo_input_xlsx)
    typed["balancing"]["balancing_enabled"] = True
    dst = tmp_path / "missing_col.xlsx"
    write_workbook(typed, dst)
    # Loader must add the missing column with the scalar default and
    # emit a warning naming it.
    with caplog.at_level(logging.WARNING):
        loaded = read_workbook(dst)
    assert "fcr_capacity_price_eur_per_mwh" in loaded["ts"].columns
    expected = float(
        BALANCING_SHEET_DEFAULTS["fcr_default_capacity_price_eur_per_mwh"]
    )
    assert (loaded["ts"]["fcr_capacity_price_eur_per_mwh"] == expected).all()
    assert any(
        "fcr_capacity_price_eur_per_mwh" in record.message
        for record in caplog.records
    )


def test_no_warning_when_balancing_disabled(tmp_path, repo_input_xlsx, caplog):
    typed = _load_typed(repo_input_xlsx)
    typed["balancing"]["balancing_enabled"] = False
    dst = tmp_path / "disabled_no_warn.xlsx"
    write_workbook(typed, dst)
    with caplog.at_level(logging.WARNING):
        loaded = read_workbook(dst)
    # No fallback warning when the gate is off.
    assert not any(
        "balancing timeseries column" in record.message
        for record in caplog.records
    )
    # And no balancing price columns get injected when disabled.
    assert "fcr_capacity_price_eur_per_mwh" not in loaded["ts"].columns


def test_balancing_sheet_workbook_writes_amber_highlight(tmp_path):
    """``write_workbook`` writes the balancing sheet alongside the rest;
    the highlight is applied by the reference-workbook builder script,
    not the writer, so here we only assert the sheet is present."""
    typed = read_workbook(Path(__file__).resolve().parent.parent / "inputs" / "input.xlsx")
    dst = tmp_path / "out.xlsx"
    write_workbook(typed, dst)
    wb = openpyxl.load_workbook(dst)
    assert "balancing" in wb.sheetnames
    sheet = wb["balancing"]
    # First column is "key", header at row 1.
    assert sheet.cell(1, 1).value == "key"
    keys_in_sheet = {
        sheet.cell(r, 1).value for r in range(2, sheet.max_row + 1)
        if sheet.cell(r, 1).value
    }
    for key in BALANCING_SHEET_DEFAULTS:
        assert key in keys_in_sheet
