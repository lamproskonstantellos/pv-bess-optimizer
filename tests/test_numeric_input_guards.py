"""Boolean-in-numeric-field guards on every input surface.

``float(True)`` coerces silently to ``1.0`` (and ``int(True)`` to ``1``),
so a stray TRUE typed into a numeric workbook field changes results
without a trace — the motivating incident was ``unavailability_pct =
TRUE`` silently becoming a 1 % availability derate.  Three locked
properties:

1. A genuine Excel boolean cell in a numeric field is rejected loudly,
   naming the sheet and key.
2. Genuinely NUMERIC 0/1 cells keep loading: the workbook kv sheets are
   read with openpyxl-faithful cell types (``_read_kv_flat``), immune to
   the pandas quirk where ``pd.read_excel`` mis-surfaces a numeric 0/1
   cell as a Python bool in a mixed-type value column (the shipped
   workbook exercises this: ``site_capex_eur`` is a numeric 0 that
   pandas would read as ``False``).
3. The YAML/dict path applies the same guard on its faithful native
   types via ``_parse_value``.

Also pins the related template-default contract: the aggregator fee
ships at 0 % (fee-free, opt-in) on the defaults map, the row template
and the shipped workbook.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from pvbess_opt.io import _parse_value, read_workbook

ROOT = Path(__file__).resolve().parent.parent
SHIPPED = ROOT / "inputs" / "input.xlsx"


def _copy_with_bool(tmp_path: Path, sheet: str, key: str, value: bool) -> Path:
    """Copy the shipped workbook and type a genuine boolean into ``key``."""
    target = tmp_path / f"bool_{sheet}_{key}.xlsx"
    shutil.copy(SHIPPED, target)
    wb = load_workbook(target)
    ws = wb[sheet]
    hit = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == key:
            row[1].value = value
            hit = True
    assert hit, f"{key!r} not found on sheet {sheet!r}"
    wb.save(target)
    return target


@pytest.mark.parametrize(("sheet", "key", "value"), [
    ("project", "unavailability_pct", True),   # the motivating incident
    ("ppa", "ppa_inflation_pct", False),       # its sibling
    ("economics", "discount_rate_pct", True),
])
def test_boolean_cell_in_numeric_field_is_rejected(tmp_path, sheet, key, value):
    path = _copy_with_bool(tmp_path, sheet, key, value)
    with pytest.raises(ValueError, match=key):
        read_workbook(path)


def test_error_message_names_sheet_and_suggests_fix(tmp_path):
    path = _copy_with_bool(tmp_path, "project", "unavailability_pct", True)
    with pytest.raises(ValueError, match=r"'project'.*expects a number"):
        read_workbook(path)


def test_boolean_in_bool_key_still_accepted(tmp_path):
    """Genuine boolean knobs (``_BOOL_KEYS``) keep accepting TRUE/FALSE."""
    path = _copy_with_bool(tmp_path, "project", "allow_bess_grid_charging", True)
    typed = read_workbook(path)
    assert typed["project"]["allow_bess_grid_charging"] is True


def test_shipped_workbook_loads_despite_pandas_bool_inference():
    """The shipped case study loads: its numeric 0 cells (e.g.
    ``site_capex_eur``, which ``pd.read_excel`` would surface as
    ``False``) are read with faithful openpyxl types and stay numeric."""
    typed = read_workbook(SHIPPED)
    assert typed["project"]["site_capex_eur"] == 0.0
    assert typed["project"]["site_devex_eur"] == 0.0


def test_parse_value_rejects_python_and_numpy_bools():
    for raw in (True, False, np.bool_(True), np.bool_(False)):
        with pytest.raises(ValueError, match="expects a number"):
            _parse_value("discount_rate_pct", raw, 7.0)


def test_parse_value_rejects_bool_for_int_keys():
    with pytest.raises(ValueError, match="expects a number"):
        _parse_value("project_lifecycle_years", True, 20)


def test_parse_value_accepts_numbers_unchanged():
    assert _parse_value("discount_rate_pct", 7.5, 7.0) == 7.5
    assert _parse_value("project_lifecycle_years", 25, 20) == 25


def test_yaml_config_rejects_boolean_for_numeric_key(tmp_path):
    """The YAML surface routes through the same guard with faithful types."""
    from pvbess_opt.io_read import load_structured_config

    cfg = tmp_path / "cfg.yaml"
    cfg.write_text(
        "economics:\n"
        "  discount_rate_pct: true\n"
    )
    with pytest.raises(ValueError, match="discount_rate_pct"):
        load_structured_config(cfg)


def test_template_default_aggregator_fee_is_zero():
    """The template no longer pre-fills a 10 % revenue fee (opt-in)."""
    from pvbess_opt.io import _ECONOMICS_ROWS, ECONOMICS_SHEET_DEFAULTS

    assert ECONOMICS_SHEET_DEFAULTS["aggregator_fee_pct_revenue"] == 0.0
    row = next(r for r in _ECONOMICS_ROWS if r[0] == "aggregator_fee_pct_revenue")
    assert row[1] == 0.0
    # The shipped case-study workbook follows the template.
    typed = read_workbook(SHIPPED)
    assert typed["economics"]["aggregator_fee_pct_revenue"] == 0.0
