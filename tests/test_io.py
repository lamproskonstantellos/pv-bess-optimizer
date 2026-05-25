"""I/O loader, schema validation, and output workbook tests."""

from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from pvbess_opt.io import (
    BESS_SHEET_DEFAULTS,
    ECONOMICS_SHEET_DEFAULTS,
    PROJECT_SHEET_DEFAULTS,
    PV_SHEET_DEFAULTS,
    SIMULATION_SHEET_DEFAULTS,
    _flat_dict_from_sheet,
    _parse_bool,
    _parse_kv_sheet,
    detect_timestep_minutes,
    read_inputs,
    read_workbook,
    write_workbook,
)


def _minimal_typed(year: int = 2026) -> dict:
    n = 24
    ts = pd.DataFrame({
        "timestamp": pd.date_range(f"{year}-01-01", periods=n, freq="h"),
        "pv_kwh": [100.0] * n,
        "load_kwh": [50.0] * n,
        "dam_price_eur_per_mwh": [80.0] * n,
    })
    return {
        "ts": ts,
        "project": dict(PROJECT_SHEET_DEFAULTS),
        "pv": dict(PV_SHEET_DEFAULTS, pv_nameplate_kwp=1000.0),
        "bess": dict(
            BESS_SHEET_DEFAULTS, bess_power_kw=500.0, bess_capacity_kwh=2000.0,
        ),
        "economics": dict(ECONOMICS_SHEET_DEFAULTS),
        "simulation": dict(SIMULATION_SHEET_DEFAULTS),
        # max-injection semantic: 73 % allowed to inject.
        "max_injection_profile": np.full(24, 73.0, dtype=float),
    }


def test_sheet_defaults_have_expected_top_level_keys():
    assert "project_lifecycle_years" in PROJECT_SHEET_DEFAULTS
    assert "pv_nameplate_kwp" in PV_SHEET_DEFAULTS
    assert "bess_power_kw" in BESS_SHEET_DEFAULTS
    assert "discount_rate_pct" in ECONOMICS_SHEET_DEFAULTS
    assert "uncertainty_enabled" in SIMULATION_SHEET_DEFAULTS


def test_parse_bool_accepts_canonical_tokens():
    assert _parse_bool("TRUE", False) is True
    assert _parse_bool("false", True) is False
    assert _parse_bool(1, False) is True
    assert _parse_bool(0, True) is False
    assert _parse_bool(None, True) is True
    assert _parse_bool("", True) is True


def test_flat_dict_skips_separator_rows():
    df = pd.DataFrame({
        "key": ["# group", "efficiency_charge", "", "bess_power_kw"],
        "value": [None, 0.95, None, 5000],
    })
    flat = _flat_dict_from_sheet(df)
    assert flat == {"efficiency_charge": 0.95, "bess_power_kw": 5000}


def test_parse_kv_sheet_routes_keys_to_correct_sheet():
    parsed = _parse_kv_sheet("bess", {
        "efficiency_charge": 0.95, "bess_power_kw": 5000.0,
    })
    assert parsed["efficiency_charge"] == 0.95
    assert parsed["bess_power_kw"] == 5000.0


def test_detect_timestep_minutes_hourly():
    ts = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=24, freq="h"),
    })
    assert detect_timestep_minutes(ts) == 60


def test_detect_timestep_irregular_raises():
    ts = pd.DataFrame({
        "timestamp": pd.to_datetime([
            "2026-01-01 00:00", "2026-01-01 01:00", "2026-01-01 03:00",
        ]),
    })
    with pytest.raises(ValueError, match="Irregular timestep"):
        detect_timestep_minutes(ts)


def test_read_workbook_repo_default(repo_input_xlsx):
    typed = read_workbook(repo_input_xlsx)
    for section in ("project", "pv", "bess", "economics", "simulation"):
        assert section in typed
    assert typed["dt_minutes"] == 15
    assert typed["project"]["mode"] == "self_consumption"


def test_read_inputs_returns_flat_params(repo_input_xlsx):
    params, ts = read_inputs(repo_input_xlsx)
    assert "efficiency_charge" in params and "bess_power_kw" in params
    assert params["mode"] == "self_consumption"
    assert "load_kwh" in ts.columns
    assert len(ts) == 35040


def test_workbook_round_trip(tmp_path, repo_input_xlsx):
    typed = read_workbook(repo_input_xlsx)
    dst = tmp_path / "round_trip.xlsx"
    write_workbook(typed, dst)
    typed2 = read_workbook(dst)
    eta1 = typed["bess"]["efficiency_charge"]
    eta2 = typed2["bess"]["efficiency_charge"]
    assert eta1 == eta2
    assert (
        typed["economics"]["discount_rate_pct"]
        == typed2["economics"]["discount_rate_pct"]
    )
    assert len(typed["ts"]) == len(typed2["ts"])


def test_self_consumption_requires_load_column(tmp_path, repo_input_xlsx):
    """self_consumption mode requires load_kwh in timeseries."""
    typed = read_workbook(repo_input_xlsx)
    typed["ts"] = typed["ts"].drop(columns=["load_kwh"])
    dst = tmp_path / "no_load.xlsx"
    write_workbook(typed, dst)
    with pytest.raises(ValueError, match="load_kwh"):
        read_workbook(dst)


def test_merchant_mode_load_optional(tmp_path, repo_input_xlsx):
    """merchant mode: load_kwh optional; if present, an info log is emitted."""
    typed = read_workbook(repo_input_xlsx)
    typed["project"]["mode"] = "merchant"
    dst = tmp_path / "merchant.xlsx"
    write_workbook(typed, dst)
    typed2 = read_workbook(dst)
    assert typed2["project"]["mode"] == "merchant"


def test_unknown_mode_raises(tmp_path):
    """An unrecognized mode value raises ValueError; no silent fallback."""
    typed = _minimal_typed()
    typed["project"]["mode"] = "vnb"
    dst = tmp_path / "test.xlsx"
    write_workbook(typed, dst)
    with pytest.raises(ValueError, match="unknown mode"):
        read_workbook(dst)


def test_write_workbook_emits_all_sheets(tmp_path, repo_input_xlsx):
    typed = read_workbook(repo_input_xlsx)
    dst = tmp_path / "out.xlsx"
    write_workbook(typed, dst)
    assert set(pd.ExcelFile(dst).sheet_names) == {
        "timeseries", "project", "pv", "bess", "economics",
        "simulation", "balancing", "max_injection_profile",
    }


# ---------------------------------------------------------------------------
# max_injection_profile schema: typed round-trip and default fallback.
# ---------------------------------------------------------------------------


def _write_minimal_workbook_with_sheet(
    tmp_path,
    sheet_name: str,
    column_name: str,
    values: np.ndarray,
):
    """Write a minimal valid workbook whose only profile sheet is
    ``sheet_name`` with column ``column_name`` and ``values``.

    Used by the schema and default-fallback tests.  All other sheets
    come from :func:`_minimal_typed` via :func:`write_workbook`, then
    the profile sheet is replaced in-place via openpyxl.
    """
    import openpyxl

    typed = _minimal_typed()
    dst = tmp_path / "shim.xlsx"
    write_workbook(typed, dst)

    wb = openpyxl.load_workbook(dst)
    # Drop the canonical max_injection_profile sheet written by
    # write_workbook so the test workbook only has the sheet we want.
    if "max_injection_profile" in wb.sheetnames:
        del wb["max_injection_profile"]
    if sheet_name and column_name is not None:
        ws = wb.create_sheet(sheet_name)
        ws.cell(1, 1, "hour_of_day")
        ws.cell(1, 2, column_name)
        for h in range(24):
            ws.cell(h + 2, 1, h)
            ws.cell(h + 2, 2, float(values[h]))
    wb.save(dst)
    return dst


def test_loader_reads_new_schema(tmp_path):
    """max_injection_profile sheet → array returned verbatim."""
    vals = np.full(24, 73.0, dtype=float)
    vals[6:8] = 40.0  # asymmetric values to prove no 100-x conversion
    dst = _write_minimal_workbook_with_sheet(
        tmp_path, "max_injection_profile", "max_injection_pct", vals,
    )
    typed = read_workbook(dst)
    np.testing.assert_array_equal(typed["max_injection_profile"], vals)


def test_loader_falls_back_to_default_when_sheet_missing(tmp_path):
    """No profile sheet at all → flat no-curtailment default at 100.0."""
    from pvbess_opt.config import DEFAULT_MAX_INJECTION_PCT_HOURLY

    assert DEFAULT_MAX_INJECTION_PCT_HOURLY == 100.0
    dst = _write_minimal_workbook_with_sheet(tmp_path, "", None, np.array([]))
    typed = read_workbook(dst)
    expected = np.full(24, DEFAULT_MAX_INJECTION_PCT_HOURLY, dtype=float)
    np.testing.assert_array_equal(typed["max_injection_profile"], expected)


def test_project_lifecycle_years_default_is_twenty():
    """Default project_lifecycle_years is 20, matching the docs."""
    assert PROJECT_SHEET_DEFAULTS["project_lifecycle_years"] == 20
