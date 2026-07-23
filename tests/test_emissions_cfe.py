"""Grid emissions and 24/7 carbon-free-energy reporting.

Pure-math checks on the dispatch-derived report plus a HiGHS full-run that
the emissions sheet appears and is styled only when a grid carbon intensity
is configured, and that the finances are untouched either way.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pvbess_opt.emissions import (
    build_emissions_report,
    cfe_score,
    grid_ci_series,
    hourly_cfe_fraction,
)
from pvbess_opt.io_read import load_structured_config
from pvbess_opt.theme import COL_WIDTH_MAX, COL_WIDTH_MIN, HEADER_FILL_HEX

ROOT = Path(__file__).resolve().parent.parent


def _highs_available() -> bool:
    try:
        import highspy
    except ImportError:
        return False
    return bool(highspy)


def _synthetic_dispatch() -> pd.DataFrame:
    """Two steps: a well-matched solar step and a fully grid-served step."""
    return pd.DataFrame({
        "load_kwh": [100.0, 100.0],
        "pv_kwh": [80.0, 0.0],
        "pv_to_load_kwh": [60.0, 0.0],
        "pv_to_grid_kwh": [10.0, 0.0],
        "pv_curtail_kwh": [0.0, 0.0],
        "grid_to_load_kwh": [20.0, 100.0],
        "bess_charge_grid_kwh": [0.0, 50.0],
        "bess_dis_load_kwh": [20.0, 0.0],
        "bess_dis_grid_kwh": [0.0, 0.0],
        "bess_dis_load_green_kwh": [20.0, 0.0],
        "bess_dis_grid_green_kwh": [0.0, 0.0],
    })


def test_cfe_score_is_time_coincident():
    res = _synthetic_dispatch()
    # carbon-free to load = [80, 0]; load = [100, 100] -> 80 / 200.
    assert cfe_score(res) == pytest.approx(40.0)


def test_cfe_score_nan_without_load():
    res = _synthetic_dispatch()
    res["load_kwh"] = [0.0, 0.0]
    assert np.isnan(cfe_score(res))


def test_24_7_is_stricter_than_annual_volumetric():
    res = _synthetic_dispatch()
    score = cfe_score(res)
    # Naive annual match also credits exported clean energy against load.
    cf_to_load = 60.0 + 20.0
    exported_clean = 10.0
    volumetric = (cf_to_load + exported_clean) / 200.0 * 100.0
    assert score < volumetric            # 40 % < 45 %


def test_build_emissions_report_math():
    res = _synthetic_dispatch()
    rep = build_emissions_report(
        res, grid_ci_kg_per_mwh=400.0, project_years=2, start_year=2030,
    )
    assert list(rep["project_year"]) == [1, 2]
    assert list(rep["calendar_year"]) == [2030, 2031]
    r0 = rep.iloc[0]
    assert r0["cfe_score_pct"] == pytest.approx(40.0)
    assert r0["load_mwh"] == pytest.approx(0.2)
    assert r0["grid_import_mwh"] == pytest.approx(0.17)
    # clean delivered = 90 kWh, grid charge = 50, grid->load = 120, import = 170.
    assert r0["avoided_emissions_t"] == pytest.approx(0.036)
    assert r0["induced_emissions_t"] == pytest.approx(0.020)
    assert r0["net_avoided_emissions_t"] == pytest.approx(0.016)
    assert r0["residual_load_emissions_t"] == pytest.approx(0.048)
    assert r0["grid_import_emissions_t"] == pytest.approx(0.068)
    # No decline -> Year 2 identical.
    assert rep.iloc[1]["avoided_emissions_t"] == pytest.approx(0.036)


def test_grid_ci_annual_decline_scales_emissions():
    res = _synthetic_dispatch()
    rep = build_emissions_report(
        res, grid_ci_kg_per_mwh=400.0, project_years=3, start_year=2030,
        grid_ci_annual_decline_pct=10.0,
    )
    base = rep.iloc[0]["avoided_emissions_t"]
    # The report rounds tonnes to 4 decimals (degradation-report convention).
    assert rep.iloc[1]["avoided_emissions_t"] == pytest.approx(round(base * 0.9, 4))
    assert rep.iloc[2]["avoided_emissions_t"] == pytest.approx(round(base * 0.81, 4))
    # CFE score is an energy match, independent of the carbon intensity.
    assert rep["cfe_score_pct"].nunique() == 1


def test_per_step_ci_column_overrides_scalar():
    res = _synthetic_dispatch()
    res["grid_co2_kg_per_mwh"] = [400.0, 800.0]
    assert grid_ci_series(res, 123.0).tolist() == [400.0, 800.0]
    rep = build_emissions_report(
        res, grid_ci_kg_per_mwh=400.0, project_years=1, start_year=2030,
    )
    # clean delivered = [90, 0] kWh weighted by [400, 800] -> 0.036 t, while a
    # flat 400 would also give 0.036; differentiate via the grid-charge step.
    # induced = grid charge [0, 50] x [400, 800] -> 50*800/1e6 = 0.040 t
    # (flat 400 would give 0.020 t), so the per-step column clearly wins.
    assert rep.iloc[0]["induced_emissions_t"] == pytest.approx(0.040)


def test_hourly_cfe_fraction_drops_loadless_steps():
    res = _synthetic_dispatch()
    res.loc[1, "load_kwh"] = 0.0
    frac = hourly_cfe_fraction(res)
    assert frac.tolist() == [pytest.approx(0.8)]


def test_grid_block_alias(tmp_path):
    ts = pd.DataFrame({
        "timestamp": pd.date_range("2019-01-01", periods=4, freq="15min"),
        "dam_price_eur_per_mwh": [50.0] * 4,
        "pv_kwh": [0.0] * 4,
    })
    ts.to_csv(tmp_path / "ts.csv", index=False)
    cfg = tmp_path / "run.yaml"
    cfg.write_text(
        "grid:\n"
        "  co2_intensity: 350\n"
        "  co2_annual_decline: 0.02\n"
        "timeseries_path: ts.csv\n",
        encoding="utf-8",
    )
    econ = load_structured_config(cfg)["economics"]
    assert econ["grid_co2_intensity_kg_per_mwh"] == pytest.approx(350.0)
    assert econ["grid_co2_annual_decline_pct"] == pytest.approx(2.0)


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_emissions_full_run_sheet_and_finance_unchanged(tmp_path):
    from pvbess_opt import RunConfig, run
    from pvbess_opt.io import read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    base = tmp_path / "base.xlsx"
    write_workbook(typed, base)

    typed_co2 = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed_co2["ts"] = typed_co2["ts"].iloc[:96].reset_index(drop=True)
    typed_co2["economics"]["grid_co2_intensity_kg_per_mwh"] = 350.0
    with_co2 = tmp_path / "co2.xlsx"
    write_workbook(typed_co2, with_co2)

    common = dict(solver="highs", mip_gap=0.05, time_limit=180)
    r0 = run(RunConfig(excel=base, outdir=tmp_path / "a", **common))
    r1 = run(RunConfig(excel=with_co2, outdir=tmp_path / "b", **common))

    # Emissions accounting never touches the dispatch or the finances.
    assert r1.financial_kpis["npv_eur"] == pytest.approx(
        r0.financial_kpis["npv_eur"], rel=1e-9, abs=1e-6,
    )
    wb1 = load_workbook(r1.out_dir / "03_results.xlsx")
    wb0 = load_workbook(r0.out_dir / "03_results.xlsx")
    assert "emissions" in wb1.sheetnames
    assert "emissions" not in wb0.sheetnames
    ws = wb1["emissions"]
    assert ws.freeze_panes == "A2"
    for cell in ws[1]:
        if cell.value is None:
            continue
        rgb = (getattr(cell.fill.fgColor, "rgb", None) or "")
        assert rgb.upper().lstrip("0").rjust(6, "0")[-6:] == HEADER_FILL_HEX
    for c in range(1, ws.max_column + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        assert dim is not None and COL_WIDTH_MIN <= float(dim.width) <= COL_WIDTH_MAX
    # The CFE curve is emitted only for the emissions-configured run;
    # the energy-flow diagram is a standard output of every run and
    # lives with the energy plots.
    assert (r1.out_dir / "04_financial_plots" / "cfe_duration_curve.pdf").exists()
    assert not (
        r0.out_dir / "04_financial_plots" / "cfe_duration_curve.pdf"
    ).exists()
    for r in (r0, r1):
        assert (r.out_dir / "05_energy_plots" / "energy_sankey.pdf").exists()
        assert not (
            r.out_dir / "04_financial_plots" / "energy_sankey.pdf"
        ).exists()


def _sankey_frame(**kw) -> pd.DataFrame:
    cols = [
        "pv_to_load_kwh", "pv_to_bess_kwh", "pv_to_grid_kwh",
        "pv_curtail_kwh", "grid_to_load_kwh", "bess_charge_grid_kwh",
        "bess_dis_load_kwh", "bess_dis_grid_kwh",
    ]
    return pd.DataFrame({c: [float(kw.get(c, 0.0)) * 1e6] for c in cols})


@pytest.mark.parametrize(("name", "frame_kwargs", "expected", "absent"), [
    ("pv_only_merchant",
     dict(pv_to_grid_kwh=21.4, pv_curtail_kwh=1.1),
     {"PV generation", "Grid export", "Curtailed PV"},
     {"BESS", "Losses", "Load", "Grid import"}),
    ("bess_only",
     dict(bess_charge_grid_kwh=19.8, bess_dis_grid_kwh=18.0),
     {"Grid import", "BESS", "Grid export", "Losses"},
     {"PV generation", "Curtailed PV", "Load"}),
    ("hybrid_with_curtailment",
     dict(pv_to_load_kwh=12.9, pv_to_bess_kwh=9.2, pv_to_grid_kwh=0.4,
          pv_curtail_kwh=1.6, grid_to_load_kwh=13.0,
          bess_charge_grid_kwh=10.6, bess_dis_load_kwh=6.3,
          bess_dis_grid_kwh=12.1),
     {"PV generation", "Grid import", "BESS", "Load", "Grid export",
      "Curtailed PV", "Losses"},
     set()),
])
def test_energy_sankey_single_asset_regimes(
    tmp_path, monkeypatch, name, frame_kwargs, expected, absent,
):
    """The energy-flow diagram collapses to the columns a regime uses:
    PV-only draws no battery column, BESS-only no PV column, and the
    curtailment sink carries the canonical "Curtailed PV" label."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    from pvbess_opt.plotting import emissions as em_mod
    from pvbess_opt.plotting.emissions import plot_energy_sankey

    captured: dict = {}
    real_save = em_mod.save_figure

    def _spy(out_path):
        captured["texts"] = [
            t.get_text().split("\n")[0]
            for t in plt.gcf().axes[0].texts
        ]
        return real_save(out_path)

    monkeypatch.setattr(em_mod, "save_figure", _spy)
    out = tmp_path / f"{name}.pdf"
    plot_energy_sankey(_sankey_frame(**frame_kwargs), out)
    assert out.exists() and out.stat().st_size > 0
    labels = set(captured["texts"])
    assert expected <= labels, (expected - labels, labels)
    assert not (absent & labels), (absent & labels)


def _out_of(flows, node):
    return sum(v for s, _t, v, _c in flows if s == node)


def _into(flows, node):
    return sum(v for _s, t, v, _c in flows if t == node)


def _srcs_sinks(flows):
    srcs = _out_of(flows, "PV generation") + _out_of(flows, "Grid import")
    sinks = sum(
        _into(flows, n)
        for n in (
            "Load", "Grid export", "Curtailed PV", "Curtailed export", "Losses"
        )
    )
    return srcs, sinks


def test_energy_sankey_availability_raises_import_and_conserves():
    """Availability rule inside the Sankey.

    Plant-side flows (PV, BESS, export) derate by the factor; grid import
    RISES to cover the load the offline plant cannot serve; the Load node
    stays at the true (never-derated) demand; and the flows conserve
    energy (sources == sinks) against that demand.
    """
    from pvbess_opt.plotting.emissions import energy_sankey_flows

    res = _sankey_frame(
        pv_to_load_kwh=12.9, pv_to_bess_kwh=9.2, pv_to_grid_kwh=0.4,
        pv_curtail_kwh=1.6, grid_to_load_kwh=13.0, bess_charge_grid_kwh=10.6,
        bess_dis_load_kwh=6.3, bess_dis_grid_kwh=12.1,
    )

    raw = energy_sankey_flows(res, 1.0)
    load = _into(raw, "Load")
    import_raw = _out_of(raw, "Grid import")
    s_raw, k_raw = _srcs_sinks(raw)
    assert s_raw == pytest.approx(k_raw)             # raw dispatch conserves

    a = 0.99
    adj = energy_sankey_flows(res, a)
    # Load is fixed demand -- unchanged by the derate.
    assert _into(adj, "Load") == pytest.approx(load)
    # Import rises to cover the downtime load: factor*import_raw + u*load.
    assert _out_of(adj, "Grid import") == pytest.approx(
        a * import_raw + (1.0 - a) * load,
    )
    assert _out_of(adj, "Grid import") > import_raw   # strictly up (load > import)
    # A plant-side sink still scales DOWN by the factor.
    assert _into(adj, "Grid export") == pytest.approx(a * _into(raw, "Grid export"))
    # Conserves energy against the true demand.
    s_adj, k_adj = _srcs_sinks(adj)
    assert s_adj == pytest.approx(k_adj)


def test_energy_sankey_curtailment_shrinks_export_and_conserves():
    """Exogenous-quota curtailment (Eq. E48) inside the Sankey.

    The two grid-export flows shrink by the curtailment factor so the
    ``Grid export`` node matches the derated ``*_export_mwh`` KPIs; the
    curtailed injection routes to a ``Curtailed export`` sink (absent by
    default); round-trip ``losses`` are unchanged (taken before the export
    haircut); and the flows still conserve energy.
    """
    from pvbess_opt.plotting.emissions import energy_sankey_flows

    res = _sankey_frame(
        pv_to_load_kwh=12.9, pv_to_bess_kwh=9.2, pv_to_grid_kwh=4.0,
        pv_curtail_kwh=1.6, grid_to_load_kwh=13.0, bess_charge_grid_kwh=10.6,
        bess_dis_load_kwh=6.3, bess_dis_grid_kwh=12.1,
    )
    raw = energy_sankey_flows(res, 1.0, 1.0)
    export_raw = _into(raw, "Grid export")
    losses_raw = _into(raw, "Losses")
    s_raw, k_raw = _srcs_sinks(raw)
    assert s_raw == pytest.approx(k_raw)
    assert "Curtailed export" not in {t for _s, t, _v, _c in raw}  # off by default

    c = 0.9
    adj = energy_sankey_flows(res, 1.0, c)
    # Export node shrinks by exactly the curtailment factor ...
    assert _into(adj, "Grid export") == pytest.approx(c * export_raw)
    # ... the removed injection lands in the Curtailed-export sink ...
    assert _into(adj, "Curtailed export") == pytest.approx(
        (1.0 - c) * export_raw,
    )
    # ... round-trip losses are unchanged (computed before the haircut) ...
    assert _into(adj, "Losses") == pytest.approx(losses_raw)
    # ... and the diagram still conserves energy.
    s_adj, k_adj = _srcs_sinks(adj)
    assert s_adj == pytest.approx(k_adj)


def test_energy_sankey_full_availability_is_raw_dispatch():
    """``availability_factor == 1`` returns the untouched raw dispatch."""
    from pvbess_opt.plotting.emissions import energy_sankey_flows

    res = _sankey_frame(
        pv_to_load_kwh=12.9, pv_to_grid_kwh=0.4, grid_to_load_kwh=13.0,
        pv_curtail_kwh=1.6,
    )
    flows = energy_sankey_flows(res, 1.0)
    # PV→grid flow equals the raw column sum (no scaling): 0.4 GWh = 400 MWh.
    pv_to_grid = next(v for s, t, v, _c in flows
                      if s == "PV generation" and t == "Grid export")
    assert pv_to_grid == pytest.approx(400.0)
