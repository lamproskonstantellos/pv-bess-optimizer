"""Capacity sizing: grid parsing, sweep, marginal value, break-even, output."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pvbess_opt.sizing import (
    SizingResult,
    compute_marginal_value_of_storage,
    find_oversizing_breakeven,
    parse_sizing_grid,
    rank_frontier,
    read_sizing_block,
    run_sizing,
    run_sizing_sweep,
    write_sizing_workbook,
)
from pvbess_opt.theme import COL_WIDTH_MAX, COL_WIDTH_MIN, HEADER_FILL_HEX

ROOT = Path(__file__).resolve().parent.parent


def _highs_available() -> bool:
    try:
        import highspy
    except ImportError:
        return False
    return bool(highspy)


def _synthetic_frontier() -> pd.DataFrame:
    mwh = np.arange(1, 10, dtype=float)
    npv = -(mwh - 5.0) ** 2 + 100.0  # peaks at 5 MWh
    return pd.DataFrame({
        "pv_nameplate_kwp": np.full(mwh.size, 1000.0),
        "bess_power_kw": np.full(mwh.size, 500.0),
        "bess_capacity_kwh": mwh * 1000.0,
        "bess_capacity_mwh": mwh,
        "npv_eur": npv,
    })


def test_parse_sizing_grid_list_range_and_duration():
    grid = parse_sizing_grid({
        "pv_nameplate_kwp": [1000, 2000],
        "bess_power_kw": {"min": 500, "max": 1000, "step": 500},
        "bess_capacity_kwh": [1000],
    })
    assert len(grid) == 2 * 2 * 1
    assert (1000.0, 500.0, 1000.0) in grid
    dur = parse_sizing_grid({
        "pv_nameplate_kwp": [1000],
        "bess_power_kw": [1000],
        "bess_duration_hours": [2, 4],
    })
    assert (1000.0, 1000.0, 2000.0) in dur
    assert (1000.0, 1000.0, 4000.0) in dur


def test_marginal_value_and_breakeven_on_synthetic_peak():
    frontier = _synthetic_frontier()
    mv = compute_marginal_value_of_storage(frontier)
    below = mv[mv["bess_capacity_mwh"] < 5.0]["marginal_npv_eur_per_mwh"]
    above = mv[mv["bess_capacity_mwh"] > 5.0]["marginal_npv_eur_per_mwh"]
    assert (below > 0).all()
    assert (above < 0).all()
    be = find_oversizing_breakeven(
        frontier["bess_capacity_mwh"], frontier["npv_eur"],
    )
    assert be == pytest.approx(5.0, abs=1e-6)


def test_breakeven_nan_when_npv_monotone_increasing():
    mwh = np.arange(1, 6, dtype=float)
    npv = mwh * 10.0
    assert np.isnan(find_oversizing_breakeven(mwh, npv))


def test_breakeven_duplicate_capacities_no_divide_by_zero():
    """Duplicate capacity points (zero spacing between sorted MWh values)
    must not raise a divide-by-zero RuntimeWarning nor invent a spurious
    crossing — the zero-spacing segment is skipped (NaN slope)."""
    import warnings

    with warnings.catch_warnings():
        warnings.simplefilter("error", RuntimeWarning)
        be = find_oversizing_breakeven([2.0, 2.0, 4.0], [10.0, 10.0, 5.0])
    # The 2->4 segment has slope (5-10)/(4-2) = -2.5 < 0, so the crossing
    # is the midpoint of that segment; the duplicate 2.0 point is skipped.
    assert np.isfinite(be)


def test_rank_frontier_sorts_by_npv_desc():
    ranked = rank_frontier(pd.DataFrame({"npv_eur": [1.0, 3.0, 2.0]}))
    assert list(ranked["npv_eur"]) == [3.0, 2.0, 1.0]


def test_read_sizing_block(tmp_path):
    cfg = tmp_path / "run.yaml"
    cfg.write_text("sizing:\n  pv_nameplate_kwp: [1000, 2000]\n", encoding="utf-8")
    assert read_sizing_block(cfg) == {"pv_nameplate_kwp": [1000, 2000]}
    assert read_sizing_block(tmp_path / "x.xlsx") is None


# ---------------------------------------------------------------------------
# Excel-driven sizing sheet (columnar, gated by an enabled TRUE/FALSE toggle)
# ---------------------------------------------------------------------------


def _minimal_typed_with_sizing(sizing_rows):
    from pvbess_opt.io import (
        BESS_SHEET_DEFAULTS,
        ECONOMICS_SHEET_DEFAULTS,
        PROJECT_SHEET_DEFAULTS,
        PV_SHEET_DEFAULTS,
        SIMULATION_SHEET_DEFAULTS,
    )
    ts = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=24, freq="h"),
        "pv_kwh": [100.0] * 24,
        "load_kwh": [50.0] * 24,
        "dam_price_eur_per_mwh": [80.0] * 24,
    })
    typed = {
        "ts": ts,
        "project": dict(PROJECT_SHEET_DEFAULTS),
        "pv": dict(PV_SHEET_DEFAULTS, pv_nameplate_kwp=1000.0),
        "bess": dict(
            BESS_SHEET_DEFAULTS, bess_power_kw=500.0, bess_capacity_kwh=2000.0,
        ),
        "economics": dict(ECONOMICS_SHEET_DEFAULTS),
        "simulation": dict(SIMULATION_SHEET_DEFAULTS),
        "max_injection_profile": np.full(24, 100.0),
    }
    if sizing_rows is not None:
        typed["sizing"] = sizing_rows
    return typed


def test_parse_sizing_sheet_columns_and_enabled():
    from pvbess_opt.sizing import _parse_sizing_sheet
    df = pd.DataFrame({
        "enabled": ["TRUE", None, None],
        "pv_nameplate_kwp": [1000, 2000, 3000],
        "bess_power_kw": [500, 1000, None],
        "bess_capacity_kwh": [None, None, None],
        "bess_duration_hours": [2, 4, None],
    })
    enabled, block = _parse_sizing_sheet(df)
    assert enabled is True
    assert block == {
        "pv_nameplate_kwp": [1000.0, 2000.0, 3000.0],
        "bess_power_kw": [500.0, 1000.0],
        "bess_duration_hours": [2.0, 4.0],
    }


def test_parse_sizing_sheet_capacity_wins_over_duration():
    from pvbess_opt.sizing import _parse_sizing_sheet
    df = pd.DataFrame({
        "enabled": ["FALSE", None],
        "pv_nameplate_kwp": [1000, 2000],
        "bess_power_kw": [500, 1000],
        "bess_capacity_kwh": [1000, 4000],
        "bess_duration_hours": [2, 4],
    })
    enabled, block = _parse_sizing_sheet(df)
    assert enabled is False
    assert "bess_capacity_kwh" in block
    assert "bess_duration_hours" not in block


def test_read_sizing_block_from_xlsx_gated_by_enabled(tmp_path):
    from pvbess_opt.io import write_workbook
    # The shipped default example is disabled -> no sweep.
    disabled = tmp_path / "disabled.xlsx"
    write_workbook(_minimal_typed_with_sizing(None), disabled)
    assert read_sizing_block(disabled) is None
    # enabled=TRUE on the first row -> the grid is parsed.
    enabled = tmp_path / "enabled.xlsx"
    write_workbook(_minimal_typed_with_sizing([
        {"enabled": "TRUE", "pv_nameplate_kwp": 1000, "bess_power_kw": 500,
         "bess_duration_hours": 2},
        {"pv_nameplate_kwp": 2000, "bess_power_kw": 1000,
         "bess_duration_hours": 4},
    ]), enabled)
    block = read_sizing_block(enabled)
    assert block == {
        "pv_nameplate_kwp": [1000.0, 2000.0],
        "bess_power_kw": [500.0, 1000.0],
        "bess_duration_hours": [2.0, 4.0],
    }
    assert len(parse_sizing_grid(block)) == 2 * 2 * 2


def test_repo_input_xlsx_ships_disabled_sizing_sheet():
    sheets = pd.ExcelFile(ROOT / "inputs" / "input.xlsx").sheet_names
    assert "sizing" in sheets
    # Shipped disabled, so a normal run is unaffected.
    assert read_sizing_block(ROOT / "inputs" / "input.xlsx") is None


def test_write_sizing_workbook_is_styled(tmp_path):
    frontier = rank_frontier(_synthetic_frontier().assign(
        irr_pct=5.0, simple_payback_years=8.0,
        lcoe_eur_per_mwh=50.0, lcos_eur_per_mwh=150.0,
    ))
    result = SizingResult(
        frontier=frontier,
        marginal_value=compute_marginal_value_of_storage(frontier),
        oversizing_breakeven_mwh=5.0,
    )
    out = write_sizing_workbook(tmp_path / "sizing.xlsx", result)
    wb = load_workbook(out)
    assert wb.sheetnames
    for sn in wb.sheetnames:
        ws = wb[sn]
        assert ws.freeze_panes == "A2"
        for cell in ws[1]:
            if cell.value is None:
                continue
            rgb = (getattr(cell.fill.fgColor, "rgb", None) or "")
            assert rgb.upper().lstrip("0").rjust(6, "0")[-6:] == HEADER_FILL_HEX
        for c in range(1, ws.max_column + 1):
            dim = ws.column_dimensions.get(get_column_letter(c))
            assert dim is not None and dim.width is not None
            assert COL_WIDTH_MIN <= float(dim.width) <= COL_WIDTH_MAX


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_evaluate_sizing_point_threads_base_dir_to_financials(
    tmp_path, monkeypatch,
):
    """A structured-config sizing sweep materializes a throwaway workbook into
    a temp dir; relative price-scenario ``store_path`` entries must resolve
    against the ORIGINAL config directory (threaded as ``base_dir``), not that
    temp dir -- mirroring ``pipeline.run`` and ``scenarios.run_scenarios`` so
    all three dispatch surfaces agree."""
    import pvbess_opt.availability as availability
    import pvbess_opt.kpis as kpis_mod
    import pvbess_opt.optimization as optimization
    import pvbess_opt.pipeline as pipeline
    import pvbess_opt.sizing as sizing

    captured: dict[str, object] = {}
    monkeypatch.setattr(
        optimization, "run_scenario",
        lambda *_a, **_k: (pd.DataFrame(), None, pd.DataFrame()),
    )
    monkeypatch.setattr(kpis_mod, "compute_kpis", lambda *_a, **_k: {})
    monkeypatch.setattr(
        availability, "apply_operating_derates", lambda kp, _p: kp,
    )

    def _fake_build_financials(excel_path, *_a, base_dir=None, **_k):
        captured["base_dir"] = base_dir
        captured["excel_parent"] = Path(excel_path).parent
        return {"fin_kpis": {}}

    monkeypatch.setattr(pipeline, "_build_financials", _fake_build_financials)

    cfg_dir = tmp_path / "project"
    cfg_dir.mkdir()
    temp_dir = tmp_path / "pvbess_sizing_tmp"
    temp_dir.mkdir()
    base_xlsx = temp_dir / "materialized.xlsx"  # the throwaway workbook

    sizing.evaluate_sizing_point(
        {}, pd.DataFrame(), {}, base_xlsx, (1.0, 1.0, 1.0),
        solver_opts={}, base_dir=cfg_dir,
    )
    # base_dir reaches _build_financials as the ORIGINAL config dir ...
    assert captured["base_dir"] == cfg_dir
    # ... not the materialization temp dir (base_xlsx's parent).
    assert captured["base_dir"] != captured["excel_parent"]


def test_sizing_sweep_runs_2x2x2(tmp_path):
    from pvbess_opt.io import _typed_to_flat, read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    short = tmp_path / "short.xlsx"
    write_workbook(typed, short)
    base = read_workbook(short)
    base_params, base_ts = _typed_to_flat(base)
    grid = parse_sizing_grid({
        "pv_nameplate_kwp": [4000, 8000],
        "bess_power_kw": [1000, 2000],
        "bess_capacity_kwh": [2000, 4000],
    })
    frontier = run_sizing_sweep(
        base_params, base_ts, base["pv"], short, grid,
        solver_opts={
            "solver_name": "highs", "mip_gap": 0.05,
            "time_limit_seconds": 180, "tee": False,
        },
    )
    assert len(frontier) == 8
    assert frontier["npv_eur"].notna().all()
    assert frontier["npv_eur"].iloc[0] >= frontier["npv_eur"].iloc[-1]


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_run_sizing_end_to_end(tmp_path):
    from pvbess_opt import RunConfig
    from pvbess_opt.io import read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    short = tmp_path / "short.xlsx"
    write_workbook(typed, short)

    config = RunConfig(
        excel=short, solver="highs", outdir=tmp_path / "out",
        mip_gap=0.05, time_limit=180,
    )
    block = {
        "pv_nameplate_kwp": [4000, 8000],
        "bess_power_kw": [1000],
        "bess_capacity_kwh": [2000, 4000],
    }
    result = run_sizing(config, block)
    assert len(result.frontier) == 4
    runs = list((tmp_path / "out").glob("*_sizing_*"))
    assert runs
    assert (runs[0] / "sizing.xlsx").exists()
    assert (runs[0] / "efficient_frontier.pdf").exists()
    assert (runs[0] / "npv_vs_capacity.pdf").exists()
