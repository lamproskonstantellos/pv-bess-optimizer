"""Debt/equity leverage: amortization, equity IRR, DSCR, no unlevered drift."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pvbess_opt.economics import (
    _amortization_schedule,
    _leverage_kpis,
    build_debt_schedule,
    calculate_irr,
)
from pvbess_opt.io_read import load_structured_config
from pvbess_opt.theme import COL_WIDTH_MAX, COL_WIDTH_MIN, HEADER_FILL_HEX

ROOT = Path(__file__).resolve().parent.parent


def _highs_available() -> bool:
    try:
        import highspy
    except ImportError:
        return False
    return bool(highspy)


@pytest.mark.parametrize("repayment", ["annuity", "linear"])
def test_amortization_balances_to_zero(repayment):
    sched = _amortization_schedule(10_000.0, 0.06, 12, repayment)
    assert len(sched) == 12
    assert sched[-1]["debt_balance_eur"] == pytest.approx(0.0, abs=1e-6)
    assert sum(r["principal_eur"] for r in sched) == pytest.approx(10_000.0, rel=1e-9)


def test_amortization_empty_without_debt_or_tenor():
    assert _amortization_schedule(0.0, 0.05, 10, "annuity") == []
    assert _amortization_schedule(1000.0, 0.05, 0, "annuity") == []


def test_sculpted_schedule_caps_principal_at_balance_on_ramp_cfads():
    """A ramp-shaped CFADS (thin early years, thick later years) inflates
    the balance carried into the thick years; the sculpted service there
    must not repay MORE principal than is outstanding, or sum(principal)
    exceeds the debt drawn and later years book phantom service against a
    zeroed balance (over-stating debt_service_eur / avg_dscr)."""
    debt, rate, tenor = 490.0, 0.06, 15
    cfads = [7.9, 6.5, 7.8, 6.5, 190.3, 192.6, 183.7, 197.7, 191.5,
             164.1, 177.3, 173.8, 191.9, 191.3, 161.7]
    sched = _amortization_schedule(debt, rate, tenor, "sculpted", cfads=cfads)
    # sum(principal) == debt (the branch's stated invariant).
    assert sum(r["principal_eur"] for r in sched) == pytest.approx(
        debt, rel=1e-9,
    )
    # No year repays more principal than the balance it started with, and
    # no service is booked once the balance is retired.
    bal = debt
    for r in sched:
        assert r["principal_eur"] <= bal + 1e-9
        if bal <= 1e-9:
            assert r["debt_service_eur"] == pytest.approx(0.0, abs=1e-9)
        bal = r["debt_balance_eur"]
    assert sched[-1]["debt_balance_eur"] == pytest.approx(0.0, abs=1e-6)


def test_equity_irr_exceeds_project_irr_on_positive_spread():
    net_cf = np.array([-1000.0] + [200.0] * 10)
    project_irr = calculate_irr(net_cf) * 100.0
    equity_irr, min_dscr, avg_dscr = _leverage_kpis(net_cf, {
        "gearing_pct": 70.0, "debt_interest_rate_pct": 5.0,
        "debt_tenor_years": 7, "debt_repayment": "annuity",
    })
    assert equity_irr > project_irr            # leverage amplifies the spread
    assert np.isfinite(min_dscr) and min_dscr > 0.0
    # The average sits between the minimum and the maximum coverage.
    assert avg_dscr >= min_dscr


def test_all_equity_returns_nan():
    net_cf = np.array([-1000.0] + [200.0] * 10)
    eq, dscr, avg = _leverage_kpis(net_cf, {"gearing_pct": 0.0})
    assert np.isnan(eq) and np.isnan(dscr) and np.isnan(avg)


def test_build_debt_schedule_columns_and_none_when_all_equity():
    yearly = pd.DataFrame({"net_cashflow_eur": [-1000.0] + [200.0] * 10})
    sched = build_debt_schedule(yearly, {
        "gearing_pct": 60.0, "debt_interest_rate_pct": 5.0,
        "debt_tenor_years": 6, "debt_repayment": "linear",
    })
    assert sched is not None and len(sched) == 6
    assert {
        "year", "debt_service_eur", "equity_cf_eur", "dscr", "debt_balance_eur",
    }.issubset(sched.columns)
    assert sched["debt_balance_eur"].iloc[-1] == pytest.approx(0.0, abs=1e-6)
    assert build_debt_schedule(yearly, {"gearing_pct": 0.0}) is None


def test_financing_block_alias(tmp_path):
    ts = pd.DataFrame({
        "timestamp": pd.date_range("2019-01-01", periods=4, freq="15min"),
        "dam_price_eur_per_mwh": [50.0] * 4,
        "pv_kwh": [0.0] * 4,
    })
    ts.to_csv(tmp_path / "ts.csv", index=False)
    cfg = tmp_path / "run.yaml"
    cfg.write_text(
        "financing:\n"
        "  gearing: 0.7\n"
        "  interest_rate: 0.05\n"
        "  tenor_years: 12\n"
        "  repayment: linear\n"
        "timeseries_path: ts.csv\n",
        encoding="utf-8",
    )
    econ = load_structured_config(cfg)["economics"]
    assert econ["gearing_pct"] == pytest.approx(70.0)
    assert econ["debt_interest_rate_pct"] == pytest.approx(5.0)
    assert econ["debt_tenor_years"] == 12
    assert econ["debt_repayment"] == "linear"


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_leverage_full_run_sheet_and_unlevered_unchanged(tmp_path):
    from pvbess_opt import RunConfig, run
    from pvbess_opt.io import read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    unlevered = tmp_path / "unlev.xlsx"
    write_workbook(typed, unlevered)

    typed_lev = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed_lev["ts"] = typed_lev["ts"].iloc[:96].reset_index(drop=True)
    typed_lev["economics"]["gearing_pct"] = 70.0
    typed_lev["economics"]["debt_tenor_years"] = 10
    levered = tmp_path / "lev.xlsx"
    write_workbook(typed_lev, levered)

    common = dict(solver="highs", mip_gap=0.05, time_limit=180)
    r0 = run(RunConfig(excel=unlevered, outdir=tmp_path / "a", **common))
    r1 = run(RunConfig(excel=levered, outdir=tmp_path / "b", **common))

    assert "equity_irr_pct" in r1.financial_kpis
    assert "min_dscr" in r1.financial_kpis
    # Leverage does not touch the unlevered metrics.
    assert r1.financial_kpis["npv_eur"] == pytest.approx(
        r0.financial_kpis["npv_eur"], rel=1e-9, abs=1e-6,
    )
    # Debt-schedule sheet only when financing is configured, and styled.
    wb1 = load_workbook(r1.out_dir / "03_results.xlsx")
    wb0 = load_workbook(r0.out_dir / "03_results.xlsx")
    assert "debt_schedule" in wb1.sheetnames
    assert "debt_schedule" not in wb0.sheetnames
    ws = wb1["debt_schedule"]
    assert ws.freeze_panes == "A2"
    for cell in ws[1]:
        if cell.value is None:
            continue
        rgb = (getattr(cell.fill.fgColor, "rgb", None) or "")
        assert rgb.upper().lstrip("0").rjust(6, "0")[-6:] == HEADER_FILL_HEX
    for c in range(1, ws.max_column + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        assert dim is not None and COL_WIDTH_MIN <= float(dim.width) <= COL_WIDTH_MAX


# ---------------------------------------------------------------------------
# Sculpted repayment profile (Eqs. E40/E40a)
# ---------------------------------------------------------------------------


IRREGULAR_CFADS = [220.0, 180.0, 60.0, 240.0, 210.0, 190.0, 230.0]


def test_sculpted_balance_amortises_to_zero():
    sched = _amortization_schedule(
        800.0, 0.06, 7, "sculpted", cfads=IRREGULAR_CFADS,
    )
    assert len(sched) == 7
    assert sched[-1]["debt_balance_eur"] == pytest.approx(0.0, abs=1e-6)
    assert sum(r["principal_eur"] for r in sched) == pytest.approx(
        800.0, rel=1e-9,
    )


def test_sculpted_dscr_is_level():
    """Per-year CFADS/service is constant across positive-CFADS years,
    incl. a replacement-style dip — the whole point of sculpting."""
    sched = _amortization_schedule(
        800.0, 0.06, 7, "sculpted", cfads=IRREGULAR_CFADS,
    )
    dscrs = [
        IRREGULAR_CFADS[int(r["year"]) - 1] / r["debt_service_eur"]
        for r in sched if r["debt_service_eur"] > 0.0
    ]
    for d in dscrs[:-1]:
        assert d == pytest.approx(dscrs[0], rel=1e-9)
    # The final-year cent sweep may deviate by float residue only.
    assert dscrs[-1] == pytest.approx(dscrs[0], rel=1e-6)
    # E40a closed form: DSCR_impl = PV(max(CFADS,0)) / B at the debt rate.
    pv = sum(
        c * 1.06 ** (-y) for y, c in enumerate(IRREGULAR_CFADS, start=1)
    )
    assert dscrs[0] == pytest.approx(pv / 800.0, rel=1e-9)


def test_sculpted_requires_cfads():
    with pytest.raises(ValueError, match="sculpted repayment requires"):
        _amortization_schedule(800.0, 0.06, 7, "sculpted")


def test_sculpted_negative_cfads_year_carries_and_amortises():
    cfads = [220.0, -50.0, 0.0, 240.0, 210.0, 190.0, 230.0]
    sched = _amortization_schedule(700.0, 0.05, 7, "sculpted", cfads=cfads)
    by_year = {int(r["year"]): r for r in sched}
    for y in (2, 3):  # non-positive CFADS: no service, balance carries
        assert by_year[y]["debt_service_eur"] == 0.0
        assert by_year[y]["principal_eur"] == 0.0
        assert by_year[y]["debt_balance_eur"] == pytest.approx(
            by_year[y - 1]["debt_balance_eur"],
        )
    assert sched[-1]["debt_balance_eur"] == pytest.approx(0.0, abs=1e-6)


def test_sculpted_min_equals_avg_dscr():
    net_cf = np.array([-1000.0, 220.0, 180.0, 60.0, 240.0, 210.0, 190.0,
                       230.0, 230.0])
    eq, min_dscr, avg_dscr = _leverage_kpis(net_cf, {
        "gearing_pct": 70.0, "debt_interest_rate_pct": 5.0,
        "debt_tenor_years": 7, "debt_repayment": "sculpted",
    })
    assert np.isfinite(eq)
    assert min_dscr == pytest.approx(avg_dscr, rel=1e-6)


def test_sculpted_debt_schedule_level_dscr_column():
    yearly = pd.DataFrame({
        "net_cashflow_eur": [-1000.0, 220.0, 180.0, 60.0, 240.0, 210.0,
                             190.0, 230.0, 230.0],
    })
    sched = build_debt_schedule(yearly, {
        "gearing_pct": 60.0, "debt_interest_rate_pct": 5.0,
        "debt_tenor_years": 7, "debt_repayment": "sculpted",
    })
    assert sched is not None
    dscrs = sched.loc[sched["debt_service_eur"] > 0.0, "dscr"]
    assert dscrs.nunique() <= 2  # 4dp rounding: level within a cent
    assert sched["debt_balance_eur"].iloc[-1] == pytest.approx(0.0, abs=0.01)
