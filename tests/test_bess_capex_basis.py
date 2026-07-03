"""BESS CAPEX energy-basis contract (v1.0.0).

BESS CAPEX is priced per kWh of nameplate energy capacity:
``capex_bess_y0 = -capex_bess_eur_per_kwh x bess_kwh``.  The
replacement charge is ``bess_replacement_cost_pct/100`` of that Year-0
figure, so it inherits the energy basis.  The legacy per-kW key
``capex_bess_eur_per_kw`` is rejected loudly on every input surface
(workbook and structured config), and the polish script migrates it.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from pvbess_opt.economics import build_yearly_cashflow, derive_asset_capacities
from pvbess_opt.io import (
    BESS_SHEET_DEFAULTS,
    ECONOMICS_SHEET_DEFAULTS,
    PROJECT_SHEET_DEFAULTS,
    PV_SHEET_DEFAULTS,
    SIMULATION_SHEET_DEFAULTS,
    read_workbook,
    write_workbook,
)
from pvbess_opt.io_read import load_structured_config

ROOT = Path(__file__).resolve().parent.parent


def _econ(**overrides):
    econ = {
        "project_lifecycle_years": 20,
        "project_start_year": 2026,
        "discount_rate_pct": 7.0,
        "opex_inflation_pct": 0.0,
        "retail_inflation_pct": 0.0,
        "dam_inflation_pct": 0.0,
        "aggregator_fee_pct_revenue": 0.0,
        "capex_pv_eur_per_kw": 0.0,
        "capex_bess_eur_per_kwh": 250.0,
        "devex_pv_eur_per_kw": 0.0,
        "devex_bess_eur_per_kw": 0.0,
        "opex_pv_eur_per_kwp": 0.0,
        "opex_bess_eur_per_kw": 0.0,
        "pv_degradation_year1_pct": 0.0,
        "pv_degradation_annual_pct": 0.0,
        "bess_degradation_annual_pct": 2.0,
        "bess_degradation_pct_per_cycle": 0.0,
        "bess_replacement_year": 10,
        "bess_replacement_cost_pct": 50.0,
    }
    econ.update(overrides)
    return econ


def _kpis():
    return {
        "profit_total_eur": 1_000_000.0,
        "bess_total_discharge_mwh": 10_000.0,
    }


def test_capex_bess_y0_is_energy_basis():
    """Year-0 BESS CAPEX = -capex_bess_eur_per_kwh x bess_kwh."""
    params = {
        "pv_nameplate_kwp": 0.0,
        "bess_power_kw": 15_000.0,
        "bess_capacity_kwh": 60_000.0,
    }
    caps = derive_asset_capacities({}, params, pd.DataFrame())
    assert caps["bess_kw"] == 15_000.0
    assert caps["bess_kwh"] == 60_000.0

    econ = _econ()
    cf = build_yearly_cashflow(_kpis(), econ, caps)
    y0_capex = float(cf.loc[cf["project_year"] == 0, "capex_eur"].iloc[0])
    expected = -econ["capex_bess_eur_per_kwh"] * caps["bess_kwh"]
    assert y0_capex == pytest.approx(expected)
    # The power rating must not enter the CAPEX: changing bess_kw alone
    # (same energy capacity) leaves the Year-0 outflow unchanged.
    caps_2h = dict(caps, bess_kw=30_000.0)
    cf_2h = build_yearly_cashflow(_kpis(), econ, caps_2h)
    assert float(
        cf_2h.loc[cf_2h["project_year"] == 0, "capex_eur"].iloc[0]
    ) == pytest.approx(y0_capex)


def test_replacement_charge_is_pct_of_energy_basis_capex():
    """Replacement CAPEX = pct/100 x capex_bess_y0, in the scheduled year."""
    params = {
        "pv_nameplate_kwp": 0.0,
        "bess_power_kw": 15_000.0,
        "bess_capacity_kwh": 60_000.0,
    }
    caps = derive_asset_capacities({}, params, pd.DataFrame())
    econ = _econ()
    cf = build_yearly_cashflow(_kpis(), econ, caps)
    capex_bess_y0 = -econ["capex_bess_eur_per_kwh"] * caps["bess_kwh"]
    repl = float(cf.loc[cf["project_year"] == 10, "capex_eur"].iloc[0])
    assert repl == pytest.approx(
        capex_bess_y0 * econ["bess_replacement_cost_pct"] / 100.0
    )
    # No other operating year carries CAPEX.
    others = cf.loc[
        (cf["project_year"] >= 1) & (cf["project_year"] != 10), "capex_eur",
    ]
    assert (others == 0.0).all()


def test_bess_absent_config_books_zero_capex():
    """bess_kwh is forced to 0 when bess_kw == 0, so no BESS CAPEX books."""
    params = {
        "pv_nameplate_kwp": 1_000.0,
        "bess_power_kw": 0.0,
        "bess_capacity_kwh": 60_000.0,
    }
    caps = derive_asset_capacities({}, params, pd.DataFrame())
    assert caps["bess_kwh"] == 0.0
    cf = build_yearly_cashflow(_kpis(), _econ(), caps)
    assert float(cf.loc[cf["project_year"] == 0, "capex_eur"].iloc[0]) == 0.0


def _minimal_typed() -> dict:
    n = 24
    ts = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=n, freq="h"),
        "pv_kwh": [100.0] * n,
        "load_kwh": [50.0] * n,
        "dam_price_eur_per_mwh": [80.0] * n,
    })
    return {
        "ts": ts,
        "project": dict(PROJECT_SHEET_DEFAULTS),
        "pv": dict(PV_SHEET_DEFAULTS, pv_nameplate_kwp=1000.0),
        "bess": dict(
            BESS_SHEET_DEFAULTS, bess_power_kw=500.0, bess_capacity_kwh=2000.0,
        ),
        "economics": dict(ECONOMICS_SHEET_DEFAULTS),
        "simulation": dict(SIMULATION_SHEET_DEFAULTS),
        "max_injection_profile": np.full(24, 100.0, dtype=float),
    }


def test_legacy_workbook_key_raises(tmp_path):
    """A workbook carrying capex_bess_eur_per_kw fails loudly with guidance."""
    dst = tmp_path / "legacy.xlsx"
    write_workbook(_minimal_typed(), dst)
    wb = load_workbook(dst)
    ws = wb["bess"]
    renamed = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "capex_bess_eur_per_kwh":
            row[0].value = "capex_bess_eur_per_kw"
            row[1].value = 200.0
            renamed = True
    assert renamed
    wb.save(dst)
    with pytest.raises(ValueError, match="capex_bess_eur_per_kwh"):
        read_workbook(dst)


def test_legacy_config_key_raises(tmp_path):
    """A YAML config carrying capex_bess_eur_per_kw fails loudly too."""
    ts_path = tmp_path / "ts.csv"
    _minimal_typed()["ts"].to_csv(ts_path, index=False)
    cfg = tmp_path / "run.yaml"
    cfg.write_text(
        "timeseries_path: ts.csv\n"
        "pv:\n"
        "  pv_nameplate_kwp: 1000\n"
        "bess:\n"
        "  bess_power_kw: 500\n"
        "  bess_capacity_kwh: 2000\n"
        "  capex_bess_eur_per_kw: 200\n",
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="capex_bess_eur_per_kwh"):
        load_structured_config(cfg)


def test_polish_script_migrates_legacy_key(tmp_path):
    """polish_input_workbook converts per-kW to per-kWh via power/capacity."""
    import sys

    sys.path.insert(0, str(ROOT / "scripts"))
    from polish_input_workbook import polish_workbook

    dst = tmp_path / "legacy.xlsx"
    write_workbook(_minimal_typed(), dst)
    wb = load_workbook(dst)
    ws = wb["bess"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "capex_bess_eur_per_kwh":
            row[0].value = "capex_bess_eur_per_kw"
            row[1].value = 200.0
    wb.save(dst)

    polish_workbook(dst)

    df = pd.read_excel(dst, sheet_name="bess")
    values = dict(zip(df["key"], df["value"], strict=True))
    assert "capex_bess_eur_per_kw" not in values
    # 200 EUR/kW x 500 kW / 2000 kWh = 50 EUR/kWh.
    assert values["capex_bess_eur_per_kwh"] == pytest.approx(50.0)
    # The migrated workbook loads warning-free.
    typed = read_workbook(dst)
    assert typed["bess"]["capex_bess_eur_per_kwh"] == pytest.approx(50.0)
