"""Programmatic API for the extracted pipeline + CLI.

``pvbess_opt.run(config)`` and ``pvbess_opt.cli.main([...])`` must run a
scenario end-to-end without the top-level ``main.py`` script — the
package is importable and testable on its own.
"""

from __future__ import annotations

from pathlib import Path

import pytest

import pvbess_opt
from pvbess_opt import Results, RunConfig, run

ROOT = Path(__file__).resolve().parent.parent


def _highs_available() -> bool:
    try:
        import highspy
    except ImportError:
        return False
    return bool(highspy)


def _short_workbook(tmp_path: Path) -> Path:
    """Write a 1-day (96-step) slice of the case-study workbook."""
    from pvbess_opt.io import read_workbook, write_workbook

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    short = tmp_path / "short.xlsx"
    write_workbook(typed, short)
    return short


def test_run_config_and_results_are_exported():
    """The programmatic surface is re-exported from the package root."""
    assert pvbess_opt.RunConfig is RunConfig
    assert pvbess_opt.Results is Results
    assert pvbess_opt.run is run
    cfg = RunConfig(excel=Path("inputs/input.xlsx"))
    assert cfg.solver == "highs"
    assert cfg.outdir == Path("results")


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_run_returns_populated_results(tmp_path):
    short = _short_workbook(tmp_path)
    config = RunConfig(
        excel=short,
        solver="highs",
        outdir=tmp_path / "results",
        mip_gap=0.05,
        time_limit=180,
    )
    result = run(config)
    assert isinstance(result, Results)
    assert result.kpis, "Results.kpis should be populated"
    assert "profit_total_eur" in result.kpis
    assert result.out_dir.exists()
    assert (result.out_dir / "03_results.xlsx").exists()
    # The advertised output layout: SUMMARY.md digest + run log.
    assert (result.out_dir / "00_summary" / "SUMMARY.md").exists()
    assert (result.out_dir / "00_summary" / "run_log.txt").exists()
    # Public Results contract: the financial bundle is surfaced too.
    assert result.financial_kpis is not None
    assert result.yearly_cashflow is not None
    _ = (result.lifetime_yearly, result.sensitivity)


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_run_snapshot_exists_and_scrubs_entsoe_token(tmp_path):
    """End-to-end lock on the round-1 secret-scrub path: a workbook that
    carries an ENTSO-E token in its market_data cell must produce a
    01_inputs snapshot with the token blanked (no fetch happened, so the
    blank_entsoe_token branch runs) and an assumptions_summary that never
    prints the token."""
    import openpyxl

    short = _short_workbook(tmp_path)
    # Inject a token into the market_data.entsoe_token cell.
    token = "SECRET-TOKEN-1234567890-abcdef"
    wb = openpyxl.load_workbook(short)
    ws = wb["market_data"]
    for row in ws.iter_rows(min_row=2, max_col=2):
        if str(row[0].value).strip() == "entsoe_token":
            row[1].value = token
            break
    wb.save(short)

    config = RunConfig(
        excel=short, solver="highs", outdir=tmp_path / "results",
        mip_gap=0.05, time_limit=180,
    )
    result = run(config)

    snapshot = result.out_dir / "01_inputs" / "input_snapshot.xlsx"
    assert snapshot.exists(), "01_inputs/input_snapshot.xlsx must be written"
    snap_wb = openpyxl.load_workbook(snapshot)
    snap_ws = snap_wb["market_data"]
    token_cell = None
    for row in snap_ws.iter_rows(min_row=2, max_col=2):
        if str(row[0].value).strip() == "entsoe_token":
            token_cell = row[1].value
            break
    assert token_cell in (None, ""), (
        f"snapshot entsoe_token must be scrubbed, got {token_cell!r}"
    )
    summary = result.out_dir / "01_inputs" / "assumptions_summary.txt"
    assert summary.exists()
    assert token not in summary.read_text(encoding="utf-8")


def test_structured_config_anchors_stores_to_config_dir_not_temp(
    tmp_path, monkeypatch,
):
    """A structured (YAML/JSON) config is materialized to a throwaway temp
    workbook, so ``config.excel`` no longer sits beside the config's own
    price-scenario stores.  ``run`` must thread the ORIGINAL config directory
    as ``base_dir`` so a relative ``store_path`` resolves against it (mirroring
    ``scenarios.run_scenarios``), never the materialization temp dir — which
    would break the documented relative-store feature with a FileNotFound
    pointing at a random temp path the user never created."""
    import pvbess_opt.pipeline as pipeline

    cfg_dir = tmp_path / "project"
    cfg_dir.mkdir()
    cfg = cfg_dir / "config.yaml"
    cfg.write_text("# structured config\n", encoding="utf-8")

    captured: dict[str, object] = {}

    # Stub the heavy materialize/read path: this test pins only the base_dir
    # threading, not the solve.  materialize_to_xlsx returns a path INSIDE the
    # real mkdtemp temp dir, so config.excel's parent is the temp dir.
    monkeypatch.setattr(pipeline, "is_structured_config", lambda _p: True)
    monkeypatch.setattr(
        pipeline, "materialize_to_xlsx",
        lambda _src, dst: Path(dst) / "materialized.xlsx",
    )
    monkeypatch.setattr(pipeline, "read_inputs", lambda _p: ({}, None))
    monkeypatch.setattr(pipeline, "apply_ieee_style", lambda: None)
    monkeypatch.setattr(pipeline, "set_show_titles", lambda _v: None)

    sentinel = object()

    def _fake_run_one(
        params, ts, config, base_name, timestamp, base_dir=None, **_kwargs,
    ):
        captured["base_dir"] = base_dir
        captured["excel_parent"] = Path(config.excel).parent
        return sentinel

    monkeypatch.setattr(pipeline, "_run_one", _fake_run_one)

    result = pipeline.run(RunConfig(excel=cfg, outdir=tmp_path / "out"))
    assert result is sentinel
    # base_dir is the config's OWN directory ...
    assert captured["base_dir"] == cfg_dir
    # ... and specifically NOT the materialization temp dir.
    assert captured["base_dir"] != captured["excel_parent"]


def test_structured_config_temp_workbook_is_cleaned_up(tmp_path, monkeypatch):
    """The throwaway workbook a structured config is materialized into has no
    consumer once the run has read it and copied its snapshot; ``run`` must
    remove the temp dir so batch/sweep invocations do not leak one per run."""
    import tempfile as _tempfile

    import pvbess_opt.pipeline as pipeline

    cfg = tmp_path / "config.yaml"
    cfg.write_text("# structured config\n", encoding="utf-8")

    made: list[Path] = []
    _real_mkdtemp = _tempfile.mkdtemp

    def _tracking_mkdtemp(*a, **k):
        d = _real_mkdtemp(*a, **k)
        made.append(Path(d))
        return d

    monkeypatch.setattr(pipeline.tempfile, "mkdtemp", _tracking_mkdtemp)
    monkeypatch.setattr(pipeline, "is_structured_config", lambda _p: True)
    monkeypatch.setattr(
        pipeline, "materialize_to_xlsx",
        lambda _src, dst: Path(dst) / "materialized.xlsx",
    )
    monkeypatch.setattr(pipeline, "read_inputs", lambda _p: ({}, None))
    monkeypatch.setattr(pipeline, "apply_ieee_style", lambda: None)
    monkeypatch.setattr(pipeline, "set_show_titles", lambda _v: None)
    monkeypatch.setattr(pipeline, "_run_one", lambda *_a, **_k: object())

    pipeline.run(RunConfig(excel=cfg, outdir=tmp_path / "out"))

    assert made, "run() should have created a materialization temp dir"
    assert not made[0].exists(), (
        f"materialization temp dir {made[0]} must be removed after the run"
    )


@pytest.mark.slow
@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_soh_plot_failure_does_not_abort_completed_run(tmp_path, monkeypatch):
    """A figure must never turn an otherwise-complete run into a reported
    failure: 03_results.xlsx and SUMMARY.md are written before any plot, so a
    raise inside ``plot_soh_trajectory`` (the one call that used to be
    unguarded, unlike its siblings) must be logged and swallowed, and the run
    must still return populated Results with the workbook on disk."""
    import pvbess_opt.pipeline as pipeline

    short = _short_workbook(tmp_path)
    called = {"n": 0}

    def _boom(*_a, **_k):
        called["n"] += 1
        raise RuntimeError("synthetic SOH-trajectory plot failure")

    monkeypatch.setattr(pipeline, "plot_soh_trajectory", _boom)

    result = run(RunConfig(
        excel=short, solver="highs", outdir=tmp_path / "results",
        mip_gap=0.05, time_limit=180,
    ))
    # The guarded plot was actually reached (the projection spans the
    # lifecycle regardless of the one-day dispatch slice) ...
    assert called["n"] >= 1, "plot_soh_trajectory must have been reached"
    # ... yet the run completed and the primary outputs are on disk.
    assert isinstance(result, Results)
    assert result.kpis
    assert (result.out_dir / "03_results.xlsx").exists()
    assert (result.out_dir / "00_summary" / "SUMMARY.md").exists()


@pytest.mark.skipif(not _highs_available(), reason="HiGHS solver not installed")
def test_cli_main_smoke(tmp_path):
    from pvbess_opt import cli

    rc = cli.main([
        str(_short_workbook(tmp_path)),
        "--solver", "highs",
        "--outdir", str(tmp_path / "cli_results"),
        "--mip-gap", "0.05",
        "--time-limit", "180",
    ])
    assert rc == 0


def test_revenue_leg_factors_and_synthetic_year_kpis():
    """Later project years' revenue plots reconcile via per-leg FLAT factors
    from Year 1 (derated KPI / raw sum) applied to each year's OWN raw sums
    — not by rescaling year-N dispatch to Year-1 totals."""
    import pandas as pd

    from pvbess_opt.pipeline import (
        _revenue_leg_factors,
        _synthetic_year_kpis,
    )

    res1 = pd.DataFrame({
        "profit_export_from_pv_eur": [60.0, 40.0],       # raw 100
        "profit_export_from_bess_eur": [30.0, 20.0],     # raw 50
        "expense_charge_bess_grid_eur": [6.0, 4.0],      # raw 10
    })
    year1_kpis = {
        # export legs derated 0.9x, withdrawal 0.95x
        "profit_export_from_pv_eur": 90.0,
        "profit_export_from_bess_eur": 45.0,
        "expense_charge_bess_grid_eur": 9.5,
    }
    f = _revenue_leg_factors(res1, year1_kpis)
    assert f["profit_export_from_pv_eur"] == pytest.approx(0.9)
    assert f["expense_charge_bess_grid_eur"] == pytest.approx(0.95)

    # Year N with its own (degraded) raw sums gets year-N targets.
    resN = pd.DataFrame({
        "profit_export_from_pv_eur": [50.0, 30.0],       # raw 80
        "profit_export_from_bess_eur": [20.0, 20.0],     # raw 40
        "expense_charge_bess_grid_eur": [5.0, 3.0],      # raw 8
    })
    synth = _synthetic_year_kpis(resN, f)
    assert synth is not None
    assert synth["profit_export_from_pv_eur"] == pytest.approx(72.0)  # 80*0.9
    assert synth["expense_charge_bess_grid_eur"] == pytest.approx(7.6)
    # No factors (no KPI dict) -> None, callers keep the raw scale.
    assert _synthetic_year_kpis(resN, {}) is None
    assert _revenue_leg_factors(res1, None) == {}


def test_materialize_external_pv_snapshot(tmp_path):
    """A snapshot from an external timeseries_path PV run must become
    self-contained: the resolved pv_kwh column written in, the path cell
    blanked — so the snapshot re-runs without the external file."""
    import numpy as np
    import pandas as pd

    from pvbess_opt.io import read_workbook, write_workbook
    from pvbess_opt.pipeline import _materialize_external_pv_snapshot

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    # Source workbook: empty pv column + a filled timeseries_path cell.
    src_typed = {k: (v.copy() if hasattr(v, "copy") else v)
                 for k, v in typed.items()}
    src_typed["pv"] = dict(typed["pv"], timeseries_path="pv_profile.csv")
    src_typed["ts"] = typed["ts"].copy()
    src_typed["ts"]["pv_kwh"] = np.nan
    src = tmp_path / "src.xlsx"
    write_workbook(src_typed, src)
    # Snapshot: a verbatim copy of the source (the pre-fix state).
    snap = tmp_path / "input_snapshot.xlsx"
    snap.write_bytes(src.read_bytes())
    # The run's RESOLVED frame carries the external profile.
    resolved_ts = typed["ts"].copy()
    resolved_ts["pv_kwh"] = 7.25

    _materialize_external_pv_snapshot(snap, src, resolved_ts)

    snap_pv = pd.read_excel(snap, sheet_name="pv")
    row = snap_pv[snap_pv["key"] == "timeseries_path"]
    assert row["value"].isna().all(), "path cell must be blanked"
    snap_ts = pd.read_excel(snap, sheet_name="timeseries")
    assert float(snap_ts["pv_kwh"].iloc[0]) == pytest.approx(7.25)
    assert int(snap_ts["pv_kwh"].notna().sum()) == 96


def test_tee_log_captures_logging_output(tmp_path):
    """run_log.txt must capture logging-module WARNINGs even when a CLI-style
    handler bound the pre-tee stderr (the docs promise full capture)."""
    import logging

    from pvbess_opt.pipeline import _tee_stdout_to_log

    log_path = tmp_path / "run_log.txt"
    logger = logging.getLogger("pvbess_opt.test_tee")
    with _tee_stdout_to_log(log_path):
        print("plain stdout line")
        logger.warning("captured-warning-marker")
    text = log_path.read_text(encoding="utf-8")
    assert "plain stdout line" in text
    assert "captured-warning-marker" in text


# --- round-12 guard refinements --------------------------------------------


def test_run_log_captures_pre_tee_loader_warnings(tmp_path):
    """Loader warnings fire in run() BEFORE the run log exists; the early
    buffer must replay them into the archived log when the tee opens (the
    round-11 handler alone only captured in-tee logging)."""
    import logging

    from pvbess_opt.pipeline import _EarlyLogBuffer, _tee_stdout_to_log

    logger = logging.getLogger("pvbess_opt.test_early")
    root = logging.getLogger()
    baseline_handlers = list(root.handlers)
    buf = _EarlyLogBuffer()
    root.addHandler(buf)
    try:
        # Fired BEFORE the log file exists (the read_inputs position).
        logger.warning("pre-tee loader warning: column X is empty")
        log_path = tmp_path / "run_log.txt"
        with _tee_stdout_to_log(log_path, early_buffer=buf):
            logger.warning("in-tee warning")
        text = log_path.read_text(encoding="utf-8")
        assert "pre-tee loader warning: column X is empty" in text
        assert "in-tee warning" in text
        # The tee detached the buffer (no double capture on a later tee).
        assert buf not in root.handlers
    finally:
        root.removeHandler(buf)
        assert list(root.handlers) == baseline_handlers


def test_materialize_external_pv_snapshot_fires_only_when_path_won(tmp_path):
    """The materialiser must mirror the resolver's precedence: a FILLED
    pv_kwh column wins over the path, and pv_source='pvgis' fetches instead
    of reading the file — in both cases the snapshot must stay untouched
    (previously the path cell was destroyed and, for pvgis, a false
    'self-contained' claim logged while the re-run still fetched)."""
    import numpy as np
    import pandas as pd

    from pvbess_opt.io import read_workbook, write_workbook
    from pvbess_opt.pipeline import _materialize_external_pv_snapshot

    typed = read_workbook(ROOT / "inputs" / "input.xlsx")
    typed["ts"] = typed["ts"].iloc[:96].reset_index(drop=True)
    resolved_ts = typed["ts"].copy()
    resolved_ts["pv_kwh"] = 7.25

    # Case 1: filled pv_kwh column + stale path -> the column won; no-op.
    src_typed = {k: (v.copy() if hasattr(v, "copy") else v)
                 for k, v in typed.items()}
    src_typed["pv"] = dict(typed["pv"], timeseries_path="stale.csv")
    src = tmp_path / "col_wins.xlsx"
    write_workbook(src_typed, src)
    snap = tmp_path / "snap1.xlsx"
    snap.write_bytes(src.read_bytes())
    _materialize_external_pv_snapshot(snap, src, resolved_ts)
    snap_pv = pd.read_excel(snap, sheet_name="pv")
    row = snap_pv[snap_pv["key"] == "timeseries_path"]
    assert row["value"].iloc[0] == "stale.csv", "path cell must survive"

    # Case 2: pv_source='pvgis' + stale path -> the fetch won; no-op.
    src_typed2 = {k: (v.copy() if hasattr(v, "copy") else v)
                  for k, v in typed.items()}
    src_typed2["pv"] = dict(
        typed["pv"], timeseries_path="stale.csv", pv_source="pvgis",
    )
    src_typed2["ts"] = typed["ts"].copy()
    src_typed2["ts"]["pv_kwh"] = np.nan
    src2 = tmp_path / "pvgis_wins.xlsx"
    write_workbook(src_typed2, src2)
    snap2 = tmp_path / "snap2.xlsx"
    snap2.write_bytes(src2.read_bytes())
    _materialize_external_pv_snapshot(snap2, src2, resolved_ts)
    snap_pv2 = pd.read_excel(snap2, sheet_name="pv")
    row2 = snap_pv2[snap_pv2["key"] == "timeseries_path"]
    assert row2["value"].iloc[0] == "stale.csv", "path cell must survive"


def test_generate_all_energy_plots_threads_synthetic_kpis_to_year2(
    tmp_path, monkeypatch,
):
    """The critic's top surviving mutant: reverting the lifetime loop's
    year-2+ ``year1_kpis`` argument back to None left every unit test
    green.  This wiring test drives _generate_all_energy_plots itself and
    asserts the year-2 plot call receives the SYNTHETIC dict (raw sums x
    Year-1 derate factor), not None."""
    import pandas as pd

    from pvbess_opt import pipeline as pl

    idx1 = pd.date_range("2026-01-01", periods=48, freq="h")
    res_year1 = pd.DataFrame({
        "timestamp": idx1,
        "profit_export_from_pv_eur": 10.0,
        "profit_export_from_bess_eur": 5.0,
        "expense_charge_bess_grid_eur": 2.0,
        "revenue_pv_ppa_eur": 0.0,
    })
    # Derated Year-1 KPIs at 90% of raw.
    year1_kpis = {
        "profit_export_from_pv_eur": 10.0 * 48 * 0.9,
        "profit_export_from_bess_eur": 5.0 * 48 * 0.9,
        "expense_charge_bess_grid_eur": 2.0 * 48 * 0.9,
        "revenue_pv_ppa_eur": 0.0,
    }
    idx2 = pd.date_range("2027-01-01", periods=48, freq="h")
    lifetime_df = pd.concat([
        res_year1.assign(project_year=1, calendar_year=2026),
        res_year1.assign(timestamp=idx2, project_year=2, calendar_year=2027),
    ], ignore_index=True)

    seen: dict[int, object] = {}

    def _spy(res, cal_year, out_dir, *args, **kwargs):
        seen[int(cal_year)] = kwargs.get("year1_kpis")

    monkeypatch.setattr(pl, "_generate_energy_plots_for_year", _spy)
    pl._generate_all_energy_plots(
        res_year1, lifetime_df, None, {"plot_daily_scope": "none"},
        tmp_path, mode="merchant", year1_kpis=year1_kpis,
    )
    assert seen[2026] is year1_kpis or seen[2026] == year1_kpis
    y2 = seen[2027]
    assert isinstance(y2, dict), "year-2 plot call must get synthetic KPIs"
    assert y2["profit_export_from_pv_eur"] == pytest.approx(10.0 * 48 * 0.9)
    assert y2["expense_charge_bess_grid_eur"] == pytest.approx(2.0 * 48 * 0.9)
