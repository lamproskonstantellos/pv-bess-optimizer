"""Style checks for the canonical input workbook.

The shipped workbook carries exactly one global accent: row 1 of every
sheet is bold + white text + filled navy ``#1F3864`` + thin ``#BFBFBF``
bottom border, with the row frozen via ``freeze_panes = "A2"``. No
other cells carry per-sheet styling. In particular, no cell anywhere
has the prior amber bootstrap fill ``#FFF2CC``.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from pvbess_opt.theme import HEADER_BORDER_HEX

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "inputs" / "input.xlsx"

EXPECTED_HEADER_FILL = "1F3864"
EXPECTED_HEADER_FONT = "FFFFFF"
# Single-sourced from the styler so the check cannot drift from what the
# workbook writer applies.
EXPECTED_HEADER_BORDER = HEADER_BORDER_HEX
FORBIDDEN_FILL_HEX = "FFF2CC"


def _normalise_rgb(rgb: object) -> str:
    if not isinstance(rgb, str):
        return ""
    return rgb.upper().lstrip("0").rjust(6, "0")[-6:]


@pytest.fixture(scope="module")
def workbook():
    return load_workbook(WORKBOOK)


def test_no_amber_fills_anywhere(workbook):
    """No cell anywhere carries the prior amber bootstrap fill."""
    hits = []
    for sn in workbook.sheetnames:
        ws = workbook[sn]
        for row in ws.iter_rows():
            for c in row:
                fill = c.fill
                if fill is None or fill.fill_type is None:
                    continue
                for attr in ("fgColor", "start_color"):
                    colour = getattr(fill, attr, None)
                    rgb = _normalise_rgb(getattr(colour, "rgb", None))
                    if rgb == FORBIDDEN_FILL_HEX:
                        hits.append((sn, c.coordinate))
                        break
    assert not hits, f"Prior amber fill remains in: {hits[:5]}..."


@pytest.mark.parametrize("sheet_name", [
    "timeseries", "project", "pv", "bess", "economics",
    "simulation", "max_injection_profile",
    "max_injection_profile_pv", "max_injection_profile_bess", "balancing",
])
def test_header_row_has_global_accent(workbook, sheet_name):
    """Row 1 of every sheet is bold + white + filled navy ``1F3864``."""
    ws = workbook[sheet_name]
    assert ws.max_row >= 1
    for cell in ws[1]:
        if cell.value is None:
            continue
        assert cell.font.bold is True, (
            f"{sheet_name}!{cell.coordinate}: header not bold"
        )
        rgb = _normalise_rgb(
            getattr(cell.fill.fgColor, "rgb", None)
            or getattr(cell.fill.start_color, "rgb", None)
        )
        assert rgb == EXPECTED_HEADER_FILL, (
            f"{sheet_name}!{cell.coordinate}: fill={rgb!r}, "
            f"expected {EXPECTED_HEADER_FILL!r}"
        )
        font_rgb = _normalise_rgb(getattr(cell.font.color, "rgb", None))
        assert font_rgb == EXPECTED_HEADER_FONT, (
            f"{sheet_name}!{cell.coordinate}: font color={font_rgb!r}, "
            f"expected {EXPECTED_HEADER_FONT!r}"
        )
        bottom = cell.border.bottom
        assert bottom is not None and bottom.style == "thin", (
            f"{sheet_name}!{cell.coordinate}: header bottom border "
            f"style={getattr(bottom, 'style', None)!r}, expected 'thin'"
        )
        border_rgb = _normalise_rgb(
            getattr(getattr(bottom, "color", None), "rgb", None)
        )
        assert border_rgb == EXPECTED_HEADER_BORDER, (
            f"{sheet_name}!{cell.coordinate}: bottom border={border_rgb!r}, "
            f"expected {EXPECTED_HEADER_BORDER!r}"
        )


@pytest.mark.parametrize("sheet_name", [
    "timeseries", "project", "pv", "bess", "economics",
    "simulation", "max_injection_profile",
    "max_injection_profile_pv", "max_injection_profile_bess", "balancing",
])
def test_header_row_is_frozen(workbook, sheet_name):
    """Every sheet freezes its header row via ``freeze_panes = 'A2'``."""
    ws = workbook[sheet_name]
    assert ws.freeze_panes == "A2", (
        f"{sheet_name}: freeze_panes={ws.freeze_panes!r}, expected 'A2'"
    )


@pytest.mark.parametrize("sheet_name", [
    "max_injection_profile",
    "max_injection_profile_pv",
    "max_injection_profile_bess",
])
def test_max_injection_trio_headers_centered(workbook, sheet_name):
    """All three max-injection sheets carry center-aligned headers.

    The combined sheet has always been centered (pandas' default header
    style at write time); the two per-asset sheets are centered by
    ``scripts/polish_input_workbook.py`` so the trio reads consistently.
    """
    ws = workbook[sheet_name]
    for cell in ws[1]:
        if cell.value is None:
            continue
        assert cell.alignment.horizontal == "center", (
            f"{sheet_name}!{cell.coordinate}: header alignment "
            f"{cell.alignment.horizontal!r}, expected 'center'"
        )


@pytest.mark.parametrize("sheet_name", [
    "timeseries", "project", "pv", "bess", "economics",
    "simulation", "max_injection_profile",
    "max_injection_profile_pv", "max_injection_profile_bess", "balancing",
])
def test_non_header_rows_carry_no_persistent_fill(workbook, sheet_name):
    """Data rows (row >= 2) carry no per-cell solid fill."""
    ws = workbook[sheet_name]
    offending = []
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 20)):
        for c in row:
            fill = c.fill
            if fill is None or fill.fill_type is None:
                continue
            for attr in ("fgColor", "start_color"):
                colour = getattr(fill, attr, None)
                rgb = _normalise_rgb(getattr(colour, "rgb", None))
                if rgb and rgb != "000000":
                    offending.append((c.coordinate, rgb))
                    break
    assert not offending, (
        f"{sheet_name}: data cells with unexpected fill: {offending[:5]}"
    )
