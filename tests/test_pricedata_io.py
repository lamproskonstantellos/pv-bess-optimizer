"""price_scenarios / scenario_engine sheets: parsing, gating, surfaces."""

from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from pvbess_opt.io import (
    SCENARIO_ENGINE_SHEET_DEFAULTS,
    _normalise_price_scenarios_block,
    read_workbook,
)
from pvbess_opt.io_read import dump_structured_config, load_structured_config
from pvbess_opt.scenarios import validate_scenario_overrides

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "inputs" / "input.xlsx"


def _workbook_with_price_scenarios(
    tmp_path: Path, rows: list[tuple], *, enabled: str = "TRUE",
) -> Path:
    """Copy the shipped workbook and rewrite the price_scenarios rows."""
    dst = tmp_path / "input_ps.xlsx"
    shutil.copy(WORKBOOK, dst)
    wb = load_workbook(dst)
    ws = wb["price_scenarios"]
    ws.delete_rows(2, ws.max_row)
    for index, row in enumerate(rows):
        ws.append([enabled if index == 0 else None, *row])
    wb.save(dst)
    return dst


_GOOD_ROWS = [
    ("Central", "parametric", "2026-07", 60, "stores/central", ""),
    ("Downside", "tyndp", "TYNDP-2026", 40, "stores/tyndp", ""),
]


# ---------------------------------------------------------------------------
# Bit-identity + gating
# ---------------------------------------------------------------------------


def test_sheets_absent_match_defaults(tmp_path):
    stripped = tmp_path / "input_no_sheets.xlsx"
    shutil.copy(WORKBOOK, stripped)
    wb = load_workbook(stripped)
    del wb["scenario_engine"]
    del wb["price_scenarios"]
    wb.save(stripped)

    typed = read_workbook(stripped)
    assert typed["scenario_engine"] == SCENARIO_ENGINE_SHEET_DEFAULTS
    assert typed["price_scenarios"] is None


def test_shipped_workbook_is_disabled_and_inert():
    typed = read_workbook(WORKBOOK)
    assert typed["scenario_engine"]["price_scenarios_enabled"] is False
    assert typed["price_scenarios"] is None


def test_enabled_sheet_parses_rows(tmp_path):
    src = _workbook_with_price_scenarios(tmp_path, _GOOD_ROWS)
    typed = read_workbook(src)
    scenarios = typed["price_scenarios"]
    assert scenarios is not None and len(scenarios) == 2
    assert scenarios[0]["name"] == "Central"
    assert scenarios[0]["provider"] == "parametric"
    assert scenarios[0]["weight_pct"] == 60.0
    assert scenarios[1]["store_path"] == "stores/tyndp"
    # An enabled sheet with the master switch off stays inert data.
    assert typed["scenario_engine"]["price_scenarios_enabled"] is False


def test_disabled_sheet_returns_none_without_validation(tmp_path):
    # Weights that do NOT sum to 100 pass while the sheet is disabled —
    # the rows are inert.
    src = _workbook_with_price_scenarios(
        tmp_path,
        [("Central", "parametric", "v", 10, "stores/central", "")],
        enabled="FALSE",
    )
    typed = read_workbook(src)
    assert typed["price_scenarios"] is None


# ---------------------------------------------------------------------------
# Validation errors (precise)
# ---------------------------------------------------------------------------


def test_weights_must_sum_to_100(tmp_path):
    src = _workbook_with_price_scenarios(tmp_path, [
        ("Central", "parametric", "v", 60, "stores/central", ""),
        ("Downside", "tyndp", "v", 60, "stores/tyndp", ""),
    ])
    with pytest.raises(ValueError, match="sum to 100"):
        read_workbook(src)


def test_duplicate_names_rejected(tmp_path):
    src = _workbook_with_price_scenarios(tmp_path, [
        ("Central", "parametric", "v", 50, "stores/a", ""),
        ("Central", "tyndp", "v", 50, "stores/b", ""),
    ])
    with pytest.raises(ValueError, match="duplicate"):
        read_workbook(src)


def test_unknown_provider_rejected(tmp_path):
    src = _workbook_with_price_scenarios(tmp_path, [
        ("Central", "vnb", "v", 100, "stores/a", ""),
    ])
    with pytest.raises(ValueError, match="vnb"):
        read_workbook(src)


def test_missing_store_path_rejected(tmp_path):
    src = _workbook_with_price_scenarios(tmp_path, [
        ("Central", "parametric", "v", 100, None, ""),
    ])
    with pytest.raises(ValueError, match="store_path"):
        read_workbook(src)


def test_negative_weight_rejected():
    with pytest.raises(ValueError, match=">= 0"):
        _normalise_price_scenarios_block(
            [{
                "name": "X", "provider": "file", "weight_pct": -5,
                "store_path": "s",
            }],
            source="test",
        )


def test_nameless_entry_rejected():
    with pytest.raises(ValueError, match="no name"):
        _normalise_price_scenarios_block(
            [{"provider": "file", "weight_pct": 100, "store_path": "s"}],
            source="test",
        )


# ---------------------------------------------------------------------------
# YAML surface + scenario overrides
# ---------------------------------------------------------------------------


def test_yaml_roundtrip_carries_both_surfaces(tmp_path):
    src = _workbook_with_price_scenarios(tmp_path, _GOOD_ROWS)
    typed = read_workbook(src)
    config_path = tmp_path / "config.yaml"
    dump_structured_config(typed, config_path)
    loaded = load_structured_config(config_path)
    assert loaded["scenario_engine"] == SCENARIO_ENGINE_SHEET_DEFAULTS
    assert loaded["price_scenarios"] == typed["price_scenarios"]


def test_yaml_without_block_roundtrips_to_none(tmp_path):
    typed = read_workbook(WORKBOOK)
    config_path = tmp_path / "config.yaml"
    dump_structured_config(typed, config_path)
    raw_text = config_path.read_text(encoding="utf-8")
    # No top-level block is emitted (the scenario_engine KEY
    # price_scenarios_enabled legitimately contains the substring).
    assert "\nprice_scenarios:" not in raw_text
    loaded = load_structured_config(config_path)
    assert loaded["price_scenarios"] is None


def test_scenario_overrides_accept_engine_targets():
    validate_scenario_overrides({
        "name": "scenario prices on",
        "scenario_engine": {
            "price_scenarios_enabled": True,
            "scenario_projection_mode": "reprice",
        },
    })
    with pytest.raises(ValueError, match="unknown"):
        validate_scenario_overrides({
            "name": "typo",
            "scenario_engine": {"price_scenario_enabled": True},
        })


def test_write_workbook_roundtrips_enabled_rows(tmp_path):
    from pvbess_opt.io import write_workbook

    src = _workbook_with_price_scenarios(tmp_path, _GOOD_ROWS)
    typed = read_workbook(src)
    dst = tmp_path / "roundtrip.xlsx"
    write_workbook(typed, dst)
    again = read_workbook(dst)
    assert again["price_scenarios"] == typed["price_scenarios"]
    sheet = pd.read_excel(dst, sheet_name="price_scenarios")
    assert str(sheet.iloc[0]["enabled"]).upper() == "TRUE"


# ---------------------------------------------------------------------------
# Output writers (results workbook sheet + SUMMARY digest)
# ---------------------------------------------------------------------------


def test_results_workbook_carries_scenario_paths_sheet(tmp_path):
    from pvbess_opt.io import write_results_workbook

    paths = pd.DataFrame({
        "scenario": ["Central", "Central"],
        "applied": [True, True],
        "weight_pct": [100.0, 100.0],
        "project_year": [1, 2],
        "dam_mean_price_eur_per_mwh": [80.0, 72.0],
    })
    res = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=4, freq="h"),
        "pv_kwh": [1.0, 2.0, 3.0, 4.0],
    })
    out = write_results_workbook(
        tmp_path / "03_results.xlsx",
        res_year1=res,
        kpis_year1={"profit_total_eur": 1.0},
        kpis_monthly_year1=None,
        scenario_price_paths=paths,
    )
    sheet = pd.read_excel(out, sheet_name="scenario_price_paths")
    assert list(sheet["scenario"]) == ["Central", "Central"]
    # Disarmed: the sheet is absent (bit-identical workbook).
    out2 = write_results_workbook(
        tmp_path / "03_results_off.xlsx",
        res_year1=res,
        kpis_year1={"profit_total_eur": 1.0},
        kpis_monthly_year1=None,
    )
    assert "scenario_price_paths" not in pd.ExcelFile(out2).sheet_names


def test_results_workbook_carries_resolve_delta_sheet(tmp_path):
    from pvbess_opt.io import write_results_workbook

    delta = pd.DataFrame({
        "project_year": [1, 3],
        "stream": ["revenue_dam_pv", "revenue_dam_pv"],
        "g_tier1_reprice": [1.0, 0.81],
        "g_tier2_resolve": [1.0, 0.84],
        "delta": [0.0, 0.03],
        "delta_pct": [0.0, 3.7037],
    })
    res = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=4, freq="h"),
        "pv_kwh": [1.0, 2.0, 3.0, 4.0],
    })
    out = write_results_workbook(
        tmp_path / "03_results.xlsx",
        res_year1=res,
        kpis_year1={"profit_total_eur": 1.0},
        kpis_monthly_year1=None,
        scenario_resolve_delta=delta,
    )
    sheet = pd.read_excel(out, sheet_name="scenario_resolve_delta")
    assert list(sheet["project_year"]) == [1, 3]
    # Reprice mode / disarmed: the sheet is absent.
    out2 = write_results_workbook(
        tmp_path / "03_results_off.xlsx",
        res_year1=res,
        kpis_year1={"profit_total_eur": 1.0},
        kpis_monthly_year1=None,
    )
    assert "scenario_resolve_delta" not in pd.ExcelFile(out2).sheet_names


def test_results_workbook_carries_ensemble_sheet(tmp_path):
    from pvbess_opt.io import write_results_workbook

    ensemble = pd.DataFrame({
        "scenario": ["Central", "Downside", "E[NPV]"],
        "weight_pct": [60.0, 40.0, None],
        "npv_eur": [300.0, 100.0, 220.0],
    })
    res = pd.DataFrame({
        "timestamp": pd.date_range("2026-01-01", periods=4, freq="h"),
        "pv_kwh": [1.0, 2.0, 3.0, 4.0],
    })
    out = write_results_workbook(
        tmp_path / "03_results.xlsx",
        res_year1=res,
        kpis_year1={"profit_total_eur": 1.0},
        kpis_monthly_year1=None,
        price_scenario_ensemble=ensemble,
    )
    sheet = pd.read_excel(out, sheet_name="price_scenario_ensemble")
    assert list(sheet["scenario"]) == ["Central", "Downside", "E[NPV]"]
    out2 = write_results_workbook(
        tmp_path / "03_results_off.xlsx",
        res_year1=res,
        kpis_year1={"profit_total_eur": 1.0},
        kpis_monthly_year1=None,
    )
    assert (
        "price_scenario_ensemble" not in pd.ExcelFile(out2).sheet_names
    )


def test_summary_gating_of_price_scenario_lines(tmp_path):
    from pvbess_opt.io import write_summary_md

    common = dict(
        kpis_year1={"profit_total_eur": 1.0},
        financial_kpis=None,
        params={"mode": "merchant", "pv_nameplate_kwp": 1000.0},
    )
    armed = write_summary_md(
        tmp_path / "SUMMARY_armed.md",
        price_scenario_lines=[
            "- Price scenarios: `reprice` on `Central` (2 enabled "
            "scenario(s))",
        ],
        **common,
    )
    assert "Price scenarios" in armed.read_text(encoding="utf-8")
    plain = write_summary_md(tmp_path / "SUMMARY_plain.md", **common)
    assert "Price scenarios" not in plain.read_text(encoding="utf-8")


def test_scenario_interp_accepts_linear(tmp_path):
    """'linear' is a selectable workbook value (not just the internal
    non-positive-factor fallback)."""
    dst = tmp_path / "linear.xlsx"
    shutil.copy(WORKBOOK, dst)
    wb = load_workbook(dst)
    ws = wb["scenario_engine"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "scenario_interp":
            row[1].value = "linear"
    wb.save(dst)
    typed = read_workbook(dst)
    assert typed["scenario_engine"]["scenario_interp"] == "linear"
