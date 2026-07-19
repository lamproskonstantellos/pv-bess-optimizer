"""ADMIE / IPTO file-API provider (Greek balancing + imbalance prices).

The Greek balancing domain on the ENTSO-E platform is effectively empty
(Greece runs a co-optimised Integrated Scheduling Process; 17.1.B/C/F/G
return no data), so Greek balancing/imbalance prices come from the
ADMIE operation-market file API::

    https://www.admie.gr/getOperationMarketFile
        ?dateStart=YYYY-MM-DD&dateEnd=YYYY-MM-DD&FileCategory=<TYPE>

which answers a JSON list of daily xlsx file URLs.  Each daily workbook
is downloaded and parsed by a **declarative header-pattern map** into
the model's balancing price columns.

PROVISIONAL CONTRACT — the exact ``FileCategory`` tokens and workbook
column headers are not publicly documented; the shipped values below
define the parsing contract the test fixtures follow and are meant to
be pinned against real files via ``scripts/probe_market_data.py
--save-dir`` on a network-enabled machine (this environment's egress
policy blocks admie.gr — see ``docs/notes/market_data_spike.md``).
Every token and pattern is a module constant (or a function argument)
so pinning is a data change, not a code change.

Timestamps in ADMIE files are Greek local wall-clock, so the assembly
is calendar-native: daily blocks concatenate directly onto the model's
local grid — no UTC round-trip.  DST days are normalised to the
nominal 24-hour block with the same semantics as the UTC path
(spring-forward: the skipped 03:00 block repeats the previous value;
fall-back: the second 03:00 block is dropped), and Feb 29 of a leap
reference year is dropped.
"""

from __future__ import annotations

import io
import json
import logging
import re
from calendar import isleap
from datetime import UTC, date, datetime, timedelta
from typing import Any

import numpy as np

from .base import (
    MarketDataCache,
    MarketDataError,
    MarketSeries,
    PriceSegment,
    utcnow_isoformat,
)

logger = logging.getLogger(__name__)

ADMIE_FILE_API_URL = "https://www.admie.gr/getOperationMarketFile"

# PROVISIONAL — pin via the probe script (see module docstring).
ADMIE_CATEGORY_BALANCING = "ISP1ISPResults"
ADMIE_CATEGORY_IMBALANCE = "IMBABE"

# Balancing price columns the GR provider serves, with the header
# pattern that locates each one inside a daily workbook (case-
# insensitive, matched against the concatenated header row text of
# every sheet).  FCR is deliberately absent: Greece procures no
# standalone FCR (co-optimised ISP), so the FCR columns stay on their
# workbook / scalar-fallback path and the resolver says so.
BALANCING_HEADER_PATTERNS: dict[str, str] = {
    "afrr_up_capacity_price_eur_per_mwh": r"afrr.*up.*capacity.*price",
    "afrr_dn_capacity_price_eur_per_mwh": r"afrr.*down.*capacity.*price",
    "mfrr_up_capacity_price_eur_per_mwh": r"mfrr.*up.*capacity.*price",
    "mfrr_dn_capacity_price_eur_per_mwh": r"mfrr.*down.*capacity.*price",
    "afrr_up_activation_price_eur_per_mwh": r"afrr.*up.*activation.*price",
    "afrr_dn_activation_price_eur_per_mwh": r"afrr.*down.*activation.*price",
    "mfrr_up_activation_price_eur_per_mwh": r"mfrr.*up.*activation.*price",
    "mfrr_dn_activation_price_eur_per_mwh": r"mfrr.*down.*activation.*price",
}

IMBALANCE_HEADER_PATTERNS: dict[str, str] = {
    "imbalance_price_eur_per_mwh": r"imbalance.*price",
}

# Local DST transition block: Greece skips/repeats the 03:00 hour.
_DST_HOUR = 3

# The local-native path never uses PriceSegment.start_utc; a fixed
# sentinel keeps the shared cache payload shape.
_LOCAL_NATIVE_SENTINEL = datetime(1970, 1, 1, tzinfo=UTC)


def _http_get(url: str, params: dict[str, str] | None,
              timeout: float) -> tuple[int, bytes]:
    """One GET; tests monkeypatch this symbol."""
    import requests

    resp = requests.get(url, params=params, timeout=timeout)
    return resp.status_code, resp.content


def list_daily_files(
    category: str, start: date, end: date, *, timeout: float = 60.0,
) -> list[str]:
    """Return the daily file URLs for ``category`` over [start, end]."""
    status, body = _http_get(
        ADMIE_FILE_API_URL,
        {
            "dateStart": start.isoformat(),
            "dateEnd": end.isoformat(),
            "FileCategory": category,
        },
        timeout,
    )
    if status != 200:
        raise MarketDataError(
            f"ADMIE file listing for {category!r} failed with HTTP "
            f"{status}."
        )
    try:
        listing = json.loads(body.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise MarketDataError(
            f"ADMIE file listing for {category!r} is not JSON "
            f"(first bytes {body[:60]!r})."
        ) from exc
    urls: list[str] = []
    for entry in listing if isinstance(listing, list) else []:
        if isinstance(entry, dict):
            url = entry.get("file_path") or entry.get("file_url")
            if isinstance(url, str) and url:
                urls.append(url)
        elif isinstance(entry, str):
            urls.append(entry)
    if not urls:
        raise MarketDataError(
            f"ADMIE published no {category!r} files between "
            f"{start.isoformat()} and {end.isoformat()}; the category "
            "token may need re-pinning via scripts/probe_market_data.py."
        )
    return urls


def parse_daily_workbook(
    content: bytes, patterns: dict[str, str], *, source_name: str,
) -> dict[str, np.ndarray]:
    """Parse one daily xlsx into ``{column: values}`` via header patterns.

    Scans every sheet for a header row matching each pattern; the
    column's numeric cells below the header become the day's values.
    All matched columns must agree on the period count (the native
    resolution is inferred from it by the caller).
    """
    from openpyxl import load_workbook

    wb = load_workbook(
        io.BytesIO(content), read_only=True, data_only=True,
    )
    compiled = {
        column: re.compile(pattern, re.IGNORECASE)
        for column, pattern in patterns.items()
    }
    out: dict[str, np.ndarray] = {}
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header: tuple[Any, ...] | None = None
            data_rows: list[tuple[Any, ...]] = []
            for row in rows:
                if header is None:
                    texts = [str(c) for c in row if isinstance(c, str)]
                    if texts and any(
                        rx.search(" ".join(texts))
                        for rx in compiled.values()
                    ):
                        header = row
                    continue
                data_rows.append(row)
            if header is None:
                continue
            for column, rx in compiled.items():
                if column in out:
                    continue
                for idx, cell in enumerate(header):
                    if isinstance(cell, str) and rx.search(cell):
                        values = [
                            float(r[idx]) for r in data_rows
                            if idx < len(r)
                            and isinstance(r[idx], (int, float))
                        ]
                        if values:
                            out[column] = np.asarray(values, dtype=float)
                        break
    finally:
        wb.close()
    if not out:
        raise MarketDataError(
            f"{source_name}: no column of the declared header map "
            f"({', '.join(patterns)}) was found; the header patterns "
            "may need re-pinning via scripts/probe_market_data.py."
        )
    lengths = {column: len(v) for column, v in out.items()}
    if len(set(lengths.values())) > 1:
        raise MarketDataError(
            f"{source_name}: matched columns disagree on the period "
            f"count: {lengths}."
        )
    return out


def _nominal_day(
    values: np.ndarray, steps_per_hour: int, *, kind: str,
    source_name: str,
) -> np.ndarray:
    """Normalise one local day block to its nominal 24-hour length.

    ``kind`` is ``normal`` / ``spring`` / ``fall``.  A spring-forward
    file legitimately carries 23 hours (the skipped 03:00 block is
    re-inserted by repeating the 02:00 values); a fall-back file 25
    hours (the second 03:00 block — the repeat — is dropped).  Files
    that already publish nominal 24-hour blocks on transition days
    pass through unchanged.
    """
    nominal = 24 * steps_per_hour
    if len(values) == nominal:
        return values
    if kind == "spring" and len(values) == nominal - steps_per_hour:
        at = _DST_HOUR * steps_per_hour
        fill = values[at - steps_per_hour:at]
        return np.concatenate([values[:at], fill, values[at:]])
    if kind == "fall" and len(values) == nominal + steps_per_hour:
        drop_from = (_DST_HOUR + 1) * steps_per_hour
        return np.concatenate(
            [values[:drop_from], values[drop_from + steps_per_hour:]],
        )
    raise MarketDataError(
        f"{source_name}: day block carries {len(values)} values, "
        f"expected {nominal} (or the +/- one-hour DST variants)."
    )


def _dst_transition_dates(year: int) -> tuple[date, date]:
    """Last Sundays of March and October (EU transition days)."""
    def last_sunday(month: int) -> date:
        d = date(year, month + 1, 1) - timedelta(days=1)
        return d - timedelta(days=(d.weekday() + 1) % 7)

    return last_sunday(3), last_sunday(10)


def assemble_year_local(
    daily: list[tuple[date, dict[str, np.ndarray]]],
    year: int,
    *,
    source_name: str,
) -> dict[str, tuple[int, np.ndarray]]:
    """Concatenate daily blocks onto the non-leap local year grid.

    Returns ``{column: (native_minutes, values)}`` with exactly
    365 nominal days per column.  Every calendar day of ``year`` must
    appear exactly once (Feb 29 of a leap year is dropped); a missing
    or duplicated day is a hard error — partial data must never be
    silently mixed.
    """
    spring, fall = _dst_transition_dates(year)
    by_day: dict[date, dict[str, np.ndarray]] = {}
    for day, columns in daily:
        if day in by_day:
            raise MarketDataError(
                f"{source_name}: duplicate daily file for "
                f"{day.isoformat()}."
            )
        by_day[day] = columns

    expected_days = [
        date(year, 1, 1) + timedelta(days=n)
        for n in range(366 if isleap(year) else 365)
    ]
    missing = [
        d for d in expected_days
        if d not in by_day and not (d.month == 2 and d.day == 29)
    ]
    if missing:
        raise MarketDataError(
            f"{source_name}: {len(missing)} daily file(s) missing for "
            f"{year}; first missing day {missing[0].isoformat()}."
        )

    column_names = sorted({c for _, cols in daily for c in cols})
    out: dict[str, tuple[int, np.ndarray]] = {}
    for column in column_names:
        blocks: list[np.ndarray] = []
        steps_per_hour: int | None = None
        for day in expected_days:
            if day.month == 2 and day.day == 29:
                continue  # industry 8760 convention, like the UTC path
            day_columns = by_day[day]
            if column not in day_columns:
                raise MarketDataError(
                    f"{source_name}: column {column!r} absent from the "
                    f"{day.isoformat()} file; refusing a partial year."
                )
            values = day_columns[column]
            if steps_per_hour is None:
                kind = (
                    "spring" if day == spring
                    else "fall" if day == fall else "normal"
                )
                # Infer the native cadence from the first day's count
                # (transition days corrected by their +/- one hour).
                n = len(values)
                if kind == "spring":
                    n += n // 23
                elif kind == "fall":
                    n -= n // 25
                if n % 24 != 0 or (60 % (n // 24)) != 0:
                    raise MarketDataError(
                        f"{source_name}: cannot infer a whole-minute "
                        f"cadence from {len(values)} values on "
                        f"{day.isoformat()}."
                    )
                steps_per_hour = n // 24
            kind = (
                "spring" if day == spring
                else "fall" if day == fall else "normal"
            )
            blocks.append(
                _nominal_day(
                    values, steps_per_hour, kind=kind,
                    source_name=f"{source_name} {day.isoformat()}",
                )
            )
        series = np.concatenate(blocks)
        assert steps_per_hour is not None
        out[column] = (60 // steps_per_hour, series)
    return out


def _fetch_category_year(
    category: str,
    patterns: dict[str, str],
    year: int,
    *,
    dataset: str,
    cache: MarketDataCache,
    fetch_mode: str,
    timeout: float,
) -> MarketSeries:
    """Cacheable fetch of one ADMIE category assembled over one year.

    The assembled per-column year series ride the shared cache as
    pseudo-segments: ``PriceSegment.start_utc`` is unused on this
    local-native path, so each column is stored as one segment whose
    metadata names the column and native cadence.
    """
    cache_key = MarketDataCache.key(dataset, "GR", year, category=category)
    mode = str(fetch_mode).strip().lower()
    if mode in ("cache_first", "offline"):
        cached = cache.load(cache_key)
        if cached is not None:
            cached.metadata["cache_state"] = "cache hit"
            cached.metadata.setdefault("cache_key", cache_key)
            return cached
        if mode == "offline":
            raise MarketDataError(
                "market_fetch_mode='offline' but the required cache "
                f"entry is missing: {cache.path(cache_key)} (cache key "
                f"{cache_key}). Run once with network access "
                "(cache_first) or copy the cache file in."
            )

    urls = list_daily_files(
        category, date(year, 1, 1), date(year, 12, 31), timeout=timeout,
    )
    daily: list[tuple[date, dict[str, np.ndarray]]] = []
    for url in urls:
        day = _date_from_admie_url(url)
        if day is None or day.year != year:
            continue
        status, content = _http_get(url, None, timeout)
        if status != 200:
            raise MarketDataError(
                f"ADMIE daily file {url} failed with HTTP {status}."
            )
        daily.append((
            day,
            parse_daily_workbook(content, patterns, source_name=url),
        ))
    assembled = assemble_year_local(
        daily, year, source_name=f"ADMIE {category}",
    )
    series = MarketSeries(
        segments=[
            PriceSegment(
                start_utc=_LOCAL_NATIVE_SENTINEL,
                resolution_minutes=native_minutes,
                values=values.tolist(),
            )
            for _column, (native_minutes, values) in sorted(
                assembled.items(),
            )
        ],
        metadata={
            "source": "admie",
            "dataset": dataset,
            "category": category,
            "zone": "GR",
            "year": int(year),
            "columns": sorted(assembled),
            "fetched_at": utcnow_isoformat(),
            "cache_key": cache_key,
            "cache_state": "live fetch" if mode != "refresh" else "refresh",
        },
    )
    cache.save(cache_key, series)
    return series


def fetch_gr_balancing_year(
    year: int,
    *,
    cache: MarketDataCache,
    fetch_mode: str = "cache_first",
    timeout: float = 60.0,
    category: str = ADMIE_CATEGORY_BALANCING,
) -> MarketSeries:
    """Greek aFRR/mFRR capacity + activation prices for one year."""
    return _fetch_category_year(
        category, BALANCING_HEADER_PATTERNS, year,
        dataset="gr-balancing", cache=cache, fetch_mode=fetch_mode,
        timeout=timeout,
    )


def fetch_gr_imbalance_year(
    year: int,
    *,
    cache: MarketDataCache,
    fetch_mode: str = "cache_first",
    timeout: float = 60.0,
    category: str = ADMIE_CATEGORY_IMBALANCE,
) -> MarketSeries:
    """Greek imbalance settlement prices for one year."""
    return _fetch_category_year(
        category, IMBALANCE_HEADER_PATTERNS, year,
        dataset="gr-imbalance", cache=cache, fetch_mode=fetch_mode,
        timeout=timeout,
    )


def series_columns(series: MarketSeries) -> dict[str, tuple[int, np.ndarray]]:
    """Rebuild ``{column: (native_minutes, values)}`` from a MarketSeries."""
    columns = [str(c) for c in series.metadata.get("columns", [])]
    if len(columns) != len(series.segments):
        raise MarketDataError(
            "ADMIE cache payload is inconsistent (columns metadata does "
            "not match the stored segments); delete the cache entry "
            f"{series.metadata.get('cache_key', '?')} and re-fetch."
        )
    return {
        column: (
            seg.resolution_minutes, np.asarray(seg.values, dtype=float),
        )
        for column, seg in zip(columns, series.segments, strict=True)
    }


_ADMIE_URL_DATE_RX = re.compile(r"(\d{8})")


def _date_from_admie_url(url: str) -> date | None:
    """Extract the delivery date from an ADMIE file URL (YYYYMMDD...)."""
    match = _ADMIE_URL_DATE_RX.search(url.rsplit("/", 1)[-1])
    if match is None:
        return None
    stamp = match.group(1)
    try:
        return date(int(stamp[:4]), int(stamp[4:6]), int(stamp[6:8]))
    except ValueError:
        return None
