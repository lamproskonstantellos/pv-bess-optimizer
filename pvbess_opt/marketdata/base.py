"""Market-data provider plumbing: zones, cache, calendar normalisation.

The market-data layer fetches historical wholesale price series (day-ahead
now; balancing/imbalance with the ADMIE phase) and lays them onto the model
grid, **replacing** the matching workbook columns when the ``market_data``
sheet selects an API source.  This module owns everything the individual
providers share:

* the bidding-zone registry (EIC code + IANA timezone per zone),
* the on-disk fetch cache (the PVGIS pattern: JSON payload keyed on the
  request identity, so a repeat run never re-hits the network),
* the calendar engine that converts provider segments (UTC, possibly
  mixed-resolution) onto the workbook's local wall-clock year grid, and
* :func:`resolve_market_data`, the single read-path hook called by
  ``io.read_workbook`` (the ``resolve_pv_source`` pattern — the YAML
  surface materialises to a workbook and re-enters the same path, so one
  hook covers every input format).

Calendar rules (normative — tested in ``tests/test_marketdata_calendar.py``):

1. Prices are **intensive** (EUR/MWh).  A coarser native resolution is
   laid onto a finer model grid by **step-hold repetition** (an hourly
   price holds for its four quarters), never by division.  A finer
   native resolution is laid onto a coarser grid by the **arithmetic
   mean** of its equal-length sub-steps, with an INFO note that
   intra-period spread is averaged away.
2. Mixed-resolution responses (e.g. a reference year straddling the
   2025-10-01 SDAC 15-minute MTU go-live: PT60M months followed by
   PT15M months) are normalised segment by segment onto the model
   cadence and stitched on the UTC axis; the stitch asserts exact
   continuity — no gap, no overlap, no double-counted step.
3. UTC → local wall clock uses the zone's IANA timezone.  On the DST
   spring-forward day the skipped local steps repeat the previous
   value; on fall-back the duplicated hour is sampled once (its first,
   summer-time occurrence) and the repeat is dropped — every day keeps
   its grid length, matching the repo's uniform-grid convention
   (``timeutils.apply_fixed_utc_offset`` documents that wall-clock DST
   alignment must be resolved by re-gridding, which happens here).
4. A leap reference year drops Feb 29 (industry 8 760-hour convention)
   AFTER the local-grid sampling, so the result always carries exactly
   one non-leap year of steps.
5. The final series must match the workbook grid exactly — length and
   Jan-1-00:00 alignment are asserted, hard error otherwise.
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
from calendar import isleap
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import UTC, datetime
from itertools import pairwise
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

DEFAULT_MARKET_CACHE_DIR = Path.home() / ".cache" / "pvbess" / "market"

#: Fetch modes accepted by the ``market_fetch_mode`` key.
FETCH_MODES: tuple[str, ...] = ("cache_first", "refresh", "offline")


class MarketDataError(ValueError):
    """A market-data configuration, fetch, or normalisation failure."""


class MarketDataUnavailableError(MarketDataError):
    """The (zone, dataset) combination publishes no usable data."""


# ---------------------------------------------------------------------------
# Bidding-zone registry
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Zone:
    """A bidding zone: canonical code, ENTSO-E EIC code, IANA timezone."""

    code: str
    eic: str
    tz: str


# Keyed by the lowercase zone token the workbook enum produces
# (``_parse_string_enum`` lowercases); ``Zone.code`` keeps the canonical
# uppercase spelling for logs and provenance.  Extendable: new zones only
# need a row here plus the ``bidding_zone`` enum entry in ``io.py``.
ZONES: dict[str, Zone] = {
    "gr": Zone("GR", "10YGR-HTSO-----Y", "Europe/Athens"),
    "de_lu": Zone("DE_LU", "10Y1001A1001A82H", "Europe/Berlin"),
    "fr": Zone("FR", "10YFR-RTE------C", "Europe/Paris"),
    "it_nord": Zone("IT_NORD", "10Y1001A1001A73I", "Europe/Rome"),
    "es": Zone("ES", "10YES-REE------0", "Europe/Madrid"),
    "bg": Zone("BG", "10YCA-BULGARIA-R", "Europe/Sofia"),
    "ro": Zone("RO", "10YRO-TEL------P", "Europe/Bucharest"),
}


def zone_from_token(token: str) -> Zone:
    """Resolve a workbook ``bidding_zone`` token to its :class:`Zone`."""
    zone = ZONES.get(str(token).strip().lower())
    if zone is None:
        raise MarketDataError(
            f"unknown bidding_zone {token!r}; supported zones: "
            f"{', '.join(z.code for z in ZONES.values())}."
        )
    return zone


# ---------------------------------------------------------------------------
# Provider result shape
# ---------------------------------------------------------------------------


@dataclass
class PriceSegment:
    """A contiguous run of prices at one native resolution.

    ``start_utc`` is timezone-aware UTC; ``values`` covers
    ``[start_utc, start_utc + len(values) * resolution_minutes)``.
    Providers may return several segments per fetch (one per document
    period); the calendar engine sorts, resamples and stitches them.
    """

    start_utc: datetime
    resolution_minutes: int
    values: list[float]

    def end_utc(self) -> datetime:
        return self.start_utc + pd.Timedelta(
            minutes=self.resolution_minutes * len(self.values)
        )


@dataclass
class MarketSeries:
    """Provider output: raw segments plus provenance metadata."""

    segments: list[PriceSegment]
    metadata: dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# On-disk cache (PVGIS pattern)
# ---------------------------------------------------------------------------


class MarketDataCache:
    """JSON cache of raw provider segments, keyed on the request identity.

    The RAW segments are cached (not the normalised model-grid series) so
    a calendar-engine change or a different workbook cadence reuses the
    same fetch.  Keys are human-readable (`dataset_zone_year_<hash>`), so
    the ``offline``-mode error can name the exact missing file.
    """

    def __init__(self, cache_dir: str | Path | None = None) -> None:
        self.cache_dir = (
            Path(cache_dir).expanduser()
            if cache_dir is not None else DEFAULT_MARKET_CACHE_DIR
        )

    @staticmethod
    def key(dataset: str, zone_code: str, year: int, **extra: Any) -> str:
        ident = {
            "dataset": dataset, "zone": zone_code, "year": int(year), **extra,
        }
        blob = json.dumps(ident, sort_keys=True).encode("utf-8")
        digest = hashlib.sha256(blob).hexdigest()[:12]
        return f"{dataset}_{zone_code.lower()}_{int(year)}_{digest}"

    def path(self, key: str) -> Path:
        return self.cache_dir / f"{key}.json"

    def load(self, key: str) -> MarketSeries | None:
        path = self.path(key)
        if not path.exists():
            return None
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            segments = [
                PriceSegment(
                    start_utc=datetime.fromisoformat(seg["start_utc"]),
                    resolution_minutes=int(seg["resolution_minutes"]),
                    values=[float(v) for v in seg["values"]],
                )
                for seg in payload["segments"]
            ]
            metadata = dict(payload.get("metadata", {}))
        except (json.JSONDecodeError, KeyError, TypeError, ValueError, OSError):
            logger.warning("Ignoring unreadable market cache file %s.", path)
            return None
        return MarketSeries(segments=segments, metadata=metadata)

    def save(self, key: str, series: MarketSeries) -> None:
        path = self.path(key)
        path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "metadata": series.metadata,
            "segments": [
                {
                    "start_utc": seg.start_utc.isoformat(),
                    "resolution_minutes": seg.resolution_minutes,
                    "values": seg.values,
                }
                for seg in series.segments
            ],
        }
        path.write_text(json.dumps(payload), encoding="utf-8")


# ---------------------------------------------------------------------------
# Token policy
# ---------------------------------------------------------------------------


def mask_token(token: str) -> str:
    """Loggable form of a security token: first 8 characters + ellipsis.

    Never log or embed a full token anywhere — error messages, INFO
    lines and provenance records all go through this helper.
    """
    return token[:8] + "…" if len(token) > 8 else "…"


def resolve_entsoe_token(market_cfg: dict[str, Any]) -> str:
    """Resolve the ENTSO-E token: workbook key first, then environment.

    Resolution order (the documented ``market_data`` contract): the
    ``entsoe_token`` cell verbatim; else the environment variable named
    by ``entsoe_token_env`` (default ``ENTSOE_API_TOKEN``); else a hard
    error with instructions.
    """
    token = str(market_cfg.get("entsoe_token") or "").strip()
    if token:
        return token
    env_name = (
        str(market_cfg.get("entsoe_token_env") or "").strip()
        or "ENTSOE_API_TOKEN"
    )
    token = os.environ.get(env_name, "").strip()
    if token:
        return token
    raise MarketDataError(
        "an ENTSO-E API token is required to fetch market data: set the "
        "'entsoe_token' key on the market_data sheet, or export the "
        f"{env_name} environment variable. Request a free token via "
        "'Web API Security Token' after registering at "
        "https://transparency.entsoe.eu."
    )


# ---------------------------------------------------------------------------
# Calendar engine
# ---------------------------------------------------------------------------


def resample_intensive(
    values: np.ndarray,
    native_minutes: int,
    dt_minutes: int,
    *,
    column: str,
    notes: set[str],
    context: str = "series",
) -> np.ndarray:
    """Resample an intensive (EUR/MWh) series between cadences.

    Rule 1 of the module contract: step-hold when the native step is
    coarser than the model step (a price is a level, NEVER divided),
    arithmetic mean of the equal-length sub-steps when finer (with a
    deduplicated note that intra-period spread is averaged away).
    """
    native = int(native_minutes)
    arr = np.asarray(values, dtype=float)
    if np.isnan(arr).any():
        raise MarketDataError(
            f"{column}: fetched {context} contains NaN prices; refusing "
            "to write gaps into the model grid."
        )
    if native == dt_minutes:
        return arr
    if native % dt_minutes == 0:
        return np.repeat(arr, native // dt_minutes)
    if dt_minutes % native == 0:
        k = dt_minutes // native
        if len(arr) % k != 0:
            raise MarketDataError(
                f"{column}: fetched {context} carries {len(arr)} values "
                f"at {native} min, not a whole number of "
                f"{dt_minutes}-min model steps."
            )
        notes.add(
            f"{native}-min market data averaged onto the {dt_minutes}-min "
            "model grid (intra-period price spread is averaged away)"
        )
        return np.asarray(arr.reshape(-1, k).mean(axis=1), dtype=float)
    raise MarketDataError(
        f"{column}: native resolution {native} min is incommensurable "
        f"with the model cadence {dt_minutes} min."
    )


def _segment_to_model_cadence(
    seg: PriceSegment, dt_minutes: int, *, column: str,
    notes: set[str],
) -> tuple[pd.DatetimeIndex, np.ndarray]:
    """Resample ONE segment onto the model cadence, on the UTC axis."""
    out = resample_intensive(
        np.asarray(seg.values, dtype=float),
        seg.resolution_minutes,
        dt_minutes,
        column=column,
        notes=notes,
        context=f"segment starting {seg.start_utc.isoformat()}",
    )
    idx = pd.date_range(
        seg.start_utc, periods=len(out), freq=f"{dt_minutes}min", tz="UTC",
    )
    return idx, out


def stitch_segments_utc(
    segments: list[PriceSegment], dt_minutes: int, *, column: str,
) -> tuple[pd.Series, list[str]]:
    """Normalise segments to the model cadence and stitch on the UTC axis.

    Segments are sorted by start; after per-segment resampling each
    segment must begin exactly where the previous one ended (rule 2) —
    a gap, an overlap, or a double-counted step is a hard error naming
    the boundary.  Returns the stitched UTC series plus any resample
    notes (deduplicated, for the consolidated INFO log).
    """
    if not segments:
        raise MarketDataError(f"{column}: provider returned no data segments.")
    notes: set[str] = set()
    parts = [
        _segment_to_model_cadence(seg, dt_minutes, column=column, notes=notes)
        for seg in sorted(segments, key=lambda s: s.start_utc)
    ]
    for (prev_idx, _), (next_idx, _) in pairwise(parts):
        prev_end = prev_idx[-1] + pd.Timedelta(minutes=dt_minutes)
        if next_idx[0] != prev_end:
            kind = "gap" if next_idx[0] > prev_end else "overlap"
            raise MarketDataError(
                f"{column}: segment stitch {kind} at the "
                f"{prev_end.isoformat()} boundary (next segment starts "
                f"{next_idx[0].isoformat()}); fetched data does not tile "
                "the request window."
            )
    index = parts[0][0].append([idx for idx, _ in parts[1:]])
    values = np.concatenate([vals for _, vals in parts])
    return pd.Series(values, index=index), sorted(notes)


def sample_local_year(
    series_utc: pd.Series,
    *,
    tz_name: str,
    year: int,
    dt_minutes: int,
    column: str,
) -> np.ndarray:
    """Sample a stitched UTC series onto the local wall-clock year grid.

    Builds the naive local grid for calendar ``year`` at the model
    cadence, maps each step to its UTC instant (DST rules 3 of the
    module contract) and samples the series there.  A leap year drops
    Feb 29 after sampling (rule 4), so the result is always exactly one
    non-leap year of steps.
    """
    steps_per_day = (24 * 60) // dt_minutes
    n_days = 366 if isleap(year) else 365
    local_idx = pd.date_range(
        f"{year}-01-01", periods=n_days * steps_per_day,
        freq=f"{dt_minutes}min",
    )
    # ambiguous=True picks the FIRST (summer-time) occurrence of the
    # duplicated fall-back hour — the repeat is never sampled, which IS
    # the drop rule.  nonexistent='NaT' marks the spring-forward hole,
    # filled below by repeating the previous value.
    localized = local_idx.tz_localize(
        tz_name, ambiguous=True, nonexistent="NaT",
    )
    utc_wanted = localized.tz_convert("UTC")
    exists = ~np.asarray(utc_wanted.isna())
    # Coverage check BEFORE the DST fill: a missing UTC instant here is
    # missing provider data, which must never be silently forward-filled.
    missing = utc_wanted[exists].difference(
        pd.DatetimeIndex(series_utc.index),
    )
    if len(missing) > 0:
        raise MarketDataError(
            f"{column}: fetched data does not cover the local {year} grid; "
            f"first missing instant {missing[0].isoformat()} "
            f"({len(missing)} step(s) missing in total)."
        )
    out = np.full(len(local_idx), np.nan)
    out[exists] = series_utc.reindex(
        utc_wanted[exists],
    ).to_numpy(dtype=float)
    # Spring-forward: the skipped local steps repeat the previous value
    # so every day keeps its full grid length.
    hole = np.isnan(out)
    if hole.any():
        if hole[0]:
            raise MarketDataError(
                f"{column}: the first step of the local {year} grid is "
                "DST-nonexistent — impossible for a Jan-1 grid; check the "
                "zone timezone."
            )
        idx_arr = np.where(~hole, np.arange(len(out)), 0)
        np.maximum.accumulate(idx_arr, out=idx_arr)
        out = out[idx_arr]
    if isleap(year):
        keep = ~((local_idx.month == 2) & (local_idx.day == 29))
        out = out[np.asarray(keep)]
    expected = 365 * steps_per_day
    if len(out) != expected:
        raise MarketDataError(
            f"{column}: normalised series carries {len(out)} steps, "
            f"expected {expected} (one non-leap year at {dt_minutes} min)."
        )
    return out


def validate_model_year_grid(
    ts: pd.DataFrame, dt_minutes: int, *, context: str,
) -> int:
    """Assert the workbook grid is one full non-leap year from Jan 1 00:00.

    The fetched reference year is laid onto the workbook grid
    positionally (calendar position, ignoring the workbook's own year
    label), which is only well-defined for a complete non-leap-year
    grid.  Regularity itself is already guaranteed upstream by
    ``detect_timestep_minutes``.  Returns the step count.
    """
    steps_per_day = (24 * 60) // dt_minutes
    expected = 365 * steps_per_day
    first = pd.Timestamp(ts["timestamp"].iloc[0])
    if (first.month, first.day, first.hour, first.minute) != (1, 1, 0, 0):
        raise MarketDataError(
            f"{context} requires a timeseries grid starting Jan 1 00:00 "
            f"(the fetched year is laid on by calendar position); the "
            f"workbook grid starts {first.isoformat()}."
        )
    if len(ts) != expected:
        raise MarketDataError(
            f"{context} requires one full non-leap model year "
            f"({expected} steps at {dt_minutes} min); the workbook grid "
            f"carries {len(ts)} steps."
        )
    return expected


# ---------------------------------------------------------------------------
# Read-path resolution hook
# ---------------------------------------------------------------------------

# Process-level memo so the several read_workbook calls inside one run
# (inputs, economics, pre-flight) fetch each (dataset, zone, year) once.
_FETCH_MEMO: dict[tuple[str, str, int], MarketSeries] = {}


def _market_cache(market_cfg: dict[str, Any]) -> MarketDataCache:
    return MarketDataCache(
        str(market_cfg.get("market_cache_dir") or "").strip() or None
    )


def _memoized_fetch(
    dataset: str,
    zone: Zone,
    year: int,
    fetcher: Callable[[], MarketSeries],
) -> MarketSeries:
    memo_key = (dataset, zone.code, int(year))
    cached = _FETCH_MEMO.get(memo_key)
    if cached is not None:
        return cached
    series = fetcher()
    _FETCH_MEMO[memo_key] = series
    return series


def _fetch_day_ahead_memoized(
    zone: Zone, year: int, market_cfg: dict[str, Any],
) -> MarketSeries:
    from .entsoe import fetch_day_ahead_year

    return _memoized_fetch(
        "dam-a44", zone, year,
        lambda: fetch_day_ahead_year(
            zone,
            int(year),
            # Lazy: a cache hit (and the whole offline mode) must work
            # without any token configured.
            token_resolver=lambda: resolve_entsoe_token(market_cfg),
            cache=_market_cache(market_cfg),
            fetch_mode=str(
                market_cfg.get("market_fetch_mode") or "cache_first"
            ).strip().lower(),
        ),
    )


# ---------------------------------------------------------------------------
# Per-(zone, dataset) source registry
# ---------------------------------------------------------------------------

# Which provider the 'auto' source picks per dataset and zone; zones
# not listed use the default.  GR is special-cased throughout: its
# balancing/imbalance domain on the ENTSO-E platform is effectively
# empty (co-optimised ISP, nationally published results), so GR routes
# to the ADMIE file API and an EXPLICIT entsoe selection for GR is an
# error rather than a silent empty fetch.
_AUTO_SOURCE: dict[str, dict[str, str]] = {
    "balancing": {"gr": "admie"},
    "imbalance": {"gr": "admie"},
}
_AUTO_DEFAULT_SOURCE = "entsoe"

# (dataset, zone) combinations known to publish nothing anywhere — the
# 'auto' source falls back to 'file' with a WARNING for these.  Empty
# today (every shipped zone has a route); kept as the documented hook
# so a future zone without any publication degrades per the contract.
_PUBLISHES_NOTHING: frozenset[tuple[str, str]] = frozenset()


def resolve_dataset_source(
    dataset: str, requested: str, zone: Zone,
) -> str | None:
    """Resolve a balancing/imbalance source selection for ``zone``.

    Returns the provider token (``entsoe`` / ``admie``) or ``None`` for
    the workbook path.  ``auto`` consults the registry (GR → admie,
    else entsoe) and degrades to ``None`` with a WARNING when the
    (dataset, zone) combination publishes nothing; an EXPLICIT
    selection of an unavailable provider raises instead — the user
    asked for something that cannot be served.
    """
    token = str(requested or "file").strip().lower()
    zone_key = zone.code.lower()
    if token == "file":
        return None
    if token == "auto":
        if (dataset, zone_key) in _PUBLISHES_NOTHING:
            logger.warning(
                "[marketdata] %s_source='auto': zone %s publishes no "
                "%s data on any registered provider; falling back to "
                "the workbook ('file') path.",
                dataset, zone.code, dataset,
            )
            return None
        return _AUTO_SOURCE.get(dataset, {}).get(
            zone_key, _AUTO_DEFAULT_SOURCE,
        )
    if token == "admie" and zone_key != "gr":
        raise MarketDataUnavailableError(
            f"{dataset}_source='admie' is the Greek TSO file API; zone "
            f"{zone.code} is not served by it. Use 'auto' (per-zone "
            "registry), 'entsoe', or 'file'."
        )
    if token == "entsoe" and zone_key == "gr":
        raise MarketDataUnavailableError(
            f"{dataset}_source='entsoe': the GR balancing/imbalance "
            "domain on the ENTSO-E platform is effectively empty "
            "(Greece runs a co-optimised Integrated Scheduling Process "
            "and publishes results nationally). Use 'admie' or 'auto'."
        )
    if token not in ("entsoe", "admie"):
        raise MarketDataError(
            f"{dataset}_source {token!r} is not one of 'file', 'auto', "
            "'entsoe', 'admie'."
        )
    return token


# ---------------------------------------------------------------------------
# Dataset application helpers
# ---------------------------------------------------------------------------


def _provenance_record(
    column: str,
    *,
    dataset: str,
    source: str,
    source_key: str,
    zone: Zone,
    year: int,
    dt_minutes: int,
    overridden: bool,
    metadata: dict[str, Any],
) -> dict[str, Any]:
    return {
        "column": column,
        "dataset": dataset,
        "source": source,
        # The market_data key that selected this source; the input
        # snapshot flips it back to 'file' after materialising the
        # fetched values, so the snapshot re-runs offline.
        "source_key": source_key,
        "bidding_zone": zone.code,
        "eic": zone.eic,
        "reference_year": year,
        "model_cadence_min": int(dt_minutes),
        "workbook_column_overridden": overridden,
        **{
            k: metadata.get(k)
            for k in ("fetched_at", "cache_state", "cache_key")
        },
    }


def _info_line(
    columns: list[str],
    *,
    label: str,
    zone: Zone,
    year: int,
    dt_minutes: int,
    metadata: dict[str, Any],
    notes: list[str],
) -> str:
    return (
        f"{', '.join(columns)} <- {label} {zone.code} ({zone.eic}), "
        f"reference year {year}, {dt_minutes}-min grid, "
        f"{metadata.get('cache_state', 'live')} "
        f"(fetched {metadata.get('fetched_at', 'unknown')}, "
        f"cache key {metadata.get('cache_key', 'n/a')})"
        + (f"; {'; '.join(notes)}" if notes else "")
    )


def _apply_utc_dataset(
    ts: pd.DataFrame,
    column_segments: dict[str, list[PriceSegment]],
    *,
    required: tuple[str, ...] | None,
    zone: Zone,
    year: int,
    dt_minutes: int,
) -> tuple[list[str], list[str]]:
    """Sample a per-column UTC segment map onto the grid; replace columns.

    ``required`` demands exactly that column set (hard error on any
    missing one — partial datasets are never silently mixed); ``None`` applies
    whatever columns the provider returned (the imbalance single/dual
    variability).  Returns (applied columns, resample notes).
    """
    if required is not None:
        missing = sorted(set(required) - set(column_segments))
        if missing:
            raise MarketDataError(
                f"fetched dataset is missing the column(s) "
                f"{', '.join(missing)} for zone {zone.code} in {year}; "
                "refusing a partial bypass — use the 'file' source for "
                "this dataset if the zone does not publish it in full."
            )
    notes: set[str] = set()
    applied: list[str] = []
    for column in (required or tuple(sorted(column_segments))):
        stitched, seg_notes = stitch_segments_utc(
            column_segments[column], dt_minutes, column=column,
        )
        notes.update(seg_notes)
        ts[column] = sample_local_year(
            stitched,
            tz_name=zone.tz,
            year=year,
            dt_minutes=dt_minutes,
            column=column,
        )
        applied.append(column)
    return applied, sorted(notes)


def _apply_local_native_dataset(
    ts: pd.DataFrame,
    columns_map: dict[str, tuple[int, np.ndarray]],
    *,
    required: tuple[str, ...],
    n_steps: int,
    dt_minutes: int,
    zone: Zone,
    year: int,
) -> tuple[list[str], list[str]]:
    """Apply an ADMIE-style local-native column map onto the grid."""
    missing = sorted(set(required) - set(columns_map))
    if missing:
        raise MarketDataError(
            f"ADMIE dataset is missing the column(s) {', '.join(missing)} "
            f"for {year}; the declared header map may need re-pinning "
            "via scripts/probe_market_data.py — refusing a partial "
            "bypass."
        )
    notes: set[str] = set()
    applied: list[str] = []
    for column in required:
        native_minutes, values = columns_map[column]
        out = resample_intensive(
            values, native_minutes, dt_minutes,
            column=column, notes=notes, context=f"{year} year series",
        )
        if len(out) != n_steps:
            raise MarketDataError(
                f"{column}: normalised ADMIE series carries {len(out)} "
                f"steps, expected {n_steps} (zone {zone.code}, "
                f"{dt_minutes}-min grid)."
            )
        ts[column] = out
        applied.append(column)
    return applied, sorted(notes)


def resolve_market_data(
    typed: dict[str, Any], ts: pd.DataFrame, dt_minutes: int,
) -> None:
    """Apply the ``market_data`` source selections to the timeseries.

    The single read-path hook (called by ``io.read_workbook`` after the
    cadence is known and BEFORE the balancing scalar fallback, so a
    fetched column suppresses the fallback exactly like a workbook
    column).  With every source at its ``file`` default this returns
    immediately — no fetch, no log line, bit-identical behaviour.

    Override semantics are total per column: a fetched series REPLACES
    the workbook column for the whole horizon, even where the workbook
    had values.  Partial data is a hard error upstream (coverage and
    stitch asserts), never a silent mix.  One consolidated INFO lists
    every bypassed column with its provenance; the same records land in
    ``typed['market_provenance']`` for the results workbook.
    """
    market_cfg = typed.get("market_data") or {}
    price_source = str(market_cfg.get("price_source") or "file").strip().lower()
    balancing_source = str(
        market_cfg.get("balancing_source") or "file"
    ).strip().lower()
    imbalance_source = str(
        market_cfg.get("imbalance_source") or "file"
    ).strip().lower()
    if (price_source, balancing_source, imbalance_source) == (
        "file", "file", "file",
    ):
        return

    fetch_mode = str(
        market_cfg.get("market_fetch_mode") or "cache_first"
    ).strip().lower()
    if fetch_mode not in FETCH_MODES:
        raise MarketDataError(
            f"market_fetch_mode {fetch_mode!r} is not one of "
            f"{', '.join(FETCH_MODES)}."
        )
    policy = str(
        market_cfg.get("price_resample_policy") or "step_hold"
    ).strip().lower()
    if policy != "step_hold":
        # The policy key is a forward-compatible selector; step_hold is
        # the only sound intensive-quantity policy and the only one
        # implemented.
        raise MarketDataError(
            f"price_resample_policy {policy!r} is not supported; the "
            "only implemented policy is 'step_hold'."
        )

    zone = zone_from_token(str(market_cfg.get("bidding_zone") or "gr"))
    year = int(market_cfg.get("price_reference_year") or 2025)
    if not 1990 <= year <= 2100:
        raise MarketDataError(
            f"price_reference_year {year} is outside the plausible "
            "1990-2100 range."
        )

    active_key = next(
        key for key, value in (
            ("price_source", price_source),
            ("balancing_source", balancing_source),
            ("imbalance_source", imbalance_source),
        ) if value != "file"
    )
    n_steps = validate_model_year_grid(
        ts, dt_minutes,
        context=f"{active_key}='{market_cfg.get(active_key)}' "
                "(market_data sheet)",
    )
    cache = _market_cache(market_cfg)

    provenance: list[dict[str, Any]] = []
    info_lines: list[str] = []
    # Captured BEFORE any bypass so the provenance can say whether the
    # workbook actually carried each column.
    pre_existing = set(ts.columns)

    if price_source == "entsoe":
        series = _fetch_day_ahead_memoized(zone, year, market_cfg)
        applied, notes = _apply_utc_dataset(
            ts,
            {"dam_price_eur_per_mwh": series.segments},
            required=("dam_price_eur_per_mwh",),
            zone=zone, year=year, dt_minutes=dt_minutes,
        )
        provenance.extend(
            _provenance_record(
                column,
                dataset="day-ahead prices (ENTSO-E 12.1.D, A44)",
                source="entsoe", source_key="price_source",
                zone=zone, year=year, dt_minutes=dt_minutes,
                overridden=column in pre_existing,
                metadata=series.metadata,
            )
            for column in applied
        )
        info_lines.append(_info_line(
            applied, label="ENTSO-E A44", zone=zone, year=year,
            dt_minutes=dt_minutes, metadata=series.metadata, notes=notes,
        ))
    elif price_source != "file":
        raise MarketDataError(
            f"price_source {price_source!r} is not one of 'file', 'entsoe'."
        )

    bal_provider = resolve_dataset_source("balancing", balancing_source, zone)
    if bal_provider == "admie":
        from .admie import (
            BALANCING_HEADER_PATTERNS,
            fetch_gr_balancing_year,
            series_columns,
        )

        series = _memoized_fetch(
            "gr-balancing", zone, year,
            lambda: fetch_gr_balancing_year(
                year, cache=cache, fetch_mode=fetch_mode,
            ),
        )
        applied, notes = _apply_local_native_dataset(
            ts, series_columns(series),
            required=tuple(BALANCING_HEADER_PATTERNS),
            n_steps=n_steps, dt_minutes=dt_minutes, zone=zone, year=year,
        )
        provenance.extend(
            _provenance_record(
                column,
                dataset="balancing prices (ADMIE ISP results)",
                source="admie", source_key="balancing_source",
                zone=zone, year=year, dt_minutes=dt_minutes,
                overridden=column in pre_existing,
                metadata=series.metadata,
            )
            for column in applied
        )
        info_lines.append(_info_line(
            applied, label="ADMIE", zone=zone, year=year,
            dt_minutes=dt_minutes, metadata=series.metadata, notes=notes,
        ))
        # Greece procures no standalone FCR (co-optimised ISP): the FCR
        # capacity column keeps its workbook / scalar-fallback path —
        # the documented registry degradation, WARNED so a GR run never
        # silently assumes a fetched FCR price.
        logger.warning(
            "[marketdata] balancing_source='%s' (GR): "
            "fcr_capacity_price_eur_per_mwh is NOT fetched — Greece "
            "procures no standalone FCR — and stays on the workbook / "
            "scalar-fallback path.",
            balancing_source,
        )
    elif bal_provider == "entsoe":
        from .entsoe import (
            ENTSOE_BALANCING_COLUMNS,
            column_segments,
            fetch_balancing_prices_year,
        )

        series = _memoized_fetch(
            "bal-prices", zone, year,
            lambda: fetch_balancing_prices_year(
                zone, year,
                token_resolver=lambda: resolve_entsoe_token(market_cfg),
                cache=cache, fetch_mode=fetch_mode,
            ),
        )
        applied, notes = _apply_utc_dataset(
            ts, column_segments(series),
            required=ENTSOE_BALANCING_COLUMNS,
            zone=zone, year=year, dt_minutes=dt_minutes,
        )
        provenance.extend(
            _provenance_record(
                column,
                dataset="balancing prices (ENTSO-E 17.1.B&C A81 + "
                        "17.1.F A84)",
                source="entsoe", source_key="balancing_source",
                zone=zone, year=year, dt_minutes=dt_minutes,
                overridden=column in pre_existing,
                metadata=series.metadata,
            )
            for column in applied
        )
        info_lines.append(_info_line(
            applied, label="ENTSO-E A81/A84", zone=zone, year=year,
            dt_minutes=dt_minutes, metadata=series.metadata, notes=notes,
        ))

    imb_provider = resolve_dataset_source("imbalance", imbalance_source, zone)
    if imb_provider == "admie":
        from .admie import (
            IMBALANCE_HEADER_PATTERNS,
            fetch_gr_imbalance_year,
            series_columns,
        )

        series = _memoized_fetch(
            "gr-imbalance", zone, year,
            lambda: fetch_gr_imbalance_year(
                year, cache=cache, fetch_mode=fetch_mode,
            ),
        )
        applied, notes = _apply_local_native_dataset(
            ts, series_columns(series),
            required=tuple(IMBALANCE_HEADER_PATTERNS),
            n_steps=n_steps, dt_minutes=dt_minutes, zone=zone, year=year,
        )
        provenance.extend(
            _provenance_record(
                column,
                dataset="imbalance prices (ADMIE IMBABE)",
                source="admie", source_key="imbalance_source",
                zone=zone, year=year, dt_minutes=dt_minutes,
                overridden=column in pre_existing,
                metadata=series.metadata,
            )
            for column in applied
        )
        info_lines.append(_info_line(
            applied, label="ADMIE", zone=zone, year=year,
            dt_minutes=dt_minutes, metadata=series.metadata, notes=notes,
        ))
    elif imb_provider == "entsoe":
        from .entsoe import column_segments, fetch_imbalance_prices_year

        series = _memoized_fetch(
            "imbalance", zone, year,
            lambda: fetch_imbalance_prices_year(
                zone, year,
                token_resolver=lambda: resolve_entsoe_token(market_cfg),
                cache=cache, fetch_mode=fetch_mode,
            ),
        )
        applied, notes = _apply_utc_dataset(
            ts, column_segments(series),
            required=None,  # single vs dual pricing varies by zone
            zone=zone, year=year, dt_minutes=dt_minutes,
        )
        provenance.extend(
            _provenance_record(
                column,
                dataset="imbalance prices (ENTSO-E 17.1.G, A85)",
                source="entsoe", source_key="imbalance_source",
                zone=zone, year=year, dt_minutes=dt_minutes,
                overridden=column in pre_existing,
                metadata=series.metadata,
            )
            for column in applied
        )
        info_lines.append(_info_line(
            applied, label="ENTSO-E A85", zone=zone, year=year,
            dt_minutes=dt_minutes, metadata=series.metadata, notes=notes,
        ))

    if provenance:
        typed["market_provenance"] = provenance
        logger.info(
            "[marketdata] bypassing workbook prices with fetched market "
            "data:\n  %s",
            "\n  ".join(info_lines),
        )


def utcnow_isoformat() -> str:
    """Timestamp helper for provenance records (UTC, second precision)."""
    return datetime.now(UTC).replace(microsecond=0).isoformat()


def materialize_bypassed_workbook(
    workbook_path: Path,
    ts: pd.DataFrame,
    provenance: list[dict[str, Any]],
) -> None:
    """Write the fetched values into a workbook copy, for reproducibility.

    Called by the pipeline on the run's input snapshot: every bypassed
    column is materialised into the snapshot's ``timeseries`` sheet and
    the selecting ``market_data`` source key is flipped back to
    ``file``, so re-running the snapshot reproduces the run offline —
    no token, no network, no cache.  The workbook is edited cell-wise
    through openpyxl (kv sheets must never round-trip through pandas —
    a mixed-type ``value`` column mis-surfaces numeric zeros as
    booleans, see ``io._read_kv_flat``).
    """
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    ws = wb["timeseries"]
    header = {
        str(cell.value).strip(): int(cell.column)
        for cell in ws[1] if cell.value is not None
    }
    for record in provenance:
        column = str(record["column"])
        values = ts[column].to_numpy(dtype=float)
        col_idx = header.get(column)
        if col_idx is None:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx).value = column
            header[column] = col_idx
        for offset, value in enumerate(values):
            ws.cell(row=2 + offset, column=col_idx).value = float(value)
    if "market_data" in wb.sheetnames:
        md = wb["market_data"]
        flipped_keys = {
            str(r.get("source_key")) for r in provenance if r.get("source_key")
        }
        for row in md.iter_rows(min_row=2, max_col=2):
            key_cell, value_cell = row[0], row[1]
            if not isinstance(key_cell.value, str):
                continue
            key = key_cell.value.strip()
            if key in flipped_keys:
                value_cell.value = "file"
            elif key == "entsoe_token":
                # The snapshot re-runs offline with 'file' sources, so
                # the token is dead weight — and results directories
                # get shared; never carry a live secret into one.
                value_cell.value = ""
    wb.save(workbook_path)
    logger.info(
        "[marketdata] materialised %d fetched column(s) into %s and reset "
        "the source key(s) to 'file' (snapshot re-runs offline).",
        len(provenance), workbook_path,
    )
