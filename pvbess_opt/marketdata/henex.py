"""HEnEx daily DAM workbook fetcher (GR cross-check for ENTSO-E A44).

The Hellenic Energy Exchange publishes a daily day-ahead results
workbook at::

    https://www.enexgroup.gr/documents/20126/366820/
        YYYYMMDD_EL-DAM_ResultsSummary_EN_vNN.xlsx

(version suffixes are probed in order — corrections republish under a
higher ``vNN``).  This module is a **diagnostic cross-check**, not a
price source: :func:`crosscheck_gr_dam` compares a fetched ENTSO-E GR
day against the HEnEx clearing prices and reports the divergence, so a
zone/EIC/timezone mistake in the primary path is caught by an
independent publication.  It is wired into
``scripts/probe_market_data.py``; the model itself never reads HEnEx.
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date
from typing import Any

import numpy as np

from .base import MarketDataError

logger = logging.getLogger(__name__)

HENEX_URL_TEMPLATE = (
    "https://www.enexgroup.gr/documents/20126/366820/"
    "{yyyymmdd}_EL-DAM_ResultsSummary_EN_v{version:02d}.xlsx"
)

# Header pattern locating the market clearing price column inside the
# results summary (matched case-insensitively; pin against a real
# workbook via the probe script when the layout drifts).
_MCP_HEADER_RX = re.compile(r"(market\s*clearing|mcp|clearing).*price", re.I)


def _http_get(url: str, timeout: float) -> tuple[int, bytes]:
    """One GET; tests monkeypatch this symbol."""
    import requests

    resp = requests.get(url, timeout=timeout)
    return resp.status_code, resp.content


def fetch_henex_dam_day(
    day: date, *, timeout: float = 60.0, max_version: int = 5,
) -> np.ndarray:
    """Fetch one day's GR DAM clearing prices from the HEnEx workbook.

    Tries version suffixes v01..vNN and parses the first one that
    answers; returns the clearing-price array at the workbook's native
    period count (24 hourly values historically, 96 quarters after the
    15-minute MTU go-live).
    """
    last_status = 0
    for version in range(1, max_version + 1):
        url = HENEX_URL_TEMPLATE.format(
            yyyymmdd=day.strftime("%Y%m%d"), version=version,
        )
        status, body = _http_get(url, timeout)
        last_status = status
        if status == 200:
            return _parse_results_summary(body, source_name=url)
    raise MarketDataError(
        f"no HEnEx DAM results workbook answered for {day.isoformat()} "
        f"(versions 1..{max_version}; last HTTP status {last_status})."
    )


def _parse_results_summary(content: bytes, *, source_name: str) -> np.ndarray:
    """Extract the clearing-price column from a results-summary workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header: tuple[Any, ...] | None = None
            prices: list[float] = []
            price_idx: int | None = None
            for row in ws.iter_rows(values_only=True):
                if price_idx is None:
                    for idx, cell in enumerate(row):
                        if isinstance(cell, str) and _MCP_HEADER_RX.search(cell):
                            header = row
                            price_idx = idx
                            break
                    continue
                cell_value = row[price_idx] if price_idx < len(row) else None
                if isinstance(cell_value, (int, float)):
                    prices.append(float(cell_value))
            if header is not None and prices:
                return np.asarray(prices, dtype=float)
    finally:
        wb.close()
    raise MarketDataError(
        f"{source_name}: no clearing-price column matched "
        f"{_MCP_HEADER_RX.pattern!r}; the workbook layout may have "
        "changed — re-pin via scripts/probe_market_data.py."
    )


def crosscheck_gr_dam(
    entsoe_day: np.ndarray,
    henex_day: np.ndarray,
    *,
    tolerance_eur_per_mwh: float = 0.01,
) -> dict[str, float]:
    """Compare one GR day from both publications; return divergence stats.

    Both arrays must cover the same day; a coarser side is step-held
    onto the finer period count first (intensive-quantity rule).  The
    returned dict carries ``max_abs_diff`` / ``mean_abs_diff`` /
    ``n_periods``; a WARNING is logged when the max exceeds the
    tolerance (the two venues publish the same auction, so any real
    divergence indicates a zone, timezone, or version mismatch).
    """
    a = np.asarray(entsoe_day, dtype=float)
    b = np.asarray(henex_day, dtype=float)
    if len(a) == 0 or len(b) == 0:
        raise MarketDataError("cross-check needs non-empty price arrays.")
    n = max(len(a), len(b))
    for side in ("entsoe", "henex"):
        arr = a if side == "entsoe" else b
        if n % len(arr) != 0:
            raise MarketDataError(
                f"cross-check period counts are incommensurable: "
                f"{len(a)} vs {len(b)}."
            )
    if len(a) != n:
        a = np.repeat(a, n // len(a))
    if len(b) != n:
        b = np.repeat(b, n // len(b))
    diff = np.abs(a - b)
    stats = {
        "max_abs_diff": float(diff.max()),
        "mean_abs_diff": float(diff.mean()),
        "n_periods": float(n),
    }
    if stats["max_abs_diff"] > tolerance_eur_per_mwh:
        logger.warning(
            "[marketdata] GR DAM cross-check divergence: ENTSO-E vs "
            "HEnEx max |diff| %.4f EUR/MWh over %d periods (tolerance "
            "%.4f). Check zone / timezone / workbook version.",
            stats["max_abs_diff"], n, tolerance_eur_per_mwh,
        )
    return stats
