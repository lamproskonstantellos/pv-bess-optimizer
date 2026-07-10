"""Excel input parsing and output writing for the PV+BESS optimizer.

The schema sheets, one logical theme per sheet:

* ``timeseries`` — per-step data with lowercase snake_case column names:
  ``timestamp``, ``load_kwh``, ``pv_kwh``, ``dam_price_eur_per_mwh``,
  optional ``retail_price_eur_per_mwh``.  ``pv_kwh`` is the single PV
  column: fill it to source PV from the timeseries (consumed verbatim as
  absolute kWh per step), or leave it empty and set ``latitude`` /
  ``longitude`` on the ``pv`` sheet to source the profile from PVGIS
  instead.  The legacy
  ``pv_kwh_override`` column is deprecated: it is read only as a fallback
  when ``pv_kwh`` is empty (:func:`pvbess_opt.io_read.resolve_pv_source`).
* ``project`` — high-level run config (lifecycle horizon, mode,
  settlement, retail tariff, grid export limit, currency / title flags).
* ``pv`` — PV source switch (``pv_source`` = auto | file | pvgis), PVGIS
  location / array geometry (``latitude``, ``longitude``, ``tilt``,
  ``azimuth``, ``losses_pct``, ``weather_year``, ``timeseries_path``),
  nameplate, degradation, CAPEX / DEVEX / OPEX.
* ``bess`` — BESS power and capacity, efficiency, SOC bounds, cycles,
  CAPEX / DEVEX / OPEX, replacement and degradation.
* ``economics`` — discount rate, inflation indices, aggregator fee,
  sensitivity deltas.
* ``simulation`` — uncertainty (rolling-horizon Monte Carlo) and plot
  scope flags.
* ``balancing`` — stochastic balancing-market participation
  (FCR / aFRR / mFRR): per-product capacity shares, acceptance /
  activation probabilities, fallback prices, and the Monte Carlo seeds.
* ``max_injection_profile`` — hour-of-day cap profile (24 rows),
  optionally with one column per calendar month, expressing the share
  of ``p_grid_export_max_kw`` available for export.  Missing → fall
  back to the no-curtailment default (a flat 100 %) and log INFO.
* ``sizing`` — optional capacity-sweep grid (columnar: one column per
  axis, one value per row) gated by an ``enabled`` TRUE / FALSE toggle.
  Shipped disabled; read by :func:`pvbess_opt.sizing.read_sizing_block`.
* ``scenarios`` — optional batch comparison (tidy/long: one override row
  per ``name`` with a dotted ``target``) gated by an ``enabled`` toggle.
  Shipped disabled; read by
  :func:`pvbess_opt.scenarios.read_scenarios_block`.

Public loader API
-----------------

* :func:`read_workbook` returns the typed nested dict:

  .. code-block:: python

     {
         "ts": pd.DataFrame,               # lowercase snake_case
         "project":            {...},
         "pv":                 {...},
         "bess":               {...},
         "economics":          {...},
         "simulation":         {...},
         "max_injection_profile": np.ndarray,  # shape (24,) or (24, 12)
         "max_injection_profile_pv": np.ndarray | None,    # optional sub-cap
         "max_injection_profile_bess": np.ndarray | None,  # optional sub-cap
         "dt_minutes": int,                # auto-detected from the timeseries
     }

* :func:`read_inputs` returns a flat ``(params, ts)`` tuple suitable for
  the optimizer / KPI / lifetime modules.

Mode-specific timeseries semantics
----------------------------------

* In ``self_consumption`` mode the ``load_kwh`` column is required; missing → ValueError.
* In ``merchant`` mode ``load_kwh`` is optional — if present, the loader
  logs an INFO message and the optimizer pins all load-coverage flows to 0.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from .balancing import resolve_balancing_config
from .constants import (
    BENCHMARK_LCOE_HIGH_EUR_PER_MWH,
    BENCHMARK_LCOE_LOW_EUR_PER_MWH,
    BENCHMARK_LCOS_HIGH_EUR_PER_MWH,
    BENCHMARK_LCOS_LOW_EUR_PER_MWH,
    DEFAULT_MAX_INJECTION_PCT_HOURLY,
    DEFAULT_SENSITIVITY_DELTA_PCT,
    DEFAULT_SENSITIVITY_DISCOUNT_RATE_DELTA_PP,
)
from .io_style import style_workbook

logger = logging.getLogger(__name__)

__all__ = [
    "BALANCING_SHEET_DEFAULTS",
    "BESS_SHEET_DEFAULTS",
    "ECONOMICS_SHEET_DEFAULTS",
    "FALSY",
    "LAYOUT_SUBDIRS",
    "PPA_SHEET_DEFAULTS",
    "PROJECT_SHEET_DEFAULTS",
    "PV_SHEET_DEFAULTS",
    "SIMULATION_SHEET_DEFAULTS",
    "TRUTHY",
    "copy_input_snapshot",
    "detect_timestep_minutes",
    "make_run_layout",
    "read_inputs",
    "read_workbook",
    "validate_pv_location_fields",
    "validate_workbook_params",
    "write_assumptions_summary",
    "write_dispatch_artifacts",
    "write_results_workbook",
    "write_summary_md",
    "write_workbook",
]

TRUTHY = {"true", "1", "yes", "y", "t"}
FALSY = {"false", "0", "no", "n", "f"}

# Tokens that disable the grid-export cap (treat as unlimited export).
# An empty cell is also treated as unlimited — see _parse_grid_export_max.
_GRID_EXPORT_UNLIMITED_TOKENS = {
    "inf", "infinity", "unlimited", "disabled", "none",
}

_COERCE_FAILED = object()


# ---------------------------------------------------------------------------
# Canonical defaults (single source of truth)
# ---------------------------------------------------------------------------

PROJECT_SHEET_DEFAULTS: dict[str, Any] = {
    "project_lifecycle_years": 20,
    "project_start_year": 2026,
    "mode": "self_consumption",
    "p_grid_export_max_kw": 5000.0,
    "retail_tariff_eur_per_mwh": 120.0,
    "allow_bess_grid_charging": False,
    # Regulated charging-side wedge (EUR/MWh) on grid-charged BESS
    # energy, entering the MILP objective AND the cashflow (Eq. E26).
    "grid_charging_fee_eur_per_mwh": 0.0,
    # Storage network-charge exemption regime: TRUE zeroes the wedge.
    "grid_charging_fee_exempt": False,
    "grid_cap_includes_load": False,
    "unavailability_pct": 1.0,
    "site_capex_eur": 0.0,
    "site_devex_eur": 0.0,
    "currency_format": "auto",
    "show_titles": False,
}

PV_SHEET_DEFAULTS: dict[str, Any] = {
    "pv_source": "auto",
    # PVGIS location / array geometry — consulted when the PV profile is
    # fetched by location instead of read from the pv_kwh column.
    # latitude / longitude / timeseries_path default to None (absent); the
    # geometry knobs carry PVGIS-sensible defaults.
    "latitude": None,
    "longitude": None,
    "tilt": "optimal",
    "azimuth": 0.0,
    "losses_pct": 14.0,
    "weather_year": 2019,
    "raddatabase": None,
    "timeseries_path": None,
    "pv_nameplate_kwp": 0.0,
    "pv_degradation_year1_pct": 2.5,
    "pv_degradation_annual_pct": 0.55,
    "capex_pv_eur_per_kw": 525.0,
    "devex_pv_eur_per_kw": 60.0,
    "opex_pv_eur_per_kwp": 7.0,
}

BESS_SHEET_DEFAULTS: dict[str, Any] = {
    "bess_power_kw": 0.0,
    "bess_capacity_kwh": 0.0,
    "efficiency_charge": 0.95,
    "efficiency_discharge": 0.95,
    "soc_min_frac": 0.20,
    "soc_max_frac": 0.95,
    "initial_soc_frac": 0.50,
    "terminal_soc_equal": True,
    "max_cycles_per_day": 1.0,
    # Full installed BESS cost per kWh of nameplate energy capacity
    # (Lazard-style band 215-315 EUR/kWh).
    "capex_bess_eur_per_kwh": 250.0,
    "devex_bess_eur_per_kw": 30.0,
    "opex_bess_eur_per_kw": 14.0,
    "bess_replacement_year": 0,
    "bess_replacement_cost_pct": 50.0,
    "bess_degradation_annual_pct": 2.0,
    # LFP cycle-fade default (matches the canonical workbook row in
    # _BESS_ROWS and the schema default); range 0.005-0.010.
    "bess_degradation_pct_per_cycle": 0.008,
    # End-of-life SOH threshold (%) driving the automatic replacement
    # when bess_replacement_year is blank / 'auto'.
    "bess_eol_soh_pct": 80.0,
    # Per-MWh-throughput cycle wear cost in the dispatch objective
    # (0 = off).  Default 10: the optimizer skips marginal cycles whose
    # spread does not beat the degradation cost.
    "bess_wear_cost_eur_per_mwh": 10.0,
}

ECONOMICS_SHEET_DEFAULTS: dict[str, Any] = {
    "discount_rate_pct": 7.0,
    "opex_inflation_pct": 1.0,
    "retail_inflation_pct": 0.0,
    "dam_inflation_pct": 0.0,
    # Default 0.0 (fee-free): market representation fees are opt-in.  Real
    # aggregator/route-to-market charges are typically EUR/MWh of sold energy
    # or a share of the market revenue only — a flat 10 % of ALL revenue
    # (the old template default) sits far above European market practice, so
    # the template no longer pre-fills it.
    "aggregator_fee_pct_revenue": 0.0,
    # Structural market-access fees (European practice), both default-off so
    # existing results stay bit-identical.  The route-to-market fee is the
    # per-MWh representation charge (Greek FoSE / FoSETeK, German
    # Direktvermarktung); the optimizer share is the BESS trading revenue
    # share (merchant / floor+share structures).
    "route_to_market_fee_eur_per_mwh": 0.0,
    "optimizer_revenue_share_pct": 0.0,
    # Optimizer floor + share-above-floor structure (Eqs. E30/E30a):
    # with the floor enabled, the share applies to the margin ABOVE the
    # guaranteed floor and shortfalls are topped up.  Default-off so the
    # plain E13d share (and existing results) stay bit-identical.
    "optimizer_floor_enabled": False,
    "optimizer_floor_eur_per_kw_year": 0.0,
    "optimizer_term_year_from": 1,
    "optimizer_term_year_to": 0,
    "optimizer_margin_basis": "dam",
    # Optional, separate route-to-market (BSP / balancing-aggregator) fee on
    # GROSS balancing revenue.  Default 0.0 so existing results are
    # bit-identical; balancing carries no energy-aggregator fee but may carry
    # this one (Gridcog-style per-stream route-to-market cost).
    "balancing_aggregator_fee_pct_revenue": 0.0,
    # BESS tolling agreement (Eqs. E29/E29a): fixed EUR/MW/yr for dispatch
    # rights over a phase window (Eq. E25).  Default-off (rate 0) so
    # existing results stay bit-identical.
    "bess_toll_eur_per_mw_year": 0.0,
    "bess_toll_year_from": 1,
    "bess_toll_year_to": 0,
    "bess_toll_merchant_treatment": "zeroed",
    "bess_toll_indexation_pct": 0.0,
    # State support with two-way clawback (Eqs. E31/E31a): a fixed
    # EUR/MW/yr support netted two-way against realised market revenue
    # relative to a threshold (RRF-style storage support; the Greek
    # Tameio Anakampsis / TAA auctions are the reference mechanism).
    # Default-off (rate 0) so existing results stay bit-identical.
    "state_support_eur_per_mw_year": 0.0,
    "state_support_year_from": 1,
    "state_support_year_to": 0,
    "state_support_clawback_threshold_eur_per_mw_year": 0.0,
    "state_support_clawback_share_pct": 0.0,
    "state_support_indexation_pct": 0.0,
    # Capacity-market payment with derating factor (Eq. E32): paid on
    # the derated power block over a contract window; counts as
    # realised market revenue for the E31a netting.  Default-off.
    "capacity_market_eur_per_mw_year": 0.0,
    "capacity_market_derating_pct": 100.0,
    "capacity_market_year_from": 1,
    "capacity_market_year_to": 0,
    "capacity_market_indexation_pct": 0.0,
    # Revenue levy on gross market turnover (Eq. E33), e.g. the 3 %
    # special RES turnover levy applied in Greece.  Default-off.
    "revenue_levy_pct": 0.0,
    "benchmark_lcoe_low_eur_per_mwh": BENCHMARK_LCOE_LOW_EUR_PER_MWH,
    "benchmark_lcoe_high_eur_per_mwh": BENCHMARK_LCOE_HIGH_EUR_PER_MWH,
    "benchmark_lcos_low_eur_per_mwh": BENCHMARK_LCOS_LOW_EUR_PER_MWH,
    "benchmark_lcos_high_eur_per_mwh": BENCHMARK_LCOS_HIGH_EUR_PER_MWH,
    "sensitivity_enabled": True,
    "sensitivity_capex_delta_pct": DEFAULT_SENSITIVITY_DELTA_PCT,
    "sensitivity_opex_delta_pct": DEFAULT_SENSITIVITY_DELTA_PCT,
    "sensitivity_revenue_delta_pct": DEFAULT_SENSITIVITY_DELTA_PCT,
    "sensitivity_discount_rate_delta_pp": DEFAULT_SENSITIVITY_DISCOUNT_RATE_DELTA_PP,
    # PPA-strike tornado driver (active only when a PPA contract is on).
    "sensitivity_ppa_price_delta_pct": DEFAULT_SENSITIVITY_DELTA_PCT,
    # Project-finance debt layer (all-equity by default: gearing 0).
    "gearing_pct": 0.0,
    "debt_interest_rate_pct": 5.0,
    "debt_tenor_years": 15,
    "debt_repayment": "annuity",
    # Grid emissions / 24/7 CFE accounting (off by default: intensity 0).
    "grid_co2_intensity_kg_per_mwh": 0.0,
    "grid_co2_annual_decline_pct": 0.0,
}

BALANCING_SHEET_DEFAULTS: dict[str, Any] = {
    # Master switch — when False the MILP and KPIs are bit-identical to
    # a workbook without the sheet.
    "balancing_enabled": False,
    # Per-product capacity shares (% of bess_power_kw). DAM keeps the
    # majority; the sum across all six lines must stay <= 100 %.
    "dam_capacity_share_pct": 70.0,
    "fcr_capacity_share_pct": 10.0,
    "afrr_up_capacity_share_pct": 8.0,
    "afrr_dn_capacity_share_pct": 7.0,
    "mfrr_up_capacity_share_pct": 3.0,
    "mfrr_dn_capacity_share_pct": 2.0,
    # Per-product bid-acceptance probabilities (% of submitted bids that
    # clear the auction).
    "fcr_bid_acceptance_pct": 70.0,
    "afrr_up_bid_acceptance_pct": 55.0,
    "afrr_dn_bid_acceptance_pct": 55.0,
    "mfrr_up_bid_acceptance_pct": 40.0,
    "mfrr_dn_bid_acceptance_pct": 40.0,
    # Per-product activation probabilities (% of cleared bids that get
    # activated within a settlement period).
    "fcr_activation_probability_pct": 15.0,
    "afrr_up_activation_probability_pct": 10.0,
    "afrr_dn_activation_probability_pct": 8.0,
    "mfrr_up_activation_probability_pct": 5.0,
    "mfrr_dn_activation_probability_pct": 4.0,
    # Per-product fallback capacity prices used when the timeseries
    # column is absent (EUR per MWh).
    "fcr_default_capacity_price_eur_per_mwh": 12.0,
    "afrr_up_default_capacity_price_eur_per_mwh": 18.0,
    "afrr_dn_default_capacity_price_eur_per_mwh": 15.0,
    "mfrr_up_default_capacity_price_eur_per_mwh": 6.0,
    "mfrr_dn_default_capacity_price_eur_per_mwh": 5.0,
    # Per-product fallback activation prices (EUR per MWh). FCR has no
    # activation payment so it is absent here.
    "afrr_up_default_activation_price_eur_per_mwh": 220.0,
    "afrr_dn_default_activation_price_eur_per_mwh": 25.0,
    "mfrr_up_default_activation_price_eur_per_mwh": 180.0,
    "mfrr_dn_default_activation_price_eur_per_mwh": 20.0,
    # FCR-specific duration requirement (hours of sustained output
    # required for the reservation to be certifiable).
    "fcr_required_duration_hours": 0.5,
    # Settlement period (minutes); must equal 60 * dt_hours when
    # balancing_enabled.
    "bm_settlement_minutes": 15,
    # Extra SOC safety buffer applied on top of the worst-case
    # activation reservation (percent of activation energy).
    "bm_soc_headroom_pct": 10.0,
    # Yearly indexation rate applied to balancing revenue lines in the
    # multi-year cashflow.
    "bm_inflation_pct": 2.0,
    # Log-normal sigmas (in percent) used by the Monte Carlo to perturb
    # capacity and activation prices around the deterministic schedule.
    "bm_price_sigma_capacity_pct": 25.0,
    "bm_price_sigma_activation_pct": 35.0,
    # Number of Monte Carlo scenarios realised post-solve for the
    # balancing-revenue distribution (P10/P50/P90 + histogram).
    "bm_mc_scenarios": 200,
    # Default Monte Carlo seed for the balancing realisation.
    "bm_random_seed": 1729,
}

PPA_SHEET_DEFAULTS: dict[str, Any] = {
    # Master switch — when False the engine, KPIs and outputs are
    # bit-identical to a workbook without the sheet.
    "ppa_enabled": False,
    # Contract structure.  'pay_as_produced' is the implemented
    # envelope; 'baseload' is reserved (designed in docs/ppa_design.md,
    # rejected by validation with guidance until shortfall pricing
    # lands).
    "ppa_structure": "pay_as_produced",
    # Settlement decomposition: 'physical' (sleeved — the covered
    # volume is paid the strike and never touches the DAM) or 'cfd'
    # (full DAM exposure plus a two-way strike-minus-DAM leg).
    "ppa_settlement": "physical",
    # Contract strike (EUR/MWh) on the covered volume.
    "ppa_price_eur_per_mwh": 65.0,
    # Covered share of the PV EXPORT (pro-rata per step).
    "ppa_volume_share_pct": 100.0,
    # Operating years 1..term under contract; the covered volume
    # reverts to the DAM afterwards.
    "ppa_term_years": 10,
    # Yearly indexation of the contract strike ((1+i)^(y-1)).
    "ppa_inflation_pct": 0.0,
    # Negative-DAM-hour clause: 'none' (pay through negative hours) or
    # 'suspend' (contract pauses in DAM < 0 steps — Eqs. P6-P8).
    "ppa_negative_price_rule": "none",
}

SIMULATION_SHEET_DEFAULTS: dict[str, Any] = {
    "uncertainty_enabled": False,
    "uncertainty_compare_sources": False,
    "uncertainty_n_seeds": 30,
    "uncertainty_window_hours": 48,
    "uncertainty_commit_hours": 24,
    "uncertainty_dam_enabled": True,
    "uncertainty_pv_enabled": True,
    "uncertainty_load_enabled": True,
    "uncertainty_sigma_dam": 0.20,
    "uncertainty_sigma_pv": 0.12,
    "uncertainty_sigma_load": 0.05,
    "uncertainty_diagnostics_enabled": True,
    # Ex-post imbalance settlement of forecast-error deviations
    # (Eqs. U6-U9); requires the rolling-horizon Monte Carlo.
    "imbalance_enabled": False,
    "imbalance_pricing": "dual",
    "imbalance_price_mult_short": 1.25,
    "imbalance_price_mult_long": 0.75,
    "plot_daily_scope": "year1_only",
    "plot_monthly_scope": "all",
    "plot_yearly_scope": "all",
}

# Sheet → defaults map.  Used by the loader to validate keys per sheet.
_SHEET_DEFAULTS: dict[str, dict[str, Any]] = {
    "project": PROJECT_SHEET_DEFAULTS,
    "pv": PV_SHEET_DEFAULTS,
    "bess": BESS_SHEET_DEFAULTS,
    "economics": ECONOMICS_SHEET_DEFAULTS,
    "simulation": SIMULATION_SHEET_DEFAULTS,
    "balancing": BALANCING_SHEET_DEFAULTS,
    "ppa": PPA_SHEET_DEFAULTS,
}

_KEY_TO_SHEET: dict[str, str] = {}
for _sheet_name, _sheet_defaults in _SHEET_DEFAULTS.items():
    for _key in _sheet_defaults:
        _KEY_TO_SHEET[_key] = _sheet_name


# ---------------------------------------------------------------------------
# Per-key parsing metadata
# ---------------------------------------------------------------------------

_BOOL_KEYS: frozenset[str] = frozenset({
    "show_titles",
    "allow_bess_grid_charging",
    "grid_charging_fee_exempt",
    "grid_cap_includes_load",
    "terminal_soc_equal",
    "sensitivity_enabled",
    "uncertainty_enabled",
    "uncertainty_compare_sources",
    "uncertainty_dam_enabled",
    "uncertainty_pv_enabled",
    "uncertainty_load_enabled",
    "uncertainty_diagnostics_enabled",
    "imbalance_enabled",
    "balancing_enabled",
    "ppa_enabled",
    "optimizer_floor_enabled",
})
_INT_KEYS: frozenset[str] = frozenset({
    "project_lifecycle_years",
    "project_start_year",
    "debt_tenor_years",
    "bess_replacement_year",
    "uncertainty_n_seeds",
    "uncertainty_window_hours",
    "uncertainty_commit_hours",
    "bm_settlement_minutes",
    "bm_mc_scenarios",
    "bm_random_seed",
    "ppa_term_years",
    "bess_toll_year_from",
    "bess_toll_year_to",
    "optimizer_term_year_from",
    "optimizer_term_year_to",
    "state_support_year_from",
    "state_support_year_to",
    "capacity_market_year_from",
    "capacity_market_year_to",
})
_STR_KEYS: frozenset[str] = frozenset({
    "mode",
    "currency_format",
    "plot_daily_scope",
    "plot_monthly_scope",
    "plot_yearly_scope",
    "pv_source",
    "debt_repayment",
    "ppa_structure",
    "ppa_settlement",
    "ppa_negative_price_rule",
    "imbalance_pricing",
    "bess_toll_merchant_treatment",
    "optimizer_margin_basis",
})
_ALLOWED_VALUES: dict[str, frozenset[str]] = {
    "mode": frozenset({"self_consumption", "merchant"}),
    "currency_format": frozenset({"auto", "millions", "raw"}),
    "plot_daily_scope": frozenset({"none", "year1_only", "all"}),
    "plot_monthly_scope": frozenset({"none", "year1_only", "all"}),
    "plot_yearly_scope": frozenset({"none", "year1_only", "all"}),
    "pv_source": frozenset({"auto", "file", "pvgis"}),
    "debt_repayment": frozenset({"annuity", "linear"}),
    # 'baseload' parses (so old workbooks load) but validation rejects
    # it with guidance while only pay_as_produced is implemented.
    "ppa_structure": frozenset({"pay_as_produced", "baseload"}),
    "ppa_settlement": frozenset({"physical", "cfd"}),
    "ppa_negative_price_rule": frozenset({"none", "suspend"}),
    "imbalance_pricing": frozenset({"single", "dual"}),
    "bess_toll_merchant_treatment": frozenset({"zeroed", "retained"}),
    "optimizer_margin_basis": frozenset({"dam", "dam_plus_balancing"}),
}


# ---------------------------------------------------------------------------
# Sheet row templates (used by the workbook writer)
# ---------------------------------------------------------------------------

_PROJECT_ROWS: tuple[tuple[str, object, str, str], ...] = (
    ("project_lifecycle_years", PROJECT_SHEET_DEFAULTS["project_lifecycle_years"], "years",
     "Total project horizon used to project Years 0..N."),
    ("project_start_year", 2026, "year",
     "Calendar year of Year 1 (first operating year). CAPEX is paid in "
     "Year 0 (calendar = project_start_year - 1)."),
    ("mode", "self_consumption", "enum",
     "Accepted values: 'self_consumption' or 'merchant'. "
     "self_consumption requires a co-located load and enforces hard load "
     "priority (pv_to_load == min(pv, load)) plus no simultaneous grid "
     "I/O via a tight big-M. merchant has no co-located load; "
     "pv_to_load / bess_dis_load / grid_to_load are pinned to zero and "
     "PV / BESS dispatch entirely to the DAM. The per-step max-injection "
     "export cap (p_grid_export_max_kw x max_injection_profile) applies "
     "unconditionally in BOTH modes — curtailment is never skipped."),
    ("p_grid_export_max_kw", 5000, "kW",
     "Max grid export (kW). Leave empty or use 'inf' / 'unlimited' / "
     "'disabled' to remove cap; no injection limit is applied."),
    ("retail_tariff_eur_per_mwh", 120, "EUR/MWh",
     "Retail tariff used in self_consumption mode for load coverage."),
    ("allow_bess_grid_charging", False, "bool",
     "If TRUE the BESS may charge from the grid in periods with pv_kwh ~ 0."),
    ("grid_charging_fee_eur_per_mwh", 0.0, "EUR/MWh",
     "Regulated charging-side wedge on grid-charged BESS energy: "
     "network use-of-system charges plus levies applied to storage "
     "charging where not exempt. Enters the MILP objective as a buy-"
     "price adder on grid-to-BESS energy (thin arbitrage spreads flip "
     "correctly) AND the cashflow as its own expense line (Eq. E26). "
     "Typical European range 10-30 EUR/MWh; 0 = no wedge. Inert unless "
     "the dispatch actually grid-charges (allow_bess_grid_charging)."),
    ("grid_charging_fee_exempt", False, "bool",
     "Storage network-charge exemption regime switch. TRUE = the "
     "project qualifies for a storage charging exemption and the wedge "
     "above is ignored (effective fee 0 in the objective AND the "
     "cashflow); FALSE = the wedge applies. Keeps the exempt / "
     "non-exempt scenario pair a one-cell switch."),
    ("grid_cap_includes_load", False, "bool",
     "Sets what the per-step grid-injection cap limits (self_consumption mode "
     "only). FALSE (default) = PHYSICAL / co-located self-consumption: the load "
     "is behind the plant meter and served directly, so only the SURPLUS reaches "
     "the grid and the cap limits surplus export. TRUE = VIRTUAL Net-Billing: the "
     "load is remote (no physical link to the plant), so the plant injects ALL "
     "generation into the grid and the offset against the remote load is computed "
     "each 15-min settlement; the cap then limits TOTAL plant injection (energy "
     "credited to the remote load PLUS any surplus). Load priority stays "
     "strict but shares the cap: its floor becomes min(pv, load, cap), so the "
     "load takes all available injection capacity before any surplus export. "
     "When the cap cannot fit the full load the uncovered remainder is grid-"
     "served at the retail tariff and surplus PV is curtailed — the run is "
     "never infeasible, it degrades to the maximum feasible coverage."),
    ("unavailability_pct", 1.0, "%",
     "Annual unavailability (outages / scheduled maintenance) applied as "
     "a post-solve derate on PV generation, BESS discharge, and revenue."),
    ("site_capex_eur", 0.0, "EUR",
     "Site-wide lump-sum CAPEX in absolute EUR for items that are not "
     "naturally per-kWp/per-kW (substation construction, MV/HV grid "
     "upgrades, interconnection works, etc.). Paid in Year 0. Excluded "
     "from LCOE/LCOS by the Lazard convention."),
    ("site_devex_eur", 0.0, "EUR",
     "Site-wide lump-sum DEVEX in absolute EUR (environmental impact "
     "studies, land acquisition fees, permits not expressed per-kW, "
     "etc.). Paid in Year 0. Excluded from LCOE/LCOS by the Lazard "
     "convention."),
    ("currency_format", "auto", "enum",
     "auto | millions | raw — financial-axis label format."),
    ("show_titles", False, "bool",
     "Render plot titles. IEEE figures normally rely on the figure caption."),
)

_PV_ROWS: tuple[tuple[str, object, str, str], ...] = (
    ("pv_source", "auto", "enum",
     "auto | file | pvgis. auto: use pv_kwh if filled, else fetch by "
     "location."),
    ("latitude", None, "deg",
     "PVGIS only: required when pv_kwh is empty. Decimal degrees, north "
     "positive, range -90..90 (e.g. Athens 37.9838, Thessaloniki 40.6401)."),
    ("longitude", None, "deg",
     "PVGIS only: required when pv_kwh is empty. Decimal degrees, east "
     "positive, range -180..180 (e.g. Athens 23.7275, Thessaloniki 22.9444)."),
    ("tilt", "optimal", "deg",
     'PVGIS only: degrees, or "optimal".'),
    ("azimuth", 0, "deg",
     "PVGIS only: 0 = south, 90 = west, -90 = east."),
    ("losses_pct", 14, "%",
     "PVGIS only: system losses."),
    ("weather_year", 2019, "year",
     'PVGIS only: non-leap year for a clean 8760, or "tmy".'),
    ("raddatabase", None, "enum",
     "PVGIS only: optional radiation-database override (e.g. "
     "'PVGIS-SARAH3' or 'PVGIS-ERA5'). Blank = let PVGIS pick the "
     "regional default for the location."),
    ("timeseries_path", None, "path",
     "file only: optional external CSV/Parquet instead of the pv_kwh "
     "column."),
    ("pv_nameplate_kwp", 0, "kWp",
     "PV nameplate capacity. 0 = no PV in this project. The pv_kwh "
     "timeseries is consumed verbatim (absolute kWh per step); nameplate "
     "is metadata used for per-kW CAPEX/OPEX and the sizing sweep axis."),
    ("pv_degradation_year1_pct", 2.5, "%",
     "Initial light-induced degradation (LID) applied at start of Year 2."),
    ("pv_degradation_annual_pct", 0.55, "%",
     "Linear PV degradation after Year 1 (Tier-1 warranty)."),
    ("capex_pv_eur_per_kw", 525, "EUR/kWp",
     "Per-kWp PV CAPEX. Set 0 if PV already exists."),
    ("devex_pv_eur_per_kw", 60, "EUR/kWp",
     "Per-kWp PV DEVEX (development / permitting). Paid in Year 0."),
    ("opex_pv_eur_per_kwp", 7, "EUR/kWp/yr",
     "Annual O&M for PV."),
)

_BESS_ROWS: tuple[tuple[str, object, str, str], ...] = (
    ("bess_power_kw", 0, "kW",
     "BESS power rating (symmetric charge / discharge limit). "
     "0 = no BESS in this project."),
    ("bess_capacity_kwh", 0, "kWh",
     "BESS energy capacity. Pinned to the workbook value (industry "
     "standard for sizing-as-input projects)."),
    ("efficiency_charge", 0.95, "-",
     "Charge efficiency (0..1). Round-trip = "
     "efficiency_charge * efficiency_discharge (0.95 x 0.95 = 0.9025)."),
    ("efficiency_discharge", 0.95, "-",
     "Discharge efficiency (0..1)."),
    ("soc_min_frac", 0.20, "-",
     "Minimum SOC as fraction of nominal capacity (0.20 = 20 %)."),
    ("soc_max_frac", 0.95, "-",
     "Maximum SOC as fraction of nominal capacity (0.95 = 95 %)."),
    ("initial_soc_frac", 0.50, "-",
     "SOC at the first timestep, as a fraction of capacity."),
    ("terminal_soc_equal", True, "bool",
     "If TRUE, force final SOC == initial SOC (closed cycle)."),
    ("max_cycles_per_day", 1.0, "-",
     "Daily equivalent-cycle cap (sum of discharge / capacity)."),
    ("bess_wear_cost_eur_per_mwh", 10.0, "EUR/MWh",
     "Cycle wear cost penalised per MWh discharged in the dispatch "
     "objective (default 10; 0 = off). A dispatch shadow price only: "
     "never charged in the cashflow (the replacement CAPEX carries "
     "degradation cost), so it is not double-counted. Applies to DAM "
     "and self-consumption discharge; expected balancing-activation "
     "throughput carries no wear penalty by design. Derive from "
     "replacement cost / cycle-life / usable energy via "
     "pvbess_opt.degradation."),
    ("capex_bess_eur_per_kwh", 250, "EUR/kWh",
     "Full installed BESS CAPEX per kWh of nameplate energy capacity "
     "(cells + PCS + BOP + EPC). Benchmark band 215-315 EUR/kWh "
     "(Lazard LCOS, utility-scale 4-hour Li-ion). Set 0 if the BESS "
     "already exists."),
    ("devex_bess_eur_per_kw", 30, "EUR/kW",
     "Per-kW BESS DEVEX (development / permitting). Paid in Year 0. "
     "Stays on the power basis: development and permitting effort "
     "scales with the power block, not the energy capacity."),
    ("opex_bess_eur_per_kw", 14, "EUR/kW/yr",
     "Annual O&M for BESS. Stays on the power basis: fixed O&M "
     "contracts are quoted per kW of the power block."),
    ("bess_replacement_year", 0, "year",
     "BESS replacement policy. N (positive integer) = replace in project "
     "year N; bess_eol_soh_pct is then ignored. Blank cell or 'auto' = "
     "replace in the first year state-of-health falls to "
     "bess_eol_soh_pct (the replacement CAPEX is charged in the "
     "cashflow in that year). 0 = never replace. Typical scheduled "
     "values 10 or 15. Only ONE replacement is ever charged; if a "
     "second SOH crossing would occur after an auto replacement the "
     "run log warns."),
    ("bess_replacement_cost_pct", 50, "%",
     "Replacement cost as percent of original BESS CAPEX. Charged in "
     "the effective replacement year (scheduled or auto)."),
    ("bess_degradation_annual_pct", 2.0, "%",
     "Linear BESS capacity fade. Approximate Tier-1 LFP cell warranty."),
    ("bess_degradation_pct_per_cycle", 0.008, "%",
     "Cycle-based BESS capacity fade per full equivalent cycle, in "
     "percent. LFP default 0.008 (range 0.005-0.010). Set to 0 to "
     "disable cycle aging (calendar-only mode)."),
    ("bess_eol_soh_pct", 80.0, "%",
     "End-of-life state-of-health threshold driving the automatic "
     "replacement when bess_replacement_year is blank or 'auto': the "
     "battery is replaced (and the replacement CAPEX charged in the "
     "cashflow) in the first project year SOH falls to this level. "
     "Ignored when a replacement year is scheduled explicitly, and "
     "when bess_replacement_year = 0 (never replace)."),
)

_ECONOMICS_ROWS: tuple[tuple[str, object, str, str], ...] = (
    ("discount_rate_pct", 7.0, "%",
     "WACC. Typical EU RES band 6-8 %."),
    ("opex_inflation_pct", 1.0, "%",
     "Annual OPEX escalation rate."),
    ("retail_inflation_pct", 0.0, "%",
     "Annual indexation of the retail tariff (load-coverage revenue). "
     "0 = no indexation. The PPA stream has its own knob "
     "(ppa_inflation_pct on the ppa sheet)."),
    ("dam_inflation_pct", 0.0, "%",
     "Annual indexation of wholesale DAM revenue (exports). Default 0 "
     "since DAM prices are driven by gas/CO2/RES penetration, not CPI. "
     "Industry tools (Lazard, Aurora, Gridcog) use exogenous price "
     "curves, not flat inflation."),
    ("aggregator_fee_pct_revenue", 0.0, "%",
     "Energy-aggregator fee on gross DAM + retail revenue (Gridcog "
     "convention; see public Gridcog cost / pricing docs). Does NOT apply "
     "to balancing or PPA revenue. Default 0 = fee-free (opt-in): real "
     "route-to-market charges are typically EUR/MWh of sold energy or a "
     "share of market revenue only, well below a flat percentage of ALL "
     "revenue. Prefer the two structural fees below "
     "(route_to_market_fee_eur_per_mwh, optimizer_revenue_share_pct) for "
     "market-faithful modelling; a warning fires if this and the optimizer "
     "share are both set (they would double-charge the battery's wholesale "
     "stream)."),
    ("route_to_market_fee_eur_per_mwh", 0.0, "EUR/MWh",
     "Route-to-market / representation fee per MWh of grid-EXPORTED energy "
     "(PV + BESS) — the charge a cumulative-representation aggregator "
     "(Greek FoSE, or the last-resort FoSETeK under regulated charges; "
     "German Direktvermarkter) levies for scheduling, forecasting, "
     "balancing responsibility and market access. Charged on SOLD energy "
     "only: never on self-consumption savings, balancing or PPA revenue, "
     "and the PPA-covered PV export share is exempt while a physical "
     "(sleeved) contract is in term (the offtaker routes that volume). "
     "Typical 0.5-5 EUR/MWh (Greek examples ~1-3.5). Flat over the project "
     "life (no indexation). 0 = off (default). Surfaces as a signed "
     "route_to_market_fee_eur cashflow column. Excluded from LCOE/LCOS."),
    ("optimizer_revenue_share_pct", 0.0, "%",
     "Battery optimizer / trading-services revenue share, charged on the "
     "POSITIVE annual BESS wholesale trading margin (DAM export revenue "
     "minus grid-charging cost); nothing is charged in years where the "
     "margin is negative (an optimizer never invoices a share of a loss). "
     "Mirrors the merchant revenue-share / floor+share structures used by "
     "BESS optimizers; typical 10-25 %. Applies ONLY to the battery's "
     "wholesale stream: never to self-consumption, PPA, or balancing "
     "revenue (the BSP fee below covers balancing). 0 = off (default). "
     "Surfaces as a signed optimizer_fee_eur cashflow column. Excluded "
     "from LCOE/LCOS."),
    ("optimizer_floor_enabled", False, "bool",
     "Enable the floor+share optimizer structure (Eq. E30): the "
     "optimizer guarantees optimizer_floor_eur_per_kw_year and "
     "optimizer_revenue_share_pct applies to the margin ABOVE the "
     "floor (not the whole positive margin). FALSE (default) = plain "
     "share, exactly the E13d behaviour. Note a floor of 0 with this "
     "enabled still guarantees a non-negative margin (losses are "
     "topped up)."),
    ("optimizer_floor_eur_per_kw_year", 0.0, "EUR/kW/yr",
     "Guaranteed annual floor per kW of BESS power under the "
     "floor+share structure (a standard BESS-optimizer contract form). "
     "Availability-scaled (x A); flat nominal, no capacity-fade "
     "scaling. Ignored unless optimizer_floor_enabled. Surfaces as an "
     "optimizer_floor_topup_eur column (>= 0). Excluded from "
     "LCOE/LCOS."),
    ("optimizer_term_year_from", 1, "year",
     "First project year of the optimizer contract (share and floor), "
     "inclusive. Default 1 preserves the whole-life share behaviour."),
    ("optimizer_term_year_to", 0, "year",
     "Last project year of the optimizer contract (inclusive); 0 = end "
     "of life (default, preserving the whole-life behaviour). After "
     "the term neither share nor floor applies."),
    ("optimizer_margin_basis", "dam", "dam | dam_plus_balancing",
     "Margin base for share and floor: 'dam' (default; the BESS DAM "
     "trading margin, the E13d base) or 'dam_plus_balancing' (adds "
     "balancing net of the BSP fee, i.e. the full E25a base - "
     "optimizers routinely manage ancillary revenue too). The share "
     "applies AFTER the BSP fee; fees never compound. Default keeps "
     "results bit-identical."),
    ("balancing_aggregator_fee_pct_revenue", 0.0, "%",
     "Optional, separate route-to-market (BSP / balancing-aggregator) fee on "
     "GROSS balancing revenue (capacity + activation). Default 0 = fee-free, "
     "bit-identical to today. Set it when balancing is routed through an "
     "aggregator/BSP that keeps a share (typical behind-the-meter / smaller "
     "assets, ~5-20 %); per-stream route-to-market cost, Gridcog convention. "
     "Surfaces as a signed balancing_aggregator_fee_eur cashflow column. "
     "Never applies to DAM / retail / PPA revenue."),
    ("bess_toll_eur_per_mw_year", 0.0, "EUR/MW/yr",
     "Tolling agreement: fixed annual payment per MW of BESS power for "
     "dispatch rights. 0 = off (default, bit-identical). Applied x "
     "availability factor; no capacity-fade scaling (payment is on the "
     "power block, not delivered energy). Surfaces as a toll_revenue_eur "
     "cashflow column. Excluded from LCOE/LCOS."),
    ("bess_toll_year_from", 1, "year",
     "First project year of the toll window (inclusive). Must be >= 1."),
    ("bess_toll_year_to", 0, "year",
     "Last project year of the toll window (inclusive). 0 = through end "
     "of project life. Must be 0 or >= bess_toll_year_from. E.g. toll "
     "years 1-5 then merchant: from=1, to=5."),
    ("bess_toll_merchant_treatment", "zeroed", "zeroed | retained",
     "'zeroed' (default): in toll years the BESS-origin merchant streams "
     "(BESS DAM margin incl. grid-charging cost and the charging-side "
     "grid fee, balancing capacity+activation, BSP fee, BESS "
     "route-to-market fee share, optimizer fee) are zeroed - the toller "
     "keeps them. 'retained': the toll stacks on top of retained "
     "merchant streams (warns: double-monetises the same MW; use only "
     "for capacity-overlay contracts)."),
    ("bess_toll_indexation_pct", 0.0, "%/yr",
     "Annual escalation of the toll rate ((1+i)^(y-1) convention, "
     "Eq. E2). Default 0 = flat nominal."),
    ("state_support_eur_per_mw_year", 0.0, "EUR/MW/yr",
     "Fixed annual state support per MW of BESS power (RRF-style "
     "storage support; the Tameio Anakampsis / TAA auctions are the "
     "Greek reference - a neutral mechanism applicable to any "
     "fixed-support scheme). 0 = off (default). Availability-scaled "
     "(x A); no capacity-fade scaling. Surfaces as state_support_eur. "
     "Excluded from LCOE/LCOS."),
    ("state_support_year_from", 1, "year",
     "First project year of the support window (inclusive)."),
    ("state_support_year_to", 0, "year",
     "Last project year of the support window (inclusive); 0 = end of "
     "life. Typical schemes run 10 years."),
    ("state_support_clawback_threshold_eur_per_mw_year", 0.0,
     "EUR/MW/yr",
     "Reference market-revenue level theta for the two-way netting: "
     "realised BESS market revenue per MW above theta is clawed back, "
     "below theta is compensated, both at "
     "state_support_clawback_share_pct. Inactive while the share is "
     "0."),
    ("state_support_clawback_share_pct", 0.0, "%",
     "Share of the (market revenue - threshold) difference netted "
     "against the support, both directions (two-way). 0 (default) = "
     "pure fixed support, no netting; 100 = full two-way settlement. "
     "The netted support can turn a year's combined support negative "
     "(a net repayment) - no floor is applied; the run log flags any "
     "such year."),
    ("state_support_indexation_pct", 0.0, "%/yr",
     "Annual escalation of the support level AND the threshold "
     "((1+i)^(y-1), Eq. E2). Default 0 = flat nominal."),
    ("capacity_market_eur_per_mw_year", 0.0, "EUR/MW/yr",
     "Capacity-market payment per DERATED MW per year. 0 = off "
     "(default). Paid on bess_power_kw x capacity_market_derating_pct, "
     "availability-scaled; no capacity-fade scaling. Counts as "
     "realised market revenue for the state-support netting "
     "(Eq. E31a). Surfaces as capacity_market_revenue_eur. Excluded "
     "from LCOE/LCOS."),
    ("capacity_market_derating_pct", 100.0, "%",
     "Derating factor applied to BESS power for the capacity payment "
     "(EU mechanisms derate storage by duration vs the stress-event "
     "window; enter the auction's published class factor, e.g. "
     "~40-95 % depending on duration). 100 = underated nameplate. In "
     "[0, 100]. The payment is on the DERATED MW (convention stated "
     "to avoid double-derating)."),
    ("capacity_market_year_from", 1, "year",
     "First project year of the capacity contract (inclusive)."),
    ("capacity_market_year_to", 0, "year",
     "Last project year (inclusive); 0 = end of life. Multi-year "
     "new-build contracts are typically 10-17 years."),
    ("capacity_market_indexation_pct", 0.0, "%/yr",
     "Annual escalation of the clearing price ((1+i)^(y-1), Eq. E2); "
     "many mechanisms CPI-index - default 0 = flat nominal."),
    ("revenue_levy_pct", 0.0, "%",
     "Levy on gross market turnover: wholesale DAM export revenue "
     "(gross, before the aggregator fee - levies charge turnover, not "
     "net-of-fee revenue), balancing capacity + activation revenue, "
     "and the PPA contract leg. Self-consumption savings are excluded "
     "(avoided cost, not invoiced turnover), as are the contracted "
     "streams (toll / state support / capacity market) and the "
     "imbalance settlement. Example: the 3 % special RES turnover levy "
     "applied in Greece. Charged before income tax and deductible from "
     "taxable income by construction. 0 = off (default; results "
     "bit-identical). Surfaces as a signed revenue_levy_eur cashflow "
     "column inside net_cashflow_eur (clamped: negative turnover never "
     "yields a rebate). Excluded from LCOE/LCOS. Validated in "
     "[0, 100]."),
    ("benchmark_lcoe_low_eur_per_mwh", BENCHMARK_LCOE_LOW_EUR_PER_MWH, "EUR/MWh",
     "Lower edge of the Lazard 2024 utility-scale PV LCOE band "
     "(EUR-equivalent at ~1.08 EUR/USD). Overrideable per project."),
    ("benchmark_lcoe_high_eur_per_mwh", BENCHMARK_LCOE_HIGH_EUR_PER_MWH, "EUR/MWh",
     "Upper edge of the Lazard 2024 utility-scale PV LCOE band "
     "(EUR-equivalent at ~1.08 EUR/USD)."),
    ("benchmark_lcos_low_eur_per_mwh", BENCHMARK_LCOS_LOW_EUR_PER_MWH, "EUR/MWh",
     "Lower edge of the Lazard 2024 utility-scale 4-hour Li-ion LCOS "
     "band (EUR-equivalent at ~1.08 EUR/USD)."),
    ("benchmark_lcos_high_eur_per_mwh", BENCHMARK_LCOS_HIGH_EUR_PER_MWH, "EUR/MWh",
     "Upper edge of the Lazard 2024 utility-scale 4-hour Li-ion LCOS "
     "band (EUR-equivalent at ~1.08 EUR/USD)."),
    ("sensitivity_enabled", True, "bool",
     "Run a one-at-a-time tornado sensitivity after the base run."),
    ("sensitivity_capex_delta_pct", DEFAULT_SENSITIVITY_DELTA_PCT, "%",
     "Symmetric +/- delta on total CAPEX (incl. DEVEX)."),
    ("sensitivity_opex_delta_pct", DEFAULT_SENSITIVITY_DELTA_PCT, "%",
     "Symmetric +/- delta on total annual OPEX."),
    ("sensitivity_revenue_delta_pct", DEFAULT_SENSITIVITY_DELTA_PCT, "%",
     "Symmetric +/- delta on Year-1 revenue base."),
    ("sensitivity_discount_rate_delta_pp", DEFAULT_SENSITIVITY_DISCOUNT_RATE_DELTA_PP, "pp",
     "Symmetric +/- delta on the discount rate, in percentage points. "
     "NPV tornado only - drops out of IRR tornado by definition."),
    ("sensitivity_ppa_price_delta_pct", DEFAULT_SENSITIVITY_DELTA_PCT, "%",
     "Symmetric +/- delta on the PPA strike. The driver appears in the "
     "tornado only when the ppa sheet's contract is enabled and the "
     "Year-1 PPA stream is non-zero."),
    ("gearing_pct", 0.0, "%",
     "Debt fraction of the initial investment (0 = all-equity, the "
     "default; unlevered results are unchanged)."),
    ("debt_interest_rate_pct", 5.0, "%",
     "Annual debt interest rate (used only when gearing_pct > 0)."),
    ("debt_tenor_years", 15, "years",
     "Debt repayment tenor in years (used only when gearing_pct > 0)."),
    ("debt_repayment", "annuity", "annuity | linear",
     "Repayment profile: annuity (level debt service) or linear "
     "(level principal)."),
    ("grid_co2_intensity_kg_per_mwh", 0.0, "kg/MWh",
     "Grid carbon intensity for emissions / 24/7 CFE accounting "
     "(0 = off, the default; an optional grid_co2_kg_per_mwh time-series "
     "column overrides this per step). Never affects dispatch or NPV."),
    ("grid_co2_annual_decline_pct", 0.0, "%",
     "Annual decline of the grid carbon intensity over the project life "
     "(grid decarbonisation). 0 = constant intensity."),
)

_SIMULATION_ROWS: tuple[tuple[str, object, str, str], ...] = (
    ("uncertainty_enabled", False, "bool",
     "Run rolling-horizon Monte Carlo. Default FALSE (perfect-foresight only)."),
    ("uncertainty_compare_sources", False, "bool",
     "When TRUE run 4 ensembles (DAM-only, PV-only, Load-only, "
     "All-combined) and emit a comparison plot."),
    ("uncertainty_n_seeds", 30, "int",
     "Monte Carlo seeds per ensemble."),
    ("uncertainty_window_hours", 48, "int",
     "Rolling window length."),
    ("uncertainty_commit_hours", 24, "int",
     "Commit slice."),
    ("uncertainty_dam_enabled", True, "bool",
     "Apply DAM noise."),
    ("uncertainty_pv_enabled", True, "bool",
     "Apply PV noise."),
    ("uncertainty_load_enabled", True, "bool",
     "Apply Load noise (ignored in merchant mode)."),
    ("uncertainty_sigma_dam", 0.20, "-",
     "Log-normal sigma for DAM. Default 0.20 (ENTSO-E D+1 benchmark)."),
    ("uncertainty_sigma_pv", 0.12, "-",
     "Log-normal sigma for PV. Default 0.12 (NREL day-ahead PV study)."),
    ("uncertainty_sigma_load", 0.05, "-",
     "Log-normal sigma for Load. Default 0.05 (predictable customer benchmark)."),
    ("uncertainty_diagnostics_enabled", True, "bool",
     "Render the forecast-calibration diagnostic plots (coverage, PIT, "
     "CRPS, residual Q-Q) into 06_uncertainty_plots/. Default TRUE."),
    ("imbalance_enabled", False, "bool",
     "Master switch for the ex-post imbalance settlement of "
     "forecast-error deviations (BRP deviation settlement). Requires "
     "uncertainty_enabled = TRUE and uncertainty_window_hours >= 2 x "
     "uncertainty_commit_hours (the day-ahead nomination is the noisy "
     "lookahead slice of each window, Eq. U6). Default FALSE - "
     "bit-identical when off."),
    ("imbalance_pricing", "dual", "enum",
     "single | dual. dual: short deviations pay the short price, long "
     "deviations are paid the long price (cost non-negative when long "
     "<= DAM <= short, Eq. U7). single: both sides settle at one "
     "imbalance price (sign-indefinite, Eq. U8); requires the "
     "imbalance_price_eur_per_mwh timeseries column."),
    ("imbalance_price_mult_short", 1.25, "-",
     "DAM-proxy multiplier for the SHORT imbalance price when the "
     "imbalance_price_short_eur_per_mwh column is absent (dual "
     "regime). Applied sign-aware to |DAM| (Eq. U8a) so negative-price "
     "hours keep long <= DAM <= short."),
    ("imbalance_price_mult_long", 0.75, "-",
     "DAM-proxy multiplier for the LONG imbalance price when the "
     "imbalance_price_long_eur_per_mwh column is absent (dual regime). "
     "Sign-aware like the short-side proxy."),
    ("plot_daily_scope", "year1_only", "scope",
     "none | year1_only | all. 'all' produces ~365 * N_years * 3 daily PDFs."),
    ("plot_monthly_scope", "all", "scope",
     "none | year1_only | all."),
    ("plot_yearly_scope", "all", "scope",
     "none | year1_only | all."),
)

_BALANCING_ROWS: tuple[tuple[str, object, str, str], ...] = (
    ("balancing_enabled", False, "bool",
     "Master switch for stochastic balancing market participation "
     "(FCR / aFRR / mFRR). When FALSE the MILP, KPIs and outputs are "
     "bit-identical to a run without the sheet."),
    ("dam_capacity_share_pct", 70.0, "%",
     "Declarative share used only by the validator to ensure the total "
     "across DAM + every balancing product stays <= 100 % of bess_power_kw. "
     "It does not actively cap DAM dispatch in the MILP — DAM consumes the "
     "residual of bess_power_kw not reserved for balancing, implicitly "
     "bounded each step by BM_POWER_UP / BM_POWER_DN."),
    ("fcr_capacity_share_pct", 10.0, "%",
     "Share of bess_power_kw available for FCR reservation (symmetric)."),
    ("afrr_up_capacity_share_pct", 8.0, "%",
     "Share of bess_power_kw available for aFRR-up reservation."),
    ("afrr_dn_capacity_share_pct", 7.0, "%",
     "Share of bess_power_kw available for aFRR-down reservation."),
    ("mfrr_up_capacity_share_pct", 3.0, "%",
     "Share of bess_power_kw available for mFRR-up reservation."),
    ("mfrr_dn_capacity_share_pct", 2.0, "%",
     "Share of bess_power_kw available for mFRR-down reservation."),
    ("fcr_bid_acceptance_pct", 70.0, "%",
     "Probability that a submitted FCR bid clears the auction."),
    ("afrr_up_bid_acceptance_pct", 55.0, "%",
     "Probability that a submitted aFRR-up bid clears the auction."),
    ("afrr_dn_bid_acceptance_pct", 55.0, "%",
     "Probability that a submitted aFRR-down bid clears the auction."),
    ("mfrr_up_bid_acceptance_pct", 40.0, "%",
     "Probability that a submitted mFRR-up bid clears the auction."),
    ("mfrr_dn_bid_acceptance_pct", 40.0, "%",
     "Probability that a submitted mFRR-down bid clears the auction."),
    ("fcr_activation_probability_pct", 15.0, "%",
     "Informational only. FCR is modelled as capacity-only (no activation "
     "payment) and as symmetric in expectation (no SOC drift), so the MILP, "
     "KPIs and Monte Carlo realisation do not consume this value. Retained "
     "for documentation and future use should an FCR activation revenue "
     "stream be added."),
    ("afrr_up_activation_probability_pct", 10.0, "%",
     "Probability a cleared aFRR-up reservation is activated within a "
     "settlement period."),
    ("afrr_dn_activation_probability_pct", 8.0, "%",
     "Probability a cleared aFRR-down reservation is activated."),
    ("mfrr_up_activation_probability_pct", 5.0, "%",
     "Probability a cleared mFRR-up reservation is activated."),
    ("mfrr_dn_activation_probability_pct", 4.0, "%",
     "Probability a cleared mFRR-down reservation is activated."),
    ("fcr_default_capacity_price_eur_per_mwh", 12.0, "EUR/MWh",
     "FCR capacity-price fallback when the timeseries column is absent."),
    ("afrr_up_default_capacity_price_eur_per_mwh", 18.0, "EUR/MWh",
     "aFRR-up capacity-price fallback when the timeseries column is absent."),
    ("afrr_dn_default_capacity_price_eur_per_mwh", 15.0, "EUR/MWh",
     "aFRR-down capacity-price fallback when the timeseries column is absent."),
    ("mfrr_up_default_capacity_price_eur_per_mwh", 6.0, "EUR/MWh",
     "mFRR-up capacity-price fallback when the timeseries column is absent."),
    ("mfrr_dn_default_capacity_price_eur_per_mwh", 5.0, "EUR/MWh",
     "mFRR-down capacity-price fallback when the timeseries column is absent."),
    ("afrr_up_default_activation_price_eur_per_mwh", 220.0, "EUR/MWh",
     "aFRR-up activation-price fallback when the timeseries column is absent."),
    ("afrr_dn_default_activation_price_eur_per_mwh", 25.0, "EUR/MWh",
     "aFRR-down activation-price fallback when the timeseries column is absent."),
    ("mfrr_up_default_activation_price_eur_per_mwh", 180.0, "EUR/MWh",
     "mFRR-up activation-price fallback when the timeseries column is absent."),
    ("mfrr_dn_default_activation_price_eur_per_mwh", 20.0, "EUR/MWh",
     "mFRR-down activation-price fallback when the timeseries column is absent."),
    ("fcr_required_duration_hours", 0.5, "hours",
     "FCR-specific sustained-output requirement; sizes the SOC headroom "
     "reserved for FCR independently of the settlement period."),
    ("bm_settlement_minutes", 15, "int",
     "Balancing-market settlement period in minutes. Must equal "
     "60 * dt_hours when balancing_enabled is TRUE."),
    ("bm_soc_headroom_pct", 10.0, "%",
     "Extra SOC safety buffer applied to the worst-case activation "
     "reservation in both directions."),
    ("bm_inflation_pct", 2.0, "%",
     "Yearly indexation of balancing revenue applied in the multi-year "
     "lifetime cashflow."),
    ("bm_price_sigma_capacity_pct", 25.0, "%",
     "Log-normal sigma for Monte Carlo perturbation of capacity prices."),
    ("bm_price_sigma_activation_pct", 35.0, "%",
     "Log-normal sigma for Monte Carlo perturbation of activation prices."),
    ("bm_mc_scenarios", 200, "int",
     "Number of Monte Carlo scenarios realised post-solve for the "
     "balancing-revenue distribution (P10/P50/P90 KPIs and the "
     "balancing_mc_distribution plot)."),
    ("bm_random_seed", 1729, "int",
     "Default seed for the balancing Monte Carlo realisation."),
)


_PPA_ROWS: tuple[tuple[str, object, str, str], ...] = (
    ("ppa_enabled", False, "bool",
     "Master switch for the pay-as-produced PPA contract. When FALSE "
     "the dispatch, KPIs and outputs are bit-identical to a workbook "
     "without the sheet. See docs/ppa_design.md."),
    ("ppa_structure", "pay_as_produced", "enum",
     "Contract structure. 'pay_as_produced' (as-generated offtake on a "
     "share of the PV export) is the implemented envelope; 'baseload' "
     "is reserved for a future shaped profile and is rejected with "
     "guidance while unimplemented."),
    ("ppa_settlement", "physical", "enum",
     "physical | cfd. physical (sleeved): the covered volume is paid "
     "the strike and never touches the DAM. cfd: all PV export sells "
     "at DAM and the covered volume adds a two-way strike-minus-DAM "
     "leg (negative when DAM exceeds the strike). Both total share x "
     "export x strike on the covered volume, so dispatch is identical "
     "and only the revenue decomposition differs."),
    ("ppa_price_eur_per_mwh", 65.0, "EUR/MWh",
     "Contract strike on the covered volume."),
    ("ppa_volume_share_pct", 100.0, "%",
     "Covered share of the PV EXPORT, applied pro-rata per step "
     "(self-consumed PV is settled at retail and is not offtake "
     "volume; BESS export is not covered)."),
    ("ppa_term_years", 10, "years",
     "Operating years 1..term under contract. After the term the "
     "stream ends; under physical settlement the covered volume's DAM "
     "value rejoins the DAM revenue stream (where the aggregator fee "
     "applies to it as market revenue)."),
    ("ppa_inflation_pct", 0.0, "%",
     "Yearly indexation of the contract strike ((1+i)^(y-1)). "
     "Independent of retail_inflation_pct (CPI-linked tariffs) and "
     "dam_inflation_pct (wholesale view)."),
    ("ppa_negative_price_rule", "none", "enum",
     "none | suspend. Negative-DAM-hour clause on the contract. "
     "'none' (default): the covered volume settles through negative "
     "hours unchanged (as-produced behaviour). 'suspend': in every "
     "step with DAM < 0 the contract pauses - physical: the covered "
     "volume is not paid the strike and faces spot; cfd: the "
     "difference leg is suspended while the market leg keeps selling. "
     "The dispatch then curtails or charges the BESS instead of "
     "exporting covered PV at a loss. Common in European "
     "pay-as-produced offtake terms and premium schemes."),
)


_SHEET_ROW_TEMPLATES: dict[
    str, tuple[tuple[str, object, str, str], ...]
] = {
    "project": _PROJECT_ROWS,
    "pv": _PV_ROWS,
    "bess": _BESS_ROWS,
    "economics": _ECONOMICS_ROWS,
    "simulation": _SIMULATION_ROWS,
    "balancing": _BALANCING_ROWS,
    "ppa": _PPA_ROWS,
}

# Default share of p_grid_export_max_kw available for export (24 hourly
# rows) applied when the workbook omits the max_injection_profile sheet.
# Single source of truth lives in pvbess_opt.constants; re-imported above.


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_kv_sheet(
    typed_section: dict[str, Any],
    rows: tuple[tuple[str, object, str, str], ...],
) -> pd.DataFrame:
    out: list[dict[str, Any]] = []
    for key, default, unit, notes in rows:
        value = typed_section.get(key, default)
        out.append(
            {"key": key, "value": value, "unit": unit, "notes": notes},
        )
    return pd.DataFrame(out, columns=["key", "value", "unit", "notes"])


def _hour_interval_labels() -> list[str]:
    """24 strings of the form ``HH:00-HH:00`` covering 00:00 → 24:00."""
    return [f"{h:02d}:00-{(h + 1):02d}:00" for h in range(24)]


def _build_max_injection_sheet(profile: Any) -> pd.DataFrame:
    """Render the ``max_injection_profile`` sheet from a 1-D or 2-D array.

    Accepts:
    * shape ``(24,)`` → single ``max_injection_pct`` column.
    * shape ``(24, 12)`` → per-month columns (``max_injection_pct_jan`` ..
      ``max_injection_pct_dec``).

    The ``hour_of_day`` column is rendered as **24-hour interval
    strings** (``"00:00-01:00"`` … ``"23:00-24:00"``) for human
    readability.  Values are interpreted as the percent of
    ``p_grid_export_max_kw`` available for export in that hour.
    """
    arr = np.asarray(profile, dtype=float)
    hour_labels = _hour_interval_labels()
    if arr.ndim == 1:
        if arr.shape[0] != 24:
            raise ValueError(
                "max_injection_profile must have 24 rows "
                f"(got {arr.shape[0]})."
            )
        return pd.DataFrame({
            "hour_of_day": hour_labels,
            "max_injection_pct": arr,
        })
    if arr.ndim == 2:
        if arr.shape != (24, 12):
            raise ValueError(
                "max_injection_profile (2-D) must be shape (24, 12) "
                f"(got {arr.shape})."
            )
        cols: dict[str, Any] = {"hour_of_day": hour_labels}
        for m_idx, m_name in enumerate(_MONTH_TOKENS):
            cols[f"max_injection_pct_{m_name}"] = arr[:, m_idx]
        return pd.DataFrame(cols)
    raise ValueError(
        "max_injection_profile must be 1-D (24,) or 2-D (24, 12); "
        f"got shape {arr.shape}."
    )


_MONTH_TOKENS: tuple[str, ...] = (
    "jan", "feb", "mar", "apr", "may", "jun",
    "jul", "aug", "sep", "oct", "nov", "dec",
)


# ---------------------------------------------------------------------------
# Optional sweep sheet: sizing (capacity sweep)
# ---------------------------------------------------------------------------

# Sizing-sweep sheet: one column per grid axis, one value per row, plus an
# ``enabled`` TRUE/FALSE toggle read from the first data row.  Parsed by
# pvbess_opt.sizing._parse_sizing_sheet and expanded into the Cartesian
# product by pvbess_opt.sizing.parse_sizing_grid.  ``bess_capacity_kwh``
# takes precedence over ``bess_duration_hours`` when both carry values.
# ``enabled`` = FALSE (the shipped default) leaves a normal single run
# untouched, so the sweep never fires unless the user opts in.
SIZING_SHEET_COLUMNS: tuple[str, ...] = (
    "enabled",
    "pv_nameplate_kwp",
    "bess_power_kw",
    "bess_capacity_kwh",
    "bess_duration_hours",
)

# Disabled example grid shipped in the workbook so the columnar format is
# self-documenting.  Blank cells are skipped; here capacity is left blank so
# the duration column drives the BESS-energy axis (power x duration).
_SIZING_EXAMPLE_ROWS: tuple[tuple[Any, ...], ...] = (
    ("FALSE", 10000.0, 10000.0, None, 2.0),
    (None, 15000.0, 15000.0, None, 4.0),
    (None, 20000.0, 20000.0, None, None),
)


def _build_sizing_sheet(
    rows: list[dict[str, Any]] | None = None,
) -> pd.DataFrame:
    """Render the optional ``sizing`` sweep sheet.

    With no explicit ``rows`` the shipped disabled example is written so the
    columnar format is self-documenting.
    """
    if rows:
        frame = pd.DataFrame(rows)
        for col in SIZING_SHEET_COLUMNS:
            if col not in frame.columns:
                frame[col] = None
        return frame[list(SIZING_SHEET_COLUMNS)]
    return pd.DataFrame(
        [
            dict(zip(SIZING_SHEET_COLUMNS, r, strict=True))
            for r in _SIZING_EXAMPLE_ROWS
        ],
        columns=list(SIZING_SHEET_COLUMNS),
    )


# ---------------------------------------------------------------------------
# Optional sweep sheet: scenarios (batch comparison)
# ---------------------------------------------------------------------------

# Scenarios sheet: tidy / long — one row per override, grouped by ``name``,
# with an optional ``inherits`` clone reference and a dotted ``target``
# (e.g. ``project.mode``, ``bess.power_kw``, or the bare specials
# ``capex_multiplier`` / ``balancing``).  Parsed by
# pvbess_opt.scenarios._parse_scenarios_sheet and run by run_scenarios.
# Gated by an ``enabled`` TRUE/FALSE toggle read from the first data row;
# shipped disabled so a normal run is unaffected.
SCENARIOS_SHEET_COLUMNS: tuple[str, ...] = (
    "enabled",
    "name",
    "inherits",
    "target",
    "value",
)

# Disabled worked example mirroring examples/scenarios.yaml so the
# tidy/long format is self-documenting.  A scenario spans the consecutive
# rows that share its ``name``.
_SCENARIOS_EXAMPLE_ROWS: tuple[tuple[Any, ...], ...] = (
    ("FALSE", "Self-consumption hybrid", None, "project.mode", "self_consumption"),
    (None, "Merchant hybrid", None, "project.mode", "merchant"),
    (None, "Merchant hybrid + balancing", "Merchant hybrid", "balancing", "on"),
    (None, "Merchant PV only", "Merchant hybrid", "bess.power_kw", 0),
    (None, "Merchant PV only", "Merchant hybrid", "bess.capacity_kwh", 0),
    (None, "Cheap CAPEX (merchant)", "Merchant hybrid", "capex_multiplier", 0.8),
    (None, "Merchant hybrid + PPA", "Merchant hybrid", "ppa.ppa_enabled", "TRUE"),
    (None, "Merchant hybrid + PPA", "Merchant hybrid", "ppa.ppa_volume_share_pct", 80),
    # A price deck selects <column>__<deck> variant columns of the base
    # timeseries (e.g. dam_price_eur_per_mwh__high) for this scenario.
    (None, "Merchant hybrid (high prices)", "Merchant hybrid", "price_deck", "high"),
)


def _build_scenarios_sheet(
    rows: list[dict[str, Any]] | None = None,
) -> pd.DataFrame:
    """Render the optional ``scenarios`` batch sheet.

    With no explicit ``rows`` the shipped disabled example is written so the
    tidy/long format is self-documenting.
    """
    if rows:
        frame = pd.DataFrame(rows)
        for col in SCENARIOS_SHEET_COLUMNS:
            if col not in frame.columns:
                frame[col] = None
        return frame[list(SCENARIOS_SHEET_COLUMNS)]
    return pd.DataFrame(
        [
            dict(zip(SCENARIOS_SHEET_COLUMNS, r, strict=True))
            for r in _SCENARIOS_EXAMPLE_ROWS
        ],
        columns=list(SCENARIOS_SHEET_COLUMNS),
    )


# ---------------------------------------------------------------------------
# Optional escalation sheet: trajectories (per-year stream multipliers)
# ---------------------------------------------------------------------------

# Trajectories sheet: tidy / long — one row per (stream, year) multiplier,
# gated by an ``enabled`` TRUE/FALSE toggle read from the first data row
# (sizing / scenarios sheet pattern; shipped disabled so a normal run is
# bit-identical).  Each stream carries a ``mode``:
#
# * ``replace`` — the per-year multiplier m_y substitutes the stream's
#   flat (1 + i)^(y-1) inflation index (Eq. E24).  The loader warns when
#   the matching ``*_inflation_pct`` is also non-zero (double
#   specification).
# * ``overlay`` — m_y multiplies ON TOP of the inflation index.
#
# Year-indexed rows are deliberately used instead of comma-separated
# values in a single cell: locales that write decimal commas make
# CSV-in-a-cell a silent-corruption footgun, while Excel drag-fill makes
# one-row-per-year cheap.  The engine consumes the parsed block through
# ``economics`` (Eq. E24/E24a); this module owns the format and its
# structural validation only.
TRAJECTORY_STREAMS: tuple[str, ...] = (
    "revenue_dam",
    "revenue_retail",
    "balancing_capacity",
    "balancing_activation",
    "opex",
    "opex_pv",
    "opex_bess",
)

# Streams that deliberately take NO trajectory: the PPA strike escalates
# contractually (``ppa_inflation_pct``) and the route-to-market fee
# (Eq. E13c) is a flat per-MWh volume charge by design.

_TRAJECTORY_MODES: tuple[str, ...] = ("replace", "overlay")

# The scalar inflation key each stream's replace-mode trajectory
# substitutes, as ``(sheet, key)`` — used for the double-specification
# warning in ``validate_workbook_params``.
_TRAJECTORY_INFLATION_KEYS: dict[str, tuple[str, str]] = {
    "revenue_dam": ("economics", "dam_inflation_pct"),
    "revenue_retail": ("economics", "retail_inflation_pct"),
    "balancing_capacity": ("balancing", "bm_inflation_pct"),
    "balancing_activation": ("balancing", "bm_inflation_pct"),
    "opex": ("economics", "opex_inflation_pct"),
    "opex_pv": ("economics", "opex_inflation_pct"),
    "opex_bess": ("economics", "opex_inflation_pct"),
}

TRAJECTORIES_SHEET_COLUMNS: tuple[str, ...] = (
    "enabled",
    "stream",
    "mode",
    "year",
    "value",
)

# Disabled worked example shipped in the workbook so the tidy/long format
# is self-documenting (a mild DAM capture-rate decline).  Blank ``stream``
# / ``mode`` cells inherit the row above; enabling it requires filling
# every operating year 1..project_lifecycle_years.
_TRAJECTORIES_EXAMPLE_ROWS: tuple[tuple[Any, ...], ...] = (
    ("FALSE", "revenue_dam", "overlay", 1, 1.0),
    (None, None, None, 2, 0.99),
    (None, None, None, 3, 0.98),
)


def _build_trajectories_sheet(
    trajectories: dict[str, dict[str, Any]] | None = None,
) -> pd.DataFrame:
    """Render the optional ``trajectories`` sheet.

    ``trajectories`` is the parsed block (``{stream: {"mode": ...,
    "values": [...]}}``); ``None`` writes the shipped disabled example so
    the tidy format is self-documenting.
    """
    if trajectories:
        rows: list[dict[str, Any]] = []
        first = True
        for stream, spec in trajectories.items():
            for idx, value in enumerate(spec["values"], start=1):
                rows.append({
                    "enabled": "TRUE" if first else None,
                    "stream": stream if idx == 1 else None,
                    "mode": str(spec["mode"]) if idx == 1 else None,
                    "year": idx,
                    "value": float(value),
                })
                first = False
        return pd.DataFrame(rows, columns=list(TRAJECTORIES_SHEET_COLUMNS))
    return pd.DataFrame(
        [
            dict(zip(TRAJECTORIES_SHEET_COLUMNS, r, strict=True))
            for r in _TRAJECTORIES_EXAMPLE_ROWS
        ],
        columns=list(TRAJECTORIES_SHEET_COLUMNS),
    )


def _normalise_trajectories_block(
    raw: Any, *, source: str,
) -> dict[str, dict[str, Any]] | None:
    """Validate and normalise a trajectories mapping (shared by surfaces).

    Accepts ``{stream: {"mode": "replace"|"overlay", "values": [...]}}``
    or the list shorthand ``{stream: [...]}`` (replace mode).  Structural
    checks only: length-vs-lifecycle coverage and the m_1 == 1 anchor are
    enforced by :func:`validate_workbook_params`, which knows
    ``project_lifecycle_years``.
    """
    if raw is None:
        return None
    if not isinstance(raw, dict):
        raise ValueError(
            f"{source}: 'trajectories' must be a mapping of stream name "
            f"to a values list or {{mode, values}} block; got "
            f"{type(raw).__name__}."
        )
    if not raw:
        return None
    out: dict[str, dict[str, Any]] = {}
    for stream, spec in raw.items():
        stream = str(stream).strip().lower()
        if stream not in TRAJECTORY_STREAMS:
            raise ValueError(
                f"{source}: unknown trajectory stream {stream!r}; known "
                f"streams: {', '.join(TRAJECTORY_STREAMS)}."
            )
        if isinstance(spec, dict):
            mode = str(spec.get("mode", "replace")).strip().lower()
            values = spec.get("values")
        else:
            mode = "replace"
            values = spec
        if mode not in _TRAJECTORY_MODES:
            raise ValueError(
                f"{source}: trajectory stream {stream!r} has unknown mode "
                f"{mode!r}; expected one of {', '.join(_TRAJECTORY_MODES)}."
            )
        if not isinstance(values, (list, tuple)) or not values:
            raise ValueError(
                f"{source}: trajectory stream {stream!r} needs a non-empty "
                f"'values' list of per-year multipliers."
            )
        floats: list[float] = []
        for idx, v in enumerate(values, start=1):
            if isinstance(v, (bool, np.bool_)):
                raise ValueError(
                    f"{source}: trajectory stream {stream!r} year {idx} "
                    f"expects a number, got boolean {bool(v)!r}."
                )
            try:
                fv = float(v)
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"{source}: trajectory stream {stream!r} year {idx} "
                    f"value {v!r} is not a number."
                ) from exc
            if not np.isfinite(fv) or fv < 0.0:
                raise ValueError(
                    f"{source}: trajectory stream {stream!r} year {idx} "
                    f"multiplier must be finite and >= 0; got {fv!r}."
                )
            floats.append(fv)
        out[stream] = {"mode": mode, "values": floats}
    return out


def _parse_trajectories_sheet(
    df: pd.DataFrame,
) -> dict[str, dict[str, Any]] | None:
    """Parse the columnar ``trajectories`` sheet.

    Returns the normalised block (``{stream: {"mode": ..., "values":
    [...]}}``) or ``None`` when the sheet is disabled or empty.  Blank
    ``stream`` / ``mode`` cells inherit the row above (tidy/long, the
    scenarios-sheet convention); years must be contiguous from 1 per
    stream.
    """
    cols = {str(c).strip().lower(): c for c in df.columns}
    for required in TRAJECTORIES_SHEET_COLUMNS:
        if required not in cols:
            raise ValueError(
                f"trajectories sheet is missing the {required!r} column; "
                f"expected columns: {', '.join(TRAJECTORIES_SHEET_COLUMNS)}."
            )

    def cell(row: Any, name: str) -> Any:
        value = row[cols[name]]
        if value is None:
            return None
        if isinstance(value, float) and np.isnan(value):
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value

    enabled = False
    nonnull = df[cols["enabled"]].dropna()
    if not nonnull.empty:
        enabled = _parse_bool(nonnull.iloc[0], False)
    if not enabled:
        return None

    per_stream: dict[str, dict[str, Any]] = {}
    current: str | None = None
    for _, row in df.iterrows():
        stream_val = cell(row, "stream")
        if stream_val is not None:
            current = str(stream_val).strip().lower()
        year_val = cell(row, "year")
        value_val = cell(row, "value")
        if current is None:
            if year_val is None and value_val is None:
                continue
            raise ValueError(
                "trajectories sheet has a value row before any 'stream' "
                "cell; name the stream on the first row of each block."
            )
        if year_val is None and value_val is None:
            continue
        bucket = per_stream.setdefault(
            current, {"mode": None, "years": {}},
        )
        mode_val = cell(row, "mode")
        if mode_val is not None and bucket["mode"] is None:
            bucket["mode"] = str(mode_val).strip().lower()
        try:
            year = int(float(year_val))
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"trajectories stream {current!r}: year {year_val!r} is "
                f"not an integer."
            ) from exc
        if year < 1:
            raise ValueError(
                f"trajectories stream {current!r}: operating years start "
                f"at 1; got {year}."
            )
        if year in bucket["years"]:
            raise ValueError(
                f"trajectories stream {current!r}: duplicate year {year}."
            )
        bucket["years"][year] = value_val

    if not per_stream:
        return None

    raw_block: dict[str, Any] = {}
    for stream, bucket in per_stream.items():
        years = bucket["years"]
        expected = set(range(1, max(years) + 1))
        missing = sorted(expected - set(years))
        if missing:
            raise ValueError(
                f"trajectories stream {stream!r}: years must be "
                f"contiguous from 1; missing {missing}."
            )
        raw_block[stream] = {
            "mode": bucket["mode"] or "replace",
            "values": [years[y] for y in sorted(years)],
        }
    return _normalise_trajectories_block(
        raw_block, source="trajectories sheet",
    )


# Output-styling contract: every workbook written below passes through
# pvbess_opt.io_style.style_workbook before save, so all outputs share the
# input workbook's navy frozen-header house style.  Never save an output
# sheet unstyled.
def write_workbook(typed: dict[str, Any], dst: str | Path) -> Path:
    """Write a workbook from a typed nested dict."""
    dst = Path(dst)
    dst.parent.mkdir(parents=True, exist_ok=True)

    project_df = _build_kv_sheet(typed["project"], _PROJECT_ROWS)
    pv_df = _build_kv_sheet(typed["pv"], _PV_ROWS)
    bess_df = _build_kv_sheet(typed["bess"], _BESS_ROWS)
    economics_df = _build_kv_sheet(typed["economics"], _ECONOMICS_ROWS)
    simulation_df = _build_kv_sheet(typed["simulation"], _SIMULATION_ROWS)
    balancing_section = typed.get("balancing") or dict(BALANCING_SHEET_DEFAULTS)
    balancing_df = _build_kv_sheet(balancing_section, _BALANCING_ROWS)
    ppa_section = typed.get("ppa") or dict(PPA_SHEET_DEFAULTS)
    ppa_df = _build_kv_sheet(ppa_section, _PPA_ROWS)

    profile = typed.get("max_injection_profile")
    if profile is None:
        profile = np.full(24, DEFAULT_MAX_INJECTION_PCT_HOURLY, dtype=float)
    max_injection_df = _build_max_injection_sheet(profile)
    sizing_df = _build_sizing_sheet(typed.get("sizing"))
    scenarios_df = _build_scenarios_sheet(typed.get("scenarios"))
    trajectories_df = _build_trajectories_sheet(typed.get("trajectories"))

    with pd.ExcelWriter(dst, engine="openpyxl") as writer:
        typed["ts"].to_excel(writer, sheet_name="timeseries", index=False)
        project_df.to_excel(writer, sheet_name="project", index=False)
        pv_df.to_excel(writer, sheet_name="pv", index=False)
        bess_df.to_excel(writer, sheet_name="bess", index=False)
        economics_df.to_excel(writer, sheet_name="economics", index=False)
        simulation_df.to_excel(writer, sheet_name="simulation", index=False)
        balancing_df.to_excel(writer, sheet_name="balancing", index=False)
        ppa_df.to_excel(writer, sheet_name="ppa", index=False)
        max_injection_df.to_excel(
            writer, sheet_name="max_injection_profile", index=False,
        )
        for _src in ("pv", "bess"):
            _src_profile = typed.get(f"max_injection_profile_{_src}")
            if _src_profile is not None:
                _build_max_injection_sheet(_src_profile).to_excel(
                    writer,
                    sheet_name=f"max_injection_profile_{_src}",
                    index=False,
                )
        sizing_df.to_excel(writer, sheet_name="sizing", index=False)
        scenarios_df.to_excel(writer, sheet_name="scenarios", index=False)
        trajectories_df.to_excel(
            writer, sheet_name="trajectories", index=False,
        )
        style_workbook(writer.book)
    return dst


# ---------------------------------------------------------------------------
# Generic value-coercion helpers
# ---------------------------------------------------------------------------


def _coerce(value: Any, cast: type, default: Any) -> Any:
    """Cast ``value`` to ``cast``; return ``default`` on empty/NaN; sentinel on error."""
    if value is None:
        return default
    if isinstance(value, float) and np.isnan(value):
        return default
    if isinstance(value, str) and value.strip() == "":
        return default
    try:
        return cast(value)
    except (TypeError, ValueError):
        return _COERCE_FAILED


def _parse_bool(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and np.isnan(value):
            return default
        return value != 0
    token = str(value).strip().lower()
    if token == "":
        return default
    if token in TRUTHY:
        return True
    if token in FALSY:
        return False
    return default


def _parse_string_enum(
    value: Any, default: str, allowed: frozenset[str], key: str,
) -> str:
    if value is None:
        return default
    if isinstance(value, float) and np.isnan(value):
        return default
    token = str(value).strip().lower()
    if token == "":
        return default
    if token not in allowed:
        # An invalid value for a KNOWN enum key is a genuine mistake (a
        # typo or a stale option), not a forward/backward-compat unknown
        # key — fail loudly rather than silently substituting the default,
        # which could otherwise change the economics (e.g. a mistyped
        # ``ppa_settlement`` quietly reverting to ``physical``).  Unknown
        # *keys* are still warned-and-ignored by ``_parse_kv_sheet``.
        if key == "mode":
            raise ValueError(
                f"unknown mode {token!r}; valid modes are "
                f"{sorted(allowed)!r}"
            )
        raise ValueError(
            f"{key!r} value {value!r} is not one of {sorted(allowed)!r}."
        )
    return token


def _parse_pv_tilt(raw: Any, default: Any) -> Any:
    """Parse ``tilt``: the literal ``optimal`` or a number in degrees."""
    if raw is None:
        return default
    if isinstance(raw, float) and np.isnan(raw):
        return default
    if isinstance(raw, str):
        token = raw.strip().lower()
        if token == "":
            return default
        if token == "optimal":
            return "optimal"
    coerced = _coerce(raw, float, default)
    if coerced is _COERCE_FAILED:
        logger.warning(
            "Workbook value for 'tilt' could not be parsed as a number or "
            "'optimal' (got %r); using default %r.", raw, default,
        )
        return default
    return coerced


def _parse_pv_weather_year(raw: Any, default: Any) -> Any:
    """Parse ``weather_year``: a calendar year or the literal ``tmy``."""
    if raw is None:
        return default
    if isinstance(raw, float) and np.isnan(raw):
        return default
    if isinstance(raw, str):
        token = raw.strip().lower()
        if token == "":
            return default
        if token == "tmy":
            return "tmy"
    coerced = _coerce(raw, int, default)
    if coerced is _COERCE_FAILED:
        logger.warning(
            "Workbook value for 'weather_year' could not be parsed as a year "
            "or 'tmy' (got %r); using default %r.", raw, default,
        )
        return default
    return coerced


def _parse_pv_path(raw: Any, default: Any) -> Any:
    """Parse ``timeseries_path``: a free-form path string (blank → default)."""
    if raw is None:
        return default
    if isinstance(raw, float) and np.isnan(raw):
        return default
    text = str(raw).strip()
    return text if text else default


# ---------------------------------------------------------------------------
# Sheet → flat-dict reduction (skips separator rows)
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Per-sheet typed parser
# ---------------------------------------------------------------------------


def _parse_value(key: str, raw: Any, default: Any) -> Any:
    if key in _BOOL_KEYS:
        return _parse_bool(raw, bool(default))
    if key in _STR_KEYS:
        return _parse_string_enum(
            raw, str(default), _ALLOWED_VALUES.get(key, frozenset()), key,
        )
    if key == "tilt":
        return _parse_pv_tilt(raw, default)
    if key == "weather_year":
        return _parse_pv_weather_year(raw, default)
    if key in ("timeseries_path", "raddatabase"):
        # Free-form PVGIS strings: a blank cell resolves to the default
        # (None); a non-blank cell is taken verbatim (stripped).
        return _parse_pv_path(raw, default)
    if key == "bess_replacement_year":
        return _parse_bess_replacement_year(raw, default)
    # A boolean in a numeric field is always an input mistake, and Python
    # would coerce it silently (float(True) == 1.0, int(True) == 1), changing
    # results without a trace — e.g. ``unavailability_pct = TRUE`` becoming a
    # 1 % derate.  Reject loudly, naming the key, matching the fee-range
    # contract in ``validate_workbook_params``.  numpy bools included: a
    # pandas-read TRUE cell may arrive as ``np.bool_``, which is NOT a
    # ``bool`` subclass.
    if isinstance(raw, (bool, np.bool_)):
        raise ValueError(
            f"{key!r} expects a number, got boolean {bool(raw)!r}; "
            "write a numeric value instead (e.g. 1 for 1 %)."
        )
    if key in _INT_KEYS:
        coerced = _coerce(raw, int, default)
        if coerced is _COERCE_FAILED:
            logger.warning(
                "Workbook value for %r could not be parsed as int "
                "(got %r); using default %r.", key, raw, default,
            )
            return default
        return coerced
    coerced = _coerce(raw, float, default)
    if coerced is _COERCE_FAILED:
        logger.warning(
            "Workbook value for %r could not be parsed as float "
            "(got %r); using default %r.", key, raw, default,
        )
        return default
    return coerced


def _parse_grid_export_max(raw: Any, default: Any) -> float:
    """Parse ``p_grid_export_max_kw``.

    Returns ``float('inf')`` when the cap is disabled (empty cell, or one
    of the ``_GRID_EXPORT_UNLIMITED_TOKENS`` strings, case-insensitive).
    A finite positive float is returned unchanged.  Negative or zero
    values are returned as-is so the loader can raise a validation error;
    unparseable values fall back to ``default`` with a warning.
    """
    if raw is None:
        return float("inf")
    if isinstance(raw, float) and np.isnan(raw):
        return float("inf")
    if isinstance(raw, str):
        token = raw.strip().lower()
        if token == "" or token in _GRID_EXPORT_UNLIMITED_TOKENS:
            return float("inf")
        try:
            value = float(raw)
        except ValueError:
            logger.warning(
                "Workbook value for 'p_grid_export_max_kw' could not be "
                "parsed (got %r); using default %r.", raw, default,
            )
            return float(default)
    else:
        try:
            value = float(raw)
        except (TypeError, ValueError):
            logger.warning(
                "Workbook value for 'p_grid_export_max_kw' could not be "
                "parsed (got %r); using default %r.", raw, default,
            )
            return float(default)
    if np.isinf(value):
        return float("inf")
    return value


#: Legacy v0.x key for the BESS CAPEX, priced per kW of power.  v1.0.0
#: prices BESS CAPEX per kWh of energy capacity instead; the old key is
#: rejected loudly so an old workbook cannot silently fall back to the
#: new per-kWh default.
LEGACY_BESS_CAPEX_KEY = "capex_bess_eur_per_kw"


def reject_legacy_bess_capex_key(keys: Any, *, source: str) -> None:
    """Raise ``ValueError`` when the legacy per-kW BESS CAPEX key is present.

    ``keys`` is any iterable of key names from the bess section of a
    workbook or structured config; ``source`` names the input for the
    error message.  Conversion rule:
    ``value_per_kwh = value_per_kw x bess_power_kw / bess_capacity_kwh``.
    """
    if LEGACY_BESS_CAPEX_KEY in set(keys):
        raise ValueError(
            f"{source} uses the legacy key '{LEGACY_BESS_CAPEX_KEY}' "
            "(BESS CAPEX per kW of power). Since v1.0.0 BESS CAPEX is "
            "priced per kWh of energy capacity via "
            "'capex_bess_eur_per_kwh'. Convert the value with "
            "value_per_kwh = value_per_kw x bess_power_kw / "
            "bess_capacity_kwh, or run "
            "scripts/polish_input_workbook.py to migrate the workbook "
            "automatically."
        )


#: Sentinel value for the SOH-threshold automatic BESS replacement.
BESS_REPLACEMENT_AUTO = "auto"


def _parse_bess_replacement_year(raw: Any, default: Any) -> int | str:
    """Parse ``bess_replacement_year`` with three-way semantics.

    * blank cell or the literal string ``auto`` (case-insensitive) →
      :data:`BESS_REPLACEMENT_AUTO`: replace in the first project year
      whose SOH falls to ``bess_eol_soh_pct`` (resolved once by
      :func:`pvbess_opt.lifetime.resolve_bess_replacement_year`);
    * ``0`` → never replace;
    * a positive integer → scheduled replacement in that project year.

    Negative or non-integer values raise ``ValueError`` — this key must
    not fall back silently (the v0.x parser collapsed a blank cell into
    the default 0, silently disabling replacement).
    """
    _ = default  # three-way semantics leave no room for a fallback
    if raw is None:
        return BESS_REPLACEMENT_AUTO
    if isinstance(raw, float) and np.isnan(raw):
        return BESS_REPLACEMENT_AUTO
    if isinstance(raw, str):
        token = raw.strip().lower()
        if token == "" or token == BESS_REPLACEMENT_AUTO:
            return BESS_REPLACEMENT_AUTO
        try:
            value = float(token)
        except ValueError:
            raise ValueError(
                f"'bess_replacement_year' must be a non-negative integer, "
                f"blank, or 'auto'; got {raw!r}."
            ) from None
    else:
        try:
            value = float(raw)
        except (TypeError, ValueError):
            raise ValueError(
                f"'bess_replacement_year' must be a non-negative integer, "
                f"blank, or 'auto'; got {raw!r}."
            ) from None
    if not float(value).is_integer():
        raise ValueError(
            f"'bess_replacement_year' must be a whole year; got {raw!r}."
        )
    year = int(value)
    if year < 0:
        raise ValueError(
            f"'bess_replacement_year' must be non-negative; got {raw!r}."
        )
    return year


def _parse_kv_sheet(
    sheet_name: str, flat: dict[str, Any],
) -> dict[str, Any]:
    defaults = _SHEET_DEFAULTS[sheet_name]
    out = dict(defaults)
    for key, raw in flat.items():
        if key in defaults:
            # Boolean in a numeric field: raise here (the sheet name is in
            # scope for a precise message) rather than letting the guard in
            # ``_parse_value`` fire without it.  The workbook reader feeds
            # this parser openpyxl-faithful cell types (``_read_kv_flat``),
            # so a bool here is a genuine TRUE/FALSE cell, never a pandas
            # inference artifact.
            if key not in _BOOL_KEYS and isinstance(raw, (bool, np.bool_)):
                raise ValueError(
                    f"Sheet {sheet_name!r}: {key!r} expects a number, got "
                    f"boolean {bool(raw)!r}; write a numeric value instead "
                    "(e.g. 1 for 1 %)."
                )
            if key == "p_grid_export_max_kw":
                out[key] = _parse_grid_export_max(raw, defaults[key])
            else:
                out[key] = _parse_value(key, raw, defaults[key])
            continue
        # Unknown key for this sheet — but maybe it belongs to another sheet?
        if key in _KEY_TO_SHEET:
            logger.warning(
                "Key %r found on %r sheet but belongs to %r sheet; ignored.",
                key, sheet_name, _KEY_TO_SHEET[key],
            )
            continue
        logger.warning(
            "%s sheet key %r is unknown; ignored.", sheet_name, key,
        )
    return out


# ---------------------------------------------------------------------------
# Max-injection profile parser
# ---------------------------------------------------------------------------


_HOUR_PARSE_RE = re.compile(r"^\s*(\d{1,2})")


def _parse_hour_of_day(value: Any) -> int:
    """Coerce an ``hour_of_day`` cell into an integer 0..23.

    Accepts an integer 0..23 and the 24-hour interval
    string format (``"00:00-01:00"`` … ``"23:00-24:00"``).  The
    parser is forgiving: any leading 1-2 digit run is taken as the
    start hour.  Out-of-range values raise ``ValueError``.
    """
    if isinstance(value, (int, np.integer)):
        h = int(value)
    elif isinstance(value, (float, np.floating)):
        if np.isnan(value):
            raise ValueError("hour_of_day cell is NaN")
        h = int(value)
    else:
        s = str(value).strip()
        m = _HOUR_PARSE_RE.match(s)
        if not m:
            raise ValueError(
                f"cannot parse hour_of_day value {value!r}; "
                "expected an integer 0..23 or an interval like '00:00-01:00'"
            )
        h = int(m.group(1))
    if h < 0 or h > 23:
        raise ValueError(
            f"hour_of_day must be in 0..23 (got {h} from {value!r})"
        )
    return h


def _normalise_hourly_profile_frame(
    df: pd.DataFrame, *, sheet_name: str,
) -> pd.DataFrame:
    """Validate columns / row count and lowercase column names."""
    if df is None or df.empty:
        raise ValueError(f"{sheet_name} sheet is empty.")
    cols = {c.strip().lower() for c in df.columns}
    if "hour_of_day" not in cols:
        raise ValueError(
            f"{sheet_name} sheet must contain a 'hour_of_day' column."
        )
    df_norm = df.rename(columns={c: c.strip().lower() for c in df.columns})
    df_norm["hour_of_day"] = df_norm["hour_of_day"].map(_parse_hour_of_day)
    df_norm = df_norm.sort_values("hour_of_day").reset_index(drop=True)
    if len(df_norm) != 24:
        raise ValueError(
            f"{sheet_name} sheet must have exactly 24 rows "
            f"(got {len(df_norm)})."
        )
    hours = df_norm["hour_of_day"].astype(int).to_numpy()
    if not np.array_equal(hours, np.arange(24)):
        raise ValueError(
            f"{sheet_name} 'hour_of_day' column must cover 0..23 "
            f"exactly once; got {hours.tolist()}."
        )
    return df_norm


def _extract_profile(
    df_norm: pd.DataFrame, *, scalar_col: str, monthly_prefix: str,
) -> np.ndarray:
    """Pull the (24,) or (24, 12) array from a normalised profile frame."""
    monthly_cols = [f"{monthly_prefix}_{m}" for m in _MONTH_TOKENS]
    if all(col in df_norm.columns for col in monthly_cols):
        arr = np.zeros((24, 12), dtype=float)
        for m_idx, m_name in enumerate(_MONTH_TOKENS):
            arr[:, m_idx] = (
                df_norm[f"{monthly_prefix}_{m_name}"]
                .astype(float).to_numpy()
            )
        return arr
    if scalar_col in df_norm.columns:
        return df_norm[scalar_col].astype(float).to_numpy()
    raise ValueError(
        f"profile sheet must contain either a '{scalar_col}' column "
        f"(24x1) or all 12 '{monthly_prefix}_<month>' columns (24x12)."
    )


def _parse_max_injection_profile_sheet(
    df: pd.DataFrame, *, sheet_name: str = "max_injection_profile",
) -> np.ndarray:
    """Parse a max-injection profile sheet (combined or per-source).

    Returns a (24,) or (24, 12) array of percent-of-grid-export values.
    The optional per-source sheets (``max_injection_profile_pv`` /
    ``max_injection_profile_bess``) share the identical column schema.
    """
    df_norm = _normalise_hourly_profile_frame(df, sheet_name=sheet_name)
    return _extract_profile(
        df_norm,
        scalar_col="max_injection_pct",
        monthly_prefix="max_injection_pct",
    )


def _read_optional_injection_profile(
    xlsx_path: Path, sheets: set[str], sheet_name: str,
) -> np.ndarray | None:
    """Read an optional per-source max-injection sheet, or None if absent.

    Per-source sub-caps are opt-in: a workbook that omits the sheet keeps
    the single combined cap and behaves exactly as before.
    """
    if sheet_name not in sheets:
        return None
    try:
        return _parse_max_injection_profile_sheet(
            pd.read_excel(xlsx_path, sheet_name=sheet_name),
            sheet_name=sheet_name,
        )
    except ValueError as exc:
        raise ValueError(f"{sheet_name}: {exc}") from exc


# ---------------------------------------------------------------------------
# Timeseries normalisation + dt auto-detection
# ---------------------------------------------------------------------------


def _normalise_timeseries(ts: pd.DataFrame, *, mode: str) -> pd.DataFrame:
    """Validate timeseries columns and forward-fill numeric NaNs."""
    if "timestamp" not in ts.columns:
        raise ValueError("timeseries sheet must contain a 'timestamp' column.")
    if "pv_kwh" not in ts.columns:
        raise ValueError("timeseries sheet must contain a 'pv_kwh' column.")

    if mode == "self_consumption" and "load_kwh" not in ts.columns:
        raise ValueError(
            "timeseries sheet must contain a 'load_kwh' column when mode='self_consumption'."
        )
    if mode == "merchant" and "load_kwh" in ts.columns:
        logger.info("merchant mode: load_kwh column ignored")

    for col in ("load_kwh", "pv_kwh", "dam_price_eur_per_mwh",
                "retail_price_eur_per_mwh",
                "imbalance_price_eur_per_mwh",
                "imbalance_price_short_eur_per_mwh",
                "imbalance_price_long_eur_per_mwh"):
        if col in ts.columns:
            numeric = ts[col].astype(float)
            nan_mask = numeric.isna()
            nan_count = int(nan_mask.sum())
            if nan_count > 0:
                first_nan = ts.loc[nan_mask.idxmax(), "timestamp"]
                logger.warning(
                    "Column '%s' had %d NaN value(s) filled via ffill/bfill; "
                    "first NaN at %s. Check the input timeseries for gaps.",
                    col, nan_count, first_nan,
                )
            ts[col] = numeric.ffill().bfill()
    # Deck variant columns (``<base>__<deck>``): validate the base name
    # against the recognised price columns and apply the same NaN
    # ffill/bfill treatment their canonical column receives.
    for col in list(ts.columns):
        name = str(col)
        if "__" not in name:
            continue
        base, deck = name.split("__", 1)
        if base not in PRICE_DECK_BASE_COLUMNS:
            raise ValueError(
                f"timeseries column {name!r}: unknown price-deck base "
                f"{base!r} left of '__' (the double underscore is "
                f"reserved for deck variant columns); recognised price "
                f"columns: {', '.join(PRICE_DECK_BASE_COLUMNS)}."
            )
        if not deck or not re.fullmatch(r"[a-z0-9_]+", deck):
            raise ValueError(
                f"timeseries column {name!r}: deck name {deck!r} must "
                f"match [a-z0-9_]+."
            )
        numeric = ts[name].astype(float)
        nan_mask = numeric.isna()
        nan_count = int(nan_mask.sum())
        if nan_count > 0:
            first_nan = ts.loc[nan_mask.idxmax(), "timestamp"]
            logger.warning(
                "Column '%s' had %d NaN value(s) filled via ffill/bfill; "
                "first NaN at %s. Check the input timeseries for gaps.",
                name, nan_count, first_nan,
            )
        ts[name] = numeric.ffill().bfill()
    # pv_kwh_override deliberately stays out of the ffill/bfill loop so
    # partial NaN survives for the resolver's deprecated-fallback check
    # (pvbess_opt.io_read._apply_override_fallback) to raise on.
    return ts


def detect_timestep_minutes(ts: pd.DataFrame) -> int:
    """Auto-detect the MILP timestep (in minutes) from the timeseries."""
    idx = pd.to_datetime(ts["timestamp"]).sort_values()
    diffs = idx.diff().dropna()
    if diffs.empty:
        raise ValueError(
            "timeseries has fewer than 2 rows; cannot determine timestep."
        )
    if diffs.nunique() > 1:
        sample = diffs.value_counts().head().to_dict()
        raise ValueError(
            "Irregular timestep detected in 'timeseries' "
            f"(distinct step sizes: {sample}). Run "
            "`python scripts/resample_timeseries.py <workbook>` to harmonise "
            "the resolution before optimising."
        )
    delta = diffs.iloc[0]
    return int(delta.total_seconds() / 60)


# ---------------------------------------------------------------------------
# Public loader API
# ---------------------------------------------------------------------------


_V08_REQUIRED_SHEETS: frozenset[str] = frozenset({
    "timeseries", "project", "pv", "bess", "economics", "simulation",
})


# ---------------------------------------------------------------------------
# Balancing market validation + timeseries fallback
# ---------------------------------------------------------------------------

# Keys whose value is read as the share of bess_power_kw allocated to a
# specific market product. The DAM line is included so the sum across
# every consumer of the BESS power budget is bounded by 100 %.
# Note: only the *sum* is enforced. The individual ``dam_capacity_share_pct``
# value is declarative — DAM dispatch is bounded indirectly by
# ``BM_POWER_UP`` / ``BM_POWER_DN`` consuming the residual of bess_power_kw
# left over after the balancing reservations in each step.
_BALANCING_SHARE_KEYS: tuple[str, ...] = (
    "dam_capacity_share_pct",
    "fcr_capacity_share_pct",
    "afrr_up_capacity_share_pct",
    "afrr_dn_capacity_share_pct",
    "mfrr_up_capacity_share_pct",
    "mfrr_dn_capacity_share_pct",
)

_BALANCING_PROBABILITY_KEYS: tuple[str, ...] = (
    "fcr_bid_acceptance_pct",
    "afrr_up_bid_acceptance_pct",
    "afrr_dn_bid_acceptance_pct",
    "mfrr_up_bid_acceptance_pct",
    "mfrr_dn_bid_acceptance_pct",
    "fcr_activation_probability_pct",
    "afrr_up_activation_probability_pct",
    "afrr_dn_activation_probability_pct",
    "mfrr_up_activation_probability_pct",
    "mfrr_dn_activation_probability_pct",
)

_BALANCING_PRICE_KEYS: tuple[str, ...] = (
    "fcr_default_capacity_price_eur_per_mwh",
    "afrr_up_default_capacity_price_eur_per_mwh",
    "afrr_dn_default_capacity_price_eur_per_mwh",
    "mfrr_up_default_capacity_price_eur_per_mwh",
    "mfrr_dn_default_capacity_price_eur_per_mwh",
    "afrr_up_default_activation_price_eur_per_mwh",
    "afrr_dn_default_activation_price_eur_per_mwh",
    "mfrr_up_default_activation_price_eur_per_mwh",
    "mfrr_dn_default_activation_price_eur_per_mwh",
)

# Mapping of timeseries column name to the scalar fallback key when the
# column is missing. FCR has no activation price by design.
_BALANCING_TS_COLUMN_DEFAULTS: dict[str, str] = {
    "fcr_capacity_price_eur_per_mwh": "fcr_default_capacity_price_eur_per_mwh",
    "afrr_up_capacity_price_eur_per_mwh":
        "afrr_up_default_capacity_price_eur_per_mwh",
    "afrr_dn_capacity_price_eur_per_mwh":
        "afrr_dn_default_capacity_price_eur_per_mwh",
    "mfrr_up_capacity_price_eur_per_mwh":
        "mfrr_up_default_capacity_price_eur_per_mwh",
    "mfrr_dn_capacity_price_eur_per_mwh":
        "mfrr_dn_default_capacity_price_eur_per_mwh",
    "afrr_up_activation_price_eur_per_mwh":
        "afrr_up_default_activation_price_eur_per_mwh",
    "afrr_dn_activation_price_eur_per_mwh":
        "afrr_dn_default_activation_price_eur_per_mwh",
    "mfrr_up_activation_price_eur_per_mwh":
        "mfrr_up_default_activation_price_eur_per_mwh",
    "mfrr_dn_activation_price_eur_per_mwh":
        "mfrr_dn_default_activation_price_eur_per_mwh",
}


# Price columns that may carry per-deck variant columns in the base
# timeseries (``<base>__<deck>``, the double underscore is reserved).
# A variant column is inert pass-through in a normal run — the MILP
# reads canonical names only — and becomes active when a scenario
# selects the deck via the ``price_deck`` special
# (pvbess_opt.scenarios._apply_price_deck), which copies the named
# variants onto the canonical columns and strips every ``__`` column
# before the per-scenario workbook is written.
PRICE_DECK_BASE_COLUMNS: tuple[str, ...] = (
    "dam_price_eur_per_mwh",
    "retail_price_eur_per_mwh",
    *_BALANCING_TS_COLUMN_DEFAULTS.keys(),
    # Optional imbalance settlement price columns (Eqs. U7/U8) — deck
    # variants allowed like every other price column.
    "imbalance_price_eur_per_mwh",
    "imbalance_price_short_eur_per_mwh",
    "imbalance_price_long_eur_per_mwh",
)


def _validate_balancing_config(
    balancing: dict[str, Any], dt_minutes: int,
) -> None:
    """Validate the balancing-market config against the rules in the design note.

    Skipped silently when ``balancing_enabled`` is False so workbooks that
    carry the sheet for documentation but disable the feature still load.
    """
    if not bool(balancing.get("balancing_enabled", False)):
        return

    # Pre-default workbook checks: catch obviously bad input (negative
    # shares) before we materialise the BalancingConfig.
    for key in _BALANCING_SHARE_KEYS:
        if key not in balancing:
            continue
        value = float(balancing.get(key, 0.0) or 0.0)
        if value < 0.0:
            raise ValueError(
                f"balancing sheet key {key!r} must be non-negative; "
                f"got {value!r}."
            )

    # Authoritative share-cap check on the RESOLVED BalancingConfig.
    # Earlier we summed only the keys the workbook supplied, treating
    # every missing key as 0 %.  The dataclass default for
    # ``dam_capacity_share_pct`` is 70 %, so a workbook that omitted
    # the DAM line and listed product shares summing to 60 % would
    # pass the raw sum check (60 % <= 100 %) and then materialise a
    # 130 % allocation once the default landed.  Resolve first, then
    # sum across every consumer of the BESS power budget so the cap
    # check sees the same numbers the optimiser will.
    resolved = resolve_balancing_config(balancing)
    cap_eps = 0.5  # percent — tolerates rounding from external workbooks.
    share_sum = sum(
        float(getattr(resolved, key)) for key in _BALANCING_SHARE_KEYS
    )
    if share_sum > 100.0 + cap_eps:
        raise ValueError(
            "balancing sheet capacity shares sum to "
            f"{share_sum:.3f} % which exceeds 100 % of bess_power_kw "
            f"(keys: {list(_BALANCING_SHARE_KEYS)}; resolved values "
            "after dataclass defaults). Reduce one or more shares "
            "so the total stays at or below 100 %."
        )

    for key in _BALANCING_PROBABILITY_KEYS:
        value = float(balancing.get(key, 0.0) or 0.0)
        if value < 0.0 or value > 100.0:
            raise ValueError(
                f"balancing sheet key {key!r} must be a probability in "
                f"[0, 100]; got {value!r}."
            )

    for key in _BALANCING_PRICE_KEYS:
        value = float(balancing.get(key, 0.0) or 0.0)
        if value < 0.0:
            raise ValueError(
                f"balancing sheet key {key!r} must be non-negative; "
                f"got {value!r}."
            )

    duration = float(balancing.get("fcr_required_duration_hours", 0.0) or 0.0)
    if duration <= 0.0:
        raise ValueError(
            "balancing sheet key 'fcr_required_duration_hours' must be "
            f"strictly positive; got {duration!r}."
        )

    settlement = int(balancing.get("bm_settlement_minutes", 0) or 0)
    if settlement != int(dt_minutes):
        raise ValueError(
            "balancing sheet key 'bm_settlement_minutes' is "
            f"{settlement} min but the timeseries cadence is "
            f"{int(dt_minutes)} min. Resample the timeseries (see "
            "scripts/resample_timeseries.py) or set "
            f"bm_settlement_minutes = {int(dt_minutes)} to match."
        )

    headroom = float(balancing.get("bm_soc_headroom_pct", 0.0) or 0.0)
    if headroom < 0.0 or headroom > 50.0:
        raise ValueError(
            "balancing sheet key 'bm_soc_headroom_pct' must be in "
            f"[0, 50]; got {headroom!r}."
        )

    mc_scenarios = int(balancing.get("bm_mc_scenarios", 200) or 0)
    if mc_scenarios < 1:
        raise ValueError(
            "balancing sheet key 'bm_mc_scenarios' must be a positive "
            f"integer; got {mc_scenarios!r}."
        )


def _validate_ppa_config(ppa: dict[str, Any]) -> None:
    """Validate the ``ppa`` sheet (skipped when the contract is disabled).

    The structure enum accepts ``baseload`` at parse time so old
    workbooks load, but an ENABLED baseload contract is rejected here
    with guidance — the shaped profile is designed (docs/ppa_design.md)
    and deliberately not implemented yet.
    """
    if not bool(ppa.get("ppa_enabled", False)):
        return

    structure = str(ppa.get("ppa_structure", "pay_as_produced") or "").lower()
    if structure == "baseload":
        raise ValueError(
            "ppa_structure='baseload' is designed but not implemented "
            "(it needs shortfall-pricing rules — see docs/ppa_design.md). "
            "Use ppa_structure='pay_as_produced' or disable the contract."
        )
    if structure not in ("pay_as_produced",):
        raise ValueError(
            f"ppa_structure must be 'pay_as_produced'; got {structure!r}."
        )

    # Settlement must be a known enum: the KPI engine and the cashflow each
    # branch on it (``kpis.add_economic_columns`` on ``== 'physical'``,
    # ``economics.build_yearly_cashflow`` on ``== 'cfd'``), so an
    # unrecognised value would decompose the PPA revenue inconsistently
    # between the two — physical in one, CfD in the other.  Reject it here
    # the same way ``ppa_structure`` is validated (the loader also raises
    # via ``_parse_string_enum``; this guards hand-built typed dicts).
    settlement = str(ppa.get("ppa_settlement", "physical") or "physical").lower()
    if settlement not in ("physical", "cfd"):
        raise ValueError(
            f"ppa_settlement must be 'physical' or 'cfd'; got {settlement!r}."
        )

    # The negative-price clause gates dispatch (optimization.py Eq. P8)
    # AND settlement (kpis.py Eq. P7); an unknown value would silently
    # behave as 'none' in a hand-built dict, so reject it loudly here
    # exactly like the settlement enum above.
    negative_rule = str(
        ppa.get("ppa_negative_price_rule", "none") or "none"
    ).lower()
    if negative_rule not in ("none", "suspend"):
        raise ValueError(
            f"ppa_negative_price_rule must be 'none' or 'suspend'; got "
            f"{negative_rule!r}."
        )

    share = float(ppa.get("ppa_volume_share_pct", 0.0) or 0.0)
    if not (0.0 <= share <= 100.0):
        raise ValueError(
            f"'ppa_volume_share_pct' must be in [0, 100]; got {share!r}."
        )

    price = float(ppa.get("ppa_price_eur_per_mwh", 0.0) or 0.0)
    if price < 0.0:
        raise ValueError(
            f"'ppa_price_eur_per_mwh' must be non-negative; got {price!r}."
        )

    term = int(ppa.get("ppa_term_years", 0) or 0)
    if term < 1:
        raise ValueError(
            "'ppa_term_years' must be >= 1 when the contract is enabled; "
            f"got {term!r}."
        )


_DT_MINUTES_VALID_DIVISORS: tuple[int, ...] = (
    1, 2, 3, 4, 5, 6, 10, 12, 15, 20, 30, 60,
)


def _coerce_optional_float(value: Any) -> float | None:
    """Return ``value`` as a float, or None when blank / non-numeric."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        try:
            return float(text)
        except ValueError:
            return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if np.isnan(result):
        return None
    return result


def validate_pv_location_fields(pv: dict[str, Any]) -> None:
    """Range-check the PVGIS location / geometry fields on the ``pv`` sheet.

    Raises ``ValueError`` naming the offending key.  Blank / absent fields
    pass; whether a field is *required* is decided separately by
    :func:`pvbess_opt.io_read.resolve_pv_source`.
    """
    lat = _coerce_optional_float(pv.get("latitude"))
    if lat is not None and not (-90.0 <= lat <= 90.0):
        raise ValueError(f"'latitude' must be in [-90, 90]; got {lat!r}.")
    lon = _coerce_optional_float(pv.get("longitude"))
    if lon is not None and not (-180.0 <= lon <= 180.0):
        raise ValueError(f"'longitude' must be in [-180, 180]; got {lon!r}.")

    tilt = pv.get("tilt")
    if isinstance(tilt, str) and tilt.strip().lower() == "optimal":
        pass
    elif tilt is not None and str(tilt).strip() != "":
        tilt_f = _coerce_optional_float(tilt)
        if tilt_f is None:
            raise ValueError(
                f"'tilt' must be a number in [0, 90] or 'optimal'; got {tilt!r}."
            )
        if not (0.0 <= tilt_f <= 90.0):
            raise ValueError(
                f"'tilt' must be in [0, 90] or 'optimal'; got {tilt!r}."
            )

    azimuth = _coerce_optional_float(pv.get("azimuth"))
    if azimuth is not None and not (-180.0 <= azimuth <= 360.0):
        raise ValueError(f"'azimuth' must be in [-180, 360]; got {azimuth!r}.")

    losses = _coerce_optional_float(pv.get("losses_pct"))
    if losses is not None and not (0.0 <= losses <= 100.0):
        raise ValueError(f"'losses_pct' must be in [0, 100]; got {losses!r}.")

    weather_year = pv.get("weather_year")
    if isinstance(weather_year, str) and weather_year.strip().lower() == "tmy":
        pass
    elif weather_year is not None and str(weather_year).strip() != "":
        year_f = _coerce_optional_float(weather_year)
        if year_f is None:
            raise ValueError(
                f"'weather_year' must be a calendar year or 'tmy'; got "
                f"{weather_year!r}."
            )
        if not (1980.0 <= year_f <= 2100.0):
            raise ValueError(
                f"'weather_year' must be a plausible year (1980-2100) or "
                f"'tmy'; got {weather_year!r}."
            )


def _phase_windows_overlap(
    a: tuple[int, int], b: tuple[int, int], n_years: int,
) -> bool:
    """True when two E25 phase windows share at least one project year.

    ``year_to == 0`` resolves to ``n_years`` (end of life) before the
    inclusive-interval intersection test.
    """
    a_to = n_years if a[1] == 0 else a[1]
    b_to = n_years if b[1] == 0 else b[1]
    return a[0] <= b_to and b[0] <= a_to


# Contract stacking-warning matrix: every warning about STACKING two
# contracted-revenue structures (or mis-stacking one against the
# merchant streams) lives in this one reviewable table —
# (rule id, fire predicate over the parsed contract context, message,
# lazy %-args).  Predicates are pure functions of the context dict
# assembled in validate_workbook_params; windows compare via
# _phase_windows_overlap so phase-disjoint configurations stay silent.
# The messages are the exact strings the per-feature warnings used
# (caplog-locked by the feature test files).  Hard validation errors
# (ranges, enums, window rules) deliberately stay inline — this table
# is warnings-only (stacking is flagged, never blocked).
_CONTRACT_STACKING_RULES: tuple[tuple[str, Any, str, Any], ...] = (
    ("toll_no_op",
     lambda c: c["toll_rate"] > 0.0 and c["bess_kw"] <= 0.0,
     "bess_toll_eur_per_mw_year is set (%.4g EUR/MW/yr) but "
     "bess_power_kw is 0: the toll stream is a no-op.",
     lambda c: (c["toll_rate"],)),
    ("toll_retained",
     lambda c: c["toll_rate"] > 0.0 and c["toll_treatment"] == "retained",
     "bess_toll_merchant_treatment = 'retained': the toll "
     "stacks on top of retained merchant streams, "
     "double-monetising the same MW. Use only for "
     "capacity-overlay contracts.",
     lambda _c: ()),
    ("toll_x_optimizer_share",
     lambda c: (
         c["toll_rate"] > 0.0 and c["opt_share_pct"] > 0.0
         and _phase_windows_overlap(
             c["toll_window"], c["opt_term_window"], c["n_years"],
         )
     ),
     "The toll window overlaps an active optimizer revenue "
     "share (optimizer_revenue_share_pct = %.4g %%): under "
     "'zeroed' treatment the optimizer share is gated to zero "
     "in toll years; the two structures double-charge the "
     "same wholesale stream otherwise.",
     lambda c: (c["opt_share_pct"],)),
    ("toll_x_optimizer_floor",
     lambda c: (
         c["floor_enabled"] and c["toll_rate"] > 0.0
         and c["toll_treatment"] == "zeroed"
         and _phase_windows_overlap(
             c["toll_window"], c["opt_term_window"], c["n_years"],
         )
     ),
     "optimizer_floor_enabled while a 'zeroed' BESS toll "
     "window overlaps the optimizer term: the toll zeroes "
     "the margin in overlap years, forcing a full floor "
     "top-up every year — double-charging the "
     "counterparties.",
     lambda _c: ()),
    ("toll_x_state_support",
     lambda c: (
         c["ss_rate"] > 0.0 and c["toll_rate"] > 0.0
         and c["toll_treatment"] == "zeroed"
         and _phase_windows_overlap(
             c["toll_window"], c["ss_window"], c["n_years"],
         )
     ),
     "A state-support window overlaps a 'zeroed' BESS toll "
     "window: realised market revenue is zero in toll "
     "years, so the two-way netting tops the support up to "
     "the threshold every overlap year — cumulating two "
     "capacity payments for the same MW.",
     lambda _c: ()),
    ("capacity_x_state_support",
     lambda c: (
         c["cm_rate"] > 0.0 and c["ss_rate"] > 0.0
         and _phase_windows_overlap(
             c["cm_window"], c["ss_window"], c["n_years"],
         )
     ),
     "A capacity-market window overlaps a state-support "
     "window: support-cumulation rules typically "
     "restrict stacking the two schemes on the same MW "
     "(the capacity revenue does count toward the "
     "support's netting base, Eq. E31a).",
     lambda _c: ()),
    ("capacity_x_toll",
     lambda c: (
         c["cm_rate"] > 0.0 and c["toll_rate"] > 0.0
         and c["toll_treatment"] == "zeroed"
         and _phase_windows_overlap(
             c["cm_window"], c["toll_window"], c["n_years"],
         )
     ),
     "A capacity-market window overlaps a 'zeroed' BESS "
     "toll window: the toller usually holds the "
     "capacity obligation too — check which party "
     "collects the capacity payment.",
     lambda _c: ()),
    # Reserved: a support-scheme (sliding FiP / two-way CfD) x
    # state_support cumulation rule lands here once those keys ship;
    # keep it adjacent to capacity_x_state_support (same cumulation
    # rationale).
)


def validate_workbook_params(
    typed: dict[str, Any], *, dt_minutes: int | None = None,
) -> None:
    """Reject out-of-range and physically impossible workbook values.

    Operates on the typed nested dict returned by :func:`read_workbook`
    (sections ``project``, ``pv``, ``bess``, ``economics``,
    ``simulation``, optional ``balancing``).  Raises ``ValueError`` with
    a single message naming the offending key and value on the first
    failure encountered.

    Callers building ``typed`` programmatically (e.g. unit tests that
    skip the workbook loader) should still route through here before
    constructing the flat ``params`` dict — every downstream consumer
    relies on these invariants.
    """
    project = typed.get("project") or {}
    pv = typed.get("pv") or {}
    bess = typed.get("bess") or {}
    economics = typed.get("economics") or {}
    balancing = typed.get("balancing") or {}
    ppa = typed.get("ppa") or {}

    validate_pv_location_fields(pv)
    _validate_ppa_config(ppa)

    def _require_non_negative(section: dict[str, Any], key: str) -> None:
        if key not in section:
            return
        raw = section[key]
        try:
            value = float(raw)
        except (TypeError, ValueError):
            return
        if value < 0.0:
            raise ValueError(
                f"{key!r} must be non-negative; got {value!r}."
            )

    # Cost / sizing keys are validated against the section they actually
    # live on in ``_SHEET_DEFAULTS`` (the PV/BESS cost block sits on the
    # ``pv`` / ``bess`` sheets, not ``economics``; the site lump sums on
    # ``project``).  Reading them from the wrong section silently no-ops
    # the check and lets a negative CAPEX flip the Year-0 outflow into an
    # inflow.
    for key in (
        "pv_nameplate_kwp",
        "capex_pv_eur_per_kw",
        "devex_pv_eur_per_kw",
        "opex_pv_eur_per_kwp",
        # Degradation percentages must be non-negative: a negative fade
        # makes the SOH / production factor climb above 100 % over the
        # project life (verified: build_degradation_report with a negative
        # annual fade returns SOH 100 -> 105 -> 110.25 %).
        "pv_degradation_year1_pct",
        "pv_degradation_annual_pct",
    ):
        _require_non_negative(pv, key)
    for key in (
        "bess_power_kw",
        "bess_capacity_kwh",
        "max_cycles_per_day",
        "capex_bess_eur_per_kwh",
        "devex_bess_eur_per_kw",
        "opex_bess_eur_per_kw",
        "bess_replacement_cost_pct",
        "bess_degradation_annual_pct",
        "bess_degradation_pct_per_cycle",
        "bess_replacement_year",
    ):
        _require_non_negative(bess, key)
    for key in (
        "site_capex_eur",
        "site_devex_eur",
        # A negative wedge would be a charging subsidy paid by the
        # network — not a fee; reject it like every other cost key.
        "grid_charging_fee_eur_per_mwh",
    ):
        _require_non_negative(project, key)

    gearing = float(economics.get("gearing_pct", 0.0) or 0.0)
    if not (0.0 <= gearing <= 100.0):
        raise ValueError(
            f"'gearing_pct' must be in [0, 100]; got {gearing!r}."
        )

    # Revenue-fee percentages are a fraction of gross revenue, so they live
    # in [0, 100].  An out-of-range value (e.g. 150 meaning a mistyped
    # "1.50 %") is rejected rather than silently clamped — the same loud
    # contract as gearing_pct.  The energy-aggregator fee, the optional
    # balancing-aggregator (BSP) fee, and the optimizer revenue share all
    # share this range.
    for fee_key in ("aggregator_fee_pct_revenue",
                    "balancing_aggregator_fee_pct_revenue",
                    "optimizer_revenue_share_pct",
                    "state_support_clawback_share_pct",
                    "revenue_levy_pct"):
        if fee_key not in economics:
            continue
        fee_val = float(economics.get(fee_key, 0.0) or 0.0)
        if not (0.0 <= fee_val <= 100.0):
            raise ValueError(
                f"{fee_key!r} must be in [0, 100]; got {fee_val!r}."
            )

    # The per-MWh route-to-market fee is a non-negative charge on exported
    # energy (a negative value would be a rebate, never a fee).
    _require_non_negative(economics, "route_to_market_fee_eur_per_mwh")

    # The legacy percentage-of-gross fee and the optimizer share both charge
    # the battery's wholesale stream: stacking them double-charges it.  Warn
    # (not raise) — a user may deliberately combine a small residual
    # representation percentage with a trading share, but should do so
    # knowingly.
    _agg_pct = float(economics.get("aggregator_fee_pct_revenue", 0.0) or 0.0)
    _opt_pct = float(
        economics.get("optimizer_revenue_share_pct", 0.0) or 0.0
    )
    if _agg_pct > 0.0 and _opt_pct > 0.0:
        logger.warning(
            "Both aggregator_fee_pct_revenue (%.4g %%) and "
            "optimizer_revenue_share_pct (%.4g %%) are set: the battery's "
            "wholesale (DAM) stream is charged by BOTH fees. Prefer one "
            "structure, or set the percentages knowingly.",
            _agg_pct, _opt_pct,
        )

    # Contract phase windows (Eq. E25): year_from >= 1 (integer);
    # year_to == 0 means end-of-life, otherwise year_to >= year_from.
    def _validate_phase_window(prefix: str) -> tuple[int, int]:
        # A blank cell (None) falls back to the default; an explicit 0
        # must NOT be swallowed by a falsy-`or`, it is rejected below.
        raw_from = economics.get(f"{prefix}_year_from", 1)
        y_from = int(1 if raw_from is None else raw_from)
        raw_to = economics.get(f"{prefix}_year_to", 0)
        y_to = int(0 if raw_to is None else raw_to)
        if y_from < 1:
            raise ValueError(
                f"'{prefix}_year_from' must be >= 1; got {y_from!r}."
            )
        if y_to != 0 and y_to < y_from:
            raise ValueError(
                f"'{prefix}_year_to' must be 0 (end of life) or >= "
                f"'{prefix}_year_from' ({y_from!r}); got {y_to!r}."
            )
        return y_from, y_to

    # BESS tolling agreement (Eqs. E29/E29a): non-negative rate and
    # indexation over a valid phase window.
    _require_non_negative(economics, "bess_toll_eur_per_mw_year")
    _require_non_negative(economics, "bess_toll_indexation_pct")
    _toll_window = _validate_phase_window("bess_toll")
    _toll_treatment = str(
        economics.get("bess_toll_merchant_treatment", "zeroed") or "zeroed"
    ).strip().lower()
    if _toll_treatment not in ("zeroed", "retained"):
        raise ValueError(
            "'bess_toll_merchant_treatment' must be 'zeroed' or "
            f"'retained'; got {_toll_treatment!r}."
        )
    _toll_rate = float(
        economics.get("bess_toll_eur_per_mw_year", 0.0) or 0.0
    )

    # Optimizer floor + share-above-floor (Eqs. E30/E30a): non-negative
    # floor over a valid term window; the margin basis is enum-checked.
    _require_non_negative(economics, "optimizer_floor_eur_per_kw_year")
    _opt_term_window = _validate_phase_window("optimizer_term")
    _margin_basis = str(
        economics.get("optimizer_margin_basis", "dam") or "dam"
    ).strip().lower()
    if _margin_basis not in ("dam", "dam_plus_balancing"):
        raise ValueError(
            "'optimizer_margin_basis' must be 'dam' or "
            f"'dam_plus_balancing'; got {_margin_basis!r}."
        )
    _floor_enabled = bool(economics.get("optimizer_floor_enabled", False))

    # State support with two-way clawback (Eqs. E31/E31a): non-negative
    # levels over a valid support window; the netting share shares the
    # fee [0, 100] range check above.
    _require_non_negative(economics, "state_support_eur_per_mw_year")
    _require_non_negative(
        economics, "state_support_clawback_threshold_eur_per_mw_year",
    )
    _require_non_negative(economics, "state_support_indexation_pct")
    _ss_window = _validate_phase_window("state_support")
    _ss_rate = float(
        economics.get("state_support_eur_per_mw_year", 0.0) or 0.0
    )

    # Capacity-market payment (Eq. E32): non-negative rate/indexation,
    # derating in [0, 100], valid contract window.
    _require_non_negative(economics, "capacity_market_eur_per_mw_year")
    _require_non_negative(economics, "capacity_market_indexation_pct")
    _cm_derating = float(
        economics.get("capacity_market_derating_pct", 100.0) or 0.0
    )
    if not (0.0 <= _cm_derating <= 100.0):
        raise ValueError(
            "'capacity_market_derating_pct' must be in [0, 100]; got "
            f"{_cm_derating!r}."
        )
    _cm_window = _validate_phase_window("capacity_market")
    _cm_rate = float(
        economics.get("capacity_market_eur_per_mw_year", 0.0) or 0.0
    )

    # Contract stacking warnings — one data-driven pass over
    # _CONTRACT_STACKING_RULES (never blocking; deliberate contract
    # overlays stay possible).  Every predicate reads this context.
    _contract_ctx: dict[str, Any] = {
        "n_years": int(project.get("project_lifecycle_years", 20) or 20),
        "bess_kw": float(bess.get("bess_power_kw", 0.0) or 0.0),
        "toll_rate": _toll_rate,
        "toll_treatment": _toll_treatment,
        "toll_window": _toll_window,
        "opt_share_pct": _opt_pct,
        "opt_term_window": _opt_term_window,
        "floor_enabled": _floor_enabled,
        "ss_rate": _ss_rate,
        "ss_window": _ss_window,
        "cm_rate": _cm_rate,
        "cm_window": _cm_window,
    }
    for _rule_id, _rule_fires, _rule_msg, _rule_args in (
        _CONTRACT_STACKING_RULES
    ):
        if _rule_fires(_contract_ctx):
            logger.warning(_rule_msg, *_rule_args(_contract_ctx))

    # Grid-emissions accounting (24/7-CFE): a non-negative intensity and a
    # decline in [0, 100] %/yr.  A decline above 100 % makes the
    # ``(1 - decline)^y`` projection factor negative, flipping the grid
    # carbon intensity (and the avoided-emissions sign) from year 2 on.
    _require_non_negative(economics, "grid_co2_intensity_kg_per_mwh")
    co2_decline = float(economics.get("grid_co2_annual_decline_pct", 0.0) or 0.0)
    if not (0.0 <= co2_decline <= 100.0):
        raise ValueError(
            "'grid_co2_annual_decline_pct' must be in [0, 100]; "
            f"got {co2_decline!r}."
        )

    soc_min = float(bess.get("soc_min_frac", 0.0) or 0.0)
    soc_max = float(bess.get("soc_max_frac", 1.0) or 1.0)
    initial_soc = float(bess.get("initial_soc_frac", 0.5) or 0.5)
    if not (0.0 <= soc_min <= 1.0):
        raise ValueError(
            f"'soc_min_frac' must be in [0, 1]; got {soc_min!r}."
        )
    if not (0.0 <= soc_max <= 1.0):
        raise ValueError(
            f"'soc_max_frac' must be in [0, 1]; got {soc_max!r}."
        )
    if soc_min > soc_max:
        raise ValueError(
            f"'soc_min_frac' ({soc_min!r}) must be <= 'soc_max_frac' "
            f"({soc_max!r})."
        )
    if not (soc_min <= initial_soc <= soc_max):
        raise ValueError(
            f"'initial_soc_frac' ({initial_soc!r}) must lie within "
            f"['soc_min_frac', 'soc_max_frac'] = [{soc_min!r}, {soc_max!r}]."
        )

    for key in ("efficiency_charge", "efficiency_discharge"):
        if key not in bess:
            continue
        value = float(bess[key])
        if not (0.0 < value <= 1.0):
            raise ValueError(
                f"{key!r} must be in (0, 1]; got {value!r}."
            )

    if "bess_eol_soh_pct" in bess:
        eol = float(bess["bess_eol_soh_pct"] or 0.0)
        if not (0.0 < eol <= 100.0):
            raise ValueError(
                f"'bess_eol_soh_pct' must be in (0, 100]; got {eol!r}."
            )

    if dt_minutes is not None:
        if dt_minutes <= 0:
            raise ValueError(
                f"'dt_minutes' must be a positive integer; got {dt_minutes!r}."
            )
        if dt_minutes not in _DT_MINUTES_VALID_DIVISORS:
            raise ValueError(
                f"'dt_minutes' = {dt_minutes!r} does not evenly divide 60; "
                f"valid cadences are {list(_DT_MINUTES_VALID_DIVISORS)!r}."
            )

    if bool(balancing.get("balancing_enabled", False)):
        # Activation / acceptance probabilities are validated by
        # :func:`_validate_balancing_config` against the [0, 100] range
        # (percent-scale).  Re-running it here keeps a single source of
        # truth.
        if dt_minutes is not None:
            _validate_balancing_config(balancing, dt_minutes)

    # Imbalance settlement (Eqs. U6-U9): the nomination is the noisy
    # lookahead slice of each rolling-horizon window, so the feature is
    # meaningless without the rolling-horizon Monte Carlo and needs a
    # lookahead at least one commit block long.
    simulation_cfg = typed.get("simulation") or {}
    if bool(simulation_cfg.get("imbalance_enabled", False)):
        if not bool(simulation_cfg.get("uncertainty_enabled", False)):
            raise ValueError(
                "imbalance_enabled requires uncertainty_enabled = TRUE: "
                "the day-ahead nomination is each rolling-horizon "
                "window's noisy lookahead slice (Eq. U6)."
            )
        window_h = int(
            simulation_cfg.get("uncertainty_window_hours", 48) or 48
        )
        commit_h = int(
            simulation_cfg.get("uncertainty_commit_hours", 24) or 24
        )
        if window_h < 2 * commit_h:
            raise ValueError(
                f"imbalance_enabled requires uncertainty_window_hours "
                f"(= {window_h}) >= 2 x uncertainty_commit_hours "
                f"(= {commit_h}): the nomination for the next commit "
                f"block is the lookahead slice [commit, 2 x commit) of "
                f"each window."
            )
        m_short = float(
            simulation_cfg.get("imbalance_price_mult_short", 1.25) or 0.0
        )
        m_long = float(
            simulation_cfg.get("imbalance_price_mult_long", 0.75) or 0.0
        )
        if m_short < 0.0 or m_long < 0.0:
            raise ValueError(
                "imbalance price multipliers must be non-negative; got "
                f"short={m_short!r}, long={m_long!r}."
            )
        if m_short < 1.0 or m_long > 1.0:
            logger.warning(
                "imbalance DAM-proxy multipliers short=%.3g / long=%.3g "
                "are not incentive-compatible (expected short >= 1 and "
                "long <= 1, so that long <= DAM <= short); a deviation "
                "could be PAID more than the DAM. Set them knowingly.",
                m_short, m_long,
            )

    # Per-year trajectory vectors (Eq. E24): structural checks live in the
    # parser / normaliser; here the lifecycle-aware invariants are
    # enforced — full coverage of the project life, the Year-1 anchor
    # (m_1 == 1 keeps the Year-1 cashflow equal to the dispatch-KPI base),
    # the shared-vs-split OPEX conflict, and the replace-mode
    # double-specification warning.
    trajectories = typed.get("trajectories")
    if trajectories:
        n_years = int(project.get("project_lifecycle_years", 0) or 0)
        for stream, spec in trajectories.items():
            values = spec["values"]
            if n_years and len(values) != n_years:
                raise ValueError(
                    f"trajectories stream {stream!r} covers "
                    f"{len(values)} year(s) but project_lifecycle_years "
                    f"is {n_years}; provide one multiplier per operating "
                    f"year 1..{n_years}."
                )
            if abs(float(values[0]) - 1.0) > 1e-9:
                raise ValueError(
                    f"trajectories stream {stream!r} must anchor at "
                    f"value 1.0 in year 1 (multipliers are relative to "
                    f"the Year-1 base); got {values[0]!r}."
                )
        if "opex" in trajectories and (
            {"opex_pv", "opex_bess"} & set(trajectories)
        ):
            raise ValueError(
                "trajectories: the shared 'opex' stream cannot be "
                "combined with the per-asset 'opex_pv' / 'opex_bess' "
                "streams; use one or the other."
            )
        for stream, spec in trajectories.items():
            if spec["mode"] != "replace":
                continue
            sheet_name, infl_key = _TRAJECTORY_INFLATION_KEYS[stream]
            section = typed.get(sheet_name) or {}
            infl = float(section.get(infl_key, 0.0) or 0.0)
            if infl != 0.0:
                logger.warning(
                    "trajectories stream %r uses mode 'replace', which "
                    "substitutes the flat %s (%.4g %%): the scalar index "
                    "is IGNORED for this stream. Set %s to 0, or use "
                    "mode 'overlay' to stack the trajectory on top of "
                    "it.",
                    stream, infl_key, infl, infl_key,
                )


def _apply_balancing_timeseries_fallback(
    ts: pd.DataFrame, balancing: dict[str, Any],
) -> pd.DataFrame:
    """Fill in any missing balancing-price column with its scalar default.

    No-op when ``balancing_enabled`` is False (the columns may be absent
    and the loader leaves them alone).  When enabled, each missing
    column is appended with the scalar value from the balancing config
    and a SINGLE WARNING is emitted enumerating every missing column
    and the default applied to it — previously this function logged one
    warning per missing column, which produced a 9-line warning storm
    for a workbook that simply hadn't carried any balancing-price
    timeseries.
    """
    if not bool(balancing.get("balancing_enabled", False)):
        return ts

    out = ts
    n_rows = len(out)
    missing_columns: list[tuple[str, float, str]] = []
    for col, default_key in _BALANCING_TS_COLUMN_DEFAULTS.items():
        if col in out.columns:
            continue
        default_value = float(balancing.get(default_key, 0.0) or 0.0)
        if out is ts:  # avoid the copy until we know we need one
            out = ts.copy()
        out[col] = np.full(n_rows, default_value, dtype=float)
        missing_columns.append((col, default_value, default_key))

    if missing_columns:
        formatted = ", ".join(
            f"{col}={value:.4f} EUR/MWh (from {default_key!r})"
            for col, value, default_key in missing_columns
        )
        logger.warning(
            "Balancing timeseries columns missing; applied per-product "
            "defaults: %s.",
            formatted,
        )
    return out


def _read_kv_flat(xlsx_path: Path, sheet_name: str) -> dict[str, Any]:
    """Read a (key, value, ...) sheet to ``{key: value}`` with faithful types.

    openpyxl reports each cell's stored type exactly as the file holds it,
    whereas ``pd.read_excel`` can mis-surface a genuinely NUMERIC 0/1 cell
    as a Python bool when it infers the mixed-type value column — which
    would make the boolean-in-numeric-field guard in ``_parse_kv_sheet``
    fire on legitimate zeros.  Requires the ``key`` / ``value`` header
    pair, skips blank and ``#`` separator keys.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb[sheet_name]
        rows = ws.iter_rows(max_col=2, values_only=True)
        header = next(rows, None)
        if (
            header is None
            or len(header) < 2
            or str(header[0]).strip() != "key"
            or str(header[1]).strip() != "value"
        ):
            return {}
        out: dict[str, Any] = {}
        for row in rows:
            raw_key = row[0] if row else None
            if not isinstance(raw_key, str):
                continue
            key = raw_key.strip()
            if not key or key.startswith("#"):
                continue
            out[key] = row[1] if len(row) > 1 else None
        return out
    finally:
        wb.close()


def read_workbook(xlsx_path: str | Path) -> dict[str, Any]:
    """Read the input workbook and return the typed nested dict."""
    xlsx_path = Path(xlsx_path)
    sheets = {str(name) for name in pd.ExcelFile(xlsx_path).sheet_names}

    missing = _V08_REQUIRED_SHEETS - sheets
    if missing:
        raise ValueError(
            f"Workbook {xlsx_path!s} is missing required sheets: "
            f"{sorted(missing)}. Found: {sorted(sheets)}."
        )

    typed: dict[str, Any] = {}
    for sheet_name in ("project", "pv", "bess", "economics", "simulation"):
        flat = _read_kv_flat(xlsx_path, sheet_name)
        if sheet_name == "bess":
            reject_legacy_bess_capex_key(
                flat, source=f"Workbook {xlsx_path!s} (bess sheet)",
            )
        typed[sheet_name] = _parse_kv_sheet(sheet_name, flat)
        if (
            sheet_name == "bess"
            and "bess_degradation_pct_per_cycle" not in flat
        ):
            # Workbooks that omit the cycle-fade coefficient default it
            # to 0.0 so the run uses calendar-only fade.
            typed["bess"]["bess_degradation_pct_per_cycle"] = 0.0
            logger.info(
                "[bess] bess_degradation_pct_per_cycle not found in "
                "workbook; defaulting to 0.0 (calendar-only mode)."
            )

    # Optional ``balancing`` sheet — when absent every key falls back to
    # the defaults declared above so ``balancing_enabled`` resolves to
    # False and the rest of the loader behaves exactly as before.
    if "balancing" in sheets:
        balancing_flat = _read_kv_flat(xlsx_path, "balancing")
        typed["balancing"] = _parse_kv_sheet("balancing", balancing_flat)
    else:
        typed["balancing"] = dict(BALANCING_SHEET_DEFAULTS)

    # Optional ``ppa`` sheet — same master-switch pattern: absent means
    # the contract is disabled and the run is bit-identical to before.
    if "ppa" in sheets:
        ppa_flat = _read_kv_flat(xlsx_path, "ppa")
        typed["ppa"] = _parse_kv_sheet("ppa", ppa_flat)
    else:
        typed["ppa"] = dict(PPA_SHEET_DEFAULTS)

    # Optional ``trajectories`` sheet — per-year stream multipliers
    # (Eq. E24).  Absent sheet, ``enabled`` = FALSE, or no data rows all
    # resolve to None and the run is bit-identical to before.
    if "trajectories" in sheets:
        try:
            typed["trajectories"] = _parse_trajectories_sheet(
                pd.read_excel(xlsx_path, sheet_name="trajectories"),
            )
        except ValueError as exc:
            raise ValueError(f"trajectories: {exc}") from exc
    else:
        typed["trajectories"] = None

    # A finite grid-export cap must be strictly positive.  An empty cell
    # or an 'unlimited' token resolves to float('inf') (cap disabled).
    grid_cap = typed["project"]["p_grid_export_max_kw"]
    if not np.isinf(grid_cap) and float(grid_cap) <= 0.0:
        raise ValueError(
            "p_grid_export_max_kw must be a positive number, or empty / "
            "'inf' / 'unlimited' / 'disabled' to remove the cap; got "
            f"{grid_cap!r}."
        )

    if "max_injection_profile" in sheets:
        try:
            profile = _parse_max_injection_profile_sheet(
                pd.read_excel(xlsx_path, sheet_name="max_injection_profile"),
            )
        except ValueError as exc:
            raise ValueError(f"max_injection_profile: {exc}") from exc
    else:
        logger.info(
            "max_injection_profile sheet not found in %s; falling back "
            "to constant %.1f %% for every hour.",
            xlsx_path, DEFAULT_MAX_INJECTION_PCT_HOURLY,
        )
        profile = np.full(
            24, DEFAULT_MAX_INJECTION_PCT_HOURLY, dtype=float,
        )

    profile_pv = _read_optional_injection_profile(
        xlsx_path, sheets, "max_injection_profile_pv",
    )
    profile_bess = _read_optional_injection_profile(
        xlsx_path, sheets, "max_injection_profile_bess",
    )

    mode = str(typed["project"]["mode"]).lower()
    ts = _normalise_timeseries(
        pd.read_excel(xlsx_path, sheet_name="timeseries", parse_dates=["timestamp"]),
        mode=mode,
    )
    # Single PV-source resolution rule (auto | file | pvgis), shared with
    # the structured-config loader.  Resolves ts['pv_kwh'] from the column,
    # an external timeseries_path, or a PVGIS fetch by latitude / longitude.
    from .io_read import resolve_pv_source

    typed["ts"] = ts
    resolve_pv_source(typed, base_dir=xlsx_path.parent)
    ts = typed.pop("ts")
    dt_minutes = detect_timestep_minutes(ts)
    validate_workbook_params(typed, dt_minutes=dt_minutes)
    ts = _apply_balancing_timeseries_fallback(ts, typed["balancing"])
    # Single-price imbalance settlement has no canonical DAM
    # relationship to proxy, so the column is mandatory (the dual
    # regime falls back per side to the U8a DAM proxy instead).
    _sim = typed.get("simulation") or {}
    if (
        bool(_sim.get("imbalance_enabled", False))
        and str(_sim.get("imbalance_pricing", "dual")).lower() == "single"
        and "imbalance_price_eur_per_mwh" not in ts.columns
    ):
        raise ValueError(
            "imbalance_pricing = 'single' requires the "
            "imbalance_price_eur_per_mwh timeseries column (a lone "
            "imbalance price cannot be proxied from the DAM); add the "
            "column or use imbalance_pricing = 'dual'."
        )
    out: dict[str, Any] = {
        "ts": ts,
        "max_injection_profile": profile,
        "max_injection_profile_pv": profile_pv,
        "max_injection_profile_bess": profile_bess,
        "dt_minutes": dt_minutes,
    }
    out.update(typed)
    return out


def _typed_to_flat(
    typed: dict[str, Any],
) -> tuple[dict[str, Any], pd.DataFrame]:
    """Translate the typed dict to the flat ``(params, ts)`` shape."""
    project = typed["project"]
    pv = typed["pv"]
    bess = typed["bess"]
    ts = typed["ts"]

    bess_power_kw = float(bess["bess_power_kw"])
    bess_capacity_kwh = float(bess["bess_capacity_kwh"])
    pv_nameplate_kwp = float(pv["pv_nameplate_kwp"])

    # Resolve the grid-export cap.  When the workbook value is empty or an
    # 'unlimited' token it parses to float('inf'); we substitute a finite
    # Big-M large enough never to bind so the MILP topology is unchanged
    # and the behaviour stays solver-agnostic (HiGHS / Gurobi / CBC).
    raw_grid_cap = float(project["p_grid_export_max_kw"])
    grid_export_unlimited = bool(np.isinf(raw_grid_cap))
    if grid_export_unlimited:
        p_grid_export_cap_milp = max(
            2.0 * (pv_nameplate_kwp + bess_power_kw),
            1.0e6,
        )
        logger.info(
            "[simulation] Grid export cap disabled (unlimited). "
            "Curtailment will be zero. Internal MILP bound: %.0f kW.",
            p_grid_export_cap_milp,
        )
    else:
        p_grid_export_cap_milp = raw_grid_cap

    params: dict[str, Any] = {
        "dt_minutes": int(typed["dt_minutes"]),
        # bess
        "efficiency_charge": float(bess["efficiency_charge"]),
        "efficiency_discharge": float(bess["efficiency_discharge"]),
        "soc_min_frac": float(bess["soc_min_frac"]),
        "soc_max_frac": float(bess["soc_max_frac"]),
        "initial_soc_frac": float(bess["initial_soc_frac"]),
        "terminal_soc_equal": bool(bess["terminal_soc_equal"]),
        "max_cycles_per_day": float(bess["max_cycles_per_day"]),
        "bess_power_kw": bess_power_kw,
        "bess_capacity_kwh": bess_capacity_kwh,
        "bess_wear_cost_eur_per_mwh": float(
            bess.get("bess_wear_cost_eur_per_mwh", 0.0) or 0.0
        ),
        # pv
        "pv_nameplate_kwp": pv_nameplate_kwp,
        # project
        "p_grid_export_max_kw": p_grid_export_cap_milp,
        # Contract fields: not consumed by the internal dispatch but part of
        # the published params schema (asserted by the test suite / available
        # to API consumers), so they are retained intentionally.
        "grid_export_unlimited": grid_export_unlimited,
        "retail_tariff_eur_per_mwh": float(project["retail_tariff_eur_per_mwh"]),
        "mode": str(project["mode"]),
        "allow_bess_grid_charging": bool(project["allow_bess_grid_charging"]),
        # Charging-side wedge (Eq. E26): consumed by the MILP objective
        # (build_model) and the per-step fee column (add_economic_columns).
        "grid_charging_fee_eur_per_mwh": float(
            project.get("grid_charging_fee_eur_per_mwh", 0.0) or 0.0
        ),
        "grid_charging_fee_exempt": bool(
            project.get("grid_charging_fee_exempt", False)
        ),
        "grid_cap_includes_load": bool(project["grid_cap_includes_load"]),
        "unavailability_pct": float(project["unavailability_pct"]),
        "site_capex_eur": float(project.get("site_capex_eur", 0.0) or 0.0),
        "site_devex_eur": float(project.get("site_devex_eur", 0.0) or 0.0),
        "show_titles": bool(project["show_titles"]),
        # Max-injection cap profile (24,) or (24, 12), in percent of
        # p_grid_export_max_kw.  Expanded to a per-step array by the
        # max-injection helper module before entering the MILP.
        "max_injection_profile": typed.get("max_injection_profile"),
        # Optional per-source injection sub-caps (PV-origin and BESS-origin
        # injection), same (24,) / (24, 12) percent shape.  None => no sub-cap
        # for that source (only the combined cap binds, exactly as before).
        "max_injection_profile_pv": typed.get("max_injection_profile_pv"),
        "max_injection_profile_bess": typed.get("max_injection_profile_bess"),
        # Balancing-market section, forwarded as a nested dict so the
        # MILP, KPI, lifetime and Monte Carlo paths can opt in without
        # changing the flat-params contract.
        "balancing": dict(
            typed.get("balancing") or BALANCING_SHEET_DEFAULTS,
        ),
        # PPA contract section — same nested-dict pattern; consumed by
        # the MILP objective and the per-step EUR columns.
        "ppa": dict(typed.get("ppa") or PPA_SHEET_DEFAULTS),
    }
    return params, ts


def read_inputs(xlsx_path: str | Path) -> tuple[dict[str, Any], pd.DataFrame]:
    """Return ``(params, ts)`` — the flat shape used by the optimizer.

    Raises ``ValueError`` when both ``pv_nameplate_kwp`` and
    ``bess_power_kw`` are zero (no asset to optimise).
    """
    typed = read_workbook(xlsx_path)
    params, ts = _typed_to_flat(typed)
    if (
        float(params.get("pv_nameplate_kwp", 0.0) or 0.0) <= 0.0
        and float(params.get("bess_power_kw", 0.0) or 0.0) <= 0.0
    ):
        raise ValueError(
            "Both pv_nameplate_kwp and bess_power_kw are zero — "
            "nothing to optimise."
        )
    return params, ts


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------


_ECON_UNITS: dict[str, str] = {}
for _rows in _SHEET_ROW_TEMPLATES.values():
    for _key, _default, _unit, _notes in _rows:
        _ECON_UNITS.setdefault(_key, _unit)


def _format_assumptions(econ: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for key, value in econ.items():
        rows.append({"key": key, "value": value, "unit": _ECON_UNITS.get(key, "")})
    return pd.DataFrame(rows, columns=["key", "value", "unit"])


def copy_input_snapshot(src_xlsx: Path, out_dir: Path, tag: str) -> Path | None:
    """Copy the input workbook into ``out_dir`` with a tag suffix."""
    src_xlsx = Path(src_xlsx)
    if not src_xlsx.exists():
        return None
    dst = out_dir / f"{src_xlsx.stem}_{tag}{src_xlsx.suffix}"
    dst.write_bytes(src_xlsx.read_bytes())
    return dst


# ---------------------------------------------------------------------------
# 00..05 numbered output layout
# ---------------------------------------------------------------------------

LAYOUT_SUBDIRS: tuple[str, ...] = (
    "00_summary",
    "01_inputs",
    "02_dispatch",
    "04_financial_plots",
    "05_energy_plots",
    "06_uncertainty_plots",
)


def make_run_layout(out_dir: Path) -> dict[str, Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}
    for name in LAYOUT_SUBDIRS:
        sub = out_dir / name
        sub.mkdir(parents=True, exist_ok=True)
        paths[name.split("_", 1)[1]] = sub
    paths["root"] = out_dir
    return paths


def write_assumptions_summary(
    out_path: Path,
    params: dict[str, Any],
    econ: dict[str, Any] | None,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("PV+BESS optimizer - assumptions snapshot")
    lines.append("=" * 60)
    lines.append("")
    lines.append("[params]")
    for key in sorted(params):
        if key.startswith("_"):
            continue
        # Hide the array-valued max-injection profiles from the snapshot —
        # they're already in the workbook's max_injection_profile* sheets.
        if key in (
            "max_injection_profile",
            "max_injection_profile_pv",
            "max_injection_profile_bess",
        ):
            continue
        lines.append(f"  {key} = {params[key]!r}")
    lines.append("")
    lines.append("[economic]")
    if econ:
        for key in sorted(econ):
            lines.append(f"  {key} = {econ[key]!r}")
    out_path.write_text("\n".join(lines) + "\n")
    return out_path


_SUMMARY_YEAR1_KEYS: tuple[tuple[str, str], ...] = (
    ("profit_total_eur", "Year-1 profit [EUR]"),
    ("pv_generation_mwh", "PV generation [MWh]"),
    ("bess_total_discharge_mwh", "BESS discharge [MWh]"),
    ("system_total_export_mwh", "Grid export [MWh]"),
    ("system_total_import_mwh", "Grid import [MWh]"),
    ("pv_energy_curtailed_mwh", "PV curtailed [MWh]"),
    ("bess_equivalent_cycles_per_day", "BESS cycles per day"),
    ("system_pv_self_consumption_frac", "PV self-consumption [-]"),
    ("system_load_green_coverage_frac", "Green load coverage [-]"),
)

_SUMMARY_FINANCIAL_KEYS: tuple[tuple[str, str], ...] = (
    ("npv_eur", "NPV [EUR]"),
    ("irr_pct", "IRR [%]"),
    ("roi_pct", "ROI [%]"),
    ("bcr", "Benefit-cost ratio [-]"),
    ("simple_payback_years", "Simple payback [years]"),
    ("discounted_payback_years", "Discounted payback [years]"),
    ("lcoe_eur_per_mwh", "LCOE [EUR/MWh]"),
    ("lcos_eur_per_mwh", "LCOS [EUR/MWh]"),
    ("initial_investment_eur", "Initial investment, Year 0 [EUR]"),
    ("total_capex_eur", "Total CAPEX incl. replacement [EUR]"),
    ("total_devex_eur", "Total DEVEX [EUR]"),
    ("bess_lifetime_cycles", "BESS lifetime cycles [-]"),
)

# Optional revenue-stream totals: rendered only when the stream is
# non-zero so a run without the feature keeps a noise-free digest.
_SUMMARY_OPTIONAL_FINANCIAL_KEYS: tuple[tuple[str, str], ...] = (
    ("lifetime_ppa_revenue_total_eur", "Lifetime PPA revenue [EUR]"),
    ("lifetime_bm_revenue_total_eur", "Lifetime balancing revenue [EUR]"),
    ("total_route_to_market_fee_eur_lifecycle",
     "Lifetime route-to-market fee [EUR]"),
    ("total_optimizer_fee_eur_lifecycle", "Lifetime optimizer fee [EUR]"),
    ("total_grid_charging_fee_eur_lifecycle",
     "Lifetime grid-charging fee [EUR]"),
    ("total_imbalance_cost_eur_lifecycle",
     "Lifetime imbalance cost [EUR]"),
    ("total_toll_revenue_eur_lifecycle", "Lifetime tolling revenue [EUR]"),
    ("total_optimizer_floor_topup_eur_lifecycle",
     "Lifetime optimizer floor top-up [EUR]"),
    ("total_state_support_eur_lifecycle", "Lifetime state support [EUR]"),
    ("total_state_support_clawback_eur_lifecycle",
     "Lifetime state-support netting [EUR]"),
    ("total_capacity_market_revenue_eur_lifecycle",
     "Lifetime capacity-market revenue [EUR]"),
    ("total_revenue_levy_eur_lifecycle", "Lifetime revenue levy [EUR]"),
)

# Rolling-horizon / benchmark digest: rendered only when the
# rolling-horizon Monte Carlo ran (``mc_n_seeds`` present).  The
# benchmark gaps are the two DISTINCT numbers a publication needs --
# the requested target vs the gap the solver actually proved.
_SUMMARY_ROLLING_KEYS: tuple[tuple[str, str], ...] = (
    ("foresight_gap_pct_p50", "Foresight gap P50 [%]"),
    ("foresight_gap_pct_p10", "Foresight gap P10 [%]"),
    ("foresight_gap_pct_p90", "Foresight gap P90 [%]"),
    ("pf_benchmark_mip_gap", "Benchmark MIP gap, requested [-]"),
    ("pf_benchmark_gap_achieved", "Benchmark MIP gap, proven [-]"),
    ("mc_n_seeds", "Monte Carlo seeds"),
    # Imbalance settlement (Eqs. U6-U9) — rendered only when the
    # feature ran (keys absent otherwise).
    ("imbalance_cost_year1_eur", "Imbalance cost, Year-1 mean [EUR]"),
    ("imbalance_cost_p50_eur", "Imbalance cost P50 [EUR]"),
    ("imbalance_cost_p10_eur", "Imbalance cost P10 [EUR]"),
    ("imbalance_cost_p90_eur", "Imbalance cost P90 [EUR]"),
    ("bess_imbalance_hedge_value_mean_eur",
     "BESS imbalance hedge value, mean [EUR]"),
)


def _summary_fmt(value: Any) -> str:
    """Human format for a SUMMARY.md value cell."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, float) and np.isnan(value):
        return "n/a"
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    return f"{value:,.4g}"


def write_summary_md(
    out_path: Path,
    *,
    kpis_year1: dict[str, Any],
    financial_kpis: dict[str, Any] | None,
    params: dict[str, Any],
    solver_name: str | None = None,
    replacement_note: str | None = None,
) -> Path:
    """Write the ``00_summary/SUMMARY.md`` headline digest.

    A short, human-first markdown digest of the run: mode / asset
    configuration, headline Year-1 KPIs, and the financial KPIs.  Full
    detail lives in ``03_results.xlsx``; this file is the at-a-glance
    entry point the output layout advertises.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    lines: list[str] = ["# Run summary", ""]
    mode = str(params.get("mode", "")) or "self_consumption"
    pv_kwp = float(params.get("pv_nameplate_kwp", 0.0) or 0.0)
    bess_kw = float(params.get("bess_power_kw", 0.0) or 0.0)
    bess_kwh = float(params.get("bess_capacity_kwh", 0.0) or 0.0)
    lines.append(f"- Mode: `{mode}`")
    lines.append(
        f"- Assets: PV {pv_kwp:,.0f} kWp | BESS {bess_kw:,.0f} kW / "
        f"{bess_kwh:,.0f} kWh"
    )
    lines.append(
        "- Grid charging: "
        f"{'on' if params.get('allow_bess_grid_charging') else 'off'}"
    )
    if replacement_note:
        lines.append(f"- BESS replacement: {replacement_note}")
    if solver_name:
        lines.append(f"- Solver: `{solver_name}`")
    lines.append("")

    lines.append("## Year-1 dispatch KPIs")
    lines.append("")
    lines.append("| Metric | Value |")
    lines.append("| --- | ---: |")
    for key, label in _SUMMARY_YEAR1_KEYS:
        if key in kpis_year1:
            lines.append(f"| {label} | {_summary_fmt(kpis_year1[key])} |")
    lines.append("")

    if financial_kpis:
        lines.append("## Financial KPIs (project lifetime)")
        lines.append("")
        lines.append("| Metric | Value |")
        lines.append("| --- | ---: |")
        for key, label in _SUMMARY_FINANCIAL_KEYS:
            if key in financial_kpis:
                lines.append(
                    f"| {label} | {_summary_fmt(financial_kpis[key])} |"
                )
        for key, label in _SUMMARY_OPTIONAL_FINANCIAL_KEYS:
            value = financial_kpis.get(key)
            if isinstance(value, (int, float)) and abs(float(value)) > 1e-9:
                lines.append(f"| {label} | {_summary_fmt(value)} |")
        lines.append("")

    if kpis_year1.get("mc_n_seeds"):
        lines.append("## Rolling-horizon foresight")
        lines.append("")
        lines.append("| Metric | Value |")
        lines.append("| --- | ---: |")
        for key, label in _SUMMARY_ROLLING_KEYS:
            if key in kpis_year1:
                lines.append(
                    f"| {label} | {_summary_fmt(kpis_year1[key])} |"
                )
        lines.append("")

    lines.append("## Where to look next")
    lines.append("")
    lines.append("- `03_results.xlsx`: KPIs, cashflows, sensitivity, lifetime")
    lines.append("- `02_dispatch/dispatch_timeseries.xlsx`: per-step dispatch")
    lines.append("- `04_financial_plots/` and `05_energy_plots/`: figures")
    lines.append("- `00_summary/run_log.txt`: full run log")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return out_path


def write_dispatch_artifacts(
    dispatch_dir: Path,
    res_year1: pd.DataFrame,
    lifetime_df: pd.DataFrame | None,
    *,
    project_start_year: int = PROJECT_SHEET_DEFAULTS["project_start_year"],
) -> dict[str, Path]:
    """Write the ``02_dispatch/`` artefacts."""
    dispatch_dir = Path(dispatch_dir)
    dispatch_dir.mkdir(parents=True, exist_ok=True)
    out = dispatch_dir / "dispatch_timeseries.xlsx"

    if lifetime_df is not None and not lifetime_df.empty:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            for cy in sorted(lifetime_df["calendar_year"].unique()):
                sheet = str(int(cy))
                lifetime_df.loc[lifetime_df["calendar_year"] == cy].to_excel(
                    writer, sheet_name=sheet, index=False,
                )
            style_workbook(writer.book)
    else:
        if pd.api.types.is_datetime64_any_dtype(res_year1["timestamp"]):
            cal_year = int(
                pd.to_datetime(res_year1["timestamp"]).dt.year.iloc[0]
            )
        else:
            cal_year = int(project_start_year)
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            res_year1.to_excel(writer, sheet_name=str(cal_year), index=False)
            style_workbook(writer.book)

    return {"hourly_xlsx": out}


def _flatten_kpis_for_sheet(kpis: dict[str, Any]) -> dict[str, Any]:
    """Flatten nested-dict KPI values into prefixed scalar rows.

    The ``kpis_year1`` sheet is a flat ``metric``/``value`` table, so a
    nested dict (e.g. ``bess_utilization_diagnostics``) would otherwise be
    stringified into a single ``{...}`` cell.  Each sub-key is hoisted to
    its own row instead; the in-memory KPI dict keeps the nested form for
    API consumers.
    """
    flat: dict[str, Any] = {}
    for key, value in kpis.items():
        if isinstance(value, dict):
            prefix = (
                "bess_util"
                if key == "bess_utilization_diagnostics"
                else key
            )
            for sub_key, sub_value in value.items():
                if key == "bess_utilization_diagnostics":
                    name = f"bess_util_{sub_key.removeprefix('bess_')}"
                else:
                    name = f"{prefix}_{sub_key}"
                flat[name] = sub_value
        else:
            flat[key] = value
    return flat


def _scalar_metric_rows(kpis: dict[str, Any]) -> list[tuple[str, Any]]:
    """Rows for a ``metric``/``value`` sheet — scalar values only.

    Sequence-valued KPI entries (e.g. ``lifetime_bm_revenue_eur_per_year``,
    a per-year list kept on the dict for API consumers) would otherwise be
    written as a Python ``repr`` string crammed into one Excel cell.  The
    same numbers already live as proper columns in their dedicated sheets
    (``cashflow_yearly['balancing_revenue_eur']``), so the metric/value
    sheets drop them.
    """
    return [
        (key, value)
        for key, value in kpis.items()
        if not isinstance(value, (list, tuple, np.ndarray))
    ]


def write_results_workbook(
    out_path: Path,
    res_year1: pd.DataFrame,
    kpis_year1: dict[str, Any],
    kpis_monthly_year1: pd.DataFrame | None,
    *,
    yearly_cf: pd.DataFrame | None = None,
    monthly_cf: pd.DataFrame | None = None,
    quarterly_cf: pd.DataFrame | None = None,
    financial_kpis: dict[str, Any] | None = None,
    sensitivity: pd.DataFrame | None = None,
    lifetime_yearly: pd.DataFrame | None = None,
    economic_assumptions: dict[str, Any] | None = None,
    rolling_horizon_mc: pd.DataFrame | None = None,
    rolling_horizon_compare_mc: pd.DataFrame | None = None,
    degradation: pd.DataFrame | None = None,
    debt_schedule: pd.DataFrame | None = None,
    emissions: pd.DataFrame | None = None,
) -> Path:
    """Write the consolidated ``03_results.xlsx`` workbook."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame(
            _scalar_metric_rows(_flatten_kpis_for_sheet(kpis_year1)),
            columns=["metric", "value"],
        ).to_excel(writer, sheet_name="kpis_year1", index=False)
        if kpis_monthly_year1 is not None and not kpis_monthly_year1.empty:
            kpis_monthly_year1.reset_index(names="month").to_excel(
                writer, sheet_name="kpis_monthly_year1", index=False,
            )
        res_year1.to_excel(writer, sheet_name="dispatch_year1", index=False)
        if yearly_cf is not None and not yearly_cf.empty:
            yearly_cf.to_excel(writer, sheet_name="cashflow_yearly", index=False)
        if quarterly_cf is not None and not quarterly_cf.empty:
            quarterly_cf.to_excel(
                writer, sheet_name="cashflow_quarterly", index=False,
            )
        if monthly_cf is not None and not monthly_cf.empty:
            monthly_cf.to_excel(writer, sheet_name="cashflow_monthly", index=False)
        if financial_kpis:
            pd.DataFrame(
                _scalar_metric_rows(financial_kpis),
                columns=["metric", "value"],
            ).to_excel(writer, sheet_name="financial_kpis", index=False)
        if sensitivity is not None and not sensitivity.empty:
            sensitivity.to_excel(
                writer, sheet_name="sensitivity_analysis", index=False,
            )
        if lifetime_yearly is not None and not lifetime_yearly.empty:
            lifetime_yearly.to_excel(
                writer, sheet_name="lifetime_dispatch_yearly", index=False,
            )
        if rolling_horizon_mc is not None and not rolling_horizon_mc.empty:
            rolling_horizon_mc.to_excel(
                writer, sheet_name="rolling_horizon_mc", index=False,
            )
        if (
            rolling_horizon_compare_mc is not None
            and not rolling_horizon_compare_mc.empty
        ):
            rolling_horizon_compare_mc.to_excel(
                writer, sheet_name="rolling_horizon_compare_mc", index=False,
            )
        if economic_assumptions:
            _format_assumptions(economic_assumptions).to_excel(
                writer, sheet_name="economic_assumptions", index=False,
            )
        if degradation is not None and not degradation.empty:
            degradation.to_excel(writer, sheet_name="degradation", index=False)
        if debt_schedule is not None and not debt_schedule.empty:
            debt_schedule.to_excel(writer, sheet_name="debt_schedule", index=False)
        if emissions is not None and not emissions.empty:
            emissions.to_excel(writer, sheet_name="emissions", index=False)
        style_workbook(writer.book)
    return out_path
