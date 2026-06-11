"""Shared Excel styling — the house style for every workbook the tool writes.

Single source of truth for the navy frozen-header look applied to the
**input** workbook (via :mod:`scripts.polish_input_workbook`) and to every
**output** workbook (via :mod:`pvbess_opt.io`).  The style constants live in
:mod:`pvbess_opt.theme`; this module is the one place that applies them.

Contract: no worksheet is saved without passing through
:func:`style_workbook` / :func:`style_worksheet`.  Centering and an explicit
font name/size are intentionally *not* applied here, so output workbooks
stay byte-stable; the one centering exception lives in
:mod:`scripts.polish_input_workbook`, which center-aligns the header row of
the per-asset ``max_injection_profile_pv`` / ``_bess`` sheets of the
shipped input workbook.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .theme import (
    COL_WIDTH_MAX,
    COL_WIDTH_MIN,
    COL_WIDTH_PADDING,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    NOTES_WRAP,
)

# Sample at most this many body rows when computing AutoFit widths.  The
# dispatch / time-series sheets carry ~35 040 rows of uniform floats, so a
# full scan is O(rows x cols) for no width gain — the first 200 rows match.
WIDTH_SAMPLE_ROWS: int = 200


def _column_text_width(value: object) -> int:
    """Return the printed character width of ``value`` (its longest line)."""
    if value is None:
        return 0
    text = str(value)
    if not text:
        return 0
    return max(len(line) for line in text.splitlines())


def _notes_column(ws: Worksheet, header_row: int) -> int | None:
    """Return the 1-based index of the ``notes`` header column, or None."""
    for cell in ws[header_row]:
        value = cell.value
        col = cell.column
        if (
            isinstance(value, str)
            and value.strip().lower() == "notes"
            and col is not None
        ):
            return int(col)
    return None


def style_worksheet(
    ws: Worksheet,
    *,
    header_row: int = 1,
    freeze: str = "A2",
    width_sample_rows: int = WIDTH_SAMPLE_ROWS,
) -> None:
    """Apply the house style to ``ws``: navy frozen header + thin bottom
    border + AutoFit widths, wrapping any ``notes`` column.

    Idempotent, and identical to the input-workbook polisher so input and
    output workbooks look the same.  Sampling the body keeps the AutoFit
    pass O(1) in the row count on the large dispatch / time-series sheets.
    """
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col < 1 or max_row < 1:
        return

    # 1) Header row: navy fill, white bold font, thin bottom border.
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER

    # 2) Freeze panes (default: freeze the header row only).
    ws.freeze_panes = freeze

    # 3) Notes column (if present): wrap text.
    notes_col = _notes_column(ws, header_row)
    if notes_col is not None:
        for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row):
            row[notes_col - 1].alignment = NOTES_WRAP

    # 4) AutoFit-style column widths.  openpyxl has no native AutoFit, so
    #    approximate from content length; the header is always measured.
    widths = [
        _column_text_width(ws.cell(row=header_row, column=c).value)
        for c in range(1, max_col + 1)
    ]
    scan_last = min(max_row, header_row + int(width_sample_rows))
    for r in range(header_row + 1, scan_last + 1):
        for c in range(1, max_col + 1):
            w = _column_text_width(ws.cell(row=r, column=c).value)
            if w > widths[c - 1]:
                widths[c - 1] = w
    for idx, raw in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(
            COL_WIDTH_MIN, min(COL_WIDTH_MAX, float(raw) + COL_WIDTH_PADDING)
        )


def style_workbook(
    wb: Workbook,
    *,
    header_row: int = 1,
    freeze: str = "A2",
    width_sample_rows: int = WIDTH_SAMPLE_ROWS,
) -> None:
    """Apply :func:`style_worksheet` to every worksheet in ``wb``."""
    for ws in wb.worksheets:
        style_worksheet(
            ws,
            header_row=header_row,
            freeze=freeze,
            width_sample_rows=width_sample_rows,
        )
