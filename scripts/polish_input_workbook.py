"""Polish / migrate ``inputs/input.xlsx`` to the canonical schema + styling.

The script is idempotent: re-running it produces a byte-identical
workbook (modulo openpyxl's metadata timestamp).  Operations are applied
in order:

1. Drop the deprecated ``pv_kwh_override`` column from the ``timeseries``
   sheet — PV now lives in the single ``pv_kwh`` column.
2. Sweep every sheet for the prior amber bootstrap fill (``FFF2CC``)
   and reset it to *no fill*.
3. Rebuild every parameter sheet (``project``, ``pv``, ``bess``,
   ``economics``, ``simulation``, ``balancing``, ``ppa``, ``intraday``,
   ``market_data``, ``scenario_engine``) from the canonical row
   templates in :mod:`pvbess_opt.io`: existing values are
   preserved by key; rows are rewritten in template order; keys removed
   from the schema are dropped; new schema keys are added with their
   defaults; missing parameter sheets are created.  A migrated workbook
   therefore carries the same rows in the same order as a freshly
   generated one.
4. Apply the shared house style via
   :func:`pvbess_opt.io_style.style_worksheet`: navy ``#1F3864`` frozen
   header (white bold font, thin ``#BFBFBF`` bottom border), AutoFit
   column widths, and wrap-text on the ``notes`` column.  The same styler
   runs on every output workbook, so input and output look identical.
5. Center-align the header row of the per-asset max-injection sheets
   (``max_injection_profile_pv`` / ``max_injection_profile_bess``) —
   their short numeric columns read better with centered headers.  The
   general house style deliberately leaves header alignment at the
   Excel default (see :mod:`pvbess_opt.io_style`).
6. Create the optional ``trajectories`` sheet (per-year stream
   multipliers, Eq. E24) with its shipped disabled example when the
   workbook predates it; an existing sheet is left untouched.
7. Attach schema-generated guardrails to the parameter sheets:
   TRUE/FALSE and enum dropdowns from :data:`pvbess_opt.io._BOOL_KEYS`
   / :data:`pvbess_opt.io._ALLOWED_VALUES`, numeric range validation
   mirroring :func:`pvbess_opt.io.validate_workbook_params` bounds,
   conditional-formatting dimming for feature blocks whose toggle is
   off, and passwordless sheet protection that leaves only the value
   column editable.  ``tests/test_input_workbook_guardrails.py`` locks
   the shipped workbook to the schemas, so a loader schema change
   fails the suite until this script is re-run.
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# Allow running as a standalone script
# (``python scripts/polish_input_workbook.py``) from a checkout where the
# package is not pip-installed: put the repo root on sys.path before the
# first-party imports (mirrors scripts/resample_timeseries.py).
REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from pvbess_opt.io import _SHEET_ROW_TEMPLATES  # noqa: E402
from pvbess_opt.io_style import style_worksheet  # noqa: E402
from pvbess_opt.theme import HEADER_CENTER  # noqa: E402

AMBER_FILL_HEXES: frozenset[str] = frozenset({
    "FFF2CC", "00FFF2CC",
})

_PARAMETER_SHEETS: tuple[str, ...] = (
    "project", "pv", "bess", "economics", "simulation", "balancing", "ppa",
    "intraday", "market_data", "scenario_engine",
)

# Sheets whose header row is center-aligned on top of the house style.
_CENTERED_HEADER_SHEETS: tuple[str, ...] = (
    "max_injection_profile_pv",
    "max_injection_profile_bess",
)

logger = logging.getLogger(__name__)


def _is_amber_fill(fill: PatternFill) -> bool:
    for attr in ("fgColor", "start_color"):
        colour = getattr(fill, attr, None)
        if colour is None:
            continue
        rgb = getattr(colour, "rgb", None)
        if not isinstance(rgb, str):
            continue
        if rgb.upper() in AMBER_FILL_HEXES:
            return True
    return False


def _clear_amber_fills(ws: Worksheet) -> int:
    cleared = 0
    blank = PatternFill(fill_type=None)
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill is None or fill.fill_type is None:
                continue
            if _is_amber_fill(fill):
                cell.fill = blank
                cleared += 1
    return cleared


def _column_index(ws: Worksheet, header_name: str) -> int | None:
    """Return the 1-based column index whose header-row cell matches."""
    for cell in ws[1]:
        if isinstance(cell.value, str) and cell.value.strip().lower() == header_name:
            return int(cell.column)
    return None


def _drop_legacy_pv_override(ws: Worksheet) -> bool:
    """Delete the deprecated ``pv_kwh_override`` column, if present."""
    col = _column_index(ws, "pv_kwh_override")
    if col is None:
        return False
    ws.delete_cols(col, 1)
    return True


def _sync_param_sheet(ws: Worksheet, sheet_name: str) -> int:
    """Rebuild a parameter sheet's rows from the canonical template.

    Existing values are preserved by key; everything else is canonical:

    * rows are rewritten in TEMPLATE ORDER, so a migrated workbook and a
      freshly generated one (:func:`pvbess_opt.io.write_workbook`) carry
      the same rows in the same order;
    * keys the template adds are written with their default value / unit
      / notes;
    * rows whose key has been removed from the schema are DROPPED (the
      loader already warns-and-ignores them; carrying them in the shipped
      workbook would advertise dead knobs);
    * the ``unit`` / ``notes`` columns are rewritten so wording changes
      in the typed dict actually reach the workbook.

    Returns the number of rows written.  Unknown sheets are a no-op.
    """
    template = _SHEET_ROW_TEMPLATES.get(sheet_name)
    if template is None:
        return 0
    key_col = _column_index(ws, "key")
    value_col = _column_index(ws, "value")
    unit_col = _column_index(ws, "unit")
    notes_col = _column_index(ws, "notes")
    if key_col is None or value_col is None or unit_col is None or notes_col is None:
        return 0

    template_keys = {key for key, _default, _unit, _notes in template}
    existing: dict[str, object] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        key_cell = row[key_col - 1]
        if not isinstance(key_cell.value, str) or not key_cell.value.strip():
            continue
        key = key_cell.value.strip()
        if key not in template_keys:
            logger.info(
                "%s: dropping removed schema key %r (row %d).",
                sheet_name, key, key_cell.row,
            )
            continue
        existing[key] = row[value_col - 1].value

    last_row = max(ws.max_row, len(template) + 1)
    for r in range(2, last_row + 1):
        for c in (key_col, value_col, unit_col, notes_col):
            ws.cell(row=r, column=c).value = None

    for idx, (key, default, unit, notes) in enumerate(template):
        if key not in existing:
            logger.info(
                "%s: appending new schema key %r with its default %r.",
                sheet_name, key, default,
            )
        r = idx + 2
        ws.cell(row=r, column=key_col).value = key
        ws.cell(row=r, column=value_col).value = existing.get(key, default)
        ws.cell(row=r, column=unit_col).value = unit
        ws.cell(row=r, column=notes_col).value = notes
    return len(template)


def _migrate_legacy_bess_capex(ws: Worksheet) -> bool:
    """Convert the legacy per-kW BESS CAPEX row to the per-kWh basis.

    v1.0.0 prices BESS CAPEX per kWh of energy capacity
    (``capex_bess_eur_per_kwh``); older workbooks carry
    ``capex_bess_eur_per_kw`` on the power basis.  The conversion is
    ``value_per_kwh = value_per_kw x bess_power_kw / bess_capacity_kwh``
    and requires ``bess_capacity_kwh > 0`` on the same sheet.  Returns
    True when a migration was performed.
    """
    key_col = _column_index(ws, "key")
    value_col = _column_index(ws, "value")
    if key_col is None or value_col is None:
        return False
    cells: dict[str, object] = {}
    legacy_row: int | None = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        key_cell = row[key_col - 1]
        if not isinstance(key_cell.value, str) or not key_cell.value.strip():
            continue
        key = key_cell.value.strip()
        cells[key] = row[value_col - 1].value
        if key == "capex_bess_eur_per_kw":
            legacy_row = key_cell.row
    if legacy_row is None or "capex_bess_eur_per_kwh" in cells:
        return False
    try:
        value_per_kw = float(cells.get("capex_bess_eur_per_kw") or 0.0)
        power_kw = float(cells.get("bess_power_kw") or 0.0)
        capacity_kwh = float(cells.get("bess_capacity_kwh") or 0.0)
    except (TypeError, ValueError):
        logger.warning(
            "bess: cannot migrate legacy 'capex_bess_eur_per_kw' "
            "(non-numeric inputs); dropping the row. Set "
            "'capex_bess_eur_per_kwh' manually."
        )
        return False
    if capacity_kwh <= 0.0:
        logger.warning(
            "bess: cannot migrate legacy 'capex_bess_eur_per_kw' "
            "(bess_capacity_kwh is 0); dropping the row. Set "
            "'capex_bess_eur_per_kwh' manually."
        )
        return False
    value_per_kwh = value_per_kw * power_kw / capacity_kwh
    ws.cell(row=legacy_row, column=key_col).value = "capex_bess_eur_per_kwh"
    ws.cell(row=legacy_row, column=value_col).value = value_per_kwh
    logger.info(
        "bess: migrated legacy capex_bess_eur_per_kw=%s to "
        "capex_bess_eur_per_kwh=%s (x %s kW / %s kWh).",
        value_per_kw, value_per_kwh, power_kw, capacity_kwh,
    )
    return True


def _center_header_row(ws: Worksheet) -> None:
    """Center-align every populated header cell of ``ws`` (row 1)."""
    for cell in ws[1]:
        if cell.value is not None:
            cell.alignment = HEADER_CENTER


def _ensure_parameter_sheets(wb) -> None:
    """Create any canonical parameter sheet the workbook does not carry.

    The schema-migration counterpart of the drop/append logic in
    :func:`_sync_param_sheet`: a NEW sheet (e.g. ``ppa``) is
    created with the ``key | value | unit | notes`` header and its
    template rows, placed after the last existing parameter sheet so
    the workbook keeps its canonical ordering.
    """
    for sheet_name in _PARAMETER_SHEETS:
        if sheet_name in wb.sheetnames:
            continue
        template = _SHEET_ROW_TEMPLATES.get(sheet_name)
        if template is None:
            continue
        anchor = max(
            (
                wb.sheetnames.index(existing)
                for existing in _PARAMETER_SHEETS
                if existing in wb.sheetnames
            ),
            default=len(wb.sheetnames) - 1,
        )
        logger.info(
            "creating missing parameter sheet %r with %d template rows.",
            sheet_name, len(template),
        )
        ws = wb.create_sheet(sheet_name, index=anchor + 1)
        ws.append(["key", "value", "unit", "notes"])
        for key, default, unit, notes in template:
            ws.append([key, default, unit, notes])


def _ensure_trajectories_sheet(wb) -> bool:
    """Create the optional ``trajectories`` sheet when absent.

    Written with the shipped disabled example
    (:data:`pvbess_opt.io._TRAJECTORIES_EXAMPLE_ROWS`) so the tidy
    per-year multiplier format is self-documenting; an existing sheet is
    left untouched (idempotent — user values are never rewritten).
    Returns True when the sheet was created.
    """
    if "trajectories" in wb.sheetnames:
        return False
    from pvbess_opt.io import (
        _TRAJECTORIES_EXAMPLE_ROWS,
        TRAJECTORIES_SHEET_COLUMNS,
    )

    logger.info(
        "creating missing 'trajectories' sheet with the disabled example.",
    )
    ws = wb.create_sheet("trajectories")
    ws.append(list(TRAJECTORIES_SHEET_COLUMNS))
    for row in _TRAJECTORIES_EXAMPLE_ROWS:
        ws.append(list(row))
    return True


def _ensure_price_scenarios_sheet(wb) -> bool:
    """Create the optional ``price_scenarios`` sheet when absent.

    Written with the shipped disabled example
    (:data:`pvbess_opt.io._PRICE_SCENARIOS_EXAMPLE_ROWS`); an existing
    sheet is left untouched (idempotent — user rows are never
    rewritten).  Returns True when the sheet was created.
    """
    if "price_scenarios" in wb.sheetnames:
        return False
    from pvbess_opt.io import (
        _PRICE_SCENARIOS_EXAMPLE_ROWS,
        PRICE_SCENARIOS_SHEET_COLUMNS,
    )

    logger.info(
        "creating missing 'price_scenarios' sheet with the disabled "
        "example.",
    )
    ws = wb.create_sheet("price_scenarios")
    ws.append(list(PRICE_SCENARIOS_SHEET_COLUMNS))
    for row in _PRICE_SCENARIOS_EXAMPLE_ROWS:
        ws.append(list(row))
    return True


# ---------------------------------------------------------------------------
# Input guardrails — dropdowns, numeric bounds, protection, dependency tint
# ---------------------------------------------------------------------------

#: Grey-out styling for rows whose feature block is switched off.
_DIMMED_FONT_COLOR = "FF9CA3AF"

#: Hard numeric bounds mirrored from ``validate_workbook_params`` — the
#: validator stays the source of truth (the guardrails test probes the
#: loader just outside every bound listed here, so this table cannot
#: silently drift).  Closed [lo, hi] bounds only; entries render as a
#: stop-style "decimal between" validation.
_NUMERIC_BOUNDS: dict[str, tuple[float, float]] = {
    # io.validate_workbook_params: "must be in [0, 100]"
    "gearing_pct": (0.0, 100.0),
    "bess_overbuild_pct": (0.0, 100.0),
    "corporate_tax_rate_pct": (0.0, 100.0),
    "curtailment_pct": (0.0, 100.0),
    "curtailment_compensated_pct": (0.0, 100.0),
    "unavailability_pct": (0.0, 100.0),
    "ppa_volume_share_pct": (0.0, 100.0),
    "grid_co2_annual_decline_pct": (0.0, 100.0),
    "aggregator_fee_pct_revenue": (0.0, 100.0),
    "balancing_aggregator_fee_pct_revenue": (0.0, 100.0),
    "optimizer_revenue_share_pct": (0.0, 100.0),
    "state_support_clawback_share_pct": (0.0, 100.0),
    "revenue_levy_pct": (0.0, 100.0),
    "capacity_market_derating_pct": (0.0, 100.0),
    # io.validate_workbook_params: "must be in [0, 30]"
    "bess_cost_decline_pct_per_year": (0.0, 30.0),
    # io.validate_workbook_params: "must be in [0, 1]"
    "soc_min_frac": (0.0, 1.0),
    "soc_max_frac": (0.0, 1.0),
}

#: Whole-number lower bounds ("must be >= 1" in the validator).
_WHOLE_MINIMUMS: dict[str, int] = {
    "project_lifecycle_years": 1,
    "uncertainty_n_seeds": 1,
}

#: Half-open (lo, hi] loader bounds — an OPEN lower bound Excel's
#: single "between" rule cannot express, so these render as a
#: WARNING-style validation (Excel lets the user proceed) and are
#: exempt from the probe-lock test; the loader remains the hard gate.
#: The rendered lower limit is lo + 1e-9, so entering lo itself (a
#: value the loader rejects) fires the warning instead of passing
#: silently, while every loader-legal value above it stays enterable
#: (warning style never blocks).
_WARNING_BOUNDS: dict[str, tuple[float, float]] = {
    # io.validate_workbook_params: "must be in (0, 1]"
    "efficiency_charge": (0.0, 1.0),
    "efficiency_discharge": (0.0, 1.0),
    # io.validate_workbook_params: "must be in (0, 100]"
    "production_p90_factor_pct": (0.0, 100.0),
    "bess_eol_soh_pct": (0.0, 100.0),
}

#: Feature blocks greyed out while their toggle is off.  Same-sheet
#: rules only (cross-sheet CF formulas are fragile across spreadsheet
#: applications).  Gates verified against the loader/pipeline:
#: balancing activation keys on balancing_enabled; id_* is the
#: intraday venue; imbalance settlement and the risk metrics ride the
#: rolling-horizon MC (loader couples them to uncertainty_enabled);
#: every scenario_engine key is engine-only; ppa_* rides ppa_enabled
#: and support_* rides support_scheme (mutually exclusive families).
#: selector: prefixes of DEPENDENT keys, or None = every other key on
#: the sheet.
_DIM_GROUPS: tuple[tuple[str, str, str, tuple[str, ...] | None], ...] = (
    # (sheet, toggle key, off-condition template, dependent prefixes)
    ("balancing", "balancing_enabled", "={toggle}=FALSE", None),
    ("intraday", "id_enabled", "={toggle}=FALSE", None),
    ("simulation", "uncertainty_enabled", "={toggle}=FALSE",
     ("uncertainty_", "imbalance_", "risk_")),
    ("scenario_engine", "price_scenarios_enabled", "={toggle}=FALSE", None),
    ("ppa", "ppa_enabled", "={toggle}=FALSE", ("ppa_",)),
    # A blanked scheme cell must dim too: the loader defaults blank to
    # "none" (the block is inert), and Excel evaluates ""="none" FALSE.
    ("ppa", "support_scheme", '=OR({toggle}="none",{toggle}="")',
     ("support_",)),
)


def _kv_rows(ws: Worksheet) -> dict[str, int]:
    """Map ``key`` (column A) to its 1-based row on a key/value sheet."""
    rows: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value is not None and str(cell.value).strip():
            rows[str(cell.value).strip()] = cell.row
    return rows


def _add_validation(
    ws: Worksheet, coord: str, dv_kwargs: dict, *, key: str,
    accepted: str, style: str = "stop",
) -> None:
    dv = DataValidation(allow_blank=True, **dv_kwargs)
    dv.errorStyle = "stop" if style == "stop" else "warning"
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid value"
    dv.error = f"'{key}' accepts: {accepted}."
    dv.showInputMessage = True
    # OOXML caps promptTitle at 32 characters (Excel's own UI enforces
    # it); the longest keys exceed that, so clamp — the full key stays
    # in column A and in the error/prompt bodies.
    dv.promptTitle = key if len(key) <= 32 else key[:29] + "..."
    dv.prompt = f"Accepted: {accepted}."
    ws.add_data_validation(dv)
    dv.add(coord)


def _apply_input_guardrails(wb: Workbook) -> dict[str, int]:
    """Attach schema-generated guardrails to every parameter sheet.

    Everything here is generated from the loader's own schemas
    (``_BOOL_KEYS`` / ``_ALLOWED_VALUES``) or mirrored from
    ``validate_workbook_params`` bounds, so the workbook UI and the
    parser cannot disagree.  The guardrails are a convenience layer
    ONLY — the loaders keep rejecting bad values on every surface
    (YAML/JSON carry no dropdowns, and spreadsheet validation is
    bypassable by paste), so nothing downstream may rely on them.
    """
    from pvbess_opt.io import _ALLOWED_VALUES, _BOOL_KEYS, _STR_KEYS

    counts = {"dropdowns": 0, "bounds": 0, "dimmed_rows": 0, "sheets": 0}
    for sheet_name in _PARAMETER_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = _kv_rows(ws)
        # Idempotence: the polisher fully owns validations and
        # conditional formatting on parameter sheets — rebuild both.
        ws.data_validations.dataValidation = []
        ws.conditional_formatting = ConditionalFormattingList()

        for key, row in rows.items():
            coord = f"B{row}"
            if key in _BOOL_KEYS:
                _add_validation(
                    ws, coord,
                    {"type": "list", "formula1": '"TRUE,FALSE"'},
                    key=key, accepted="TRUE or FALSE",
                )
                counts["dropdowns"] += 1
            elif key in _STR_KEYS and _ALLOWED_VALUES.get(key):
                values = sorted(_ALLOWED_VALUES[key])
                joined = ",".join(values)
                # Excel's inline list source caps at 255 characters; the
                # longest current list (bidding_zone) sits well under it.
                if len(joined) > 255:  # pragma: no cover - schema guard
                    raise ValueError(
                        f"enum list for {key!r} exceeds Excel's inline "
                        "255-char validation limit; move it to a helper "
                        "sheet."
                    )
                _add_validation(
                    ws, coord,
                    {"type": "list", "formula1": f'"{joined}"'},
                    key=key, accepted=", ".join(values),
                )
                counts["dropdowns"] += 1
            elif key in _NUMERIC_BOUNDS:
                lo, hi = _NUMERIC_BOUNDS[key]
                _add_validation(
                    ws, coord,
                    {"type": "decimal", "operator": "between",
                     "formula1": str(lo), "formula2": str(hi)},
                    key=key, accepted=f"a number in [{lo:g}, {hi:g}]",
                )
                counts["bounds"] += 1
            elif key in _WHOLE_MINIMUMS:
                lo_int = _WHOLE_MINIMUMS[key]
                _add_validation(
                    ws, coord,
                    {"type": "whole", "operator": "greaterThanOrEqual",
                     "formula1": str(lo_int)},
                    key=key, accepted=f"a whole number >= {lo_int}",
                )
                counts["bounds"] += 1
            elif key in _WARNING_BOUNDS:
                lo, hi = _WARNING_BOUNDS[key]
                _add_validation(
                    ws, coord,
                    {"type": "decimal", "operator": "between",
                     # lo itself is loader-illegal on these half-open
                     # bounds: nudge the rendered limit so entering lo
                     # fires the (non-blocking) warning.
                     "formula1": f"{lo + 1e-9:.9f}", "formula2": str(hi)},
                    key=key,
                    accepted=f"a number in ({lo:g}, {hi:g}] — {lo:g} "
                             "itself is rejected by the loader",
                    style="warning",
                )
                counts["bounds"] += 1

        # Dependency greying: dim the whole row of a dependent key while
        # its feature toggle is off.  Visual only — values stay editable
        # and the loaders keep validating them.
        dim_font = Font(color=_DIMMED_FONT_COLOR)
        for group_sheet, toggle, template, prefixes in _DIM_GROUPS:
            if group_sheet != sheet_name or toggle not in rows:
                continue
            formula = template.format(toggle=f"$B${rows[toggle]}")
            for key, row in rows.items():
                if key == toggle:
                    continue
                if prefixes is not None and not key.startswith(prefixes):
                    continue
                rule = FormulaRule(  # type: ignore[no-untyped-call]
                    formula=[formula[1:]], font=dim_font,
                )
                ws.conditional_formatting.add(f"A{row}:D{row}", rule)
                counts["dimmed_rows"] += 1

        # Protection: the key/unit/notes columns and the header are the
        # schema — lock them so a stray edit cannot silently rename a
        # key (the kv reader requires exact keys; a damaged key column
        # historically made every value fall back to its default).
        # Only the value column stays editable.  No password: the
        # protection is a guardrail, not a lock — Review > Unprotect
        # lifts it deliberately.
        for row_idx, row_cells in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4),
            start=1,
        ):
            for cell in row_cells:
                cell.protection = Protection(
                    locked=not (cell.column == 2 and row_idx >= 2),
                )
        ws.protection.sheet = True
        counts["sheets"] += 1
    return counts


def polish_workbook(path: Path) -> dict[str, int]:
    """Polish ``path`` in place and return per-sheet diagnostics.

    Returned dict maps sheet name to the number of amber-fill cells
    cleared on that sheet (kept for backward compatibility with the
    earlier polish script's logging).
    """
    wb = load_workbook(path)
    if "timeseries" in wb.sheetnames:
        _drop_legacy_pv_override(wb["timeseries"])
    if "bess" in wb.sheetnames:
        _migrate_legacy_bess_capex(wb["bess"])
    _ensure_parameter_sheets(wb)
    _ensure_trajectories_sheet(wb)
    _ensure_price_scenarios_sheet(wb)
    cleared_by_sheet: dict[str, int] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cleared_by_sheet[sheet_name] = _clear_amber_fills(ws)
        if sheet_name in _PARAMETER_SHEETS:
            _sync_param_sheet(ws, sheet_name)
        style_worksheet(ws)
        if sheet_name in _CENTERED_HEADER_SHEETS:
            _center_header_row(ws)
    guardrails = _apply_input_guardrails(wb)
    logging.getLogger(__name__).info(
        "[guardrails] %d dropdowns, %d numeric bounds, %d dimmed rows "
        "across %d protected parameter sheets.",
        guardrails["dropdowns"], guardrails["bounds"],
        guardrails["dimmed_rows"], guardrails["sheets"],
    )
    wb.save(path)
    return cleared_by_sheet


def _main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "path", nargs="?", default="inputs/input.xlsx", type=Path,
        help="Workbook to polish (default: inputs/input.xlsx).",
    )
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    cleared = polish_workbook(args.path)
    for sheet, n in cleared.items():
        logger.info(
            "%s: polished (cleared %d amber-highlighted cells, "
            "AutoFit applied, header styled, frozen at A2).",
            sheet, n,
        )


if __name__ == "__main__":
    _main()
