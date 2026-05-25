"""Update ``inputs/input.xlsx`` with the balancing-market sheet.

One-off utility for the v0.9.0 release. Re-running it is idempotent:
the script overwrites the ``balancing`` sheet and the nine optional
timeseries columns with the canonical defaults and a deterministic
synthetic-price series seeded with ``1729``.

The script intentionally writes ``balancing_enabled = FALSE`` so the
shipped workbook does not change the headline dispatch behaviour —
users opt in by flipping the flag in the workbook.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

from pvbess_opt.balancing import (
    BalancingConfig,
    generate_synthetic_balancing_timeseries,
)
from pvbess_opt.io import (
    _BALANCING_ROWS,
    BALANCING_SHEET_DEFAULTS,
    detect_timestep_minutes,
)

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "inputs" / "input.xlsx"

AMBER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

logger = logging.getLogger(__name__)


def _write_balancing_sheet(wb: openpyxl.Workbook) -> None:
    if "balancing" in wb.sheetnames:
        del wb["balancing"]
    target_index = wb.sheetnames.index("economics") + 1 if (
        "economics" in wb.sheetnames
    ) else len(wb.sheetnames)
    ws = wb.create_sheet("balancing", index=target_index)
    ws.cell(1, 1, "key").font = Font(bold=True)
    ws.cell(1, 2, "value").font = Font(bold=True)
    ws.cell(1, 3, "unit").font = Font(bold=True)
    ws.cell(1, 4, "notes").font = Font(bold=True)
    for row_idx, (key, default, unit, notes) in enumerate(_BALANCING_ROWS, start=2):
        ws.cell(row_idx, 1, key).fill = AMBER_FILL
        cell_value = default
        if isinstance(default, bool):
            cell_value = bool(default)
        ws.cell(row_idx, 2, cell_value).fill = AMBER_FILL
        ws.cell(row_idx, 3, unit)
        ws.cell(row_idx, 4, notes)


def _append_balancing_timeseries(wb: openpyxl.Workbook) -> None:
    ts_sheet = wb["timeseries"]
    header_row = [cell.value for cell in ts_sheet[1]]
    n_rows = ts_sheet.max_row - 1

    # Detect cadence from the timestamp column.
    ts_df = pd.read_excel(WORKBOOK_PATH, sheet_name="timeseries", parse_dates=["timestamp"])
    dt_minutes = detect_timestep_minutes(ts_df)
    dt_hours = dt_minutes / 60.0

    cfg = BalancingConfig(**BALANCING_SHEET_DEFAULTS)
    prices = generate_synthetic_balancing_timeseries(
        n_rows, dt_hours, cfg, seed=1729,
    )

    next_col = len(header_row) + 1
    for offset, column in enumerate(prices.columns):
        col_idx = next_col + offset
        header_cell = ts_sheet.cell(1, col_idx, column)
        header_cell.fill = AMBER_FILL
        header_cell.font = Font(bold=True)
        values = prices[column].to_numpy(dtype=float)
        for row_offset, value in enumerate(values, start=2):
            ts_sheet.cell(row_offset, col_idx, float(value))


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"workbook not found: {WORKBOOK_PATH}")
    logger.info("Opening %s", WORKBOOK_PATH)
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    _write_balancing_sheet(wb)
    _append_balancing_timeseries(wb)
    wb.save(WORKBOOK_PATH)
    logger.info(
        "Updated %s — added the balancing sheet (highlighted amber) "
        "and the nine synthetic balancing-price timeseries columns.",
        WORKBOOK_PATH,
    )


if __name__ == "__main__":
    main()
